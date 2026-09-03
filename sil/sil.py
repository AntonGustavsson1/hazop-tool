"""
SIL PFD-kalkylator — Desktop GUI
ProSa Process Safety Consulting AB  |  IEC 61508 / IEC 61511
"""

import copy, json, webbrowser, datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from pathlib import Path

from calc import (
    Architecture, ComponentParams, SubsystemParams,
    calc_sif, calc_subsystem, SIFResult, SIL_LIMITS,
    sil_from_pfd, validate_component, pfd_simplified, pfd_all_architectures,
)
import components_db as cdb
try:
    import verification_db as vdb
    _VDB_AVAILABLE = True
except ImportError:
    _VDB_AVAILABLE = False

# ── Konstanter ─────────────────────────────────────────────────────────────────
ARCH_OPTIONS  = [a.value for a in Architecture]
SIL_OPTIONS   = [1, 2, 3, 4]
COMP_TYPES    = ["A (enkel)", "B (komplex)"]
CCF_MODELS    = ["beta", "mooNbeta"]
CALC_METHODS  = ["markov", "simplified"]
BG       = "#f5f5f5"
BG_FRAME = "#ffffff"
ACCENT   = "#1a5276"
GREEN    = "#1e8449"
RED      = "#c0392b"
ORANGE   = "#d35400"
FONT_H   = ("Segoe UI", 11, "bold")
FONT_N   = ("Segoe UI", 10)
FONT_S   = ("Segoe UI", 9)
FONT_M   = ("Consolas", 9)
SIL_COL  = {4: "#1a5276", 3: GREEN, 2: "#2e86c1", 1: ORANGE, 0: RED}

DEFAULT_SUBS = {
    "sensor": {"name":"Sensor",       "arch":"1oo1","lambda_d":"1e-6","dc":"0.0",
               "beta":"0.02","ti":"8760","mttr":"8","ptc":"1.0","sff":"0.0","comp_type":"A",
               "sc":"0","st":"0","mission_time":"175200","pst_coverage":"0.0",
               "pst_interval":"720","ccf_model":"beta","calc_method":"markov","channels":[]},
    "logic":  {"name":"Logic solver", "arch":"1oo1","lambda_d":"1e-7","dc":"0.99",
               "beta":"0.02","ti":"8760","mttr":"8","ptc":"1.0","sff":"0.99","comp_type":"B",
               "sc":"0","st":"0","mission_time":"175200","pst_coverage":"0.0",
               "pst_interval":"720","ccf_model":"beta","calc_method":"markov","channels":[]},
    "fe":     {"name":"Slutelement",  "arch":"1oo1","lambda_d":"1e-5","dc":"0.0",
               "beta":"0.02","ti":"8760","mttr":"8","ptc":"1.0","sff":"0.1","comp_type":"A",
               "sc":"0","st":"0","mission_time":"175200","pst_coverage":"0.0",
               "pst_interval":"720","ccf_model":"beta","calc_method":"markov","channels":[]},
}
DEFAULT_DATASHEET = {
    "tag":"","hazard":"","consequence":"","process_safety_time":"",
    "response_time":"","sil_basis":"LOPA / Riskgraf / Annat",
    "sif_description":"","safe_state":"","reset":"Manuell",
    "bypass":"","pt_procedure":"","notes":"",
}


# ── Tooltip ────────────────────────────────────────────────────────────────────
class _Tip:
    def __init__(self, w, t):
        self._w, self._t, self._pop = w, t, None
        w.bind("<Enter>", self._show); w.bind("<Leave>", self._hide)
    def _show(self, _):
        x = self._w.winfo_rootx() + 20
        y = self._w.winfo_rooty() + self._w.winfo_height() + 4
        self._pop = tk.Toplevel(self._w); self._pop.wm_overrideredirect(True)
        self._pop.wm_geometry(f"+{x}+{y}")
        tk.Label(self._pop, text=self._t, bg="#ffffcc", relief="solid",
                 borderwidth=1, font=FONT_S, padx=4, pady=2).pack()
    def _hide(self, _):
        if self._pop: self._pop.destroy(); self._pop = None


# ── Komponent-väljare ──────────────────────────────────────────────────────────
class ComponentPickerDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Välj komponent från databas")
        self.geometry("860x500"); self.resizable(True, True)
        self.transient(parent); self.grab_set()
        self.result = None
        self._build(); self._refresh(); self.wait_window()

    def _build(self):
        top = tk.Frame(self, bg=BG, padx=8, pady=6); top.pack(fill="x")
        tk.Label(top, text="Sök:", font=FONT_N, bg=BG).pack(side="left")
        self._vs = tk.StringVar(); self._vs.trace_add("write", lambda *_: self._refresh())
        ttk.Entry(top, textvariable=self._vs, width=22, font=FONT_N).pack(side="left", padx=(4,14))
        tk.Label(top, text="Kategori:", font=FONT_N, bg=BG).pack(side="left")
        self._vc = tk.StringVar(value="Alla"); self._vc.trace_add("write", lambda *_: self._refresh())
        ttk.Combobox(top, textvariable=self._vc, values=cdb.get_categories(),
                     width=24, state="readonly", font=FONT_N).pack(side="left", padx=(4,0))
        frm = tk.Frame(self); frm.pack(fill="both", expand=True, padx=8)
        cols = ("cat","mfr","model","lam","dc","beta","src")
        hdrs = ("Kategori","Tillverkare","Modell","λ_D [1/h]","DC","β","Källa")
        ws   = (140,120,160,90,55,55,65)
        self._tv = ttk.Treeview(frm, columns=cols, show="headings", selectmode="browse")
        for c,h,w in zip(cols,hdrs,ws):
            self._tv.heading(c,text=h); self._tv.column(c,width=w,minwidth=30)
        self._tv.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(frm, orient="vertical", command=self._tv.yview)
        sb.pack(side="right", fill="y"); self._tv.configure(yscrollcommand=sb.set)
        self._tv.bind("<Double-Button-1>", lambda _: self._ok())
        self._tv.bind("<<TreeviewSelect>>", self._on_sel)
        self._dv = tk.StringVar()
        tk.Label(self, textvariable=self._dv, font=FONT_S, fg="#555", bg=BG,
                 anchor="w", wraplength=820).pack(fill="x", padx=8, pady=(2,0))
        bf = tk.Frame(self, bg=BG, padx=8, pady=6); bf.pack(fill="x")
        ttk.Button(bf, text="Avbryt", command=self.destroy).pack(side="right", padx=(4,0))
        ttk.Button(bf, text="Välj",   command=self._ok).pack(side="right")

    def _refresh(self):
        self._tv.delete(*self._tv.get_children())
        rows = cdb.search(self._vc.get(), self._vs.get())
        self._rows = {}
        for r in rows:
            self._tv.insert("","end", iid=str(r["id"]), values=(
                r["category"],r["manufacturer"],r["model"],
                f"{r['lambda_d']:.2e}",f"{r['dc']:.2f}",f"{r['beta']:.2f}",r["source"]))
            self._rows[str(r["id"])] = dict(r)

    def _on_sel(self, _):
        sel = self._tv.selection()
        if sel and sel[0] in self._rows:
            self._dv.set(self._rows[sel[0]].get("description",""))

    def _ok(self):
        sel = self._tv.selection()
        if not sel: messagebox.showinfo("Välj","Markera en rad.", parent=self); return
        self.result = self._rows.get(sel[0]); self.destroy()


# ── Komponent-hanterare ────────────────────────────────────────────────────────
class ComponentManagerDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Hantera komponentdatabas")
        self.geometry("920x540"); self.transient(parent); self.grab_set()
        self._build(); self._refresh(); self.wait_window()

    def _build(self):
        frm = tk.Frame(self); frm.pack(fill="both", expand=True, padx=8, pady=6)
        cols = ("cat","mfr","model","lam","dc","beta","custom")
        hdrs = ("Kategori","Tillverkare","Modell","λ_D [1/h]","DC","β","Eget")
        ws   = (140,120,160,90,55,55,45)
        self._tv = ttk.Treeview(frm, columns=cols, show="headings", selectmode="browse")
        for c,h,w in zip(cols,hdrs,ws):
            self._tv.heading(c,text=h); self._tv.column(c,width=w,minwidth=30)
        self._tv.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(frm, orient="vertical", command=self._tv.yview)
        sb.pack(side="right", fill="y"); self._tv.configure(yscrollcommand=sb.set)
        bf = tk.Frame(self, bg=BG, padx=8, pady=6); bf.pack(fill="x")
        ttk.Button(bf, text="+ Lägg till",    command=self._add).pack(side="left", padx=(0,4))
        ttk.Button(bf, text="Ta bort (egna)", command=self._delete).pack(side="left")
        ttk.Button(bf, text="Stäng",          command=self.destroy).pack(side="right")

    def _refresh(self):
        self._tv.delete(*self._tv.get_children())
        self._rows = {}
        for r in cdb.search():
            iid = str(r["id"])
            self._tv.insert("","end", iid=iid, values=(
                r["category"],r["manufacturer"],r["model"],
                f"{r['lambda_d']:.2e}",f"{r['dc']:.2f}",f"{r['beta']:.2f}",
                "Ja" if r["custom"] else ""))
            self._rows[iid] = dict(r)

    def _add(self):
        AddComponentDialog(self); self._refresh()

    def _delete(self):
        sel = self._tv.selection()
        if not sel: return
        row = self._rows.get(sel[0])
        if not row or not row.get("custom"):
            messagebox.showinfo("Info","Bara egna komponenter kan tas bort.", parent=self); return
        if messagebox.askyesno("Ta bort", f"Ta bort '{row['model']}'?", parent=self):
            conn = cdb.get_connection()
            conn.execute("DELETE FROM components WHERE id=?", (row["id"],))
            conn.commit(); conn.close()
            self._refresh()


class AddComponentDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Lägg till komponent"); self.geometry("420x340")
        self.transient(parent); self.grab_set(); self._build(); self.wait_window()

    def _build(self):
        f = tk.Frame(self, padx=16, pady=12); f.pack(fill="both", expand=True)
        fields = [("Kategori","cat"),("Tillverkare","mfr"),("Modell","model"),
                  ("λ_D [1/h]","lam"),("DC [0-1]","dc"),("β [0-1]","beta"),("Beskrivning","desc")]
        self._vs = {}
        defaults = {"lam":"1e-6","dc":"0.0","beta":"0.02"}
        for i,(lbl,key) in enumerate(fields):
            tk.Label(f, text=lbl+":", font=FONT_N, anchor="w").grid(row=i,column=0,sticky="w",pady=2)
            v = tk.StringVar(value=defaults.get(key,""))
            self._vs[key] = v
            ttk.Entry(f, textvariable=v, width=28, font=FONT_N).grid(row=i,column=1,sticky="ew",pady=2)
        f.columnconfigure(1, weight=1)
        bf = tk.Frame(self, padx=8, pady=6); bf.pack(fill="x")
        ttk.Button(bf, text="Avbryt", command=self.destroy).pack(side="right", padx=(4,0))
        ttk.Button(bf, text="Spara",  command=self._save).pack(side="right")

    def _save(self):
        try:
            cdb.add_custom(
                self._vs["cat"].get(), self._vs["mfr"].get(), self._vs["model"].get(),
                float(self._vs["lam"].get()), float(self._vs["dc"].get()),
                float(self._vs["beta"].get()), self._vs["desc"].get())
            self.destroy()
        except Exception as e:
            messagebox.showerror("Fel", str(e), parent=self)


# ── Heterogena kanaler ─────────────────────────────────────────────────────────
class MixedChannelsDialog(tk.Toplevel):
    """Dialog för per-kanal parametrar i 1oo2/2oo3."""

    def __init__(self, parent, arch: str, existing: list[dict]):
        super().__init__(parent)
        n = 2 if arch in ("1oo2","1oo2D","2oo2") else 3
        self.title(f"Heterogena kanaler — {arch} ({n} kanaler)")
        self.geometry(f"600x{180 + n*90}"); self.resizable(False, False)
        self.transient(parent); self.grab_set()
        self.result: list[dict] | None = None
        self._n = n; self._rows: list[dict[str,tk.StringVar]] = []
        self._build(n, existing); self.wait_window()

    def _build(self, n: int, existing: list[dict]):
        tk.Label(self, text="Ange parametrar per kanal. Gemensamma parametrar (TI, MTTR, PTC m.fl.)\nhämtas från delsystemets huvudfält.",
                 font=FONT_S, fg="#555", justify="left").pack(padx=12, pady=(10,4), anchor="w")

        hdrs_frame = tk.Frame(self); hdrs_frame.pack(fill="x", padx=12)
        for ci, hdr in enumerate(["","Namn","λ_D [1/h]","DC [0–1]","β [0–1]"]):
            tk.Label(hdrs_frame, text=hdr, font=("Segoe UI",9,"bold"), width=14 if ci>0 else 6,
                     anchor="w").grid(row=0, column=ci, padx=2)

        for i in range(n):
            row_frame = tk.Frame(self); row_frame.pack(fill="x", padx=12, pady=2)
            tk.Label(row_frame, text=f"Kanal {i+1}:", font=FONT_S, width=7,
                     anchor="w").grid(row=0, column=0, padx=2)
            ex = existing[i] if i < len(existing) else {}
            rv = {}
            for ci,(key,default) in enumerate([("name",f"Kanal {i+1}"),
                                                ("lambda_d","1e-6"),("dc","0.0"),("beta","0.02")]):
                v = tk.StringVar(value=ex.get(key,default))
                rv[key] = v
                ttk.Entry(row_frame, textvariable=v, width=14, font=FONT_M).grid(
                    row=0, column=ci+1, padx=2, sticky="ew")
            self._rows.append(rv)

        bf = tk.Frame(self, padx=12, pady=10); bf.pack(fill="x")
        ttk.Button(bf, text="Avbryt",          command=self.destroy).pack(side="right", padx=(4,0))
        ttk.Button(bf, text="Använd heterogena kanaler", command=self._ok).pack(side="right")
        ttk.Button(bf, text="Rensa (använd identiska)", command=self._clear).pack(side="left")

    def _ok(self):
        result = []
        for rv in self._rows:
            try:
                result.append({"name": rv["name"].get(),
                                "lambda_d": rv["lambda_d"].get(),
                                "dc": rv["dc"].get(),
                                "beta": rv["beta"].get()})
            except Exception as e:
                messagebox.showerror("Fel", str(e), parent=self); return
        self.result = result; self.destroy()

    def _clear(self):
        self.result = []; self.destroy()


# ── SIF-datablad ───────────────────────────────────────────────────────────────
class SIFDatasheetDialog(tk.Toplevel):
    """IEC 61511 kl. 11 dokumentationsfält för en SIF."""

    FIELDS = [
        ("Tag-nummer:",               "tag",                ""),
        ("Hazardbeskrivning:",        "hazard",             ""),
        ("Konsekvens:",               "consequence",        ""),
        ("Process Safety Time [s]:",  "process_safety_time",""),
        ("Tillåten responstid [s]:",  "response_time",      ""),
        ("SIL-bestämningsmetod:",     "sil_basis",          "LOPA / Riskgraf / Annat"),
        ("SIF-beskrivning:",          "sif_description",    ""),
        ("Säkert tillstånd:",         "safe_state",         ""),
        ("Återstart:",                "reset",              "Manuell"),
        ("Bypass-policy:",            "bypass",             ""),
        ("Provtestprocedur (ref.):",  "pt_procedure",       ""),
        ("Övrigt / Noteringar:",      "notes",              ""),
    ]

    def __init__(self, parent, data: dict, on_save):
        super().__init__(parent)
        self.title("SIF-datablad (IEC 61511 kl. 11)")
        self.geometry("620x560"); self.resizable(True, True)
        self.transient(parent); self.grab_set()
        self._on_save = on_save
        self._vars: dict[str, tk.StringVar] = {}
        self._build(data); self.wait_window()

    def _build(self, data: dict):
        canvas = tk.Canvas(self, highlightthickness=0)
        sb = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        inner = tk.Frame(canvas, padx=14, pady=10)
        canvas.create_window((0,0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        for i,(lbl,key,default) in enumerate(self.FIELDS):
            tk.Label(inner, text=lbl, font=FONT_S, anchor="w").grid(
                row=i, column=0, sticky="w", pady=(4,0))
            v = tk.StringVar(value=data.get(key, default))
            self._vars[key] = v
            ttk.Entry(inner, textvariable=v, width=48, font=FONT_N).grid(
                row=i, column=1, sticky="ew", padx=(8,0), pady=(4,0))
        inner.columnconfigure(1, weight=1)

        bf = tk.Frame(self, padx=12, pady=8); bf.pack(fill="x", side="bottom")
        ttk.Button(bf, text="Avbryt", command=self.destroy).pack(side="right", padx=(4,0))
        ttk.Button(bf, text="Spara",  command=self._save).pack(side="right")

    def _save(self):
        self._on_save({key: v.get() for _, key, _ in self.FIELDS for v in [self._vars[key]]})
        self.destroy()


# ── Revisionshistorik ──────────────────────────────────────────────────────────
class RevisionDialog(tk.Toplevel):
    def __init__(self, parent, revisions: list[dict], on_save):
        super().__init__(parent)
        self.title("Revisionshistorik")
        self.geometry("640x400"); self.resizable(True, True)
        self.transient(parent); self.grab_set()
        self._revisions = list(revisions)
        self._on_save = on_save
        self._build(); self.wait_window()

    def _build(self):
        cols = ("date","user","comment")
        hdrs = ("Datum","Utförd av","Kommentar")
        frm = tk.Frame(self); frm.pack(fill="both", expand=True, padx=8, pady=8)
        self._tv = ttk.Treeview(frm, columns=cols, show="headings", selectmode="browse")
        for c,h,w in zip(cols,hdrs,[160,120,280]):
            self._tv.heading(c,text=h); self._tv.column(c,width=w)
        self._tv.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(frm, orient="vertical", command=self._tv.yview)
        sb.pack(side="right", fill="y"); self._tv.configure(yscrollcommand=sb.set)
        self._refresh_list()

        add_frm = tk.Frame(self, padx=8, pady=4); add_frm.pack(fill="x")
        tk.Label(add_frm, text="Utförd av:", font=FONT_S).pack(side="left")
        self._v_user = tk.StringVar()
        ttk.Entry(add_frm, textvariable=self._v_user, width=18, font=FONT_N).pack(side="left", padx=(4,12))
        tk.Label(add_frm, text="Kommentar:", font=FONT_S).pack(side="left")
        self._v_comment = tk.StringVar()
        ttk.Entry(add_frm, textvariable=self._v_comment, width=30, font=FONT_N).pack(side="left", padx=(4,8))
        ttk.Button(add_frm, text="+ Lägg till", command=self._add).pack(side="left")

        bf = tk.Frame(self, padx=8, pady=6); bf.pack(fill="x")
        ttk.Button(bf, text="Stäng och spara", command=self._close).pack(side="right")

    def _refresh_list(self):
        self._tv.delete(*self._tv.get_children())
        for r in reversed(self._revisions):
            self._tv.insert("","end", values=(r.get("date",""),r.get("user",""),r.get("comment","")))

    def _add(self):
        if not self._v_comment.get().strip():
            messagebox.showinfo("Info","Ange en kommentar.", parent=self); return
        self._revisions.append({
            "date": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
            "user": self._v_user.get(),
            "comment": self._v_comment.get(),
        })
        self._v_comment.set(""); self._refresh_list()

    def _close(self):
        self._on_save(self._revisions); self.destroy()


# ── Delsystemspanel ────────────────────────────────────────────────────────────
class SubsystemFrame(ttk.LabelFrame):
    # Primära parametrar (3 per rad)
    _PARAMS = [
        ("λ_D [1/h]", "lambda_d","1e-6", "Farlig felfrekvens totalt (DU+DD)"),
        ("DC [0–1]",  "dc",      "0.0",  "Diagnostiktäckning. 0=ingen, 0.9=90%"),
        ("β [0–1]",   "beta",    "0.02", "CCF-faktor. Typvärde 0.02–0.10"),
        ("TI [h]",    "ti",      "8760", "Provtestintervall i timmar"),
        ("MTTR [h]",  "mttr",    "8",    "Reparationstid för detekterade fel"),
        ("PTC [0–1]", "ptc",     "1.0",  "Provtesttäckning. 1.0=alla DU-fel hittas"),
        ("SFF [0–1]", "sff",     "0.0",  "Säker felfraktion (HFT-kontroll + STR)"),
    ]
    # Avancerade parametrar
    _ADV = [
        ("SC (SIL Claim Limit)", "sc",           "0",      "SIL Claim Limit från leverantör. 0=ej angiven"),
        ("ST [h]",               "st",           "0",      "Self-test intervall [h]. 0=kontinuerlig diagnostik"),
        ("Livslängd [h]",        "mission_time", "175200", "Utrustningslivslängd i timmar (175200=20 år)"),
        ("PST-täckning [0–1]",   "pst_coverage", "0.0",    "Partial stroke test coverage (för ventiler)"),
        ("PST-intervall [h]",    "pst_interval", "720",    "Intervall för partial stroke test (720h=1 mån)"),
        # exida-parametrar A1–A6
        ("DTI [h]",              "dti",          "0",      "A3: Diagnostisk testintervall. 0=kontinuerlig. MDT_DD += DTI/2"),
        ("PTD [h]",              "ptd",          "8",      "A1: Proof Test Duration — tid utan skydd under provtest. PFD += PTD/TI. Standard 8h. Sensor ~1h, liten ventil ~2h, stor ventil ~8h"),
        ("PIF [0–1]",            "pif",          "0",      "A2: Initial haveri-sannolikhet vid idriftsättning. Typvärde 0.01=1%"),
        ("SSI [0–4]",            "ssi",          "2",      "A6: Site Safety Index. 0=svag(×2.0), 2=standard(×1.0), 4=perfekt(×0.5)"),
        ("Nyttjoliv [h]",        "useful_life",  "0",      "A8: Slitage startar vid nyttjoliv. 0=inaktivt. Typvärde 175200=20år"),
    ]

    def __init__(self, parent, title):
        super().__init__(parent, text=title, padding=(8,4,8,8))
        self._vars: dict[str,tk.StringVar] = {}
        self._adv_open = False
        self._channels: list[dict] = []
        self._build()

    def _build(self):
        NCOLS = 3

        # ── Rad 0: namn + arkitektur + välj + metod + CCF ─────────────────────
        ttk.Label(self, text="Namn:", font=FONT_S).grid(row=0,column=0,sticky="w",pady=(0,3))
        self.v_name = tk.StringVar()
        ttk.Entry(self, textvariable=self.v_name, width=18,
                  font=FONT_N).grid(row=0,column=1,sticky="ew",pady=(0,3))

        ttk.Label(self, text="Arkitektur:", font=FONT_S).grid(
            row=0,column=2,sticky="w",padx=(10,4),pady=(0,3))
        self.v_arch = tk.StringVar(value="1oo1")
        ttk.Combobox(self, textvariable=self.v_arch, values=ARCH_OPTIONS,
                     width=9, state="readonly", font=FONT_N).grid(
            row=0,column=3,sticky="ew",pady=(0,3))

        ttk.Button(self, text="Välj från databas...",
                   command=self._pick).grid(row=0,column=4,sticky="e",padx=(8,0),pady=(0,3))

        # Rad 1: Beräkningsmetod + CCF-modell
        meta = tk.Frame(self); meta.grid(row=1,column=0,columnspan=5,sticky="w",pady=(0,4))
        ttk.Label(meta, text="Metod:", font=FONT_S).pack(side="left")
        self.v_method = tk.StringVar(value="markov")
        ttk.Combobox(meta, textvariable=self.v_method, values=CALC_METHODS,
                     width=10, state="readonly", font=FONT_S).pack(side="left", padx=(2,16))
        ttk.Label(meta, text="CCF-modell:", font=FONT_S).pack(side="left")
        self.v_ccf = tk.StringVar(value="beta")
        ttk.Combobox(meta, textvariable=self.v_ccf, values=CCF_MODELS,
                     width=10, state="readonly", font=FONT_S).pack(side="left", padx=(2,16))
        self._mixed_btn = ttk.Button(meta, text="Heterogena kanaler...", command=self._mixed)
        self._mixed_btn.pack(side="left")

        ttk.Separator(self, orient="horizontal").grid(
            row=2,column=0,columnspan=5,sticky="ew",pady=(0,5))

        # ── Primära parametrar ─────────────────────────────────────────────────
        for idx,(lbl,key,default,tip) in enumerate(self._PARAMS):
            gr = 3 + (idx // NCOLS)*2
            gc = idx % NCOLS
            px = (0,0) if gc==0 else (14,0)
            l = ttk.Label(self, text=lbl, font=FONT_S, foreground="#444")
            l.grid(row=gr,column=gc,sticky="w",padx=px)
            _Tip(l,tip)
            v = tk.StringVar(value=default); self._vars[key]=v
            ttk.Entry(self, textvariable=v, width=13,
                      font=FONT_M).grid(row=gr+1,column=gc,sticky="ew",padx=px,pady=(1,4))

        # TI-hint
        self._ti_hint = tk.Label(self, text="", font=FONT_S, fg="#888")
        self._ti_hint.grid(row=5,column=0,sticky="w")
        self._vars["ti"].trace_add("write", lambda *_: self._upd_ti())
        self._upd_ti()

        # ── FIT-ingång ─────────────────────────────────────────────────────────
        fit_frm = tk.Frame(self); fit_frm.grid(row=9, column=0, columnspan=5, sticky="ew", pady=(2,4))
        tk.Label(fit_frm, text="FIT-ingång:", font=FONT_S, fg="#666").pack(side="left")
        tk.Label(fit_frm, text="λDU:", font=FONT_S).pack(side="left", padx=(8,2))
        self.v_fit_du = tk.StringVar()
        ttk.Entry(fit_frm, textvariable=self.v_fit_du, width=9, font=FONT_M).pack(side="left")
        tk.Label(fit_frm, text="λDD:", font=FONT_S).pack(side="left", padx=(8,2))
        self.v_fit_dd = tk.StringVar()
        ttk.Entry(fit_frm, textvariable=self.v_fit_dd, width=9, font=FONT_M).pack(side="left")
        tk.Label(fit_frm, text="FIT", font=FONT_S, fg="#888").pack(side="left", padx=(2,8))
        ttk.Button(fit_frm, text="→ Applicera", command=self._apply_fit).pack(side="left")
        lbl_help = tk.Label(fit_frm, text="(?)", font=FONT_S, fg="#aaa")
        lbl_help.pack(side="left")
        _Tip(lbl_help, "Ange λDU och λDD direkt i FIT (fel per 10⁹ h) från FMEDA-data.\n"
             "Beräknar λD = (λDU+λDD)×10⁻⁹ och DC = λDD/(λDU+λDD).")

        # Komponenttyp
        ttk.Separator(self, orient="horizontal").grid(
            row=10,column=0,columnspan=5,sticky="ew",pady=(2,4))
        ttk.Label(self, text="Komponenttyp (IEC 61508):", font=FONT_S,
                  foreground="#666").grid(row=11,column=0,columnspan=2,sticky="w")
        self.v_ctype = tk.StringVar(value="A (enkel)")
        ttk.Combobox(self, textvariable=self.v_ctype, values=COMP_TYPES,
                     width=14, state="readonly", font=FONT_S).grid(
            row=12,column=0,columnspan=2,sticky="w",pady=(2,4))
        ttk.Label(self, text="A=enkel/mekanisk  B=komplex/mikroprocessor",
                  font=FONT_S, foreground="#999").grid(
            row=12,column=2,columnspan=3,sticky="w",padx=(12,0))

        # ── Avancerade parametrar (fällbara) ───────────────────────────────────
        self._toggle_btn = ttk.Button(self, text="▶ Avancerade parametrar",
                                      command=self._toggle_adv)
        self._toggle_btn.grid(row=13,column=0,columnspan=5,sticky="w",pady=(4,0))

        self._adv_frame = tk.Frame(self, bg=BG_FRAME)
        # byggs men visas inte förrän toggle
        self._build_adv()

        for c in range(5): self.columnconfigure(c,weight=1)

    def _build_adv(self):
        f = self._adv_frame
        NCOLS = 3
        for idx,(lbl,key,default,tip) in enumerate(self._ADV):
            gr = (idx // NCOLS)*2
            gc = idx % NCOLS
            px = (0,0) if gc==0 else (14,0)
            l = ttk.Label(f, text=lbl, font=FONT_S, foreground="#555")
            l.grid(row=gr,column=gc,sticky="w",padx=px)
            _Tip(l,tip)
            v = tk.StringVar(value=default); self._vars[key]=v
            ttk.Entry(f, textvariable=v, width=13,
                      font=FONT_M).grid(row=gr+1,column=gc,sticky="ew",padx=px,pady=(1,4))

        # Online PT-checkbox (D1)
        last_row = ((len(self._ADV)-1) // NCOLS)*2 + 2
        self._v_pt_online = tk.BooleanVar(value=False)
        pt_frm = tk.Frame(f); pt_frm.grid(row=last_row, column=0, columnspan=3, sticky="w", pady=(4,2))
        cb = ttk.Checkbutton(pt_frm, text="Online provtest (SIF på bypass)",
                             variable=self._v_pt_online)
        cb.pack(side="left")
        _Tip(cb, "D1: Kryssa i om provtestet körs med processen igång (SIF bypassad).\n"
                 "PTD/TI-bidraget är NOLL vid offline-provtest (process nedstängd).\n"
                 "Källa: exida white paper 'Key Variables for PFDavg', sidan 8.")

        for c in range(NCOLS): f.columnconfigure(c,weight=1)

    def _toggle_adv(self):
        self._adv_open = not self._adv_open
        if self._adv_open:
            self._toggle_btn.config(text="▼ Avancerade parametrar")
            self._adv_frame.grid(row=14,column=0,columnspan=5,sticky="ew",pady=(2,4))
        else:
            self._toggle_btn.config(text="▶ Avancerade parametrar")
            self._adv_frame.grid_remove()

    def _apply_fit(self):
        """Konverterar FIT-värden (λDU, λDD) till λD och DC och fyller i fälten."""
        try:
            ldu = float(self.v_fit_du.get()) if self.v_fit_du.get().strip() else 0.0
            ldd = float(self.v_fit_dd.get()) if self.v_fit_dd.get().strip() else 0.0
        except ValueError:
            messagebox.showerror("Fel", "Ange giltiga tal för λDU och λDD i FIT.", parent=self.winfo_toplevel())
            return
        if ldu < 0 or ldd < 0:
            messagebox.showerror("Fel", "FIT-värden måste vara ≥ 0.", parent=self.winfo_toplevel())
            return
        total_fit = ldu + ldd
        if total_fit <= 0:
            messagebox.showinfo("Info", "Ange minst ett FIT-värde > 0.", parent=self.winfo_toplevel())
            return
        lambda_d_h = total_fit * 1e-9    # [1/h]
        dc = ldd / total_fit if total_fit > 0 else 0.0
        self._vars["lambda_d"].set(f"{lambda_d_h:.3g}")
        self._vars["dc"].set(f"{dc:.4f}")

    def _upd_ti(self):
        try:
            self._ti_hint.config(text=f"≈ {float(self._vars['ti'].get())/8760:.2f} år")
        except Exception:
            self._ti_hint.config(text="")

    def _pick(self):
        dlg = ComponentPickerDialog(self.winfo_toplevel())
        if dlg.result:
            r = dlg.result
            self.v_name.set(f"{r['manufacturer']} {r['model']}")
            self._vars["lambda_d"].set(f"{r['lambda_d']:.3g}")
            self._vars["dc"].set(f"{r['dc']:.3g}")
            self._vars["beta"].set(f"{r['beta']:.3g}")

    def _mixed(self):
        arch = self.v_arch.get()
        dlg = MixedChannelsDialog(self.winfo_toplevel(), arch, self._channels)
        if dlg.result is not None:
            self._channels = dlg.result

    def _make_comp(self) -> ComponentParams:
        ct = self.v_ctype.get()[0]
        def _f(key, default="0"): return float(self._vars.get(key, tk.StringVar(value=default)).get() or default)
        def _i(key, default="0"): return int(float(self._vars.get(key, tk.StringVar(value=default)).get() or default))
        try:
            return ComponentParams(
                name=self.v_name.get(),
                lambda_d=float(self._vars["lambda_d"].get()),
                dc=float(self._vars["dc"].get()),
                beta=float(self._vars["beta"].get()),
                ti=float(self._vars["ti"].get()),
                mttr=float(self._vars["mttr"].get()),
                ptc=float(self._vars["ptc"].get()),
                sff=float(self._vars["sff"].get()),
                comp_type=ct,
                sc=_i("sc"), st=_f("st"),
                mission_time=_f("mission_time", "175200"),
                pst_coverage=_f("pst_coverage"), pst_interval=_f("pst_interval", "720"),
                ccf_model=self.v_ccf.get(),
                # exida-parametrar A1–A6
                dti=_f("dti"),
                ptd=_f("ptd"),
                pif=_f("pif"),
                ssi=_i("ssi", "2"),
                pt_online=getattr(self, "_v_pt_online", tk.BooleanVar()).get(),
                useful_life=_f("useful_life"),
            )
        except ValueError as e:
            raise ValueError(f"Ogiltigt värde i {self.cget('text')}: {e}")

    def get_params(self) -> SubsystemParams:
        comp = self._make_comp()
        channels = []
        if self._channels:
            base = comp
            for ch in self._channels:
                try:
                    cp = ComponentParams(
                        name=ch.get("name",""),
                        lambda_d=float(ch.get("lambda_d", base.lambda_d)),
                        dc=float(ch.get("dc", base.dc)),
                        beta=float(ch.get("beta", base.beta)),
                        ti=base.ti, mttr=base.mttr, ptc=base.ptc,
                        sff=base.sff, comp_type=base.comp_type,
                        sc=base.sc, st=base.st, mission_time=base.mission_time,
                        pst_coverage=base.pst_coverage, pst_interval=base.pst_interval,
                    )
                    channels.append(cp)
                except Exception:
                    pass
        return SubsystemParams(
            name=self.v_name.get() or self.cget("text"),
            architecture=Architecture(self.v_arch.get()),
            component=comp,
            channels=channels,
            calc_method=self.v_method.get(),
        )

    def to_dict(self) -> dict:
        d = {"name":self.v_name.get(),"arch":self.v_arch.get(),
             "comp_type":self.v_ctype.get(),"ccf_model":self.v_ccf.get(),
             "calc_method":self.v_method.get(),
             "fit_du":self.v_fit_du.get(),"fit_dd":self.v_fit_dd.get(),
             "pt_online": getattr(self, "_v_pt_online", tk.BooleanVar()).get(),
             "channels":[dict(ch) for ch in self._channels],
             **{k:v.get() for k,v in self._vars.items()}}
        return d

    def from_dict(self, d: dict):
        self.v_name.set(d.get("name",""))
        self.v_arch.set(d.get("arch","1oo1"))
        self.v_ctype.set(d.get("comp_type","A (enkel)"))
        self.v_ccf.set(d.get("ccf_model","beta"))
        self.v_method.set(d.get("calc_method","markov"))
        self.v_fit_du.set(d.get("fit_du",""))
        self.v_fit_dd.set(d.get("fit_dd",""))
        if hasattr(self, "_v_pt_online"):
            self._v_pt_online.set(bool(d.get("pt_online", False)))
        self._channels = [dict(ch) for ch in d.get("channels",[])]
        for k,v in self._vars.items():
            if k in d: v.set(d[k])


# ── Känslighetsgraf ────────────────────────────────────────────────────────────
class SensitivityWindow(tk.Toplevel):
    def __init__(self, parent, sensor_p, logic_p, fe_p, sil_req):
        super().__init__(parent)
        self.title("Känslighetsgraf — PFD vs provtestintervall")
        self.geometry("720x480"); self.resizable(True, True)
        self._build(sensor_p, logic_p, fe_p, sil_req)

    def _build(self, sp, lp, fp, sil_req):
        try:
            import numpy as np
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            from matplotlib.figure import Figure
        except ImportError:
            tk.Label(self, text="Kräver matplotlib", font=FONT_N).pack(pady=40); return

        ti_yr = np.linspace(0.25, 4.0, 60)
        ti_h  = ti_yr * 8760.0
        fig = Figure(figsize=(7.2, 4.5), dpi=96, facecolor=BG_FRAME)
        ax  = fig.add_subplot(111)
        colors = {"Sensor":"#2980b9","Logic solver":"#27ae60","Slutelement":"#e67e22","Totalt":"#c0392b"}
        for label, base_p in [("Sensor",sp),("Logic solver",lp),("Slutelement",fp)]:
            pfds = []
            for ti in ti_h:
                p2 = copy.deepcopy(base_p); p2.component.ti = float(ti)
                pfds.append(calc_subsystem(p2).pfd)
            ax.plot(ti_yr, pfds, label=label, color=colors[label], linewidth=1.5)
        tot_pfds = []
        for ti in ti_h:
            s2,l2,f2 = copy.deepcopy(sp),copy.deepcopy(lp),copy.deepcopy(fp)
            s2.component.ti = l2.component.ti = f2.component.ti = float(ti)
            tot_pfds.append(calc_sif("",s2,l2,f2,sil_req).pfd_total)
        ax.plot(ti_yr, tot_pfds, label="Totalt", color=colors["Totalt"],
                linewidth=2.5, linestyle="--")
        for lim,lbl in [(1e-1,"SIL 1"),(1e-2,"SIL 2"),(1e-3,"SIL 3"),(1e-4,"SIL 4")]:
            ax.axhline(lim, color="#aaa", linestyle=":", linewidth=0.8)
            ax.text(4.05, lim, lbl, va="center", fontsize=8, color="#999")
        req_lo = SIL_LIMITS.get(sil_req,(1e-9,1e-9))[0]
        ax.axhline(req_lo, color=RED, linestyle="--", linewidth=1.2, label=f"Krav SIL {sil_req}")
        ax.set_yscale("log"); ax.set_xlabel("Provtestintervall [år]", fontsize=10)
        ax.set_ylabel("PFD_avg", fontsize=10)
        ax.set_title("PFD vs TI (övriga param. konstanta)", fontsize=10)
        ax.legend(fontsize=9); ax.grid(True, which="both", linestyle=":", linewidth=0.5)
        fig.tight_layout()
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        canvas = FigureCanvasTkAgg(fig, master=self)
        canvas.draw(); canvas.get_tk_widget().pack(fill="both", expand=True)


# ── Resultpanel ────────────────────────────────────────────────────────────────
class ResultsFrame(ttk.LabelFrame):
    def __init__(self, parent):
        super().__init__(parent, text="Beräkningsresultat", padding=(8,4,8,8))
        self._build()

    def _build(self):
        hdrs = ["Delsystem","Arkitektur","Metod","PFD_avg","PFH [1/h]","STR [1/h]",
                "SIL (PFD)","HFT/SFF","SC"]
        for c,h in enumerate(hdrs):
            ttk.Label(self,text=h,font=("Segoe UI",9,"bold"),foreground="#555").grid(
                row=0,column=c,sticky="w",padx=(0 if c==0 else 6,0))
        ttk.Separator(self,orient="horizontal").grid(
            row=1,column=0,columnspan=len(hdrs),sticky="ew",pady=3)
        self._rows: list[list[tk.Label]] = []
        for r in range(3):
            rl = []
            for c in range(len(hdrs)):
                lbl = tk.Label(self,text="—",font=FONT_M,bg=BG_FRAME,anchor="w")
                lbl.grid(row=r+2,column=c,sticky="ew",padx=(0 if c==0 else 6,0),pady=1)
                rl.append(lbl)
            self._rows.append(rl)
        ttk.Separator(self,orient="horizontal").grid(
            row=5,column=0,columnspan=len(hdrs),sticky="ew",pady=3)
        self._tot = tk.Label(self,text="",font=("Segoe UI",10,"bold"),bg=BG_FRAME,anchor="w")
        self._tot.grid(row=6,column=0,columnspan=len(hdrs),sticky="ew")
        self._str_lbl = tk.Label(self,text="",font=FONT_N,bg=BG_FRAME,anchor="w",fg="#555")
        self._str_lbl.grid(row=7,column=0,columnspan=len(hdrs),sticky="ew")
        self._verd = tk.Label(self,text="",font=("Segoe UI",13,"bold"),bg=BG_FRAME,anchor="w")
        self._verd.grid(row=8,column=0,columnspan=len(hdrs),sticky="ew")
        self._det = tk.Text(self,height=5,font=("Consolas",9),state="disabled",
                            bg="#f8f8f8",relief="flat")
        self._det.grid(row=9,column=0,columnspan=len(hdrs),sticky="ew",pady=(6,0))
        self._det.grid_remove()
        for c in range(len(hdrs)): self.columnconfigure(c,weight=1)

    def show(self, r: SIFResult):
        subs = [("Sensor",r.sensor),("Logic solver",r.logic),("Slutelement",r.final_element)]
        for i,(name,s) in enumerate(subs):
            hft_txt = f"SIL {s.sil_hft}" if s.sil_hft>0 else "–"
            hft_col = GREEN if s.sil_hft>=r.sil_required else (RED if s.sil_hft>0 else "#888")
            sc_txt  = f"SIL {s.sil_sc}"  if s.sil_sc>0  else "–"
            sc_col  = GREEN if s.sil_sc>=r.sil_required  else (RED if s.sil_sc>0  else "#888")
            self._rows[i][0].config(text=name)
            self._rows[i][1].config(text=s.architecture)
            self._rows[i][2].config(text=getattr(s,"calc_method","markov"))
            self._rows[i][3].config(text=f"{s.pfd:.3e}")
            self._rows[i][4].config(text=f"{s.pfh:.3e}")
            self._rows[i][5].config(text=f"{s.str_rate:.2e}" if s.str_rate>0 else "–")
            self._rows[i][6].config(text=f"SIL {s.sil_pfd}" if s.sil_pfd>0 else "<SIL1",
                                    fg=GREEN if s.sil_pfd>=r.sil_required else RED)
            self._rows[i][7].config(text=hft_txt, fg=hft_col)
            self._rows[i][8].config(text=sc_txt,  fg=sc_col)

        mttfs = f"{r.mttfs/8760:.1f} år" if r.mttfs<1e9 else "∞"
        self._tot.config(text=(f"PFD = {r.pfd_total:.3e}   PFH = {r.pfh_total:.3e}"
                               f"   → SIL {r.sil_achieved}  (krav SIL {r.sil_required})"))
        self._str_lbl.config(text=(f"STR = {r.str_total:.2e} [1/h]   "
                                   f"MTTFS = {mttfs}  (medelid för falskt utlösning)"))
        fg = GREEN if r.passed else RED
        self._verd.config(
            text=(f"GODKANT — SIL {r.sil_achieved} ≥ SIL {r.sil_required}" if r.passed
                  else f"EJ GODKANT — SIL {r.sil_achieved} < SIL {r.sil_required}"),
            fg=fg)
        lines = []
        for name,s in subs:
            lines.append(f"  {name} ({s.architecture}  {getattr(s,'calc_method','markov')})")
            for j,st in enumerate(s.markov.states):
                m = " <-- SIF-fel" if j in s.markov.pfd_states else ""
                lines.append(f"    π[{j}] {st:30s} = {s.markov.steady_state[j]:.4e}{m}")
        self._det.config(state="normal")
        self._det.delete("1.0","end"); self._det.insert("end","\n".join(lines))
        self._det.config(state="disabled"); self._det.grid()

    def clear(self):
        for row in self._rows:
            for l in row: l.config(text="—",fg="black")
        self._tot.config(text=""); self._str_lbl.config(text=""); self._verd.config(text="")
        self._det.grid_remove()


# ── PFD-budgetdiagram ──────────────────────────────────────────────────────────
class BudgetChart(ttk.LabelFrame):
    def __init__(self, parent):
        super().__init__(parent, text="PFD-budget", padding=(8,4,8,8))
        self._ok = False
        try:
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            from matplotlib.figure import Figure
            fig = Figure(figsize=(7,2.2), dpi=96, facecolor=BG_FRAME)
            self._ax = fig.add_subplot(111)
            fig.subplots_adjust(left=0.18,right=0.78,top=0.80,bottom=0.20)
            self._canvas = FigureCanvasTkAgg(fig, master=self)
            self._canvas.get_tk_widget().pack(fill="both", expand=True)
            self._fig = fig; self._ok = True
        except ImportError:
            tk.Label(self, text="pip install matplotlib för diagram",
                     font=FONT_S, fg="#888").pack(pady=8)

    def update(self, r: SIFResult):
        if not self._ok: return
        ax = self._ax; ax.clear()
        names = ["Sensor","Logic solver","Slutelement"]
        pfds  = [r.sensor.pfd, r.logic.pfd, r.final_element.pfd]
        total = r.pfd_total
        colors = [SIL_COL.get(sil_from_pfd(p),"#888") for p in pfds]
        bars = ax.barh(names, pfds, color=colors, height=0.45, edgecolor="white")
        xmax = max(pfds+[1e-7])
        for bar,pfd in zip(bars,pfds):
            pct = pfd/total*100 if total>0 else 0
            ax.text(xmax*2.8, bar.get_y()+bar.get_height()/2,
                    f"{pfd:.2e}  ({pct:.0f}%)", va="center", ha="left",
                    fontsize=8.5, color="#333")
        for x,lbl in [(1e-1,"SIL 1"),(1e-2,"SIL 2"),(1e-3,"SIL 3"),(1e-4,"SIL 4")]:
            ax.axvline(x=x, color="#bbb", linestyle="--", linewidth=0.8)
            ax.text(x, len(names)-0.05, lbl, ha="center", va="bottom", fontsize=7.5, color="#999")
        ax.set_xscale("log"); ax.set_xlim(min(pfds+[1e-8])/5, 2.0)
        ax.set_xlabel("PFD_avg", fontsize=9)
        sil_str = f"SIL {r.sil_achieved}" if r.sil_achieved>0 else "<SIL1"
        ax.set_title(f"PFD={total:.2e} → {sil_str} | {'GODKANT' if r.passed else 'EJ GODKANT'}",
                     fontsize=9.5, color=GREEN if r.passed else RED)
        ax.tick_params(labelsize=8.5); self._fig.canvas.draw()


# ── HTML-rapport ───────────────────────────────────────────────────────────────
def generate_html_report(sifs_data: list[dict], project_file: str = "") -> str:
    date_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

    def row(cells, header=False):
        tag = "th" if header else "td"
        return "<tr>" + "".join(f"<{tag}>{c}</{tag}>" for c in cells) + "</tr>"

    body = (f"<h1>SIL PFD-beräkningsunderlag</h1>"
            f"<p><b>Projekt:</b> {project_file or '—'} &nbsp;&nbsp; <b>Datum:</b> {date_str}<br>"
            f"<b>Standard:</b> IEC 61511 / IEC 61508 &nbsp;&nbsp; <b>Metod:</b> Markov steady-state</p><hr>")

    for sd in sifs_data:
        name = sd.get("name","—"); sil_req = sd.get("sil_req",2)
        ds = sd.get("datasheet", DEFAULT_DATASHEET)
        revs = sd.get("revisions",[])

        # Datablad
        body += f"<h2>{name}</h2>"
        body += "<table><tbody>"
        for lbl,key,_ in SIFDatasheetDialog.FIELDS:
            val = ds.get(key,"—") or "—"
            body += f"<tr><th style='width:200px'>{lbl}</th><td>{val}</td></tr>"
        body += f"<tr><th>SIL-krav</th><td>SIL {sil_req}</td></tr>"
        body += "</tbody></table>"

        # Beräkning
        try:
            from calc import SubsystemParams, Architecture, ComponentParams, calc_sif
            def mk(d,n):
                ct = d.get("comp_type","A (enkel)")[0]
                return SubsystemParams(
                    name=d.get("name",n),
                    architecture=Architecture(d.get("arch","1oo1")),
                    calc_method=d.get("calc_method","markov"),
                    component=ComponentParams(
                        name=d.get("name",""), lambda_d=float(d.get("lambda_d",1e-6)),
                        dc=float(d.get("dc",0)), beta=float(d.get("beta",0.02)),
                        ti=float(d.get("ti",8760)), mttr=float(d.get("mttr",8)),
                        ptc=float(d.get("ptc",1.0)), sff=float(d.get("sff",0)), comp_type=ct,
                        sc=int(float(d.get("sc",0))), st=float(d.get("st",0)),
                        mission_time=float(d.get("mission_time",175200)),
                        pst_coverage=float(d.get("pst_coverage",0)),
                        pst_interval=float(d.get("pst_interval",720)),
                        ccf_model=d.get("ccf_model","beta")))
            r = calc_sif(name, mk(sd["sensor"],"Sensor"), mk(sd["logic"],"Logic"),
                         mk(sd["fe"],"Slutelement"), sil_req)
            verdict_col = "#1e8449" if r.passed else "#c0392b"
            verdict_txt = (f"GODKANT — SIL {r.sil_achieved} ≥ SIL {sil_req}" if r.passed
                           else f"EJ GODKANT — SIL {r.sil_achieved} &lt; SIL {sil_req}")
            mttfs = f"{r.mttfs/8760:.1f} år" if r.mttfs<1e9 else "∞"
        except Exception as e:
            body += f"<p>Fel: {e}</p><hr>"; continue

        body += f"<p><b>Utfall:</b> <span style='color:{verdict_col};font-weight:bold'>{verdict_txt}</span></p>"

        body += "<table><thead>" + row(["Delsystem","Arkitektur","Metod","λ_D","DC","β","TI [h]","PTC","SFF","SC","PST-täckning"],True) + "</thead><tbody>"
        for sk,sn,sr in [("sensor","Sensor",r.sensor),("logic","Logic solver",r.logic),("fe","Slutelement",r.final_element)]:
            d = sd.get(sk,{})
            sc_str = d.get("sc","0"); pst_str = d.get("pst_coverage","0")
            body += row([sn,d.get("arch",""),d.get("calc_method","markov"),
                         d.get("lambda_d",""),d.get("dc",""),d.get("beta",""),
                         d.get("ti",""),d.get("ptc",""),d.get("sff",""),sc_str,pst_str])
        body += "</tbody></table>"

        body += "<table><thead>" + row(["Delsystem","PFD_avg","PFH [1/h]","STR [1/h]","SIL (PFD)","HFT/SFF","SC-begr."],True) + "</thead><tbody>"
        for lbl,sr in [("Sensor",r.sensor),("Logic solver",r.logic),("Slutelement",r.final_element)]:
            body += row([lbl, f"{sr.pfd:.3e}", f"{sr.pfh:.3e}",
                         f"{sr.str_rate:.2e}" if sr.str_rate>0 else "–",
                         f"SIL {sr.sil_pfd}" if sr.sil_pfd>0 else "&lt;SIL1",
                         f"SIL {sr.sil_hft}" if sr.sil_hft>0 else "–",
                         f"SIL {sr.sil_sc}"  if sr.sil_sc>0  else "–"])
        body += row([f"<b>TOTALT</b>",f"<b>{r.pfd_total:.3e}</b>",f"<b>{r.pfh_total:.3e}</b>",
                     f"<b>{r.str_total:.2e}</b>",f"<b>SIL {r.sil_achieved}</b>",
                     "",""]); body += "</tbody></table>"
        body += f"<p><i>MTTFS = {mttfs}</i></p>"

        # Revisioner
        if revs:
            body += "<h3>Revisionshistorik</h3><table><thead>" + row(["Datum","Utförd av","Kommentar"],True) + "</thead><tbody>"
            for rv in revs:
                body += row([rv.get("date",""),rv.get("user",""),rv.get("comment","")])
            body += "</tbody></table>"
        body += "<hr>"

    html = (f"<!DOCTYPE html><html lang='sv'><head><meta charset='utf-8'>"
            f"<title>SIL PFD-rapport — ProSa</title>"
            f"<style>"
            f"body{{font-family:'Segoe UI',Arial,sans-serif;margin:32px 48px;color:#222;font-size:13px}}"
            f"h1{{color:#1a5276;border-bottom:2px solid #1a5276;padding-bottom:6px}}"
            f"h2{{color:#1a5276;margin-top:28px}}h3{{color:#444}}"
            f"table{{border-collapse:collapse;width:100%;margin:10px 0 18px 0;font-size:12px}}"
            f"th{{background:#1a5276;color:#fff;padding:5px 10px;text-align:left}}"
            f"td{{padding:4px 10px;border-bottom:1px solid #ddd}}"
            f"tr:nth-child(even){{background:#f5f8fc}}"
            f"hr{{border:none;border-top:1px solid #ccc;margin:24px 0}}"
            f"@media print{{body{{margin:10px}}}}"
            f"</style></head><body>{body}"
            f"<p style='color:#999;font-size:11px;margin-top:32px'>"
            f"Genererad av ProSa SIL-kalkylator | {date_str}</p>"
            f"</body></html>")
    return html



# ── Verifieringsdialog ─────────────────────────────────────────────────────────
class VerificationDialog(tk.Toplevel):
    """
    Verifieringsdialog med två flikar:
      Flik 1 — Formelverifiering mot SIF-001 (alla 4 arkitekturer per delsystem)
      Flik 2 — Hybrit-konsistens (SIF-001 t.o.m. SIF-015, additiv modell + SIL-klass)
    """

    def __init__(self, parent, sifs_data: list, current_sif_index: int):
        super().__init__(parent)
        self.title("Verifiera beräkningsmotor — ProSa SIL-kalkylator")
        self.geometry("900x640"); self.resizable(True, True)
        self.transient(parent); self.grab_set()
        self._data = sifs_data
        self._idx  = current_sif_index
        self._build(); self.wait_window()

    def _build(self):
        if not _VDB_AVAILABLE:
            tk.Label(self, text="verification_db.py saknas.", font=FONT_N, fg=RED).pack(pady=20)
            ttk.Button(self, text="Stäng", command=self.destroy).pack()
            return

        nb = ttk.Notebook(self); nb.pack(fill="both", expand=True, padx=6, pady=6)

        # ── Flik 1: Formelverifiering (SIF-001 PDF) ────────────────────────────
        tab1 = tk.Frame(nb, bg=BG); nb.add(tab1, text="Formelverifiering (SIF-001 PDF)")
        top1 = tk.Frame(tab1, bg=BG, padx=8, pady=6); top1.pack(fill="x")
        tk.Label(top1, text="Referensfall:", font=FONT_N, bg=BG).pack(side="left")
        self._v_case = tk.StringVar()
        try:
            cases = vdb.get_case_ids()
        except Exception:
            vdb.init_db(); cases = vdb.get_case_ids()
        self._v_case.set(cases[0] if cases else "")
        ttk.Combobox(top1, textvariable=self._v_case, values=cases,
                     width=18, state="readonly", font=FONT_N).pack(side="left", padx=(4,12))
        ttk.Button(top1, text="Kör verifiering", command=self._run_formula).pack(side="left")
        tk.Label(top1, text="Verifierar att alla formler (1oo1/1oo2/2oo2/2oo3) stämmer mot PDF-referens.",
                 font=FONT_S, fg="#666", bg=BG).pack(side="left", padx=8)

        frm1 = tk.Frame(tab1); frm1.pack(fill="both", expand=True, padx=8, pady=(2,4))
        self._tv1, self._sum1 = self._make_tree(frm1, tab1)

        # ── Flik 2: Hybrit-konsistens (SIF-001 – SIF-015) ─────────────────────
        tab2 = tk.Frame(nb, bg=BG); nb.add(tab2, text="Hybrit-konsistens (SIF-001 till SIF-015)")
        top2 = tk.Frame(tab2, bg=BG, padx=8, pady=6); top2.pack(fill="x")
        ttk.Button(top2, text="Kör konsistenskontroll", command=self._run_hybrit).pack(side="left")
        tk.Label(top2,
                 text="Kontrollerar: PFD-summa, RRF och SIL-klassificering mot Hybrit SILver-rapport (2022).",
                 font=FONT_S, fg="#666", bg=BG).pack(side="left", padx=8)

        frm2 = tk.Frame(tab2); frm2.pack(fill="both", expand=True, padx=8, pady=(2,4))
        self._tv2, self._sum2 = self._make_tree(frm2, tab2)

        # ── Stängknapp ─────────────────────────────────────────────────────────
        bf = tk.Frame(self, bg=BG, padx=8, pady=4); bf.pack(fill="x", side="bottom")
        ttk.Button(bf, text="Stäng", command=self.destroy).pack(side="right")

    def _make_tree(self, frm: tk.Frame, parent_tab: tk.Frame):
        """Skapar en Treeview + scrollbar + summarad; returnerar (tv, summary_label)."""
        cols = ("label", "calc", "ref", "diff", "ok")
        hdrs = ("Kontroll", "Beräknat", "Referens", "Avv. %", "Status")
        ws   = (300, 110, 110, 80, 60)
        tv = ttk.Treeview(frm, columns=cols, show="headings", selectmode="none")
        for c, h, w in zip(cols, hdrs, ws):
            tv.heading(c, text=h); tv.column(c, width=w, minwidth=30)
        tv.tag_configure("ok",     foreground="#1e8449")
        tv.tag_configure("fail",   foreground="#c0392b")
        tv.tag_configure("header", foreground="#1a5276", font=("Segoe UI", 9, "bold"))
        tv.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(frm, orient="vertical", command=tv.yview)
        sb.pack(side="right", fill="y"); tv.configure(yscrollcommand=sb.set)
        sum_lbl = tk.Label(parent_tab, text="", font=FONT_N, bg=BG, anchor="w")
        sum_lbl.pack(fill="x", padx=8, pady=(0, 2))
        return tv, sum_lbl

    def _populate_tree(self, tv: ttk.Treeview, summary_lbl, rows: list):
        """Fyller en Treeview med verifieringsrader och uppdaterar summaraden."""
        tv.delete(*tv.get_children())
        pass_count = fail_count = 0
        for row in rows:
            calc_v = row.get("calc")
            ref_v  = row.get("ref")
            diff_v = row.get("diff_pct")
            ok     = row.get("ok")

            if calc_v is None and ref_v is None:
                tv.insert("", "end", values=(row["label"], "", "", "", ""),
                          tags=("header",))
                continue

            if isinstance(calc_v, float):
                calc_str = f"{calc_v:.4e}" if 0 < calc_v < 1 else f"{calc_v:.1f}"
            else:
                calc_str = str(calc_v) if calc_v is not None else "—"
            if isinstance(ref_v, float):
                ref_str = f"{ref_v:.4e}" if 0 < ref_v < 1 else f"{ref_v:.1f}"
            else:
                ref_str = str(ref_v) if ref_v is not None else "—"

            diff_str = f"{diff_v:.2f}%" if diff_v is not None else "—"
            ok_str   = "OK" if ok else ("FAIL" if ok is not None else "—")
            tag      = "ok" if ok else ("fail" if ok is not None else "")
            tv.insert("", "end",
                      values=(row["label"], calc_str, ref_str, diff_str, ok_str),
                      tags=(tag,))
            if ok is True:  pass_count += 1
            elif ok is False: fail_count += 1

        total = pass_count + fail_count
        col   = GREEN if fail_count == 0 else RED
        suffix = "  — Alla godkanda" if fail_count == 0 else f"  — {fail_count} FEL"
        summary_lbl.config(text=f"{pass_count}/{total} kontroller OK{suffix}", fg=col)

    # ── Flik 1: formelverifiering ──────────────────────────────────────────────
    def _run_formula(self):
        if not _VDB_AVAILABLE:
            return
        case_id = self._v_case.get()
        if not case_id:
            messagebox.showinfo("Info", "Välj ett referensfall.", parent=self)
            return

        sif = self._data[self._idx] if 0 <= self._idx < len(self._data) else {}

        def mk_comp(d: dict) -> "ComponentParams":
            ct = d.get("comp_type", "A (enkel)")[0]
            try:
                fit_du = float(d.get("fit_du", "") or 0)
            except ValueError:
                fit_du = 0.0
            try:
                fit_dd = float(d.get("fit_dd", "") or 0)
            except ValueError:
                fit_dd = 0.0
            return ComponentParams(
                name=d.get("name", ""),
                lambda_d=float(d.get("lambda_d", 1e-6)),
                dc=float(d.get("dc", 0)),
                beta=float(d.get("beta", 0.02)),
                ti=float(d.get("ti", 8760)),
                mttr=float(d.get("mttr", 8)),
                ptc=float(d.get("ptc", 1.0)),
                sff=float(d.get("sff", 0)),
                comp_type=ct,
                sc=int(float(d.get("sc", 0))),
                st=float(d.get("st", 0)),
                mission_time=float(d.get("mission_time", 87600)),
                pst_coverage=float(d.get("pst_coverage", 0)),
                pst_interval=float(d.get("pst_interval", 720)),
                ccf_model=d.get("ccf_model", "beta"),
                lambda_du_fit=fit_du,
                lambda_dd_fit=fit_dd,
            )

        try:
            p_s  = mk_comp(sif.get("sensor", {}))
            p_l  = mk_comp(sif.get("logic",  {}))
            p_fe = mk_comp(sif.get("fe",     {}))
            s_all  = pfd_all_architectures(p_s)
            l_all  = pfd_all_architectures(p_l)
            fe_all = pfd_all_architectures(p_fe)
            arch_s  = sif.get("sensor", {}).get("arch", "1oo1")
            arch_l  = sif.get("logic",  {}).get("arch", "1oo1")
            arch_fe = sif.get("fe",     {}).get("arch", "1oo1")
            s_sel  = pfd_simplified(Architecture(arch_s),  p_s)
            l_sel  = pfd_simplified(Architecture(arch_l),  p_l)
            fe_sel = pfd_simplified(Architecture(arch_fe), p_fe)
            total  = 1.0 - (1.0 - s_sel)*(1.0 - l_sel)*(1.0 - fe_sel)
            sil_a  = sil_from_pfd(total)
            rows   = vdb.run_verification(case_id, s_all, s_sel,
                                          l_all, l_sel, fe_all, fe_sel, total, sil_a)
        except Exception as e:
            messagebox.showerror("Fel", str(e), parent=self); return

        self._populate_tree(self._tv1, self._sum1, rows)

    # ── Flik 2: Hybrit-konsistens ──────────────────────────────────────────────
    def _run_hybrit(self):
        if not _VDB_AVAILABLE:
            return
        try:
            vdb.init_hybrit_db()
            rows = vdb.run_hybrit_consistency_check()
        except Exception as e:
            messagebox.showerror("Fel", str(e), parent=self); return
        self._populate_tree(self._tv2, self._sum2, rows)


# ── Huvud-app ──────────────────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("ProSa SIL-kalkylator")
        self.geometry("1180x960"); self.configure(bg=BG)
        cdb.init_db()
        if _VDB_AVAILABLE:
            try:
                vdb.init_db()
            except Exception:
                pass
        self._sifs: list[dict] = []
        self._current = -1; self._file: Path | None = None
        self._build_menu(); self._build_ui(); self._new_sif()

    def _build_menu(self):
        mb = tk.Menu(self)
        fm = tk.Menu(mb, tearoff=0)
        fm.add_command(label="Nytt projekt",           command=self._new_project)
        fm.add_command(label="Öppna projekt...",       command=self._open_project)
        fm.add_command(label="Spara projekt",          command=self._save_project)
        fm.add_command(label="Spara som...",           command=self._save_project_as)
        fm.add_separator()
        fm.add_command(label="Exportera Excel...",     command=self._export_excel)
        fm.add_command(label="Generera HTML-rapport",  command=self._export_html)
        fm.add_separator()
        fm.add_command(label="Avsluta",                command=self.quit)
        mb.add_cascade(label="Projekt", menu=fm)
        dm = tk.Menu(mb, tearoff=0)
        dm.add_command(label="Hantera komponenter...", command=lambda: ComponentManagerDialog(self))
        dm.add_separator()
        dm.add_command(label="Verifiera mot referensfall...", command=self._open_verification)
        mb.add_cascade(label="Databas", menu=dm)
        self.config(menu=mb)

    def _build_ui(self):
        pane = tk.PanedWindow(self, orient="horizontal", bg=BG, sashwidth=4)
        pane.pack(fill="both", expand=True, padx=8, pady=8)

        # Vänster: SIF-lista
        left = tk.Frame(pane, bg=BG, width=200); pane.add(left, minsize=180)
        tk.Label(left, text="SIF-lista", font=FONT_H, bg=BG, fg=ACCENT).pack(anchor="w",pady=(0,4))
        self._lb = tk.Listbox(left, font=FONT_N, selectmode="single", activestyle="none",
                               relief="flat", bg="white", selectbackground=ACCENT,
                               selectforeground="white")
        self._lb.pack(fill="both", expand=True)
        self._lb.bind("<<ListboxSelect>>", self._on_select)
        self._lb.bind("<Button-3>", self._on_right_click)
        bar = tk.Frame(left, bg=BG); bar.pack(fill="x", pady=(4,0))
        ttk.Button(bar, text="+ Ny SIF", command=self._new_sif).pack(side="left",expand=True,fill="x")
        ttk.Button(bar, text="Ta bort",  command=self._delete_sif).pack(side="left",expand=True,fill="x")

        # Höger
        right = tk.Frame(pane, bg=BG); pane.add(right, minsize=720)

        # Metadata
        meta = tk.Frame(right, bg=BG); meta.pack(fill="x", pady=(0,4))
        tk.Label(meta, text="SIF-namn:", font=FONT_N, bg=BG).pack(side="left")
        self.v_sif_name = tk.StringVar()
        ttk.Entry(meta, textvariable=self.v_sif_name, width=26,
                  font=FONT_N).pack(side="left", padx=(4,16))
        tk.Label(meta, text="Krav SIL:", font=FONT_N, bg=BG).pack(side="left")
        self.v_sil_req = tk.IntVar(value=2)
        ttk.Combobox(meta, textvariable=self.v_sil_req, values=SIL_OPTIONS,
                     width=4, state="readonly", font=FONT_N).pack(side="left", padx=(4,12))
        ttk.Button(meta, text="SIF-datablad...", command=self._open_datasheet).pack(side="left",padx=(0,6))
        ttk.Button(meta, text="Revisioner...",   command=self._open_revisions).pack(side="left")

        # Delsystems-frames (scrollbart)
        canvas = tk.Canvas(right, bg=BG, highlightthickness=0)
        vsb = ttk.Scrollbar(right, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        inner = tk.Frame(canvas, bg=BG)
        canvas.create_window((0,0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")))
        canvas.bind("<MouseWheel>", lambda e: canvas.yview_scroll(-1*(e.delta//120),"units"))

        self.frm_sensor = SubsystemFrame(inner, "Sensor / Transmitter")
        self.frm_sensor.pack(fill="x", pady=2)
        self.frm_logic  = SubsystemFrame(inner, "Logic Solver (SIS/PLC)")
        self.frm_logic.pack(fill="x", pady=2)
        self.frm_fe     = SubsystemFrame(inner, "Slutelement (ventil / aktuator)")
        self.frm_fe.pack(fill="x", pady=2)

        # Varningsrad
        self._warn_lbl = tk.Label(inner, text="", font=FONT_S, fg=ORANGE, bg=BG,
                                   anchor="w", wraplength=980)
        self._warn_lbl.pack(fill="x")

        # Knappar
        cb = tk.Frame(inner, bg=BG); cb.pack(fill="x", pady=4)
        ttk.Button(cb, text="  BERAKNA  ",        command=self._calculate).pack(side="left")
        ttk.Button(cb, text="Känslighetsgraf...", command=self._sensitivity).pack(side="left",padx=(8,0))
        tk.Label(cb, text="Markov  |  IEC 61508 / 61511",
                 font=FONT_S, fg="#888", bg=BG).pack(side="left", padx=10)

        self.frm_results = ResultsFrame(inner); self.frm_results.pack(fill="x", pady=2)
        self.frm_chart   = BudgetChart(inner);  self.frm_chart.pack(fill="x", pady=2)

    # ── Högerklick ─────────────────────────────────────────────────────────────
    def _on_right_click(self, event):
        self._lb.selection_clear(0,"end")
        self._lb.selection_set(self._lb.nearest(event.y))
        self._on_select(None)
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="Kopiera SIF", command=self._copy_sif)
        menu.tk_popup(event.x_root, event.y_root)

    def _copy_sif(self):
        if self._current<0: return
        self._save_current()
        new = copy.deepcopy(self._sifs[self._current])
        new["name"] += " (kopia)"
        self._sifs.append(new)
        self._lb.insert("end", new["name"])
        idx = len(self._sifs)-1
        self._lb.selection_clear(0,"end"); self._lb.selection_set(idx)
        self._current = idx; self._load_sif(new)

    # ── SIF-hantering ───────────────────────────────────────────────────────────
    def _new_sif(self):
        self._save_current()
        sif = {"name":f"SIF-{len(self._sifs)+1}","sil_req":2,
               "sensor":dict(DEFAULT_SUBS["sensor"]),
               "logic": dict(DEFAULT_SUBS["logic"]),
               "fe":    dict(DEFAULT_SUBS["fe"]),
               "datasheet": dict(DEFAULT_DATASHEET),
               "revisions": [{"date":datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
                              "user":"","comment":"Nytt dokument"}]}
        self._sifs.append(sif); idx = len(self._sifs)-1
        self._lb.insert("end", sif["name"])
        self._lb.selection_clear(0,"end"); self._lb.selection_set(idx)
        self._current = idx; self._load_sif(sif)

    def _delete_sif(self):
        if not self._sifs or self._current<0: return
        if len(self._sifs)==1:
            messagebox.showinfo("Info","Minst en SIF måste finnas."); return
        if not messagebox.askyesno("Ta bort",f"Ta bort '{self._sifs[self._current]['name']}'?"): return
        self._sifs.pop(self._current); self._lb.delete(self._current)
        self._current = min(self._current, len(self._sifs)-1)
        self._lb.selection_set(self._current); self._load_sif(self._sifs[self._current])

    def _on_select(self, _):
        sel = self._lb.curselection()
        if not sel or sel[0]==self._current: return
        self._save_current(); self._current=sel[0]; self._load_sif(self._sifs[sel[0]])

    def _save_current(self):
        if self._current<0 or self._current>=len(self._sifs): return
        sif = self._sifs[self._current]
        sif["name"]    = self.v_sif_name.get() or sif["name"]
        sif["sil_req"] = int(self.v_sil_req.get())
        sif["sensor"]  = self.frm_sensor.to_dict()
        sif["logic"]   = self.frm_logic.to_dict()
        sif["fe"]      = self.frm_fe.to_dict()
        self._lb.delete(self._current)
        self._lb.insert(self._current, sif["name"])
        self._lb.selection_set(self._current)

    def _load_sif(self, sif: dict):
        self.v_sif_name.set(sif.get("name",""))
        self.v_sil_req.set(sif.get("sil_req",2))
        self.frm_sensor.from_dict(sif.get("sensor",DEFAULT_SUBS["sensor"]))
        self.frm_logic.from_dict(sif.get("logic", DEFAULT_SUBS["logic"]))
        self.frm_fe.from_dict(sif.get("fe",    DEFAULT_SUBS["fe"]))
        self.frm_results.clear(); self._warn_lbl.config(text="")

    # ── Datablad + revisioner ───────────────────────────────────────────────────
    def _open_datasheet(self):
        self._save_current()
        if self._current<0: return
        sif = self._sifs[self._current]
        ds = sif.setdefault("datasheet", dict(DEFAULT_DATASHEET))
        def save(new_ds): sif["datasheet"] = new_ds
        SIFDatasheetDialog(self, ds, save)

    def _open_revisions(self):
        self._save_current()
        if self._current<0: return
        sif = self._sifs[self._current]
        revs = sif.setdefault("revisions",[])
        def save(new_revs): sif["revisions"] = new_revs
        RevisionDialog(self, revs, save)

    # ── Validering + beräkning ──────────────────────────────────────────────────
    def _calculate(self):
        self._save_current()
        warns = []
        for frm,lbl in [(self.frm_sensor,"Sensor"),(self.frm_logic,"Logic"),(self.frm_fe,"Slutelement")]:
            try:
                for w in validate_component(frm.get_params().component):
                    warns.append(f"{lbl}: {w}")
            except ValueError:
                pass
        self._warn_lbl.config(text="Varningar: " + " | ".join(warns) if warns else "")
        try:
            result = calc_sif(
                name=self.v_sif_name.get(),
                sensor=self.frm_sensor.get_params(),
                logic=self.frm_logic.get_params(),
                final_element=self.frm_fe.get_params(),
                sil_required=int(self.v_sil_req.get()),
            )
        except ValueError as e:
            messagebox.showerror("Inmatningsfel", str(e)); return
        self.frm_results.show(result); self.frm_chart.update(result)

    def _sensitivity(self):
        self._save_current()
        try:
            sp = self.frm_sensor.get_params()
            lp = self.frm_logic.get_params()
            fp = self.frm_fe.get_params()
        except ValueError as e:
            messagebox.showerror("Inmatningsfel", str(e)); return
        SensitivityWindow(self, sp, lp, fp, int(self.v_sil_req.get()))

    def _open_verification(self):
        if not _VDB_AVAILABLE:
            messagebox.showinfo(
                "Saknas",
                "verification_db.py saknas. Kopiera filen till samma mapp som sil.py.",
                parent=self)
            return
        self._save_current()
        VerificationDialog(self, self._sifs, self._current)

    # ── Fil ─────────────────────────────────────────────────────────────────────
    def _new_project(self):
        if messagebox.askyesno("Nytt projekt","Osparade ändringar försvinner."):
            self._sifs.clear(); self._lb.delete(0,"end")
            self._current=-1; self._file=None
            self.title("ProSa SIL-kalkylator"); self._new_sif()

    def _open_project(self):
        path = filedialog.askopenfilename(title="Öppna projekt",
            filetypes=[("SIL-projekt","*.silj"),("Alla","*.*")])
        if not path: return
        try:
            data = json.loads(Path(path).read_text(encoding="utf-8"))
            self._sifs = data.get("sifs",[])
            self._lb.delete(0,"end")
            for s in self._sifs: self._lb.insert("end", s.get("name","SIF"))
            self._current=0; self._lb.selection_set(0); self._load_sif(self._sifs[0])
            self._file=Path(path)
            self.title(f"ProSa SIL-kalkylator — {self._file.name}")
        except Exception as e:
            messagebox.showerror("Fel",f"Kunde inte öppna:\n{e}")

    def _save_project(self):
        self._save_current()
        if self._file: self._write_file(self._file)
        else: self._save_project_as()

    def _save_project_as(self):
        self._save_current()
        path = filedialog.asksaveasfilename(title="Spara projekt",defaultextension=".silj",
            filetypes=[("SIL-projekt","*.silj"),("Alla","*.*")])
        if not path: return
        self._file=Path(path); self._write_file(self._file)

    def _write_file(self, path: Path):
        path.write_text(json.dumps({"sifs":self._sifs},indent=2,ensure_ascii=False),
                        encoding="utf-8")
        self.title(f"ProSa SIL-kalkylator — {path.name}")

    def _export_html(self):
        self._save_current()
        path = filedialog.asksaveasfilename(title="Spara HTML-rapport",
            defaultextension=".html", filetypes=[("HTML","*.html"),("Alla","*.*")])
        if not path: return
        html = generate_html_report(self._sifs, str(self._file or ""))
        Path(path).write_text(html, encoding="utf-8")
        webbrowser.open(f"file:///{Path(path).as_posix()}")
        messagebox.showinfo("Klar",f"Öppnad i webbläsaren:\n{path}")

    def _export_excel(self):
        self._save_current()
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            messagebox.showerror("Saknar paket","pip install openpyxl"); return
        path = filedialog.asksaveasfilename(title="Exportera Excel",
            defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
        if not path: return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "SIF-beräkningar"
        ws.merge_cells("A1:N1")
        ws["A1"] = "ProSa SIL PFD-kalkylator — Beräkningsunderlag"
        ws["A1"].font = Font(bold=True,size=13)
        hf = PatternFill("solid",fgColor="1a5276"); hfont = Font(bold=True,color="FFFFFF")
        hdrs = ["SIF","SIL-krav","Delsystem","Arkitektur","Metod",
                "λ_D","DC","β","TI [h]","MTTR","PTC","SFF","SC","PFD_avg","SIL"]
        for c,h in enumerate(hdrs,1):
            cell=ws.cell(row=3,column=c,value=h); cell.fill=hf; cell.font=hfont
        from calc import SubsystemParams, Architecture, ComponentParams, calc_subsystem
        row=4
        for sif in self._sifs:
            for sk,sn in [("sensor","Sensor"),("logic","Logic solver"),("fe","Slutelement")]:
                d=sif.get(sk,{})
                try:
                    ct=d.get("comp_type","A (enkel)")[0]
                    sub=SubsystemParams(name=d.get("name",sn),
                        architecture=Architecture(d.get("arch","1oo1")),
                        calc_method=d.get("calc_method","markov"),
                        component=ComponentParams(
                            name=d.get("name",""),lambda_d=float(d.get("lambda_d",1e-6)),
                            dc=float(d.get("dc",0)),beta=float(d.get("beta",0.02)),
                            ti=float(d.get("ti",8760)),mttr=float(d.get("mttr",8)),
                            ptc=float(d.get("ptc",1.0)),sff=float(d.get("sff",0)),comp_type=ct,
                            sc=int(float(d.get("sc",0))),st=float(d.get("st",0)),
                            mission_time=float(d.get("mission_time",175200)),
                            pst_coverage=float(d.get("pst_coverage",0)),
                            pst_interval=float(d.get("pst_interval",720)),
                            ccf_model=d.get("ccf_model","beta")))
                    res=calc_subsystem(sub)
                    pfd_s=f"{res.pfd:.3e}"; sil_s=f"SIL {res.sil}" if res.sil>0 else "<SIL1"
                except Exception: pfd_s="FEL"; sil_s="—"
                ws.append([sif.get("name",""),f"SIL {sif.get('sil_req',2)}",sn,
                            d.get("arch",""),d.get("calc_method","markov"),
                            d.get("lambda_d",""),d.get("dc",""),d.get("beta",""),
                            d.get("ti",""),d.get("mttr",""),d.get("ptc",""),
                            d.get("sff",""),d.get("sc",""),pfd_s,sil_s]); row+=1
            ws.append([]); row+=1
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=13
        wb.save(path); messagebox.showinfo("Exporterat",f"Sparat:\n{path}")


if __name__ == "__main__":
    App().mainloop()
