"""Excel export for the complete HAZOP worksheet.

This module intentionally has no Qt dependency.  The worksheet UI and this
export both consume the database's sort order, but the exporter materialises
the same shared row grid for consequence categories, safeguards and
recommendations so an exported workbook remains editable and readable.
"""

import json
import math

from database import get_matrix, risk_info


_CHAIN_LABELS = {
    'loc': 'LOC – Utsläpp / läcka',
    'fire': 'Brand (pool fire / jet fire)',
    'flash_fire': 'Flash fire',
    'explosion': 'Explosion (VCE / BLEVE)',
    'toxic': 'Toxisk exponering',
    'environmental': 'Miljöutsläpp',
    'personnel': 'Personskador',
    'fatality': 'Dödsfall',
    'equipment': 'Utrustningsskador',
    'production': 'Driftstopp / produktionsbortfall',
    'custom': 'Övrigt (se text)',
}


def freq_axis_label(value):
    matrix = get_matrix()
    codes = matrix.get('x_codes', [])
    labels = matrix.get('x_labels', [])
    index = max(0, min(int(value) + 1, matrix.get('cols', 7) - 1))
    code = str(codes[index]).strip() if index < len(codes) else ''
    if code:
        return code
    label = labels[index] if index < len(labels) else f'F={value}'
    return label.split()[0] if label.strip() else f'F={value}'


def cons_axis_label(value):
    matrix = get_matrix()
    labels = matrix.get('y_labels', [])
    index = max(0, min(int(value) - 1, matrix.get('rows', 5) - 1))
    label = labels[index] if index < len(labels) else f'C={value}'
    return label.split()[0] if label.strip() else f'C={value}'


def parse_chain_from_json(raw):
    if not raw:
        return {}
    try:
        value = json.loads(raw)
        return value if isinstance(value, dict) else {}
    except (TypeError, ValueError, json.JSONDecodeError):
        return {}


def build_consequence_text(base, chain):
    parts = [str(base or '').strip()] if str(base or '').strip() else []
    for key, label in _CHAIN_LABELS.items():
        if chain.get(key):
            short = label.split('(')[0].strip().split(' – ')[-1].strip()
            parts.append(short)
    return ' → '.join(parts)


def _probability_reduction_steps(probability):
    try:
        value = float(probability)
    except (TypeError, ValueError):
        return 0
    if value <= 0 or value >= 100:
        return 0
    return int(math.floor(-math.log10(value / 100.0)))


def _rrf_reduction_steps(rrf):
    try:
        value = float(rrf)
    except (TypeError, ValueError):
        return 0
    return int(math.floor(math.log10(value))) if value > 1 else 0


def total_freq_reduction(base_frequency, safeguard_rrf, fa_active,
                         fa_probability, ignition_active, ignition_probability,
                         extra_factors):
    safeguard_steps = (int(math.log10(max(1, safeguard_rrf)))
                       if safeguard_rrf > 1 else 0)
    frequency_steps = safeguard_steps
    if fa_active:
        frequency_steps += _probability_reduction_steps(fa_probability)
    if ignition_active:
        frequency_steps += _probability_reduction_steps(ignition_probability)
    frequency_steps += sum(
        _rrf_reduction_steps(factor.get('rrf', 10))
        for factor in extra_factors if factor.get('active'))
    return (max(-1, base_frequency - frequency_steps),
            10 ** frequency_steps if frequency_steps > 0 else 1,
            frequency_steps)


def _number(value, number):
    value = str(value or '').strip()
    return f'{number}. {value}' if value else ''


def _worksheet_rows(db):
    """Yield rows in node -> deviation -> cause -> worksheet-row order."""
    nodes = [dict(node) for node in db.nodes()]
    deviations_by_node = {
        node_id: [dict(row) for row in rows]
        for node_id, rows in db.deviations_for_nodes(
            node['id'] for node in nodes).items()
    }
    deviation_ids = [dev['id'] for devs in deviations_by_node.values() for dev in devs]
    causes_by_deviation = {
        deviation_id: [dict(row) for row in rows]
        for deviation_id, rows in db.causes_for_deviations(deviation_ids).items()
    }
    cause_ids = [cause['id'] for causes in causes_by_deviation.values() for cause in causes]
    consequences_by_cause = {
        cause_id: [dict(row) for row in rows]
        for cause_id, rows in db.consequences_for_causes(cause_ids).items()
    }
    consequence_ids = [cons['id'] for conss in consequences_by_cause.values() for cons in conss]

    safeguards_by_consequence = {
        consequence_id: [dict(row) for row in rows]
        for consequence_id, rows in db.safeguards_for_consequences(
            consequence_ids).items()
    }
    categories_by_consequence = db.get_consequence_severities_for_consequences(
        consequence_ids)
    final_severity_by_consequence = (
        db.get_final_consequence_severities_for_consequences(consequence_ids))
    factors_by_consequence = db.reduction_factors_for_consequences(consequence_ids)
    recommendations_by_consequence = db.recommendations_for_consequences(consequence_ids)

    severity_ids = [category['id']
                    for categories in categories_by_consequence.values()
                    for category in categories]
    excluded_safeguards_by_severity = (
        db.get_severity_excluded_sgs_for_severities(severity_ids))
    excluded_factors_by_severity = (
        db.get_severity_excluded_reduction_factors_for_severities(severity_ids))
    safeguard_ids = [sg['id']
                     for safeguards in safeguards_by_consequence.values()
                     for sg in safeguards]
    excluded_causes_by_safeguard = (
        db.get_safeguard_excluded_causes_for_safeguards(safeguard_ids))

    deviation_numbers = {}
    for node in nodes:
        numbers_by_description = {}
        for deviation in deviations_by_node.get(node['id'], []):
            description = deviation.get('description') or ''
            numbers_by_description.setdefault(
                description, len(numbers_by_description) + 1)
            deviation_numbers[deviation['id']] = numbers_by_description[description]

    def cause_text(cause):
        try:
            equipment_ids = db.group_equipment_ids_for_cause(cause)
        except Exception:
            equipment_ids = []
        if len(equipment_ids) >= 2:
            return '\n'.join(db.group_cause_description_lines(cause, equipment_ids))
        return (cause.get('description') or '').strip()

    def consequence_text(consequence):
        chain = parse_chain_from_json(consequence.get('consequence_chain', ''))
        return (build_consequence_text(consequence.get('description') or '', chain)
                or (consequence.get('description') or '').strip())

    for node_number, node in enumerate(nodes, 1):
        node_label = _number(node.get('name'), node_number)
        for deviation in deviations_by_node.get(node['id'], []):
            deviation_label = _number(
                deviation.get('description'), deviation_numbers[deviation['id']])
            causes = [dict(cause) for cause in
                      causes_by_deviation.get(deviation['id'], [])]

            if not causes:
                yield _row(
                    node_label, deviation_label, '', '', '', None, '', '', '', None, '',
                    (node['id'], deviation['id'], None, None),
                )
                continue

            for cause_number, cause in enumerate(causes, 1):
                frequency = db.cause_frequency_level(cause)
                frequency_label = ('' if cause.get('frequency_cleared')
                                   else freq_axis_label(frequency))
                cause_label = _number(cause_text(cause), cause_number)
                consequences = [dict(cons) for cons in
                                consequences_by_cause.get(cause['id'], [])]

                if not consequences:
                    yield _row(
                        node_label, deviation_label, cause_label, frequency_label, '',
                        None, '', '', '', None, '',
                        (node['id'], deviation['id'], cause['id'], None),
                    )
                    continue

                for consequence_number, consequence in enumerate(consequences, 1):
                    consequence_id = consequence['id']
                    safeguards = [dict(sg) for sg in safeguards_by_consequence.get(
                        consequence_id, [])]
                    categories = [dict(category) for category in
                                  categories_by_consequence.get(consequence_id, [])]
                    recommendations = [dict(rec) for rec in
                                       recommendations_by_consequence.get(
                                           consequence_id, [])]
                    factors = [dict(factor) for factor in
                               factors_by_consequence.get(consequence_id, [])]
                    n_rows = max(len(safeguards), len(categories),
                                 len(recommendations), 1)
                    final_overrides = final_severity_by_consequence.get(
                        consequence_id, {})

                    active_factors = [factor for factor in factors
                                      if factor.get('active')]
                    aggregate_rrf = 1.0
                    for factor in active_factors:
                        try:
                            aggregate_rrf *= max(1.0, float(factor.get('rrf') or 1))
                        except (TypeError, ValueError):
                            continue
                    enablers_label = f'{len(active_factors)} ({aggregate_rrf:g})'

                    excluded_cause_safeguards = {
                        sg['id'] for sg in safeguards
                        if cause['id'] in excluded_causes_by_safeguard.get(
                            sg['id'], set())
                    }
                    consequence_label = _number(
                        consequence_text(consequence), consequence_number)

                    for row_index in range(n_rows):
                        sg_index = (
                            min(len(safeguards) - 1,
                                (row_index * len(safeguards)) // n_rows)
                            if safeguards else None)
                        safeguard = (safeguards[sg_index]
                                     if sg_index is not None else None)
                        category_index = (
                            min(len(categories) - 1,
                                (row_index * len(categories) + n_rows - 1) // n_rows)
                            if categories else None)
                        category = (categories[category_index]
                                    if category_index is not None else None)
                        recommendation = (recommendations[row_index]
                                          if row_index < len(recommendations) else None)

                        if category:
                            severity = category.get('severity') or 1
                            excluded_safeguards = excluded_safeguards_by_severity.get(
                                category['id'], set())
                            effective_safeguards = [
                                sg for sg in safeguards
                                if sg['id'] not in excluded_safeguards
                                and sg['id'] not in excluded_cause_safeguards]
                            sg_rrf = 1
                            for sg in effective_safeguards:
                                sg_rrf *= sg.get('rrf') or 1
                            effective_factors = [
                                factor for factor in factors
                                if factor['id'] not in excluded_factors_by_severity.get(
                                    category['id'], set())]
                            final_severity = final_overrides.get(
                                category.get('category_id'), severity)
                            category_short = (category.get('name') or '')[:3]
                            risk_before = (
                                f'{category_short}  {frequency_label}  '
                                f'{cons_axis_label(severity)}')
                        else:
                            severity = consequence.get('severity') or 1
                            sg_rrf = 1
                            for sg in safeguards:
                                if sg['id'] not in excluded_cause_safeguards:
                                    sg_rrf *= sg.get('rrf') or 1
                            effective_factors = factors
                            final_severity = severity
                            category_short = ''
                            risk_before = None

                        final_frequency, _, _ = total_freq_reduction(
                            frequency, sg_rrf, False, 10, False, 10,
                            effective_factors)
                        risk_after = (
                            f'{category_short}  {freq_axis_label(final_frequency)}  '
                            f'{cons_axis_label(final_severity)}'
                            if category else None)
                        safeguard_label = (
                            _number(safeguard.get('description'), sg_index + 1)
                            if safeguard else '')
                        rrf_label = str(safeguard.get('rrf') or 1) if safeguard else ''
                        recommendation_label = (
                            f"{int(recommendation['display_number']):03d}. "
                            f"{(recommendation.get('description') or '').strip()}"
                            if recommendation else '')

                        yield _row(
                            node_label, deviation_label, cause_label, frequency_label,
                            consequence_label,
                            (risk_before, frequency, severity) if risk_before else None,
                            safeguard_label, rrf_label, enablers_label,
                            (risk_after, final_frequency, final_severity)
                            if risk_after else None,
                            recommendation_label,
                            (node['id'], deviation['id'], cause['id'], consequence_id),
                        )


def _row(node, deviation, cause, frequency, consequence, risk_before,
         safeguard, rrf, enablers, risk_after, recommendation, merge_key):
    return {
        'values': [node, deviation, cause, frequency, consequence,
                   risk_before[0] if risk_before else '', safeguard, rrf,
                   enablers, risk_after[0] if risk_after else '', recommendation],
        'risk_before': risk_before,
        'risk_after': risk_after,
        'merge_key': merge_key,
    }


def export_worksheet_excel(db, filepath, merge_identical=False):
    """Export the complete worksheet in the reference workbook's layout."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        return False, 'openpyxl saknas.\nKör: pip install openpyxl'

    try:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'HAZOP Scenario'
        widths = [17, 23, 36, 10, 54, 19, 46, 9, 14, 19, 53]
        headers = [
            'Nod', 'Avvikelse', 'Orsak', 'Frekvens', 'Konsekvens',
            'Riskklass före barriärer', 'Barriär', 'RRF', 'Enablers',
            'Riskklass efter barriärer', 'Recommendation',
        ]
        for column, width in enumerate(widths, 1):
            worksheet.column_dimensions[
                openpyxl.utils.get_column_letter(column)].width = width

        thin = Side(border_style='thin', color='CBD5E1')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill('solid', fgColor='EEECE1')
        title_fill = PatternFill('solid', fgColor='1F2937')
        enabler_fill = PatternFill('solid', fgColor='F3F4F6')
        header_font = Font(name='Carlito', size=10, bold=True, color='111827')
        data_font = Font(name='Carlito', size=10, color='111827')
        centered = Alignment(horizontal='center', vertical='center', wrap_text=True)
        wrapped = Alignment(horizontal='left', vertical='top', wrap_text=True)

        worksheet.merge_cells('A1:K1')
        title = worksheet['A1']
        title.value = 'HAZOP Scenario – redigerbar sammanställning'
        title.fill = title_fill
        title.font = Font(name='Carlito', size=16, bold=True, color='FFFFFF')
        title.alignment = Alignment(horizontal='left', vertical='center')
        worksheet.row_dimensions[1].height = 30

        for column, value in enumerate(headers, 1):
            cell = worksheet.cell(3, column, value)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = centered
            cell.border = border
        for column, value in ((3, 'Orsak'), (4, 'Frekvens'),
                              (7, 'Barriär'), (8, 'RRF')):
            cell = worksheet.cell(4, column, value)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = centered
            cell.border = border
        for merged in ('A3:A4', 'B3:B4', 'C3:D3', 'E3:E4', 'F3:F4',
                       'G3:H3', 'I3:I4', 'J3:J4', 'K3:K4'):
            worksheet.merge_cells(merged)
        worksheet.row_dimensions[3].height = 25.05
        worksheet.row_dimensions[4].height = 25.05

        rows = list(_worksheet_rows(db))
        recommendation_order = []
        seen_recommendations = set()
        for row in rows:
            value = row['values'][10]
            if value:
                number, description = value.split('.', 1)
                if number not in seen_recommendations:
                    seen_recommendations.add(number)
                    recommendation_order.append((number, description.strip()))

        def colour(value, fallback):
            return str(value or fallback).lstrip('#').upper()

        def row_height(values):
            max_lines = 1
            for value, width in zip(values, widths):
                for line in str(value or '').splitlines() or ['']:
                    max_lines = max(max_lines,
                                    1 + len(line) // max(12, int(width * 1.25)))
            return min(180, max(38, 15 * max_lines + 8))

        for excel_row, row in enumerate(rows, 5):
            values = row['values']
            for column, value in enumerate(values, 1):
                cell = worksheet.cell(excel_row, column, value)
                cell.font = data_font
                cell.border = border
                cell.alignment = centered if column in (4, 6, 8, 9, 10) else wrapped
                if column == 9:
                    cell.fill = enabler_fill
                    cell.font = Font(name='Carlito', size=10, bold=True,
                                     color='111827')

            for column, risk in ((6, row['risk_before']),
                                 (10, row['risk_after'])):
                cell = worksheet.cell(excel_row, column)
                if risk:
                    _, frequency, severity = risk
                    _, background, foreground = risk_info(frequency, severity)
                    cell.fill = PatternFill('solid',
                                            fgColor=colour(background, 'FFFFFF'))
                    cell.font = Font(name='Carlito', size=10, bold=True,
                                     color=colour(foreground, '111827'))
                else:
                    cell.fill = PatternFill('solid', fgColor='FFFFFF')
                    cell.font = Font(name='Carlito', size=10, bold=True,
                                     color='8D9299')
            worksheet.row_dimensions[excel_row].height = row_height(values)

        if merge_identical and rows:
            # These are hierarchy columns only.  Risk, barriers, enablers and
            # recommendations remain independent rows within the shared grid.
            merge_key_columns = (0, 1, 2, 2, 3)
            for column, key_column in enumerate(merge_key_columns, 1):
                start = 5
                previous_value = worksheet.cell(start, column).value
                previous_key = rows[0]['merge_key'][key_column]
                for row_number in range(6, worksheet.max_row + 2):
                    index = row_number - 5
                    current_value = (worksheet.cell(row_number, column).value
                                     if row_number <= worksheet.max_row else object())
                    current_key = (rows[index]['merge_key'][key_column]
                                   if index < len(rows) else object())
                    if current_value != previous_value or current_key != previous_key:
                        if (row_number - start > 1
                                and previous_value not in (None, '')):
                            worksheet.merge_cells(
                                start_row=start, start_column=column,
                                end_row=row_number - 1, end_column=column)
                        start = row_number
                        previous_value = current_value
                        previous_key = current_key

        recommendations = workbook.create_sheet('Rekommendationer')
        recommendations.merge_cells('A1:B1')
        recommendations['A1'] = 'Unika rekommendationer'
        recommendations['A1'].fill = title_fill
        recommendations['A1'].font = Font(name='Carlito', size=14,
                                           bold=True, color='FFFFFF')
        recommendations['A1'].alignment = Alignment(horizontal='left', vertical='center')
        for column, value in ((1, 'ID'), (2, 'Recommendation')):
            cell = recommendations.cell(3, column, value)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = centered
            cell.border = border
        catalog = {
            f"{int(rec['display_number']):03d}": dict(rec)
            for rec in db.all_recommendations()
        }
        ordered_numbers = [number for number, _ in recommendation_order]
        ordered_numbers.extend(
            number for number in sorted(catalog, key=lambda item: int(item))
            if number not in ordered_numbers)
        for row_number, number in enumerate(ordered_numbers, 4):
            description = next(
                (text for rec_number, text in recommendation_order
                 if rec_number == number),
                (catalog.get(number) or {}).get('description', ''),
            )
            for column, value in ((1, number), (2, description or '')):
                cell = recommendations.cell(row_number, column, value)
                cell.font = data_font
                cell.border = border
                cell.alignment = centered if column == 1 else wrapped
            recommendations.row_dimensions[row_number].height = row_height(
                [number, description])
        recommendations.column_dimensions['A'].width = 12
        recommendations.column_dimensions['B'].width = 92

        workbook.save(filepath)
        return True, ''
    except Exception as error:
        return False, str(error)
