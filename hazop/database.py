#!/usr/bin/env python3
"""Database layer — SQLite schema, risk matrix and Database class (split out
of hazop.py 2026-08-17, see NOTES.md "Förenkla koden + dela upp hazop.py i
fler filer"). No Qt — importable standalone."""

import json
import logging
import sqlite3
import datetime
import re
import html
import os
import tempfile
from contextlib import contextmanager
from pathlib import Path

from constants import (
    DEVIATION_TYPES, _app_dir,
    NODE_T, CAUSE_T, CONS_T, SG_T, DEV_T, SYSTEM_T,
)
from lopa_models import (
    DEFAULT_SAFEGUARD_TYPES,
    calculate_lopa,
    normalise_lopa_config,
)


def append_tag_to_text(description: str, tag: str) -> str:
    """Append an equipment tag to the end of a free-text description with
    a single separating space (2026-08-09, see NOTES.md) — used when an
    equipment marker is drag-and-dropped onto a KON/SG cell, building a
    running sentence ("hög nivå i" + drop TA-1 -> "hög nivå i TA-1", then
    "... => överbreddning till" + drop TA-2 -> "... => överbreddning
    till TA-2") instead of overwriting a separate tag field. Starting
    from the still-untouched default placeholder text replaces it
    outright rather than appending to boilerplate nobody wrote."""
    description = description or ''
    tag = (tag or '').strip()
    if not tag:
        return description
    stripped = description.strip()
    if not stripped or stripped in ('Ny konsekvens', 'Ny safeguard', 'Ny orsak'):
        return tag
    if description[-1].isspace():
        return description + tag
    return description + ' ' + tag


def normalize_arrows(text: str) -> str:
    """Convert ASCII arrows typed by users to the UI's standard arrow."""
    if not text:
        return text or ''
    return re.sub(r'\s*(?:=>|->)\s*', ' → ', str(text)).strip()


def parse_tag_refs(raw: str) -> list:
    """Decode tagged_refs (comma-separated, order preserved) into a list
    of tag strings — every tag ever drag-appended into a KON/SG cell's
    free text, used to bold those substrings when rendering the cell
    (2026-08-09, see NOTES.md)."""
    if not raw:
        return []
    return [t for t in (s.strip() for s in raw.split(',')) if t]


def add_tag_ref(raw: str, tag: str) -> str:
    """Append tag to the tagged_refs list, deduplicated, order preserved
    (the most recent drop moves to the end)."""
    tag = (tag or '').strip()
    if not tag:
        return raw or ''
    refs = [t for t in parse_tag_refs(raw) if t != tag]
    refs.append(tag)
    return ','.join(refs)


def _tag_letter_prefix(tag: str) -> str:
    """Return the last alphabetic tag code, ignoring serial digits.

    This deliberately lives in the database layer as a tiny dependency-free
    fallback.  Tag memory must remain usable before the Qt helper modules have
    finished importing (and compound tags such as ``E1.M1.PU103`` should be
    keyed by ``PU`` rather than by their numbers).
    """
    value = str(tag or '').upper()
    tokens = re.findall(r'[A-ZÅÄÖ]+', value)
    return tokens[-1] if tokens else ''


# ══════════════════════════════════════════════════════════════════════════════
# DATABASE
# ══════════════════════════════════════════════════════════════════════════════

DB_PATH = _app_dir() / "hazop_project.db"

SCHEMA = """
PRAGMA foreign_keys = ON;

CREATE TABLE IF NOT EXISTS systems (
    id         INTEGER PRIMARY KEY AUTOINCREMENT,
    name       TEXT NOT NULL DEFAULT 'Nytt system',
    sort_order INTEGER DEFAULT 0
);

CREATE TABLE IF NOT EXISTS nodes (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    name        TEXT NOT NULL DEFAULT 'Ny nod',
    description TEXT DEFAULT '',
    pid_ref     TEXT DEFAULT '',
    media       TEXT DEFAULT '',
    pressure    TEXT DEFAULT '',
    temperature TEXT DEFAULT '',
    sort_order  INTEGER NOT NULL DEFAULT 0
);

CREATE TABLE IF NOT EXISTS causes (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    node_id     INTEGER NOT NULL REFERENCES nodes(id) ON DELETE CASCADE,
    description TEXT NOT NULL DEFAULT 'Ny orsak',
    likelihood  INTEGER NOT NULL DEFAULT 1,
    frequency_cleared INTEGER NOT NULL DEFAULT 0,
    sort_order  INTEGER NOT NULL DEFAULT 0
);

CREATE TABLE IF NOT EXISTS consequences (
    id                INTEGER PRIMARY KEY AUTOINCREMENT,
    cause_id          INTEGER NOT NULL REFERENCES causes(id) ON DELETE CASCADE,
    description       TEXT NOT NULL DEFAULT 'Ny konsekvens',
    severity          INTEGER NOT NULL DEFAULT 1,
    category          TEXT DEFAULT '',
    consequence_chain TEXT DEFAULT '',
    sort_order        INTEGER NOT NULL DEFAULT 0
);

CREATE TABLE IF NOT EXISTS safeguards (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    consequence_id  INTEGER NOT NULL REFERENCES consequences(id) ON DELETE CASCADE,
    description     TEXT NOT NULL DEFAULT '',
    rrf             INTEGER NOT NULL DEFAULT 1,
    source_id       INTEGER DEFAULT NULL,
    sort_order      INTEGER NOT NULL DEFAULT 0
);

CREATE TABLE IF NOT EXISTS reduction_factors (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    consequence_id  INTEGER NOT NULL REFERENCES consequences(id) ON DELETE CASCADE,
    description     TEXT NOT NULL DEFAULT '',
    rrf             INTEGER NOT NULL DEFAULT 10,
    active          INTEGER NOT NULL DEFAULT 1
);

-- An enabler can be active for a consequence but intentionally not apply to
-- one of its assessed consequence categories.  This mirrors the existing
-- safeguard/category exclusion model without mixing the two object types.
CREATE TABLE IF NOT EXISTS reduction_factor_severity_exclusions (
    severity_id          INTEGER NOT NULL,
    reduction_factor_id  INTEGER NOT NULL,
    PRIMARY KEY (severity_id, reduction_factor_id)
);

-- 2026-08-25: replaces the old actions table (one row per consequence,
-- no reuse) with a shared catalog + many-to-many link, so the same
-- recommendation text can be linked to several consequences instead of
-- being duplicated (see NOTES.md "Rekommendationshantering — delad
-- katalog med återanvändning"). `id` is the stable internal key; the
-- user-visible recommendation number is a separate compact display sequence.
CREATE TABLE IF NOT EXISTS recommendations (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    display_number  INTEGER NOT NULL DEFAULT 0,
    description     TEXT NOT NULL DEFAULT '',
    responsible     TEXT DEFAULT '',
    due_date        TEXT DEFAULT '',
    status          TEXT DEFAULT 'Öppen'
);

CREATE TABLE IF NOT EXISTS consequence_recommendations (
    consequence_id    INTEGER NOT NULL REFERENCES consequences(id) ON DELETE CASCADE,
    recommendation_id INTEGER NOT NULL REFERENCES recommendations(id) ON DELETE CASCADE,
    PRIMARY KEY (consequence_id, recommendation_id)
);

CREATE TABLE IF NOT EXISTS app_config (
    key   TEXT PRIMARY KEY,
    value TEXT
);

-- A risk-matrix template can change the semantic meaning of every stored
-- frequency/consequence level.  Keep an auditable record of each deliberate
-- conversion, separate from the editable current matrix configuration.
CREATE TABLE IF NOT EXISTS risk_matrix_migrations (
    id             INTEGER PRIMARY KEY AUTOINCREMENT,
    created_at     TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    source_matrix  TEXT NOT NULL,
    target_matrix  TEXT NOT NULL,
    mapping_json   TEXT NOT NULL,
    backup_path    TEXT NOT NULL DEFAULT ''
);

CREATE TABLE IF NOT EXISTS consequence_categories (
    id         INTEGER PRIMARY KEY AUTOINCREMENT,
    name       TEXT NOT NULL,
    sort_order INTEGER DEFAULT 0
);

CREATE TABLE IF NOT EXISTS component_types (
    id         INTEGER PRIMARY KEY AUTOINCREMENT,
    name       TEXT NOT NULL,
    sort_order INTEGER DEFAULT 0
);

CREATE TABLE IF NOT EXISTS failure_modes (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    component_id INTEGER NOT NULL REFERENCES component_types(id) ON DELETE CASCADE,
    description  TEXT NOT NULL DEFAULT '',
    freq_per_year REAL DEFAULT NULL,
    sort_order   INTEGER DEFAULT 0
);

CREATE TABLE IF NOT EXISTS reduction_factors (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    consequence_id  INTEGER NOT NULL REFERENCES consequences(id) ON DELETE CASCADE,
    description     TEXT NOT NULL DEFAULT '',
    rrf             INTEGER NOT NULL DEFAULT 10,
    active          INTEGER NOT NULL DEFAULT 1
);

CREATE TABLE IF NOT EXISTS equipment_types (
    prefix          TEXT PRIMARY KEY,
    equipment_type  TEXT NOT NULL,
    display_name    TEXT DEFAULT ''
);

CREATE TABLE IF NOT EXISTS tag_database (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    tag_code    TEXT NOT NULL,
    name_sv     TEXT DEFAULT '',
    name_en     TEXT DEFAULT '',
    category    TEXT DEFAULT '',
    standard    TEXT DEFAULT '',
    source      TEXT DEFAULT 'excel',
    active      INTEGER DEFAULT 1
);

CREATE TABLE IF NOT EXISTS tag_database_settings (
    key   TEXT PRIMARY KEY,
    value TEXT
);

CREATE TABLE IF NOT EXISTS pid_identified_tags (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    tag_code     TEXT NOT NULL UNIQUE,
    examples     TEXT DEFAULT '',
    name_sv      TEXT DEFAULT '',
    comp_type    TEXT DEFAULT '',
    confirmed    INTEGER DEFAULT 0,
    source       TEXT DEFAULT 'scan'
);

CREATE TABLE IF NOT EXISTS equipment_catalog (
    id             INTEGER PRIMARY KEY AUTOINCREMENT,
    tag            TEXT NOT NULL,
    original_tag   TEXT DEFAULT '',
    prefix         TEXT DEFAULT '',
    pid_page       INTEGER DEFAULT 0,
    equipment_type TEXT DEFAULT '',
    description    TEXT DEFAULT '',
    is_ocr         INTEGER DEFAULT 0,
    include        INTEGER DEFAULT 1
);
"""

# Frequency axis: F=-1..5 (7 levels, logarithmic events/year)
# Consequence axis: C=1..5 (5 levels)
#
# A matrix template is deliberately self-contained.  The category key is a
# stable template identifier, the name is what users see, and the colour is
# metadata for the category (separate from the colour of a risk cell).  The
# severity descriptions belong to the template as well; projects get their
# own category ids only when a template is applied.
DEFAULT_TEMPLATE_CATEGORIES = [
    {'key': 'person', 'name': 'Person', 'color': '#2563eb',
     'descriptions': ['Första hjälpen', 'Medicinsk behandling',
                      'Allvarlig personskada', 'Enstaka dödsfall',
                      'Flera dödsfall']},
    {'key': 'miljo', 'name': 'Miljö', 'color': '#16a34a',
     'descriptions': ['Ingen bestående påverkan', 'Lokal och kortvarig påverkan',
                      'Begränsad sanering', 'Omfattande sanering',
                      'Långvarig eller omfattande miljöskada']},
    {'key': 'ekonomi', 'name': 'Ekonomi', 'color': '#d97706',
     'descriptions': ['Obetydlig kostnad', 'Mindre kostnad', 'Betydande kostnad',
                      'Stor ekonomisk skada', 'Mycket stor ekonomisk skada']},
    {'key': 'anlaggning', 'name': 'Anläggning', 'color': '#7c3aed',
     'descriptions': ['Ingen skada', 'Mindre skada', 'Begränsad produktionstörning',
                      'Allvarlig anläggningsskada', 'Förlust av anläggning eller långvarigt stopp']},
    {'key': 'rykte', 'name': 'Rykte', 'color': '#475569',
     'descriptions': ['Ingen extern påverkan', 'Lokal uppmärksamhet',
                      'Regional uppmärksamhet', 'Nationell uppmärksamhet',
                      'Internationell eller långvarig uppmärksamhet']},
]

DEFAULT_MATRIX = {
    'rows': 5,   # consequence rows, index 0 = C1 (lowest)
    'cols': 7,   # frequency columns, index 0 = F-1 (lowest)
    'x_axis': 'frequency',
    # Short, editable level identifiers shown directly on the matrix.  The
    # longer explanatory text stays in x_labels/y_labels and is shown in the
    # Axlar editor and tooltips.
    'x_codes': ['F-1', 'F0', 'F1', 'F2', 'F3', 'F4', 'F5'],
    'y_codes': ['1', '2', '3', '4', '5'],
    'x_labels': [
        'Otänkbar (<1/100 000 år)',
        'Extremt sällan (1/100 000 år)',
        'Sällan (1/10 000 år)',
        'Osannolik (1/1 000 år)',
        'Möjlig (1/100 år)',
        'Trolig (1–10 år)',
        'Frekvent (>1/år)',
    ],
    'y_labels': [
        'Försumbar',
        'Liten',
        'Måttlig',
        'Allvarlig',
        'Katastrofal',
    ],
    'cell_colors': [
        # C=1: F-1 → F5
        ['#27ae60', '#27ae60', '#27ae60', '#27ae60', '#f39c12', '#e67e22', '#e74c3c'],
        # C=2
        ['#27ae60', '#27ae60', '#27ae60', '#f39c12', '#e67e22', '#e74c3c', '#e74c3c'],
        # C=3
        ['#27ae60', '#27ae60', '#f39c12', '#e67e22', '#e74c3c', '#e74c3c', '#e74c3c'],
        # C=4
        ['#27ae60', '#f39c12', '#e67e22', '#e74c3c', '#e74c3c', '#e74c3c', '#e74c3c'],
        # C=5
        ['#f39c12', '#e67e22', '#e74c3c', '#e74c3c', '#e74c3c', '#e74c3c', '#e74c3c'],
    ],
    'cell_labels': [
        ['Låg',    'Låg',    'Låg',    'Låg',    'Medium', 'Hög',    'Kritisk'],
        ['Låg',    'Låg',    'Låg',    'Medium', 'Hög',    'Kritisk','Kritisk'],
        ['Låg',    'Låg',    'Medium', 'Hög',    'Kritisk','Kritisk','Kritisk'],
        ['Låg',    'Medium', 'Hög',    'Kritisk','Kritisk','Kritisk','Kritisk'],
        ['Medium', 'Hög',    'Kritisk','Kritisk','Kritisk','Kritisk','Kritisk'],
    ],
    'consequence_categories': DEFAULT_TEMPLATE_CATEGORIES,
}

# ── Risk Matrix Caching with Automatic Invalidation ──────────────────────────
# The risk matrix is read from the database and cached for performance. When a
# matrix is updated via set_risk_matrix(), the cache must be invalidated to ensure
# all subsequent get_matrix() calls reflect the new state. Using a dedicated
# manager class ensures invalidation is automatic when the setter is called.

class _RiskMatrixCache:
    """Manager for risk matrix caching with automatic invalidation support."""
    def __init__(self):
        self._current_matrix = None
        self._db = None

    def load(self, db):
        """Load and cache the risk matrix from database."""
        self._db = db
        cfg = db.get_risk_matrix()
        if cfg:
            self._current_matrix = _normalise_matrix(cfg)
        else:
            self._current_matrix = DEFAULT_MATRIX

    def invalidate(self):
        """Invalidate the cache; next get() will reload from DB."""
        self._current_matrix = None

    def get(self):
        """Get the current cached matrix (or DEFAULT if not loaded)."""
        return self._current_matrix or DEFAULT_MATRIX

    def reload_from_db(self):
        """Force reload from database (used after set_risk_matrix)."""
        if self._db:
            self.load(self._db)


_risk_matrix_cache = _RiskMatrixCache()


def _normalise_matrix(cfg: dict) -> dict:
    """Ensure a stored matrix config is internally consistent.

    Pads level codes, x_labels / y_labels and cell arrays to match rows/cols.
    Used once on load so the rest of the code can trust the structure.
    """
    rows = int(cfg.get('rows', 5))
    cols = int(cfg.get('cols', 7))

    # Pad or trim x_labels
    x = list(cfg.get('x_labels', []))
    while len(x) < cols:
        x.append(f'F{len(x) - 1}')
    cfg['x_labels'] = x[:cols]

    # Pad or trim y_labels
    y = list(cfg.get('y_labels', []))
    while len(y) < rows:
        y.append(f'C{len(y) + 1}')
    cfg['y_labels'] = y[:rows]

    # `x_codes`/`y_codes` were added after projects already contained the
    # complete label text. Derive a conservative code from that text when
    # loading an older project, then store the code separately from its
    # description from here on.
    def _code_from_label(label, fallback):
        text = str(label or '').strip()
        if not text:
            return fallback
        for separator in (' – ', ' — ', ' - '):
            if separator in text:
                return text.split(separator, 1)[0].strip() or fallback
        return text.split()[0] if text.split() else fallback

    x_codes = list(cfg.get('x_codes', []))
    while len(x_codes) < cols:
        index = len(x_codes)
        x_codes.append(_code_from_label(
            cfg['x_labels'][index] if index < len(cfg['x_labels']) else '',
            f'F{index - 1}'))
    cfg['x_codes'] = [str(code) for code in x_codes[:cols]]

    y_codes = list(cfg.get('y_codes', []))
    while len(y_codes) < rows:
        index = len(y_codes)
        y_codes.append(_code_from_label(
            cfg['y_labels'][index] if index < len(cfg['y_labels']) else '',
            str(index + 1)))
    cfg['y_codes'] = [str(code) for code in y_codes[:rows]]

    # Pad or trim cell_colors / cell_labels
    def _pad_grid(grid, default_val):
        result = []
        for r in range(rows):
            row = list(grid[r]) if r < len(grid) else []
            while len(row) < cols:
                row.append(default_val)
            result.append(row[:cols])
        return result

    cfg['cell_colors']    = _pad_grid(cfg.get('cell_colors', []), '#27ae60')
    cfg['cell_labels']    = _pad_grid(cfg.get('cell_labels', []), 'Låg')
    cfg['cell_fg_colors'] = _pad_grid(cfg.get('cell_fg_colors', []), '#ffffff')
    cfg['rows'] = rows
    cfg['cols'] = cols
    categories = cfg.get('consequence_categories')
    if not isinstance(categories, list):
        categories = []
    normalised_categories = []
    seen_keys = set()
    for index, raw in enumerate(categories):
        raw = raw if isinstance(raw, dict) else {}
        key = str(raw.get('key') or raw.get('name') or f'category-{index + 1}').strip()
        key = key.casefold().replace(' ', '-').replace('_', '-') or f'category-{index + 1}'
        if key in seen_keys:
            key = f'{key}-{index + 1}'
        seen_keys.add(key)
        descriptions = list(raw.get('descriptions') or [])
        descriptions = [str(descriptions[i]) if i < len(descriptions) else ''
                        for i in range(rows)]
        normalised_categories.append({
            'key': key,
            'name': str(raw.get('name') or key),
            'color': str(raw.get('color') or '#64748b'),
            'descriptions': descriptions,
        })
    cfg['consequence_categories'] = normalised_categories
    # LOPA shares this project/template risk matrix.  Its own settings are
    # nested here so copied matrix templates and locked LOPA revisions retain
    # the category names, colours and TEL scale that were actually used.
    cfg['lopa'] = normalise_lopa_config(cfg.get('lopa'), cfg)
    return cfg


def load_matrix(db):
    """Load risk matrix from database into the cache."""
    _risk_matrix_cache.load(db)


def get_matrix():
    """Get the currently cached risk matrix."""
    return _risk_matrix_cache.get()


def risk_info(frequency, consequence):
    """Return (label, bg_color, fg_color) from matrix lookup.

    Data is always stored as cell_colors[cons_idx][freq_idx].
    x_axis only controls display orientation — not data access.
    """
    cfg   = get_matrix()
    rows  = cfg.get('rows', 5)   # consequence levels
    cols  = cfg.get('cols', 7)   # frequency levels
    c_idx = max(0, min(int(consequence) - 1, rows - 1))   # C=1 → 0
    f_idx = max(0, min(int(frequency)  + 1, cols - 1))   # F=-1 → 0
    try:
        color = cfg['cell_colors'][c_idx][f_idx]   # always [cons][freq]
        label = cfg['cell_labels'][c_idx][f_idx]
        if not color:
            color = '#27ae60'
        if not label:
            label = 'Låg'
    except (IndexError, KeyError, TypeError):
        color, label = '#27ae60', 'Låg'
    try:
        fg = cfg['cell_fg_colors'][c_idx][f_idx] or '#ffffff'
    except (IndexError, KeyError, TypeError):
        fg = '#ffffff'
    return label, color, fg


# Frequency F=-1..5, stored as integer in causes.likelihood.
# FREQ_LABELS/freq_to_idx/idx_to_freq now live in pid_viewer.py (imported
# above) since EquipmentDeviationBar needs them too and pid_viewer.py
# cannot import back from hazop.py without a circular import.
_FREQ_VALUES = [-1, 0, 1, 2, 3, 4, 5]

# Default frequency boundaries (events/year) between each F-column.
# 6 boundaries for 7 columns (F=-1..F5).
# freq < boundaries[0]       → F=-1
# boundaries[i] <= freq < boundaries[i+1] → F=i
# freq >= boundaries[5]      → F=5
DEFAULT_FREQ_BOUNDARIES = [1e-5, 1e-4, 1e-3, 1e-2, 0.1, 1.0]


def freq_to_f_level(freq_per_year, boundaries=None) -> int:
    """Convert numeric frequency (events/year) to F-level (-1..5).

    0.05/year → F=3  (10-100 year interval)
    0.5/year  → F=4  (1-10 year interval)
    """
    if boundaries is None:
        cfg = get_matrix()
        boundaries = cfg.get('freq_boundaries', DEFAULT_FREQ_BOUNDARIES)
    boundaries = sorted(float(b) for b in boundaries)
    if not freq_per_year or freq_per_year <= 0:
        return -1
    for i, b in enumerate(boundaries):
        if float(freq_per_year) < b:
            return i - 1
    return len(boundaries) - 1   # above all → F=5


# Component-specific standard causes seeded on first run.
# comp_type must match keys in COMPONENT_TYPES (pid_viewer.py).
# ── Standardorsaker per avvikelse och objekt ──────────────────────────────────
# Format: {avvikelse: {objektnamn: [(beskrivning, frekvens_per_år | None)]}}
# Frekvenser är typvärden (OREDA / processsindustri) — justera per projekt.
# Generiska beskrivningar: täcker orsaken, inte det specifika scenariot.
_COMP_STD_CAUSES = {
    # ── Lågt flöde ────────────────────────────────────────────────────────────
    "Lågt flöde": {
        "Manuell ventil":     [("Ventil stängd / delvis stängd",       1e-3),
                               ("Ventil blockerad (igensättning)",      5e-4),
                               ("Blind platta / blindning kvarglömd",   1e-4)],
        "On-off ventil":      [("Ventil felar stängd (fail-closed)",    1e-2),
                               ("Ventil fastnar i stängt läge",         5e-3),
                               ("Manöversignal uteblir",                1e-2)],
        "Reglerventil":       [("Reglerventil felar stängd",            2e-2),
                               ("Ventil fastnar / stiction",            1e-2),
                               ("Felaktig styrsignal — lågt utflöde",   5e-3)],
        "Backventil":         [("Backventil fastnar stängd",            1e-2),
                               ("Backventil monterad baklänges",        1e-4)],
        "Pump":               [("Pump stopp",                           2e-2),
                               ("Reducerad pumpkapacitet",              1e-2),
                               ("Kavitation",                           5e-3),
                               ("Inlopp blockerat",                     5e-3)],
        "Kompressor / fläkt": [("Kompressor / fläkt stopp",            2e-2),
                               ("Reducerad kapacitet",                  1e-2),
                               ("Inloppsfilter igensatt",               5e-2)],
        "Filter / sil":       [("Filter / sil igensatt",               0.1),
                               ("Filterelement felaktigt monterat",     1e-3)],
        "Värmeväxlare / kylare / värmare": [
                               ("Rör igensatta — fouling",              5e-2),
                               ("Vakuumbrott / tömning",                1e-3)],
        "Tank / kärl / kolonn":[("Låg nivå i matningskärl",            5e-2),
                               ("Utlopp stängt / nivåstyrning",         1e-2)],
        "Rörledning / slang": [("Igensatt rörledning",                  5e-3),
                               ("Luftlås / hydrater / is",              1e-3)],
        "Instrument":         [("Flödesgivare felar — styrventil stänger", 0.1),
                               ("Börvärde felaktigt inställt",          1e-2)],
        "Styrsystem / PLC / DCS": [("Styrsignal felar — ventil stänger", 5e-3),
                               ("Kommunikationsavbrott",                1e-2)],
    },

    # ── Högt flöde ────────────────────────────────────────────────────────────
    "Högt flöde": {
        "Manuell ventil":     [("Ventil öppnad felaktigt",              1e-3),
                               ("Bypassventil öppnad",                  5e-4)],
        "On-off ventil":      [("Ventil felar öppen (fail-open)",       1e-2),
                               ("Ventil fastnar i öppet läge",          5e-3)],
        "Reglerventil":       [("Reglerventil felar öppen",             2e-2),
                               ("Felaktig styrsignal — högt utflöde",   5e-3)],
        "Pump":               [("Pumpkapacitet för hög",                5e-3),
                               ("Frekvensomformare — fel varvtal",      1e-2)],
        "Kompressor / fläkt": [("Kompressor — för hög kapacitet",       5e-3)],
        "Tank / kärl / kolonn":[("Övertryck driver högre flöde",        1e-2)],
        "Instrument":         [("Flödesgivare felar — styrventil öppnar", 0.1),
                               ("Börvärde felaktigt högt",              1e-2)],
        "Styrsystem / PLC / DCS": [("Styrsignal felar — ventil öppnar", 5e-3)],
    },

    # ── Högt tryck ────────────────────────────────────────────────────────────
    "Högt tryck": {
        "Manuell ventil":     [("Utloppsventil stängd",                 1e-3),
                               ("Ventil blockerad",                     5e-4)],
        "On-off ventil":      [("Utloppsventil felar stängd",           1e-2),
                               ("Ventil fastnar stängd på utlopp",      5e-3)],
        "Reglerventil":       [("Reglerventil på utlopp felar stängd",  2e-2),
                               ("Felaktig tryckreglering",              5e-3)],
        "Pump":               [("Pump deadhead — utlopp blockerat",     5e-3)],
        "Värmeväxlare / kylare / värmare": [
                               ("Kylningsbortfall",                     5e-3),
                               ("Termisk expansion utan ventilering",   1e-3)],
        "Tank / kärl / kolonn":[("Blockerat avluftningssystem",         5e-4)],
        "Säkerhetsventil / sprängbleck": [
                               ("Säkerhetsventil felar stängd",         1e-3),
                               ("Sprängbleck defekt",                   5e-4)],
        "Instrument":         [("Trycktransmitter felar — styrventil stänger", 0.1),
                               ("Börvärde tryckreglering felaktigt",    1e-2)],
        "Styrsystem / PLC / DCS": [("Tryckreglering felar",             5e-3)],
        "Rörledning / slang": [("Blockerad utloppsledning",             5e-4)],
        "Kompressor / fläkt": [("Kompressorsurge",                      1e-2)],
        "Backventil":         [("Backventil blockerar utflöde",         5e-3)],
        "Filter / sil":       [("Filter igensatt — tryckstegring uppströms", 0.1)],
    },

    # ── Lågt tryck ────────────────────────────────────────────────────────────
    "Lågt tryck": {
        "Manuell ventil":     [("Dräneringsventil öppnad",              5e-4),
                               ("Läckage via öppen ventil",             1e-3)],
        "On-off ventil":      [("Utloppsventil felar öppen",            1e-2),
                               ("Avblåsningsventil fastnar öppen",      5e-3)],
        "Reglerventil":       [("Reglerventil felar öppen",             2e-2)],
        "Pump":               [("Pump stopp — tryckfall",               2e-2)],
        "Rörledning / slang": [("Rörläckage / slangbrott",              5e-4),
                               ("Packningsläckage",                     1e-3)],
        "Fläns / koppling / packning": [
                               ("Packningsläckage",                     2e-3),
                               ("Flänsläckage",                         5e-4)],
        "Säkerhetsventil / sprängbleck": [
                               ("Säkerhetsventil öppnar för tidigt",    1e-3),
                               ("Sprängbleck utlöst",                   1e-4)],
        "Instrument":         [("Tryckmätare felar — styrventil öppnar", 0.1)],
        "Tank / kärl / kolonn":[("Kärl dränerat",                       5e-3)],
    },

    # ── Hög nivå ──────────────────────────────────────────────────────────────
    "Hög nivå": {
        "Manuell ventil":     [("Utloppsventil stängd",                 1e-3),
                               ("Inloppsventil öppnad utan utlopp",     5e-4)],
        "On-off ventil":      [("Utloppsventil felar stängd",           1e-2),
                               ("Inloppsventil felar öppen",            1e-2)],
        "Reglerventil":       [("Utloppsreglering felar stängd",        2e-2),
                               ("Inloppsreglering felar öppen",         1e-2)],
        "Pump":               [("Utloppspump stopp",                    2e-2)],
        "Instrument":         [("Nivågivare felar — reglering stänger utlopp", 0.1),
                               ("Börvärde nivå felaktigt",              1e-2)],
        "Tank / kärl / kolonn":[("Inflöde > utflöde",                  5e-3)],
        "Styrsystem / PLC / DCS": [("Nivåreglering felar",              5e-3)],
        "Backventil":         [("Backventil läcker — backflöde till kärl", 5e-3)],
    },

    # ── Låg nivå ──────────────────────────────────────────────────────────────
    "Låg nivå": {
        "Manuell ventil":     [("Inloppsventil stängd",                 1e-3),
                               ("Dräneringsventil öppnad",              5e-4)],
        "On-off ventil":      [("Inloppsventil felar stängd",           1e-2),
                               ("Utloppsventil felar öppen",            1e-2)],
        "Reglerventil":       [("Inloppsreglering felar stängd",        2e-2),
                               ("Utloppsreglering felar öppen",         1e-2)],
        "Pump":               [("Inloppspump stopp",                    2e-2),
                               ("Pumpläckage / tätningsfel",            5e-3)],
        "Rörledning / slang": [("Rörläckage",                           5e-4)],
        "Instrument":         [("Nivågivare felar — reglering öppnar utlopp", 0.1)],
        "Tank / kärl / kolonn":[("Läckage via botten / sida",           5e-4)],
    },

    # ── Hög temperatur ────────────────────────────────────────────────────────
    "Hög temperatur": {
        "Manuell ventil":     [("Kylmediumventil stängd",               5e-4),
                               ("Värmemediumventil öppnad",             5e-4)],
        "Reglerventil":       [("Kylventil felar stängd",               2e-2),
                               ("Värmeventil felar öppen",              1e-2)],
        "Värmeväxlare / kylare / värmare": [
                               ("Kylningsbortfall",                     5e-3),
                               ("Värmetillförsel okontrollerad",        1e-3)],
        "Instrument":         [("Temperaturgivare felar — kylning stängs", 0.1)],
        "Tank / kärl / kolonn":[("Exoterm reaktion",                    1e-4),
                               ("Extern värmetillförsel",               1e-4)],
        "Rörledning / slang": [("Isolationsfel / brandpåverkan",        5e-4)],
        "Styrsystem / PLC / DCS": [("Temperaturreglering felar",        5e-3)],
        "Pump":               [("Pumpfriktionsvärme",                   5e-3)],
        "Kompressor / fläkt": [("Kompressionsöverhettning",             1e-2)],
    },

    # ── Låg temperatur ────────────────────────────────────────────────────────
    "Låg temperatur": {
        "Manuell ventil":     [("Värmemediumventil stängd",             5e-4),
                               ("Kylmediumventil öppnad",               5e-4)],
        "Reglerventil":       [("Värmeventil felar stängd",             2e-2),
                               ("Kylventil felar öppen",                1e-2)],
        "Värmeväxlare / kylare / värmare": [
                               ("Värmebortfall",                        5e-3),
                               ("Överkylning",                          1e-3)],
        "Instrument":         [("Temperaturgivare felar — värmning stängs", 0.1)],
        "Rörledning / slang": [("Frysrisk — isolationsbortfall",        1e-3)],
        "Tank / kärl / kolonn":[("Endoterm reaktion / avdunstning",     1e-4)],
    },

    # ── Omvänt flöde ─────────────────────────────────────────────────────────
    "Omvänt flöde": {
        "Backventil":         [("Backventil defekt — läcker",           1e-2),
                               ("Backventil saknas",                    1e-4)],
        "Manuell ventil":     [("Ventil öppnas mot tryckkälla",         5e-4)],
        "Pump":               [("Pump stopp — backflöde via pump",      2e-2),
                               ("Pump roterar baklänges",               1e-3)],
        "Rörledning / slang": [("Felkopplad ledning",                   1e-4)],
        "Styrsystem / PLC / DCS": [("Ventilstyrning felar", 5e-3)],
    },

    # ── Missriktat flöde ──────────────────────────────────────────────────────
    "Missriktat flöde": {
        "Manuell ventil":     [("Fel ventil öppnad",                    1e-3),
                               ("Bypassventil öppnad",                  5e-4)],
        "On-off ventil":      [("Automatstyrd ventil öppnar fel väg",   1e-2)],
        "Reglerventil":       [("Styrventil öppnar alternativ väg",     5e-3)],
        "Rörledning / slang": [("Felkopplad ledning",                   1e-4)],
        "Styrsystem / PLC / DCS": [("Felaktig ventilstyrning",          5e-3)],
        "Instrument":         [("Flödesgivare i fel linje",             0.1)],
    },

    # ── Avvikande sammansättning ──────────────────────────────────────────────
    "Avvikande sammansättning": {
        "Manuell ventil":     [("Fel ventil öppnad — korsflöde",        1e-3)],
        "Reglerventil":       [("Dos- / blandningsventil i fel läge",   5e-3)],
        "Tank / kärl / kolonn":[("Kontamination i kärl",                5e-4),
                               ("Fel råmaterial / kemikalie",           1e-3)],
        "Rörledning / slang": [("Felkopplad ledning",                   1e-4)],
        "Instrument":         [("Analysgivare felar — doseringsstyrning", 0.1)],
        "Pump":               [("Felaktigt pumpmedium",                 5e-4)],
    },

    # ── Bortfall av hjälpsystem ───────────────────────────────────────────────
    "Bortfall av hjälpsystem": {
        "Elförsörjning":      [("Strömavbrott",                         0.1),
                               ("Säkring / skydd löser ut",             0.5)],
        "Tryckluft / instrumentluft": [
                               ("Lufttrycksfall",                       5e-2),
                               ("Luftkompressor stopp",                 0.1)],
        "Kylsystem / värmesystem": [
                               ("Kylvattenpump stopp",                  2e-2),
                               ("Kylvattentryck faller",                5e-2)],
        "Styrsystem / PLC / DCS": [("DCS / PLC haveri",                 1e-2),
                               ("Kommunikationsavbrott",                0.1)],
    },

    # ── Drift ─────────────────────────────────────────────────────────────────
    "Drift": {
        "Manuell ventil":     [("Felaktig manöver — fel ventil",        1e-2),
                               ("Ventil glömd i fel läge",              5e-3)],
        "Operatör / procedur / underhåll": [
                               ("Felaktig procedur / fel sekvens",      5e-2),
                               ("Procedur saknas eller otydlig",        None),
                               ("Kommunikationsfel",                    None)],
        "Instrument":         [("Felläsning av mätvärde",               5e-2)],
    },

    # ── Underhåll ─────────────────────────────────────────────────────────────
    "Underhåll": {
        "Manuell ventil":     [("Isolationsventil felaktigt ställd",    5e-3),
                               ("Ventil i fel läge efter arbete",       1e-3)],
        "Fläns / koppling / packning": [
                               ("Felaktig packning installerad",        1e-3),
                               ("Flansbultar ej åtdragna",              5e-4)],
        "Operatör / procedur / underhåll": [
                               ("Felaktig isolering (LOTO)",            5e-3),
                               ("Arbete på trycksatt system",           1e-3),
                               ("Fel komponent installerad",            5e-4)],
        "Rörledning / slang": [("Blind platta kvarglömd",               5e-4)],
        "Instrument":         [("Instrument ej återdriftsatt",          1e-2)],
    },

    # ── Start-up / Shut-down ──────────────────────────────────────────────────
    "Start-up / Shut-down": {
        "Manuell ventil":     [("Fel ventilsekvens",                    1e-2),
                               ("Ventil stängd vid pumpstart",          5e-3)],
        "Rörledning / slang": [("Kondensatbank — vätskeslag",           1e-3),
                               ("Luftlås vid start",                    5e-4)],
        "Pump":               [("Pump startas mot stängt utlopp",       5e-3),
                               ("Pump startas utan inloppstryck",       5e-3)],
        "Operatör / procedur / underhåll": [
                               ("Felaktig start-/stoppsekvens",         1e-2),
                               ("Procedur ej följd",                    5e-2)],
        "Tank / kärl / kolonn":[("Kärl ej förberett vid start",         1e-3)],
        "Reglerventil":       [("Reglerventil i manuellt läge vid start", 5e-3)],
        "Värmeväxlare / kylare / värmare": [
                               ("Termisk chock vid uppstart",           1e-3)],
    },
}

_STD_OBJECTS = [
    "Manuell ventil", "On-off ventil", "Reglerventil", "Backventil",
    "Säkerhetsventil / sprängbleck", "Pump", "Kompressor / fläkt",
    "Filter / sil", "Värmeväxlare / kylare / värmare", "Tank / kärl / kolonn",
    "Rörledning / slang", "Fläns / koppling / packning", "Instrument",
    "Styrsystem / PLC / DCS", "Elförsörjning", "Tryckluft / instrumentluft",
    "Kylsystem / värmesystem", "Blandare / omrörare",
    "Operatör / procedur / underhåll", "Övrigt",
]

_COMP_TYPE_TO_OBJ: dict = {
    'Pump': 'Pump', 'Kompressor': 'Kompressor / fläkt',
    'Ventil': 'Reglerventil', 'Rörledning': 'Rörledning / slang',
    'Instrument / Sensor': 'Instrument', 'Tank / Kärl': 'Tank / kärl / kolonn',
    'Värmeväxlare': 'Värmeväxlare / kylare / värmare',
}

# The compact, project-facing standard-cause catalogue.  The former large
# catalogue is deliberately retained in the archive during migration; this
# list is what new cause pickers should expose.
_REDUCED_STD_CATALOG = [
    ('Manuell ventil', 'På samtliga avvikelser', 'Ventil felaktigt stängd'),
    ('Manuell ventil', 'På samtliga avvikelser', 'Ventil felaktigt öppnad'),
    ('On-off ventil', 'På samtliga avvikelser', 'Ventil felaktigt stängd'),
    ('On-off ventil', 'På samtliga avvikelser', 'Ventil felaktigt öppnad'),
    ('Reglerventil', 'På samtliga avvikelser', 'Reglerventil felar stängd'),
    ('Reglerventil', 'På samtliga avvikelser', 'Reglerventil felar öppen'),
    ('Säkerhetsventil / sprängbleck', 'Missriktat', 'Säkerhetsventil öppnar för tidigt/läcker'),
    ('Backventil', 'Lågt flöde', 'Fastnar stängd'),
    ('Backventil', 'Omvänt flöde', 'Felar öppen'),
    ('Kompressor / fläkt', 'Lågt flöde', 'Kompressor / fläkt stopp'),
    ('Kompressor / fläkt', 'Högt flöde', 'Kompressor, för hög kapacitet'),
    ('Filter / sil', 'Lågt flöde', 'Filter / sil igensatt'),
    ('Filter / sil', 'Högt tryck', 'Filter / sil igensatt'),
    ('Filter / sil', 'Högt flöde', 'Filter skadat'),
    ('Filter / sil', 'Avvikande sammansättning', 'Filter skadat'),
    ('Värmeväxlare / kylare / värmare', 'Lågt flöde', 'Igensatt värmeväxlare'),
    ('Värmeväxlare / kylare / värmare', 'Hög temperatur', 'Kylningsbortfall'),
    ('Värmeväxlare / kylare / värmare', 'Hög temperatur', 'För hög värmning'),
    ('Värmeväxlare / kylare / värmare', 'Låg temperatur', 'Värmebortfall'),
    ('Värmeväxlare / kylare / värmare', 'Låg temperatur', 'För låg kylning'),
    ('Värmeväxlare / kylare / värmare', 'Missriktat flöde', 'Läckage i värmeväxlare'),
    ('Instrument', 'Lågt flöde', 'Givare felar, styrventil stänger'),
    ('Instrument', 'Lågt flöde', 'Börvärde felaktigt inställt'),
    ('Instrument', 'Högt flöde', 'Givare felar, styrventil öppnar'),
    ('Instrument', 'Högt flöde', 'Börvärde felaktigt högt'),
    ('Instrument', 'Missriktat flöde', 'Givare felar, styrventil öppnar'),
    ('Instrument', 'Högt tryck', 'Givare felar, styrventil stänger'),
    ('Instrument', 'Högt tryck', 'Börvärde tryckreglering felaktigt'),
    ('Instrument', 'Lågt tryck', 'Givare felar, styrventil öppnar'),
    ('Instrument', 'Hög nivå', 'Givare felar, reglering stänger ventil'),
    ('Instrument', 'Hög nivå', 'Börvärde nivå felaktigt'),
    ('Instrument', 'Låg nivå', 'Börvärde nivå felaktigt'),
    ('Instrument', 'Låg nivå', 'Givare felar, reglering öppnar utlopp'),
    ('Instrument', 'Hög temperatur', 'Givare felar, max kylning'),
    ('Instrument', 'Låg temperatur', 'Givare felar, max värmning'),
    ('Instrument', 'Avvikande sammansättning', 'Analysgivare felar'),
    ('Instrument', 'Drift', 'Felläsning av mätvärde'),
    ('Instrument', 'Underhåll', 'Instrument ej återdriftsatt'),
    ('Pump', 'Lågt flöde', 'Pump stopp'),
    ('Pump', 'Högt flöde', 'Pumpkapacitet för hög'),
    ('Pump', 'Omvänt flöde', 'Pump stopp, backflöde via pump'),
    ('Pump', 'Högt tryck', 'Utlopp blockerat'),
    ('Pump', 'Start-up / Shut-down', 'Pump startas mot stängt utlopp'),
    ('Pump', 'Start-up / Shut-down', 'Pump startas utan inloppstryck'),
    ('Elförsörjning', 'Bortfall av hjälpsystem', 'Strömavbrott'),
    ('Tryckluft / instrumentluft', 'Bortfall av hjälpsystem', 'Lufttrycksfall'),
    ('Kylsystem / värmesystem', 'Bortfall av hjälpsystem', 'Bortfall av kylvatten'),
]

_COMP_KEY_TO_OBJ: dict = {
    # Legacy comp_type keys from old _COMP_STD_CAUSES
    'Pump':                    'Pump',
    'Kompressor':              'Kompressor / fläkt',
    'Ventil':                  'Reglerventil',
    'Rörledning':              'Rörledning / slang',
    'Instrument / Sensor':     'Instrument',
    'Tank / Kärl':             'Tank / kärl / kolonn',
    'Värmeväxlare':            'Värmeväxlare / kylare / värmare',
    # New exact keys matching _STD_OBJECTS names (no mapping needed — identity)
    'Manuell ventil':               'Manuell ventil',
    'On-off ventil':                'On-off ventil',
    'Reglerventil':                 'Reglerventil',
    'Backventil':                   'Backventil',
    'Säkerhetsventil / sprängbleck':'Säkerhetsventil / sprängbleck',
    'Kompressor / fläkt':           'Kompressor / fläkt',
    'Filter / sil':                 'Filter / sil',
    'Värmeväxlare / kylare / värmare':'Värmeväxlare / kylare / värmare',
    'Tank / kärl / kolonn':         'Tank / kärl / kolonn',
    'Rörledning / slang':           'Rörledning / slang',
    'Fläns / koppling / packning':  'Fläns / koppling / packning',
    'Instrument':                   'Instrument',
    'Styrsystem / PLC / DCS':       'Styrsystem / PLC / DCS',
    'Elförsörjning':                'Elförsörjning',
    'Tryckluft / instrumentluft':   'Tryckluft / instrumentluft',
    'Kylsystem / värmesystem':      'Kylsystem / värmesystem',
    'Blandare / omrörare':          'Blandare / omrörare',
    'Operatör / procedur / underhåll':'Operatör / procedur / underhåll',
    'Övrigt':                       'Övrigt',
}


def _fix_instrument_causes_v2(conn):
    """No-op: instrument causes now seeded correctly via _COMP_STD_CAUSES."""
    pass


def _fix_instrument_causes_v3(conn):
    """No-op: instrument causes now seeded correctly via _COMP_STD_CAUSES."""
    pass


def _seed_standard_objects(conn):
    for i, name in enumerate(_STD_OBJECTS):
        conn.execute(
            "INSERT OR IGNORE INTO standard_objects (name, sort_order) VALUES (?,?)",
            (name, i))
    conn.commit()


def _seed_component_causes(conn):
    """Insert standard causes (idempotent). Entries can be str or (desc, freq) tuples."""
    for dev_name, by_type in _COMP_STD_CAUSES.items():
        row = conn.execute(
            "SELECT id FROM standard_deviations WHERE description=?", (dev_name,)).fetchone()
        if not row:
            continue
        dev_id = row[0]
        max_sort = (conn.execute(
            "SELECT COALESCE(MAX(sort_order),0) FROM standard_causes WHERE deviation_id=?",
            (dev_id,)).fetchone()[0] or 0)
        sort_i = max_sort + 1
        for comp_key, causes in by_type.items():
            obj_name = _COMP_KEY_TO_OBJ.get(comp_key, comp_key)
            obj_row  = conn.execute(
                "SELECT id FROM standard_objects WHERE name=?", (obj_name,)).fetchone()
            obj_id   = obj_row[0] if obj_row else None
            for entry in causes:
                c_desc = entry[0] if isinstance(entry, tuple) else entry
                c_freq = entry[1] if isinstance(entry, tuple) and len(entry) > 1 else None
                exists = conn.execute(
                    "SELECT id FROM standard_causes "
                    "WHERE deviation_id=? AND description=?",
                    (dev_id, c_desc)).fetchone()
                if not exists:
                    conn.execute(
                        "INSERT INTO standard_causes "
                        "(deviation_id, description, sort_order, comp_type, object_id, frequency)"
                        " VALUES (?,?,?,?,?,?)",
                        (dev_id, c_desc, sort_i, comp_key, obj_id, c_freq))
                    sort_i += 1
                else:
                    updates, vals = [], []
                    if obj_id is not None:
                        updates.append("object_id=?"); vals.append(obj_id)
                    if c_freq is not None:
                        updates.append("frequency=?"); vals.append(c_freq)
                    if updates:
                        vals.append(exists[0])
                        conn.execute(
                            f"UPDATE standard_causes SET {','.join(updates)} WHERE id=?", vals)
    conn.commit()


def _migrate_causes_to_object_id(conn):
    """Populate standard_causes.object_id from comp_type using _COMP_TYPE_TO_OBJ mapping."""
    for comp, obj_name in _COMP_TYPE_TO_OBJ.items():
        row = conn.execute(
            "SELECT id FROM standard_objects WHERE name=?", (obj_name,)).fetchone()
        if not row:
            continue
        conn.execute(
            "UPDATE standard_causes SET object_id=? WHERE comp_type=? AND object_id IS NULL",
            (row[0], comp))
    conn.commit()


def _sync_f_levels_from_base_frequency(conn):
    """Set causes.likelihood (F-level) from standard_cause/base_frequency when frequency data exists."""
    updated = 0
    rows = conn.execute("""
        SELECT c.id, c.base_frequency, c.likelihood, c.frequency_cleared,
               sc.frequency AS sc_freq
        FROM causes c
        LEFT JOIN standard_causes sc ON sc.id = c.standard_cause_id
    """).fetchall()
    for row in rows:
        if row['frequency_cleared']:
            continue
        base_freq_per_year = row['sc_freq'] if row['sc_freq'] is not None else row['base_frequency']
        if base_freq_per_year is None or base_freq_per_year <= 0:
            continue
        f_level = freq_to_f_level(base_freq_per_year)
        if row['likelihood'] != f_level or (row['base_frequency'] is None and row['sc_freq'] is not None):
            conn.execute(
                "UPDATE causes SET likelihood=?, base_frequency=COALESCE(base_frequency, ?) WHERE id=?",
                (f_level, base_freq_per_year, row['id']))
            updated += 1
    return updated


# Keep old name as alias for backward compatibility
_sync_cause_likelihoods_from_frequency = _sync_f_levels_from_base_frequency



class Database:
    # The undo history deliberately lives outside the project schema.  It is
    # an in-memory session history: saving/opening a project starts a new
    # baseline, while the existing timestamped backups remain the recovery
    # mechanism across application restarts.
    _HISTORY_LIMIT = 100

    def __init__(self, path=DB_PATH):
        # These flags must exist before the first schema/migration commit.
        # Startup seeding and migrations establish the initial baseline, not
        # user-visible undo entries.
        self._history_initialized = False
        self._history_restoring = False
        self._history_snapshot = None
        self._undo_stack = []
        self._redo_stack = []
        self._history_listeners = []
        self._history_group_depth = 0
        self._history_group_before = None
        self._history_group_changed = False
        self.path = Path(path)
        # A pre-existing, non-empty DB file means _migrate() below may run real
        # ALTER TABLE/CREATE TABLE statements against live user data. Snapshot
        # it *before* touching the schema so a buggy/failed migration can
        # always be recovered from. Brand-new (empty) DBs have nothing to lose.
        pre_existing_db = self.path.exists() and self.path.stat().st_size > 0
        # Python's sqlite3 default busy-timeout is only 5s — too short for
        # real-world lock contention (the online-backup-API copy every
        # commit() does, a previous instance still releasing its WAL lock
        # on exit, or the .db file living in a OneDrive-synced folder as it
        # does here). A DDL statement mid-migration hitting that window
        # raised sqlite3.OperationalError: database is locked (real crash
        # report, crash_20260807_162445_OperationalError.json). Raised to
        # 30s so transient contention gets retried instead of crashing.
        self.conn = sqlite3.connect(str(self.path), timeout=30.0)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA foreign_keys = ON")
        self.conn.execute("PRAGMA journal_mode = WAL")   # faster concurrent reads
        self.conn.executescript(SCHEMA)
        self.commit()
        if pre_existing_db:
            try:
                self._write_backup(startup=True)   # pre-migration safety snapshot
            except Exception:
                logging.warning("Pre-migration backup failed", exc_info=True)
        self._migrate()
        self._write_backup(startup=True)   # unconditional post-migration snapshot
        if not pre_existing_db:
            # A brand-new study (first-ever launch with no DB file yet, or
            # "Nytt projekt" which deletes the old file before reconstructing
            # this class) should start with one node already there — the
            # user shouldn't have to click "+ Nod" before any work is
            # possible. Does NOT fire for a real _load_hzp() open (DB_PATH
            # already holds the copied project's data there, or — on a
            # failed copy — the untouched original), so a study where the
            # user deliberately deleted their last node is left alone.
            # 2026-08-24 (see NOTES.md "Ny toppnivå System"): also seed one
            # System containing that default node, since System is now the
            # top of the intended hierarchy — a brand new study shouldn't
            # start with a lone, ungrouped node.
            system_id = self.add_system()
            self.add_node(system_id=system_id)

        # Everything above is database setup (including the default node),
        # not a user action.  Establish the first undo baseline only after
        # the constructor has finished all migrations and seed writes.
        self._reset_history_baseline()

    def __del__(self):
        """Clean up database connection on object destruction.

        Ensures the SQLite connection is properly closed even if the Database
        object is replaced or goes out of scope without explicit cleanup.
        """
        try:
            if hasattr(self, 'conn') and self.conn:
                # Flush WAL checkpoint before closing
                self.conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
                self.conn.close()
        except Exception:
            # Silently ignore errors during cleanup
            pass

    def _migrate(self):
        # Kolumnmigreringarna körs TVÅ gånger: före executescript (befintliga
        # databaser) och efter (färska databaser där CREATE TABLE just skapat
        # bastabellerna utan de migrerade kolumnerna — annars kraschar
        # seedningen på t.ex. standard_causes.comp_type). Alla satser är
        # idempotenta; fel ignoreras.
        logging.info("Database: starting migration...")
        self._column_migrations()
        self._migrate_tables_and_seed()
        self._drop_legacy_consequence_likelihood_column()
        self._migrate_actions_to_recommendations()
        self._normalize_recommendation_display_numbers()
        self._clean_existing_recommendation_markup()
        # analysis_sessions is part of the base schema but some project files
        # are created from an older schema snapshot.  Ensure the new metadata
        # columns exist after all base-table creation passes have completed.
        for statement in (
                "ALTER TABLE analysis_sessions ADD COLUMN date TEXT DEFAULT ''",
                "ALTER TABLE analysis_sessions ADD COLUMN location TEXT DEFAULT ''",
                "ALTER TABLE analysis_sessions ADD COLUMN is_digital INTEGER NOT NULL DEFAULT 0",
                "ALTER TABLE analysis_sessions ADD COLUMN start_time TEXT DEFAULT ''",
            "ALTER TABLE analysis_sessions ADD COLUMN end_time TEXT DEFAULT ''",
            "ALTER TABLE participant_attendance ADD COLUMN note TEXT DEFAULT ''",
            "ALTER TABLE pid_sheets ADD COLUMN drawing_number TEXT DEFAULT ''",
            "ALTER TABLE pid_sheets ADD COLUMN drawing_name TEXT DEFAULT ''",
            "ALTER TABLE pid_sheets ADD COLUMN drawing_revision TEXT DEFAULT ''",
            "ALTER TABLE pid_sheets ADD COLUMN drawing_date TEXT DEFAULT ''",
            # LOPA document fields are stored on the revision so a locked
            # revision is a complete historic record, not merely a snapshot
            # of the calculation inputs.
            "ALTER TABLE lopa_revisions ADD COLUMN document_date TEXT NOT NULL DEFAULT ''",
            "ALTER TABLE lopa_revisions ADD COLUMN dimensioning_category_key TEXT NOT NULL DEFAULT ''",
            "ALTER TABLE lopa_revisions ADD COLUMN additional_actions TEXT NOT NULL DEFAULT ''",
            "ALTER TABLE lopa_revisions ADD COLUMN additional_requirements TEXT NOT NULL DEFAULT ''",
            "ALTER TABLE lopa_revisions ADD COLUMN process_safety_time REAL DEFAULT NULL",
            "ALTER TABLE lopa_source_scenarios ADD COLUMN control_frequency TEXT NOT NULL DEFAULT ''",
            "ALTER TABLE lopa_source_scenarios ADD COLUMN assumption_reason TEXT NOT NULL DEFAULT ''",
            "ALTER TABLE lopa_records ADD COLUMN sif_number TEXT NOT NULL DEFAULT ''"):
            try:
                self.conn.execute(statement)
            except sqlite3.OperationalError:
                pass
        self.conn.execute(
            "CREATE TABLE IF NOT EXISTS reduction_factor_catalog ("
            "id INTEGER PRIMARY KEY AUTOINCREMENT,"
            "description TEXT NOT NULL UNIQUE, rrf INTEGER NOT NULL DEFAULT 10,"
            "active INTEGER NOT NULL DEFAULT 1)")
        self.conn.execute(
            "CREATE TABLE IF NOT EXISTS reduction_factor_severity_exclusions ("
            "severity_id INTEGER NOT NULL, reduction_factor_id INTEGER NOT NULL,"
            "PRIMARY KEY (severity_id, reduction_factor_id))")
        self._seed_standard_enablers()
        self._migrate_legacy_consequence_factors_to_enablers()
        self.commit()
        logging.info("Database: migration complete")
        self._validate_schema()

    def _seed_standard_enablers(self):
        """Keep the common event-tree enablers available in every project."""
        for description, rrf in (("Antändning", 10), ("Eskalering", 10)):
            self.conn.execute(
                "INSERT OR IGNORE INTO reduction_factor_catalog(description,rrf,active) "
                "VALUES (?,?,1)", (description, rrf))

    def _migrate_legacy_consequence_factors_to_enablers(self):
        """Move old FA/ignition checkboxes into the unified enabler list.

        Older projects stored a percentage in the consequence row, whereas
        the unified list stores the inverse RRF.  Converting ``10 %`` to
        ``RRF 10`` (and ``1 %`` to ``RRF 100``) preserves the existing number
        of frequency-reduction steps exactly.  The legacy switches are then
        cleared so they can never be counted a second time.
        """
        try:
            rows = self.conn.execute(
                "SELECT id,fa_active,fa_rrf,ignition_active,ignition_rrf "
                "FROM consequences WHERE fa_active=1 OR ignition_active=1").fetchall()
        except sqlite3.OperationalError:
            return
        for row in rows:
            for active_key, probability_key, description in (
                    ('fa_active', 'fa_rrf', 'Eskalering'),
                    ('ignition_active', 'ignition_rrf', 'Antändning')):
                if not row[active_key]:
                    continue
                try:
                    probability = float(row[probability_key] or 10)
                except (TypeError, ValueError):
                    probability = 10.0
                rrf = max(1.0, 100.0 / max(0.001, min(99.9, probability)))
                existing = self.conn.execute(
                    "SELECT id FROM reduction_factors WHERE consequence_id=? "
                    "AND lower(description)=lower(?) ORDER BY id LIMIT 1",
                    (row['id'], description)).fetchone()
                if existing:
                    self.conn.execute(
                        "UPDATE reduction_factors SET rrf=?,active=1 WHERE id=?",
                        (rrf, existing['id']))
                else:
                    self.conn.execute(
                        "INSERT INTO reduction_factors(consequence_id,description,rrf,active) "
                        "VALUES (?,?,?,1)", (row['id'], description, rrf))
            self.conn.execute(
                "UPDATE consequences SET fa_active=0,ignition_active=0 WHERE id=?",
                (row['id'],))

    def _clean_existing_recommendation_markup(self):
        """One-time-safe cleanup for imported HTML document fragments."""
        try:
            rows = self.conn.execute("SELECT id, description FROM recommendations").fetchall()
        except sqlite3.OperationalError:
            return
        changed = []
        for row in rows:
            clean = self._clean_recommendation_text(row['description'])
            if clean != (row['description'] or ''):
                changed.append((clean, row['id']))
        if changed:
            self.conn.executemany(
                "UPDATE recommendations SET description=? WHERE id=?", changed)
            self.commit()

    def _normalize_recommendation_display_numbers(self):
        """Backfill and compact the user-visible recommendation-number sequence.

        Recommendation ids remain stable foreign-key targets.  The separate
        display number can safely close gaps left by deletion without changing
        a shared recommendation's consequence links.
        """
        try:
            rows = self.conn.execute(
                "SELECT id, display_number FROM recommendations "
                "ORDER BY CASE WHEN display_number > 0 THEN display_number ELSE id END, id").fetchall()
        except sqlite3.OperationalError:
            return
        updates = [(number, row['id']) for number, row in enumerate(rows, start=1)
                   if row['display_number'] != number]
        if updates:
            self.conn.executemany(
                "UPDATE recommendations SET display_number=? WHERE id=?", updates)
            self.commit()

    def _migrate_actions_to_recommendations(self):
        """One-time data migration for the 2026-08-25 recommendations
        catalog rework (see NOTES.md "Rekommendationshantering — delad
        katalog med återanvändning"): the old actions table (one row per
        consequence, no reuse) is gone from SCHEMA for brand-new
        databases, but a pre-existing database file still physically has
        it until this runs. Copies every row into the new
        recommendations catalog plus a matching consequence_recommendations
        link back to its original consequence_id, then drops the old
        table — after that, 'actions' not in tables is true forever, so
        this is a no-op on every later launch. Same idempotent-via-table-
        existence pattern as _drop_legacy_consequence_likelihood_column."""
        tables = {r['name'] for r in self.conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'")}
        if 'actions' not in tables:
            return
        old_rows = self.conn.execute("SELECT * FROM actions ORDER BY id").fetchall()
        for a in old_rows:
            cur = self.conn.execute(
                "INSERT INTO recommendations (description,responsible,due_date,status) "
                "VALUES (?,?,?,?)",
                (a['description'], a['responsible'], a['due_date'], a['status']))
            self.conn.execute(
                "INSERT INTO consequence_recommendations (consequence_id,recommendation_id) "
                "VALUES (?,?)", (a['consequence_id'], cur.lastrowid))
        self.conn.execute("DROP TABLE actions")
        self.commit()
        logging.info(f"Migrated {len(old_rows)} actions -> recommendations catalog")

    def _drop_legacy_consequence_likelihood_column(self):
        """consequences.likelihood predates the schema redesign that moved
        likelihood onto causes (see CLAUDE.md — 'Likelihood lives on
        causes, severity on consequences'). Nothing reads or writes it
        anymore, but old database files still carry it with stale values.
        A no-op on any database created after the redesign; harmless if
        the installed SQLite predates DROP COLUMN support (3.35+, 2021)."""
        try:
            cols = [r['name'] for r in self.conn.execute("PRAGMA table_info(consequences)")]
            if 'likelihood' in cols:
                self.conn.execute("ALTER TABLE consequences DROP COLUMN likelihood")
                self.commit()
                logging.info("Dropped legacy consequences.likelihood column")
        except sqlite3.OperationalError as e:
            logging.warning(f"Could not drop legacy consequences.likelihood column: {e}")

    def _column_migrations(self):
        """Execute idempotent column migrations with proper error handling.

        Distinguishes between benign errors (column already exists) and real failures
        (syntax errors, permission issues) for accurate logging and diagnostics.
        """
        migration_count = 0
        skipped_count = 0
        error_count = 0

        migrations = [
            "ALTER TABLE nodes ADD COLUMN markup_points TEXT DEFAULT ''",
            "ALTER TABLE nodes ADD COLUMN markup_style TEXT DEFAULT ''",
            "ALTER TABLE nodes ADD COLUMN pid_page INTEGER DEFAULT 0",
            "ALTER TABLE nodes ADD COLUMN media TEXT DEFAULT ''",
            "ALTER TABLE nodes ADD COLUMN pressure TEXT DEFAULT ''",
            "ALTER TABLE nodes ADD COLUMN temperature TEXT DEFAULT ''",
            "ALTER TABLE nodes ADD COLUMN updated_at TEXT DEFAULT ''",
            "ALTER TABLE nodes ADD COLUMN updated_by TEXT DEFAULT ''",
            "ALTER TABLE nodes ADD COLUMN system_id INTEGER REFERENCES systems(id)",
            "ALTER TABLE nodes ADD COLUMN sort_order INTEGER NOT NULL DEFAULT 0",
            "ALTER TABLE deviations ADD COLUMN sort_order INTEGER NOT NULL DEFAULT 0",
            "ALTER TABLE causes ADD COLUMN likelihood INTEGER NOT NULL DEFAULT 1",
            "ALTER TABLE causes ADD COLUMN source_id INTEGER DEFAULT NULL",
            "ALTER TABLE causes ADD COLUMN base_frequency REAL DEFAULT NULL",
            "ALTER TABLE causes ADD COLUMN frequency_cleared INTEGER NOT NULL DEFAULT 0",
            "ALTER TABLE causes ADD COLUMN deviation_id INTEGER REFERENCES deviations(id)",
            "ALTER TABLE causes ADD COLUMN sort_order INTEGER NOT NULL DEFAULT 0",
            "ALTER TABLE safeguards ADD COLUMN rrf INTEGER NOT NULL DEFAULT 1",
            "ALTER TABLE safeguards ADD COLUMN source_id INTEGER DEFAULT NULL",
            "ALTER TABLE safeguards ADD COLUMN sort_order INTEGER NOT NULL DEFAULT 0",
            "ALTER TABLE consequences ADD COLUMN category TEXT DEFAULT ''",
            "ALTER TABLE consequences ADD COLUMN sort_order INTEGER NOT NULL DEFAULT 0",
            "ALTER TABLE consequences ADD COLUMN consequence_chain TEXT DEFAULT ''",
            "ALTER TABLE consequences ADD COLUMN source_id INTEGER DEFAULT NULL",
            "ALTER TABLE consequences ADD COLUMN fa_active INTEGER DEFAULT 0",
            "ALTER TABLE consequences ADD COLUMN fa_rrf INTEGER DEFAULT 10",
            "ALTER TABLE consequences ADD COLUMN ignition_active INTEGER DEFAULT 0",
            "ALTER TABLE consequences ADD COLUMN ignition_rrf INTEGER DEFAULT 10",
            "ALTER TABLE cause_markers ADD COLUMN component_tag TEXT DEFAULT ''",
            "ALTER TABLE standard_causes ADD COLUMN comp_type TEXT DEFAULT ''",
            "ALTER TABLE standard_causes ADD COLUMN frequency REAL DEFAULT NULL",
            "ALTER TABLE standard_causes ADD COLUMN use_in_cause_form INTEGER DEFAULT 1",
            "ALTER TABLE causes ADD COLUMN standard_cause_id INTEGER DEFAULT NULL",
            "ALTER TABLE causes ADD COLUMN comp_type TEXT DEFAULT ''",
            "ALTER TABLE causes ADD COLUMN comp_tag  TEXT DEFAULT ''",
            "ALTER TABLE causes ADD COLUMN linked_consequence_id INTEGER DEFAULT NULL",
            "ALTER TABLE safeguards ADD COLUMN sg_type TEXT DEFAULT 'Övrigt'",
            "ALTER TABLE recommendations ADD COLUMN display_number INTEGER NOT NULL DEFAULT 0",
            "ALTER TABLE node_markups ADD COLUMN font_size INTEGER DEFAULT 12",
            "ALTER TABLE cause_markers ADD COLUMN rect_w REAL DEFAULT NULL",
            "ALTER TABLE cause_markers ADD COLUMN rect_h REAL DEFAULT NULL",
            "ALTER TABLE consequence_markers ADD COLUMN rect_w REAL DEFAULT NULL",
            "ALTER TABLE consequence_markers ADD COLUMN rect_h REAL DEFAULT NULL",
            "ALTER TABLE safeguard_markers ADD COLUMN rect_w REAL DEFAULT NULL",
            "ALTER TABLE safeguard_markers ADD COLUMN rect_h REAL DEFAULT NULL",
            "ALTER TABLE off_page_connector ADD COLUMN ref_page INTEGER DEFAULT NULL",
            "ALTER TABLE off_page_connector ADD COLUMN dot_scene_x REAL DEFAULT NULL",
            "ALTER TABLE off_page_connector ADD COLUMN dot_scene_y REAL DEFAULT NULL",
            "ALTER TABLE consequence_steps ADD COLUMN node_key TEXT DEFAULT ''",
            "ALTER TABLE standard_causes ADD COLUMN object_id INTEGER REFERENCES standard_objects(id)",
            "ALTER TABLE causes ADD COLUMN comment TEXT DEFAULT ''",
            "ALTER TABLE nodes ADD COLUMN approved_by TEXT DEFAULT ''",
            "ALTER TABLE nodes ADD COLUMN approved_at TEXT DEFAULT ''",
            "ALTER TABLE nodes ADD COLUMN study_status TEXT DEFAULT 'draft'",
            "ALTER TABLE study_tag_memory ADD COLUMN active INTEGER DEFAULT 1",
            "ALTER TABLE analysis_sessions ADD COLUMN date TEXT DEFAULT ''",
            "ALTER TABLE analysis_sessions ADD COLUMN location TEXT DEFAULT ''",
            # Smart object recognition — composite key so the same prefix can
            # map to multiple types (e.g. HV→Handventil×5, HV→Backventil×2).
            # The type with the highest usage_count wins on lookup.
            """CREATE TABLE IF NOT EXISTS study_tag_memory (
                tag         TEXT NOT NULL,
                comp_type   TEXT NOT NULL DEFAULT '',
                comp_tag    TEXT NOT NULL DEFAULT '',
                phash       TEXT NOT NULL DEFAULT '',
                usage_count INTEGER NOT NULL DEFAULT 1,
                updated     TEXT NOT NULL DEFAULT '',
                active      INTEGER NOT NULL DEFAULT 1,
                PRIMARY KEY (tag, comp_type)
            )""",
            """CREATE TABLE IF NOT EXISTS symbol_fingerprints (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                phash       TEXT NOT NULL,
                comp_type   TEXT NOT NULL DEFAULT '',
                tag_example TEXT NOT NULL DEFAULT '',
                usage_count INTEGER NOT NULL DEFAULT 1
            )""",
            # Fas 1+2 valve-detection improvements (2026-08-06, see NOTES.md) —
            # per-field confidence + line/medium/DN tracing + untagged-valve
            # status on equipment_markers. Existing 'confidence' column is
            # kept and populated with the weakest-link min() of the four new
            # ones, so every existing reader keeps working unchanged.
            "ALTER TABLE equipment_markers ADD COLUMN detection_confidence REAL DEFAULT NULL",
            "ALTER TABLE equipment_markers ADD COLUMN tag_reading_confidence REAL DEFAULT NULL",
            "ALTER TABLE equipment_markers ADD COLUMN tag_assignment_confidence REAL DEFAULT NULL",
            "ALTER TABLE equipment_markers ADD COLUMN line_assignment_confidence REAL DEFAULT NULL",
            "ALTER TABLE equipment_markers ADD COLUMN line_number TEXT DEFAULT ''",
            "ALTER TABLE equipment_markers ADD COLUMN medium_code TEXT DEFAULT ''",
            "ALTER TABLE equipment_markers ADD COLUMN medium_code_verified INTEGER DEFAULT 0",
            "ALTER TABLE equipment_markers ADD COLUMN nominal_size TEXT DEFAULT ''",
            "ALTER TABLE equipment_markers ADD COLUMN tag_status TEXT DEFAULT 'tagged'",
            # The tag/counter strip has its own optional anchor.  NULL means
            # automatic collision-free placement; non-NULL values are a user
            # override set through P&ID -> right-click -> Flytta etikett.
            "ALTER TABLE equipment_markers ADD COLUMN label_x REAL DEFAULT NULL",
            "ALTER TABLE equipment_markers ADD COLUMN label_y REAL DEFAULT NULL",
            # Nod → Utrustning → Avvikelse (2026-08-07, se NOTES.md) — kopplar
            # en utrustning till en nod, och en avvikelse till en specifik
            # utrustning. Båda nullable: befintliga rader/avvikelser lämnas
            # helt orörda (equipment_id/node_id=NULL), inget backfill behövs.
            "ALTER TABLE equipment_catalog ADD COLUMN node_id INTEGER REFERENCES nodes(id)",
            "ALTER TABLE deviations ADD COLUMN equipment_id INTEGER REFERENCES equipment_catalog(id)",
            # Live tag-länk mellan Orsak-cellens taggremsa och objektet på
            # P&ID (2026-08-13, se NOTES.md) — samma equipment_id-FK-mönster
            # som deviations.equipment_id ovan redan använder, istället för
            # en frusen comp_tag/comp_type-strängkopia. Backfill (matcha
            # comp_tag mot equipment_catalog.tag) körs separat efter denna
            # lista, se _backfill_cause_equipment_ids().
            "ALTER TABLE causes ADD COLUMN equipment_id INTEGER REFERENCES equipment_catalog(id)",
            # A functional group cause keeps the affected/secondary object as
            # a real P&ID link as well as the primary controlling object.
            "ALTER TABLE causes ADD COLUMN secondary_equipment_id INTEGER REFERENCES equipment_catalog(id)",
            "ALTER TABLE causes ADD COLUMN group_equipment_ids TEXT DEFAULT ''",
            "ALTER TABLE causes ADD COLUMN group_choices_set INTEGER NOT NULL DEFAULT 0",
            # Drag-and-drop tagg från P&ID till konsekvens (2026-08-07, se
            # NOTES.md) — en konsekvens kan nu bära ett eget taggnummer
            # (t.ex. en pump nedströms orsaken), visat högst upp i
            # KON-kolumnen precis som orsakskolumnen redan visar sin egen
            # tagg. Fri text (description) rörs inte av detta — taggen är
            # ett komplement, inte en ersättning.
            "ALTER TABLE consequences ADD COLUMN comp_type TEXT DEFAULT ''",
            "ALTER TABLE consequences ADD COLUMN comp_tag  TEXT DEFAULT ''",
            # Drag-and-drop tagg från P&ID till safeguard (2026-08-08, se
            # NOTES.md) — samma komplement-inte-ersättning-mönster som
            # consequences.comp_tag/comp_type ovan.
            "ALTER TABLE safeguards ADD COLUMN comp_type TEXT DEFAULT ''",
            "ALTER TABLE safeguards ADD COLUMN comp_tag  TEXT DEFAULT ''",
            # Drag-and-drop taggar in i fritexten (2026-08-09, se NOTES.md) —
            # comp_tag ovan bara visar det SENAST dragna objektet, men flera
            # olika objekt kan nu byggas in i samma fritext. tagged_refs
            # (komma-separerad lista, dedup, ordning bevarad) håller reda på
            # ALLA taggar som någonsin dragits in, så cellen kan fetmarkera
            # varje förekomst av dem i texten.
            "ALTER TABLE consequences ADD COLUMN tagged_refs TEXT DEFAULT ''",
            "ALTER TABLE safeguards ADD COLUMN tagged_refs TEXT DEFAULT ''",
            # Nodtyper i Avvikelser & Orsaker (2026-08-17, se NOTES.md) —
            # nullable, seedas mot "Processnod" lazily av Database.node_types().
            "ALTER TABLE standard_deviations ADD COLUMN node_type_id INTEGER REFERENCES node_types(id)",
            "ALTER TABLE standard_deviations ADD COLUMN active INTEGER NOT NULL DEFAULT 1",
            "ALTER TABLE standard_causes ADD COLUMN active INTEGER NOT NULL DEFAULT 1",
        ]

        for sql in migrations:
            try:
                logging.debug(f"Attempting migration: {sql[:70]}...")
                self.conn.execute(sql)
                migration_count += 1
                logging.debug(f"Migration complete: {sql[:70]}...")
            except sqlite3.OperationalError as e:
                error_msg = str(e).lower()
                # Benign errors: column/table already exists
                if "already exists" in error_msg or "no such table" in error_msg or "duplicate column" in error_msg:
                    logging.debug(f"Skipping (already exists): {sql[:70]}")
                    skipped_count += 1
                else:
                    # Real operational error (constraint violation, syntax error, etc.)
                    logging.error(f"Operational error during migration: {str(e)}")
                    logging.error(f"SQL: {sql[:100]}")
                    error_count += 1
            except sqlite3.DatabaseError as e:
                # Database errors (corruption, permission, etc.)
                logging.error(f"Database error during migration: {str(e)}")
                logging.error(f"SQL: {sql[:100]}")
                error_count += 1
            except Exception as e:
                # Unexpected errors
                logging.error(f"Unexpected error during migration: {type(e).__name__}: {str(e)}")
                logging.error(f"SQL: {sql[:100]}")
                error_count += 1

        logging.info(f"Column migrations: {migration_count} applied, {skipped_count} skipped, {error_count} errors")

        if error_count > 0:
            logging.warning(f"Migration had {error_count} real errors — database may be in inconsistent state")

        self._backfill_cause_equipment_ids()

    def _backfill_cause_equipment_ids(self):
        """Best-effort backfill for the new causes.equipment_id FK
        (2026-08-13, see NOTES.md) — matches each still-unlinked cause's
        frozen comp_tag string against equipment_catalog.tag
        (case-insensitive). Only links when EXACTLY ONE catalog row
        matches; a tag with zero or multiple matches is left NULL rather
        than guessed. Safe to re-run on every startup: the WHERE clause
        only ever touches rows that are still unlinked, and re-deriving
        the same match is a no-op."""
        try:
            rows = self.conn.execute(
                "SELECT id, comp_tag FROM causes "
                "WHERE equipment_id IS NULL AND comp_tag IS NOT NULL AND comp_tag != ''").fetchall()
            for row in rows:
                matches = self.conn.execute(
                    "SELECT id FROM equipment_catalog WHERE LOWER(tag)=LOWER(?)",
                    (row['comp_tag'],)).fetchall()
                if len(matches) == 1:
                    self.conn.execute(
                        "UPDATE causes SET equipment_id=? WHERE id=?",
                        (matches[0]['id'], row['id']))
            if rows:
                self.commit()
        except sqlite3.OperationalError as e:
            logging.warning(f"Could not backfill causes.equipment_id: {e}")

    def _migrate_reduced_standard_catalog(self):
        """Archive the former broad standard-cause catalogue and install the
        compact project catalogue once.  Rows are never deleted: existing
        ``causes.standard_cause_id`` references therefore remain valid, while
        the archive tables provide an explicit recovery/export source."""
        key = 'reduced_standard_catalog_v1'
        if self.conn.execute("SELECT value FROM app_config WHERE key=?", (key,)).fetchone():
            # The first reduced-catalog migration could be followed by the
            # legacy duplicate-deviation cleanup.  That cleanup selected the
            # old (now archived/inactive) row as the canonical duplicate and
            # accidentally left the entire active deviation list empty while
            # the active causes still referenced those original rows.  Heal
            # existing project files in place; never delete or rewrite data.
            active_count = self.conn.execute(
                "SELECT COUNT(*) FROM standard_deviations WHERE active=1").fetchone()[0]
            if not active_count:
                self.conn.execute(
                    "UPDATE standard_deviations SET active=1 WHERE id IN "
                    "(SELECT DISTINCT deviation_id FROM standard_causes WHERE active=1)")
                self.conn.commit()
            # Reduced standard causes use one conservative default frequency
            # when no explicit rate was supplied.
            self.conn.execute(
                "UPDATE standard_causes SET frequency=0.02 "
                "WHERE active=1 AND frequency IS NULL")
            self.conn.commit()
            return
        self.conn.executescript("""
            CREATE TABLE IF NOT EXISTS standard_deviations_archive (
                archive_id INTEGER PRIMARY KEY AUTOINCREMENT,
                original_id INTEGER, description TEXT, sort_order INTEGER,
                node_type_id INTEGER, archived_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS standard_causes_archive (
                archive_id INTEGER PRIMARY KEY AUTOINCREMENT,
                original_id INTEGER, deviation_id INTEGER, description TEXT,
                sort_order INTEGER, object_id INTEGER, comp_type TEXT,
                frequency REAL, use_in_cause_form INTEGER, archived_at TEXT NOT NULL
            );
        """)
        stamp = datetime.datetime.now(datetime.timezone.utc).isoformat()
        self.conn.execute(
            "INSERT INTO standard_deviations_archive "
            "(original_id,description,sort_order,node_type_id,archived_at) "
            "SELECT id,description,sort_order,node_type_id,? FROM standard_deviations",
            (stamp,))
        self.conn.execute(
            "INSERT INTO standard_causes_archive "
            "(original_id,deviation_id,description,sort_order,object_id,comp_type,frequency,use_in_cause_form,archived_at) "
            "SELECT id,deviation_id,description,sort_order,object_id,comp_type,frequency,use_in_cause_form,? "
            "FROM standard_causes", (stamp,))
        self.conn.execute("UPDATE standard_deviations SET active=0")
        self.conn.execute("UPDATE standard_causes SET active=0")

        # Universal rows are copied into every usable deviation template;
        # this keeps the existing picker model (deviation -> causes) intact.
        specific = {}
        for obj, dev, desc in _REDUCED_STD_CATALOG:
            specific.setdefault((obj, dev), []).append(desc)
        usable_devs = [d for d in DEVIATION_TYPES if d != 'Övrigt']
        for dev_name in usable_devs:
            row = self.conn.execute(
                "SELECT id FROM standard_deviations WHERE description=? AND active=1 LIMIT 1",
                (dev_name,)).fetchone()
            if row:
                dev_id = row[0]
            else:
                sort_order = usable_devs.index(dev_name)
                dev_id = self.conn.execute(
                    "INSERT INTO standard_deviations(description,sort_order,active) VALUES (?,?,1)",
                    (dev_name, sort_order)).lastrowid
            entries = [(o, d, c) for o, d, c in _REDUCED_STD_CATALOG
                       if d == 'På samtliga avvikelser' or d == dev_name]
            for obj, _d, desc in entries:
                obj_row = self.conn.execute(
                    "SELECT id FROM standard_objects WHERE name=? LIMIT 1", (obj,)).fetchone()
                self.conn.execute(
                    "INSERT INTO standard_causes(deviation_id,description,sort_order,object_id,comp_type,frequency,active) "
                    "VALUES (?,?,?,?,?,?,1)",
                    (dev_id, desc, self.conn.execute(
                        "SELECT COALESCE(MAX(sort_order),-1)+1 FROM standard_causes WHERE deviation_id=? AND active=1",
                        (dev_id,)).fetchone()[0], obj_row[0] if obj_row else None, obj, 0.02))
        self.conn.execute(
            "INSERT OR REPLACE INTO app_config(key,value) VALUES (?, '1')", (key,))
        self.conn.commit()
        logging.info("Archived old standard catalogue and installed reduced catalogue (%d entries)",
                     len(_REDUCED_STD_CATALOG))

    def _migrate_tables_and_seed(self):
        self.conn.executescript("""
            CREATE TABLE IF NOT EXISTS pid_config (
                key TEXT PRIMARY KEY, value TEXT
            );
            CREATE TABLE IF NOT EXISTS equipment_types (
                prefix         TEXT PRIMARY KEY,
                equipment_type TEXT NOT NULL,
                display_name   TEXT DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS pid_identified_tags (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tag_code TEXT NOT NULL UNIQUE,
                examples TEXT DEFAULT '', name_sv TEXT DEFAULT '',
                comp_type TEXT DEFAULT '', confirmed INTEGER DEFAULT 0,
                source TEXT DEFAULT 'scan'
            );
            CREATE TABLE IF NOT EXISTS tag_database (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tag_code TEXT NOT NULL, name_sv TEXT DEFAULT '',
                name_en TEXT DEFAULT '', category TEXT DEFAULT '',
                standard TEXT DEFAULT '', source TEXT DEFAULT 'excel',
                active INTEGER DEFAULT 1
            );
            CREATE TABLE IF NOT EXISTS tag_database_settings (
                key TEXT PRIMARY KEY, value TEXT
            );
            CREATE TABLE IF NOT EXISTS reduction_factors (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                consequence_id  INTEGER NOT NULL REFERENCES consequences(id) ON DELETE CASCADE,
                description     TEXT NOT NULL DEFAULT '',
                rrf             INTEGER NOT NULL DEFAULT 10,
                active          INTEGER NOT NULL DEFAULT 1
            );
            CREATE TABLE IF NOT EXISTS equipment_catalog (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                tag            TEXT NOT NULL,
                original_tag   TEXT DEFAULT '',
                prefix         TEXT DEFAULT '',
                pid_page       INTEGER DEFAULT 0,
                equipment_type TEXT DEFAULT '',
                description    TEXT DEFAULT '',
                is_ocr         INTEGER DEFAULT 0,
                include        INTEGER DEFAULT 1
            );
            CREATE TABLE IF NOT EXISTS app_config (
                key TEXT PRIMARY KEY, value TEXT
            );
            CREATE TABLE IF NOT EXISTS consequence_categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL, sort_order INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS severity_definitions (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                severity_level INTEGER NOT NULL,
                category_id    INTEGER NOT NULL REFERENCES consequence_categories(id) ON DELETE CASCADE,
                description    TEXT    DEFAULT '',
                UNIQUE(severity_level, category_id)
            );
            CREATE TABLE IF NOT EXISTS deviations (
                id      INTEGER PRIMARY KEY AUTOINCREMENT,
                node_id INTEGER NOT NULL REFERENCES nodes(id) ON DELETE CASCADE,
                description TEXT NOT NULL DEFAULT 'Övrigt'
            );
            CREATE TABLE IF NOT EXISTS cause_markers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cause_id INTEGER NOT NULL REFERENCES causes(id) ON DELETE CASCADE,
                pid_page INTEGER DEFAULT 0, x REAL DEFAULT 0, y REAL DEFAULT 0,
                component_type TEXT DEFAULT '', component_tag TEXT DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS consequence_markers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                consequence_id INTEGER NOT NULL REFERENCES consequences(id) ON DELETE CASCADE,
                pid_page INTEGER DEFAULT 0, x REAL DEFAULT 0, y REAL DEFAULT 0,
                target_name TEXT DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS safeguard_markers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                safeguard_id INTEGER NOT NULL REFERENCES safeguards(id) ON DELETE CASCADE,
                pid_page INTEGER DEFAULT 0, x REAL DEFAULT 0, y REAL DEFAULT 0,
                tag TEXT DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS equipment_markers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                equipment_id INTEGER REFERENCES equipment_catalog(id) ON DELETE CASCADE,
                tag TEXT DEFAULT '',
                pid_page INTEGER DEFAULT 0, x REAL DEFAULT 0, y REAL DEFAULT 0,
                comp_type TEXT DEFAULT '',
                shape_outline TEXT DEFAULT '',
                confidence REAL DEFAULT 0,
                link_method TEXT DEFAULT '',
                label_x REAL DEFAULT NULL,
                label_y REAL DEFAULT NULL
            );
            CREATE TABLE IF NOT EXISTS pid_revisions (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                revision    TEXT NOT NULL DEFAULT '',
                notes       TEXT DEFAULT '',
                created_at  TEXT DEFAULT '',
                pdf_path    TEXT DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS pid_sheets (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                display_order INTEGER NOT NULL,
                physical_page INTEGER NOT NULL,
                sheet_name    TEXT DEFAULT '',
                revision_id   INTEGER DEFAULT NULL
            );
            CREATE TABLE IF NOT EXISTS standard_deviations (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                description TEXT NOT NULL,
                sort_order  INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS standard_causes (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                deviation_id INTEGER NOT NULL REFERENCES standard_deviations(id) ON DELETE CASCADE,
                description  TEXT NOT NULL,
                sort_order   INTEGER DEFAULT 0,
                object_id    INTEGER REFERENCES standard_objects(id) ON DELETE SET NULL
            );
            CREATE TABLE IF NOT EXISTS standard_objects (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                name       TEXT NOT NULL UNIQUE,
                sort_order INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS symbol_templates (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                name          TEXT NOT NULL UNIQUE,
                features_json TEXT NOT NULL,
                comp_type     TEXT DEFAULT '',
                created       TEXT DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS consequence_history (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                description TEXT NOT NULL UNIQUE
            );
            CREATE TABLE IF NOT EXISTS cause_descriptions (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                cause_id        INTEGER NOT NULL REFERENCES standard_causes(id) ON DELETE CASCADE,
                description     TEXT NOT NULL,
                sort_order      INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS board_annotations (
                id     INTEGER PRIMARY KEY AUTOINCREMENT,
                x      REAL DEFAULT 0,
                y      REAL DEFAULT 0,
                w      REAL DEFAULT 200,
                h      REAL DEFAULT 80,
                text   TEXT DEFAULT '',
                color  TEXT DEFAULT '#fff9c4'
            );
            CREATE TABLE IF NOT EXISTS consequence_steps (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                consequence_id INTEGER NOT NULL REFERENCES consequences(id) ON DELETE CASCADE,
                step           INTEGER NOT NULL,   -- 1..5
                text           TEXT    NOT NULL DEFAULT '',
                ref_tag        TEXT    DEFAULT '',
                node_key       TEXT    DEFAULT ''  -- konsekvensgraf-nod (för beroende kolumner)
            );
            CREATE TABLE IF NOT EXISTS node_markups (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                node_id    INTEGER NOT NULL REFERENCES nodes(id) ON DELETE CASCADE,
                type       TEXT NOT NULL DEFAULT 'polygon',
                points     TEXT DEFAULT '[]',
                label      TEXT DEFAULT '',
                color      TEXT DEFAULT '#1565C0',
                opacity    REAL DEFAULT 0.45,
                line_width INTEGER DEFAULT 12,
                font_size  INTEGER DEFAULT 12,
                visible    INTEGER DEFAULT 1,
                pid_page   INTEGER DEFAULT 0,
                sort_order INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS safeguard_cause_exclusions (
                safeguard_id INTEGER NOT NULL REFERENCES safeguards(id) ON DELETE CASCADE,
                cause_id     INTEGER NOT NULL REFERENCES causes(id)     ON DELETE CASCADE,
                PRIMARY KEY (safeguard_id, cause_id)
            );
            CREATE TABLE IF NOT EXISTS node_red_markups (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                node_id    INTEGER NOT NULL REFERENCES nodes(id) ON DELETE CASCADE,
                type       TEXT NOT NULL DEFAULT 'polygon',
                points     TEXT DEFAULT '[]',
                label      TEXT DEFAULT '',
                color      TEXT DEFAULT '#CC0000',
                opacity    REAL DEFAULT 1.0,
                line_width INTEGER DEFAULT 4,
                font_size  INTEGER DEFAULT 12,
                visible    INTEGER DEFAULT 1,
                pid_page   INTEGER DEFAULT 0,
                sort_order INTEGER DEFAULT 0,
                symbol_w   REAL DEFAULT 40,
                symbol_h   REAL DEFAULT 40,
                symbol_rot REAL DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS off_page_connector (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                pid_page   INTEGER NOT NULL,
                x_pdf      REAL,
                y_pdf      REAL,
                direction  TEXT,
                edge       TEXT,
                ref_text   TEXT,
                ref_sheet  TEXT,
                ref_line_id TEXT,
                media_type TEXT,
                weight     REAL DEFAULT 0.0,
                confidence REAL DEFAULT 0.5,
                raw_text   TEXT,
                ocr_used   INTEGER DEFAULT 0,
                analyzed_at TEXT
            );
            CREATE TABLE IF NOT EXISTS pid_connection (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                from_page       INTEGER,
                to_page         INTEGER,
                from_connector  INTEGER,
                to_connector    INTEGER,
                media_type      TEXT,
                weight          REAL DEFAULT 0.0,
                confidence      REAL DEFAULT 0.5,
                is_bidirectional INTEGER DEFAULT 0,
                is_ghost        INTEGER DEFAULT 0,
                ghost_ref       TEXT,
                warning         TEXT
            );
        """)

        # Second pass: fresh DBs now have all base tables — add migrated columns
        self._column_migrations()

        if not self.conn.execute("SELECT COUNT(*) FROM consequence_categories").fetchone()[0]:
            for i, name in enumerate(['Person', 'Miljö', 'Ekonomi', 'Anläggning', 'Rykte']):
                self.conn.execute(
                    "INSERT INTO consequence_categories (name, sort_order) VALUES (?,?)", (name, i))

        # Seed component types from hardcoded COMPONENT_TYPES if table is empty
        if not self.conn.execute("SELECT COUNT(*) FROM component_types").fetchone()[0]:
            from pid_viewer import COMPONENT_TYPES as _CT
            for sort_i, (comp_name, modes) in enumerate(_CT.items()):
                cur = self.conn.execute(
                    "INSERT INTO component_types (name, sort_order) VALUES (?,?)",
                    (comp_name, sort_i))
                comp_id = cur.lastrowid
                for mode_i, mode in enumerate(modes):
                    self.conn.execute(
                        "INSERT INTO failure_modes (component_id, description, sort_order)"
                        " VALUES (?,?,?)", (comp_id, mode, mode_i))

        self.conn.executescript("""
            CREATE TABLE IF NOT EXISTS component_types (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL, sort_order INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS failure_modes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                component_id INTEGER NOT NULL REFERENCES component_types(id) ON DELETE CASCADE,
                description TEXT NOT NULL DEFAULT '',
                freq_per_year REAL DEFAULT NULL,
                sort_order INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS consequence_severities (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                consequence_id INTEGER NOT NULL,
                category_id    INTEGER NOT NULL,
                severity       INTEGER NOT NULL DEFAULT 1,
                UNIQUE(consequence_id, category_id)
            );
            -- An optional post-barrier consequence level.  No row means
            -- "same as consequence_severities", which keeps the original
            -- severity as the default for Slutkonsekvens.
            CREATE TABLE IF NOT EXISTS consequence_final_severities (
                consequence_id INTEGER NOT NULL,
                category_id    INTEGER NOT NULL,
                severity       INTEGER NOT NULL DEFAULT 1,
                UNIQUE(consequence_id, category_id)
            );
            CREATE TABLE IF NOT EXISTS consequence_severity_exclusions (
                severity_id  INTEGER NOT NULL,
                safeguard_id INTEGER NOT NULL,
                PRIMARY KEY (severity_id, safeguard_id)
            );
            -- Deltagarmatris (2026-08-11, user request: "byggde en till
            -- flik med deltagare istället där man definerar förnamn,
            -- efternamn, roll på y axel och analystillfälen på x axeln
            -- så det blir en matris") — replaces the old free-text
            -- 'project_participants' app_config field. sort_order on
            -- both participants and sessions mirrors the reorderability
            -- convention used by consequence_categories/standard_deviations
            -- (no reorder UI yet, but the column is ready for it).
            CREATE TABLE IF NOT EXISTS participants (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                first_name TEXT NOT NULL DEFAULT '',
                last_name  TEXT NOT NULL DEFAULT '',
                role       TEXT DEFAULT '',
                sort_order INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS analysis_sessions (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                label      TEXT NOT NULL DEFAULT '',
                sort_order INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS participant_attendance (
                participant_id INTEGER NOT NULL REFERENCES participants(id) ON DELETE CASCADE,
                session_id     INTEGER NOT NULL REFERENCES analysis_sessions(id) ON DELETE CASCADE,
                attended       INTEGER NOT NULL DEFAULT 0,
                PRIMARY KEY (participant_id, session_id)
            );
            CREATE TABLE IF NOT EXISTS participant_columns (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                name       TEXT NOT NULL DEFAULT '',
                sort_order INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS participant_column_values (
                participant_id INTEGER NOT NULL REFERENCES participants(id) ON DELETE CASCADE,
                column_id      INTEGER NOT NULL REFERENCES participant_columns(id) ON DELETE CASCADE,
                value          TEXT DEFAULT '',
                PRIMARY KEY (participant_id, column_id)
            );
            CREATE TABLE IF NOT EXISTS project_revisions (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                label       TEXT NOT NULL DEFAULT '',
                date        TEXT DEFAULT '',
                description TEXT DEFAULT '',
                sort_order  INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS project_custom_fields (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                name       TEXT NOT NULL DEFAULT '',
                value      TEXT DEFAULT '',
                sort_order INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS node_types (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                name       TEXT NOT NULL DEFAULT '',
                sort_order INTEGER DEFAULT 0
            );
            -- Manuell sidrotation (2026-08-12, see NOTES.md) — a user-chosen
            -- extra rotation (0/90/180/270, clockwise) for one physical P&ID
            -- page, composed on top of (not replacing) the PDF's own /Rotate
            -- flag. Keyed by physical_page like the marker tables (not by
            -- pid_sheets.id) so it survives re-sorting the virtual sheet
            -- order, same rationale as cause_markers/consequence_markers.
            CREATE TABLE IF NOT EXISTS pid_page_rotation (
                physical_page INTEGER PRIMARY KEY,
                rotation      INTEGER NOT NULL DEFAULT 0
            );
            -- LOPA is deliberately revisioned and stores HAZOP ids as soft
            -- references.  A deleted HAZOP row must leave historical LOPA
            -- evidence visible (with a missing-source warning), never delete
            -- an approved analysis through a foreign-key cascade.
            CREATE TABLE IF NOT EXISTS safeguard_equipment_links (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                safeguard_id INTEGER NOT NULL REFERENCES safeguards(id) ON DELETE CASCADE,
                equipment_id INTEGER REFERENCES equipment_catalog(id) ON DELETE SET NULL,
                trigger_code TEXT NOT NULL DEFAULT '',
                trigger_custom TEXT NOT NULL DEFAULT '',
                tag_snapshot TEXT NOT NULL DEFAULT '',
                sort_order   INTEGER NOT NULL DEFAULT 0,
                UNIQUE(safeguard_id, equipment_id, trigger_code, trigger_custom)
            );
            CREATE TABLE IF NOT EXISTS lopa_records (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                display_number TEXT NOT NULL,
                sif_name       TEXT NOT NULL DEFAULT '',
                sis_name       TEXT NOT NULL DEFAULT '',
                archived       INTEGER NOT NULL DEFAULT 0,
                created_at     TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at     TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );
            CREATE UNIQUE INDEX IF NOT EXISTS idx_lopa_active_display_number
                ON lopa_records(display_number) WHERE archived=0;
            CREATE TABLE IF NOT EXISTS lopa_revisions (
                id                    INTEGER PRIMARY KEY AUTOINCREMENT,
                lopa_id               INTEGER NOT NULL REFERENCES lopa_records(id) ON DELETE CASCADE,
                label                 TEXT NOT NULL,
                status                TEXT NOT NULL DEFAULT 'Utkast',
                created_at            TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                created_by            TEXT NOT NULL DEFAULT '',
                locked_at             TEXT NOT NULL DEFAULT '',
                locked_by             TEXT NOT NULL DEFAULT '',
                unlock_reason         TEXT NOT NULL DEFAULT '',
                performed_by_text     TEXT NOT NULL DEFAULT '',
                approved_by_text      TEXT NOT NULL DEFAULT '',
                matrix_snapshot_json  TEXT NOT NULL DEFAULT '{}',
                notes                 TEXT NOT NULL DEFAULT '',
                UNIQUE(lopa_id, label)
            );
            CREATE TABLE IF NOT EXISTS lopa_source_scenarios (
                id                    INTEGER PRIMARY KEY AUTOINCREMENT,
                revision_id           INTEGER NOT NULL REFERENCES lopa_revisions(id) ON DELETE CASCADE,
                hazop_node_id         INTEGER REFERENCES nodes(id) ON DELETE SET NULL,
                hazop_deviation_id    INTEGER REFERENCES deviations(id) ON DELETE SET NULL,
                hazop_cause_id        INTEGER REFERENCES causes(id) ON DELETE SET NULL,
                origin_safeguard_id   INTEGER REFERENCES safeguards(id) ON DELETE SET NULL,
                equipment_id          INTEGER REFERENCES equipment_catalog(id) ON DELETE SET NULL,
                trigger_code          TEXT NOT NULL DEFAULT '',
                trigger_custom        TEXT NOT NULL DEFAULT '',
                cause_text            TEXT NOT NULL DEFAULT '',
                scenario_text         TEXT NOT NULL DEFAULT '',
                base_frequency        REAL DEFAULT NULL,
                frequency_origin      TEXT NOT NULL DEFAULT '',
                assumption_percent    REAL DEFAULT NULL,
                active                INTEGER NOT NULL DEFAULT 1,
                follows_hazop         INTEGER NOT NULL DEFAULT 1,
                detached_reason       TEXT NOT NULL DEFAULT '',
                source_missing        INTEGER NOT NULL DEFAULT 0,
                created_at            TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );
            CREATE UNIQUE INDEX IF NOT EXISTS idx_lopa_source_origin
                ON lopa_source_scenarios(revision_id, origin_safeguard_id)
                WHERE origin_safeguard_id IS NOT NULL;
            CREATE TABLE IF NOT EXISTS lopa_source_consequences (
                id                    INTEGER PRIMARY KEY AUTOINCREMENT,
                source_id             INTEGER NOT NULL REFERENCES lopa_source_scenarios(id) ON DELETE CASCADE,
                hazop_consequence_id  INTEGER REFERENCES consequences(id) ON DELETE SET NULL,
                hazop_category_id     INTEGER REFERENCES consequence_categories(id) ON DELETE SET NULL,
                category_key          TEXT NOT NULL DEFAULT '',
                category_name         TEXT NOT NULL DEFAULT '',
                severity              INTEGER NOT NULL DEFAULT 0,
                description           TEXT NOT NULL DEFAULT '',
                active                INTEGER NOT NULL DEFAULT 1,
                follows_hazop         INTEGER NOT NULL DEFAULT 1,
                detached_reason       TEXT NOT NULL DEFAULT '',
                source_missing        INTEGER NOT NULL DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS lopa_sensor_groups (
                id                    INTEGER PRIMARY KEY AUTOINCREMENT,
                revision_id           INTEGER NOT NULL REFERENCES lopa_revisions(id) ON DELETE CASCADE,
                voting                TEXT NOT NULL DEFAULT '1oo1',
                sort_order            INTEGER NOT NULL DEFAULT 0,
                needs_voting_review   INTEGER NOT NULL DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS lopa_sensor_members (
                id                    INTEGER PRIMARY KEY AUTOINCREMENT,
                group_id              INTEGER NOT NULL REFERENCES lopa_sensor_groups(id) ON DELETE CASCADE,
                equipment_id          INTEGER REFERENCES equipment_catalog(id) ON DELETE SET NULL,
                origin_safeguard_id   INTEGER REFERENCES safeguards(id) ON DELETE SET NULL,
                trigger_code          TEXT NOT NULL DEFAULT '',
                trigger_custom        TEXT NOT NULL DEFAULT '',
                tag_snapshot          TEXT NOT NULL DEFAULT '',
                active                INTEGER NOT NULL DEFAULT 1,
                sort_order            INTEGER NOT NULL DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS lopa_final_groups (
                id                    INTEGER PRIMARY KEY AUTOINCREMENT,
                revision_id           INTEGER NOT NULL REFERENCES lopa_revisions(id) ON DELETE CASCADE,
                voting                TEXT NOT NULL DEFAULT '1oo1',
                sort_order            INTEGER NOT NULL DEFAULT 0,
                needs_voting_review   INTEGER NOT NULL DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS lopa_final_members (
                id                    INTEGER PRIMARY KEY AUTOINCREMENT,
                group_id              INTEGER NOT NULL REFERENCES lopa_final_groups(id) ON DELETE CASCADE,
                equipment_id          INTEGER REFERENCES equipment_catalog(id) ON DELETE SET NULL,
                name_snapshot         TEXT NOT NULL DEFAULT '',
                action_text           TEXT NOT NULL DEFAULT '',
                active                INTEGER NOT NULL DEFAULT 1,
                sort_order            INTEGER NOT NULL DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS lopa_barriers (
                id                    INTEGER PRIMARY KEY AUTOINCREMENT,
                revision_id           INTEGER NOT NULL REFERENCES lopa_revisions(id) ON DELETE CASCADE,
                source_id             INTEGER REFERENCES lopa_source_scenarios(id) ON DELETE CASCADE,
                source_safeguard_id   INTEGER REFERENCES safeguards(id) ON DELETE SET NULL,
                sg_type               TEXT NOT NULL DEFAULT 'Övrigt',
                description           TEXT NOT NULL DEFAULT '',
                rrf                   REAL NOT NULL DEFAULT 1,
                independent            INTEGER NOT NULL DEFAULT 1,
                manual                INTEGER NOT NULL DEFAULT 0,
                applies_all_categories INTEGER NOT NULL DEFAULT 0,
                active                INTEGER NOT NULL DEFAULT 1,
                follows_hazop         INTEGER NOT NULL DEFAULT 1,
                detached_reason       TEXT NOT NULL DEFAULT '',
                source_missing        INTEGER NOT NULL DEFAULT 0,
                sort_order            INTEGER NOT NULL DEFAULT 0
            );
            CREATE UNIQUE INDEX IF NOT EXISTS idx_lopa_barrier_source_safeguard
                ON lopa_barriers(source_id, source_safeguard_id)
                WHERE source_safeguard_id IS NOT NULL;
            CREATE TABLE IF NOT EXISTS lopa_barrier_categories (
                barrier_id            INTEGER NOT NULL REFERENCES lopa_barriers(id) ON DELETE CASCADE,
                category_key          TEXT NOT NULL,
                active                INTEGER NOT NULL DEFAULT 1,
                PRIMARY KEY (barrier_id, category_key)
            );
            CREATE TABLE IF NOT EXISTS lopa_escalation_rows (
                id                    INTEGER PRIMARY KEY AUTOINCREMENT,
                source_id             INTEGER NOT NULL REFERENCES lopa_source_scenarios(id) ON DELETE CASCADE,
                category_key          TEXT NOT NULL,
                factor_values_json    TEXT NOT NULL DEFAULT '{}',
                reason                TEXT NOT NULL DEFAULT '',
                active                INTEGER NOT NULL DEFAULT 1,
                UNIQUE(source_id, category_key)
            );
            CREATE TABLE IF NOT EXISTS lopa_change_log (
                id                    INTEGER PRIMARY KEY AUTOINCREMENT,
                lopa_id               INTEGER NOT NULL REFERENCES lopa_records(id) ON DELETE CASCADE,
                revision_id           INTEGER REFERENCES lopa_revisions(id) ON DELETE SET NULL,
                action                TEXT NOT NULL,
                detail                TEXT NOT NULL DEFAULT '',
                created_at            TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                actor                 TEXT NOT NULL DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS lopa_comments (
                id                    INTEGER PRIMARY KEY AUTOINCREMENT,
                revision_id           INTEGER NOT NULL REFERENCES lopa_revisions(id) ON DELETE CASCADE,
                author                TEXT NOT NULL DEFAULT '',
                body                  TEXT NOT NULL DEFAULT '',
                created_at            TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );
        """)

        # Seed missing deviation_id for existing causes
        orphan_nodes = [r[0] for r in self.conn.execute(
            "SELECT DISTINCT node_id FROM causes WHERE deviation_id IS NULL").fetchall()]
        for nid in orphan_nodes:
            row = self.conn.execute(
                "SELECT id FROM deviations WHERE node_id=? AND description='Övrigt' LIMIT 1",
                (nid,)).fetchone()
            if row:
                dev_id = row[0]
            else:
                cur = self.conn.execute(
                    "INSERT INTO deviations (node_id, description) VALUES (?, 'Övrigt')", (nid,))
                dev_id = cur.lastrowid
            self.conn.execute(
                "UPDATE causes SET deviation_id=? WHERE node_id=? AND deviation_id IS NULL",
                (dev_id, nid))

        # Seed standard_deviations template library if empty
        if not self.conn.execute("SELECT COUNT(*) FROM standard_deviations").fetchone()[0]:
            _STD_CAUSES = {
                "Lågt flöde":    ["Stängd ventil", "Delvis stängd ventil", "Igensatt filter/sil",
                                   "Stoppad pump", "Igensatt rör/ledning", "Läckage uppströms",
                                   "Fel på reglerventil (ej öppnar)"],
                "Högt flöde":    ["Felöppen ventil", "Fel på reglerventil (ej stänger)",
                                   "Ökat drifttryck uppströms", "Ökad pumpkapacitet"],
                "Missriktat flöde": ["Felaktig rörledningsdragning", "Fel rörkoppling",
                                     "Backventil saknas / ur funktion"],
                "Omvänt flöde":  ["Backventil saknas / ur funktion", "Pumpfel – flöde vänds",
                                   "Tryckfall uppströms"],
                "Högt tryck":    ["Stängd utloppsventil", "Blockerat utlopp",
                                   "Ökat inflöde", "Övervärmd gas/vätska", "Felaktig tryckreglering"],
                "Lågt tryck":    ["Läckage i system", "Otäta flänsar/koppling",
                                   "Öppet/läckande utlopp", "Pumphaveri"],
                "Hög nivå":      ["Öppet inlopp", "Stängd utloppsventil", "Felaktig nivåreglering",
                                   "Läckage till kärl"],
                "Låg nivå":      ["Läckage i botten/sida", "Felaktig nivåreglering",
                                   "Stängd inloppsventil", "Pumphaveri"],
                "Hög temperatur": ["Värmeväxlare ur funktion", "Övervärmd inkommande fluid",
                                    "Felaktig temperaturreglering", "Exoterm reaktion"],
                "Låg temperatur": ["Kylmedelfel", "Underkylning av inkommande fluid",
                                    "Felaktig temperaturreglering"],
                "Avvikande sammansättning": ["Fel råvara", "Förorenad råvara",
                                              "Felaktig dosering", "Läckage av annat medium"],
                "Bortfall av hjälpsystem": ["Strömavbrott", "Instrumentluftsfel",
                                             "Kylarfel", "Automatikfel"],
                "Drift":         ["Mänskligt fel vid drift", "Felaktig procedur",
                                   "Kommunikationsfel"],
                "Underhåll":     ["Arbete på trycksatt system", "Felaktig isolering",
                                   "Verktyg kvar i system"],
                "Start-up / Shut-down": ["Felaktig sekvens", "Valves i fel läge",
                                          "Instrument ej kalibrerade"],
                "Övrigt":        [],
            }
            for sort_i, dev_name in enumerate(DEVIATION_TYPES):
                cur = self.conn.execute(
                    "INSERT INTO standard_deviations (description, sort_order) VALUES (?,?)",
                    (dev_name, sort_i))
                dev_tmpl_id = cur.lastrowid
                for cause_i, c_desc in enumerate(_STD_CAUSES.get(dev_name, [])):
                    self.conn.execute(
                        "INSERT INTO standard_causes (deviation_id, description, sort_order)"
                        " VALUES (?,?,?)", (dev_tmpl_id, c_desc, cause_i))

        # Seed standard objects FIRST — causes need object IDs
        if not self.conn.execute(
                "SELECT value FROM app_config WHERE key='std_objects_seeded_v1'").fetchone():
            _seed_standard_objects(self.conn)
            self.conn.execute(
                "INSERT OR REPLACE INTO app_config (key,value) VALUES ('std_objects_seeded_v1','1')")

        # Seed v1/v2/v3 legacy sentinels (mark as done so old code doesn't re-run)
        for sv in ('comp_causes_seeded_v1', 'comp_causes_seeded_v2', 'comp_causes_seeded_v3',
                   'causes_object_id_migrated_v1'):
            self.conn.execute(
                f"INSERT OR IGNORE INTO app_config (key,value) VALUES ('{sv}','legacy')")

        # Seed full object-keyed causes (v4) — replaces old comp_type seeding
        if not self.conn.execute(
                "SELECT value FROM app_config WHERE key='comp_causes_seeded_v4'").fetchone():
            _seed_component_causes(self.conn)
            _migrate_causes_to_object_id(self.conn)   # backfill existing rows
            self.conn.execute(
                "INSERT OR REPLACE INTO app_config (key,value) VALUES ('comp_causes_seeded_v4','1')")

        # v5: replace verbose/specific causes with generic ones + seed frequencies
        if not self.conn.execute(
                "SELECT value FROM app_config WHERE key='comp_causes_seeded_v5'").fetchone():
            # Delete all seeded standard causes and reseed with generic versions
            self.conn.execute("DELETE FROM standard_causes")
            _seed_component_causes(self.conn)
            self.conn.execute(
                "INSERT OR REPLACE INTO app_config (key,value) VALUES ('comp_causes_seeded_v5','1')")

        self._migrate_reduced_standard_catalog()

        # Older migrations could seed the deviation template library twice.
        # That made the per-node seeding loop create two equal deviations in
        # the worksheet.  Collapse only generic (equipment_id IS NULL)
        # duplicates; equal descriptions tied to different equipment are
        # intentional and must remain separate.
        if not self.conn.execute(
                "SELECT 1 FROM app_config WHERE key='dedupe_deviations_v1'").fetchone():
            try:
                template_groups = self.conn.execute(
                    "SELECT description, GROUP_CONCAT(id) ids FROM standard_deviations "
                    "GROUP BY description HAVING COUNT(*) > 1").fetchall()
                for group in template_groups:
                    ids = [int(v) for v in str(group['ids']).split(',') if v]
                    keep = min(ids)
                    for duplicate in ids:
                        if duplicate == keep:
                            continue
                        self.conn.execute(
                            "UPDATE standard_causes SET deviation_id=? WHERE deviation_id=?",
                            (keep, duplicate))
                        self.conn.execute("DELETE FROM standard_deviations WHERE id=?", (duplicate,))

                deviation_groups = self.conn.execute(
                    "SELECT node_id, description, GROUP_CONCAT(id) ids "
                    "FROM deviations WHERE equipment_id IS NULL "
                    "GROUP BY node_id, description HAVING COUNT(*) > 1").fetchall()
                for group in deviation_groups:
                    ids = [int(v) for v in str(group['ids']).split(',') if v]
                    cause_counts = {
                        dev_id: self.conn.execute(
                            "SELECT COUNT(*) FROM causes WHERE deviation_id=?", (dev_id,)
                        ).fetchone()[0]
                        for dev_id in ids
                    }
                    keep = max(ids, key=lambda dev_id: (cause_counts[dev_id], -dev_id))
                    for duplicate in ids:
                        if duplicate == keep:
                            continue
                        self.conn.execute(
                            "UPDATE causes SET deviation_id=? WHERE deviation_id=?",
                            (keep, duplicate))
                        self.conn.execute("DELETE FROM deviations WHERE id=?", (duplicate,))
                self.conn.execute(
                    "INSERT OR REPLACE INTO app_config (key,value) VALUES "
                    "('dedupe_deviations_v1','1')")
                self.commit()
            except Exception:
                logging.warning("Deviation duplicate cleanup failed", exc_info=True)

        # A second edge case appeared after equipment-specific deviations were
        # introduced: an old empty generic template row could remain beside a
        # populated equipment-bound row with the same description.  The empty
        # generic row is only a seed artefact, so remove it; populated rows
        # for different equipment are retained.
        if not self.conn.execute(
                "SELECT 1 FROM app_config WHERE key='dedupe_deviations_v2'").fetchone():
            try:
                candidates = self.conn.execute(
                    "SELECT d.id, d.node_id, d.description "
                    "FROM deviations d WHERE d.equipment_id IS NULL "
                    "AND NOT EXISTS (SELECT 1 FROM causes c WHERE c.deviation_id=d.id) "
                    "AND EXISTS (SELECT 1 FROM deviations d2 "
                    "            WHERE d2.node_id=d.node_id AND d2.description=d.description "
                    "              AND d2.equipment_id IS NOT NULL "
                    "              AND EXISTS (SELECT 1 FROM causes c2 WHERE c2.deviation_id=d2.id))"
                ).fetchall()
                for row in candidates:
                    self.conn.execute("DELETE FROM deviations WHERE id=?", (row['id'],))
                self.conn.execute(
                    "INSERT OR REPLACE INTO app_config (key,value) VALUES "
                    "('dedupe_deviations_v2','1')")
                self.commit()
            except Exception:
                logging.warning("Equipment deviation duplicate cleanup failed", exc_info=True)

        # Prevent the same migration regression from reintroducing generic
        # duplicates, while allowing equipment-specific deviations to coexist.
        try:
            self.conn.execute(
                "CREATE UNIQUE INDEX IF NOT EXISTS idx_deviations_node_description_generic "
                "ON deviations(node_id, description) WHERE equipment_id IS NULL")
            self.conn.execute(
                "CREATE UNIQUE INDEX IF NOT EXISTS idx_standard_deviations_description "
                "ON standard_deviations(description)")
        except Exception:
            logging.warning("Could not create deviation uniqueness indexes", exc_info=True)

        # Ensure every node has all standard deviations from template library.
        # dict.fromkeys also protects fresh databases if a legacy template
        # source ever contains the same description twice.
        # Archived deviations remain available in the library, but must not
        # be copied into newly created nodes.  The old unfiltered query made
        # the retired library reappear after a fresh database migration.
        std_devs = list(dict.fromkeys(r[0] for r in self.conn.execute(
            "SELECT description FROM standard_deviations "
            "WHERE active=1 ORDER BY sort_order").fetchall()))
        if not std_devs:
            std_devs = DEVIATION_TYPES
        all_nodes = [r[0] for r in self.conn.execute("SELECT id FROM nodes").fetchall()]
        for nid in all_nodes:
            existing = {r[0] for r in self.conn.execute(
                "SELECT description FROM deviations WHERE node_id=?", (nid,)).fetchall()}
            for dev_type in std_devs:
                if dev_type not in existing:
                    self.conn.execute(
                        "INSERT INTO deviations (node_id, description) VALUES (?,?)",
                        (nid, dev_type))

        _sync_cause_likelihoods_from_frequency(self.conn)

        # Migration v1: collapse __PFX__ sentinels and full tags into bare prefixes
        if not self.conn.execute(
                "SELECT value FROM app_config WHERE key='tag_memory_prefix_only_v1'").fetchone():
            try:
                rows = self.conn.execute(
                    "SELECT tag, comp_type, phash, usage_count FROM study_tag_memory").fetchall()
                self.conn.execute("DELETE FROM study_tag_memory")
                merged: dict = {}
                for r in rows:
                    raw = r[0]
                    if raw.upper().startswith('__PFX__'):
                        raw = raw[7:]
                    pfx = _tag_letter_prefix(raw) if raw else ''
                    if not pfx:
                        continue
                    prev = merged.get(pfx)
                    if prev is None or r[3] > prev[2]:
                        merged[pfx] = (r[1], r[2] or '', r[3] or 1)
                now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
                for pfx, (ct, ph, cnt) in merged.items():
                    self.conn.execute(
                        "INSERT OR IGNORE INTO study_tag_memory "
                        "(tag,comp_type,phash,usage_count,updated) VALUES (?,?,?,?,?)",
                        (pfx, ct, ph, cnt, now))
            except Exception:
                pass
            self.conn.execute(
                "INSERT OR REPLACE INTO app_config (key,value)"
                " VALUES ('tag_memory_prefix_only_v1','1')")

        # Migration v2: change from single-key (tag PK) to composite key (tag,comp_type).
        # Old DBs have tag as sole PRIMARY KEY — recreate with composite key so the
        # same prefix can accumulate counts for multiple types independently.
        if not self.conn.execute(
                "SELECT value FROM app_config WHERE key='tag_memory_composite_v1'").fetchone():
            try:
                old_rows = self.conn.execute(
                    "SELECT tag, comp_type, comp_tag, phash, usage_count, updated, "
                    "COALESCE(active,1) FROM study_tag_memory").fetchall()
                self.conn.executescript("""
                    DROP TABLE IF EXISTS study_tag_memory_old;
                    ALTER TABLE study_tag_memory RENAME TO study_tag_memory_old;
                    CREATE TABLE study_tag_memory (
                        tag         TEXT NOT NULL,
                        comp_type   TEXT NOT NULL DEFAULT '',
                        comp_tag    TEXT NOT NULL DEFAULT '',
                        phash       TEXT NOT NULL DEFAULT '',
                        usage_count INTEGER NOT NULL DEFAULT 1,
                        updated     TEXT NOT NULL DEFAULT '',
                        active      INTEGER NOT NULL DEFAULT 1,
                        PRIMARY KEY (tag, comp_type)
                    );
                """)
                for r in old_rows:
                    self.conn.execute(
                        "INSERT OR IGNORE INTO study_tag_memory "
                        "(tag,comp_type,comp_tag,phash,usage_count,updated,active) "
                        "VALUES (?,?,?,?,?,?,?)",
                        (r[0], r[1], r[2], r[3], r[4], r[5], r[6]))
                self.conn.execute("DROP TABLE IF EXISTS study_tag_memory_old")
            except Exception:
                pass
            self.conn.execute(
                "INSERT OR REPLACE INTO app_config (key,value)"
                " VALUES ('tag_memory_composite_v1','1')")

        # Fix accidental active=0 from a previous bad implementation that
        # deactivated entries in upsert_tag_memory.  Restore active=1 for
        # all rows so the count-based winner logic works correctly.
        # Only runs once; user's intentional active=0 (via panel checkbox)
        # is re-applied afterward if they choose to.
        if not self.conn.execute(
                "SELECT value FROM app_config "
                "WHERE key='tag_memory_restore_active_v1'").fetchone():
            try:
                self.conn.execute(
                    "UPDATE study_tag_memory SET active=1 WHERE usage_count > 0")
            except Exception:
                pass
            self.conn.execute(
                "INSERT OR REPLACE INTO app_config (key,value)"
                " VALUES ('tag_memory_restore_active_v1','1')")

        self.commit()

    def _validate_schema(self):
        """Validate that all expected tables and critical columns exist.

        Runs after migrations to ensure the database schema is complete.
        Logs warnings for missing columns that are required for functionality.
        """
        logging.info("Database: validating schema...")

        # Critical tables that must exist
        critical_tables = {
            'nodes', 'causes', 'consequences', 'safeguards', 'deviations',
            'cause_markers', 'consequence_markers', 'safeguard_markers',
            'standard_causes', 'standard_deviations', 'app_config'
        }

        # Map of table -> list of critical columns that should exist
        critical_columns = {
            'nodes': ['id', 'name', 'markup_points', 'markup_style', 'pid_page'],
            'causes': ['id', 'description', 'likelihood', 'deviation_id', 'comp_type'],
            'consequences': ['id', 'description', 'severity', 'category'],
            'safeguards': ['id', 'description', 'rrf', 'sg_type'],
            'deviations': ['id', 'node_id', 'description'],
            'cause_markers': ['id', 'cause_id', 'pid_page', 'x', 'y'],
            'consequence_markers': ['id', 'consequence_id', 'pid_page', 'x', 'y'],
            'safeguard_markers': ['id', 'safeguard_id', 'pid_page', 'x', 'y'],
            'standard_causes': ['id', 'deviation_id', 'description', 'comp_type'],
            'standard_deviations': ['id', 'description'],
            'app_config': ['key', 'value'],
        }

        missing_tables = []
        for table in critical_tables:
            try:
                cursor = self.conn.execute(f"PRAGMA table_info({table})")
                if not cursor.fetchall():
                    missing_tables.append(table)
            except sqlite3.OperationalError:
                missing_tables.append(table)

        if missing_tables:
            logging.error(f"Database validation: missing tables: {', '.join(missing_tables)}")

        # Check for critical columns
        missing_columns = {}
        for table, columns in critical_columns.items():
            if table in missing_tables:
                continue  # Skip tables that are already missing

            try:
                cursor = self.conn.execute(f"PRAGMA table_info({table})")
                existing_cols = {row[1] for row in cursor.fetchall()}  # Column name is index 1
                missing = [col for col in columns if col not in existing_cols]
                if missing:
                    missing_columns[table] = missing
            except sqlite3.OperationalError as e:
                logging.error(f"Database validation: cannot check {table}: {e}")

        if missing_columns:
            for table, cols in missing_columns.items():
                logging.warning(f"Database validation: {table} missing columns: {', '.join(cols)}")

        if not missing_tables and not missing_columns:
            logging.info("Database: schema validation passed — all critical tables and columns present")
        else:
            logging.warning("Database: schema validation found issues — app may have reduced functionality")

    # ── Config ────────────────────────────────────────────────────────────────
    def get_config(self, key, default=None):
        try:
            row = self.conn.execute(
                "SELECT value FROM app_config WHERE key=?", (key,)).fetchone()
            return row['value'] if row else default
        except Exception:
            return default

    def set_config(self, key, value):
        try:
            self.conn.execute(
                "INSERT OR REPLACE INTO app_config (key,value) VALUES (?,?)", (key, value))
            self.commit()
        except Exception:
            pass

    _DEFAULT_PALETTE = [
        {'name': 'Kritisk', 'color': '#e74c3c', 'fg_color': '#ffffff'},
        {'name': 'Hög',     'color': '#e67e22', 'fg_color': '#ffffff'},
        {'name': 'Medium',  'color': '#f39c12', 'fg_color': '#000000'},
        {'name': 'Låg',     'color': '#27ae60', 'fg_color': '#ffffff'},
    ]

    def get_color_palette(self):
        val = self.get_config('color_palette')
        if val:
            try:
                return json.loads(val)
            except Exception:
                pass
        return list(self._DEFAULT_PALETTE)

    def set_color_palette(self, palette):
        self.set_config('color_palette', json.dumps(palette))

    def get_risk_matrix(self):
        val = self.get_config('risk_matrix')
        if val:
            try:
                return json.loads(val)
            except Exception:
                pass
        return None

    def set_risk_matrix(self, cfg):
        """Store risk matrix and immediately bind the cache to this project.

        The risk popup reads the process-wide cache. Invalidating alone left
        its database reference pointing to whichever project was loaded
        earlier, so a save could redraw the popup with old labels/boundaries.
        """
        cfg = self._risk_matrix_copy(cfg)
        # Older projects did not embed categories in the matrix.  Every newly
        # saved matrix does, so a saved configuration is also a complete
        # reusable template rather than only a coloured grid.
        if not cfg.get('consequence_categories'):
            cfg['consequence_categories'] = self._project_category_template(cfg['rows'])
        self.set_config('risk_matrix', json.dumps(cfg))
        _risk_matrix_cache.load(self)

    # ── LOPA configuration and revisioned data ──────────────────────────────
    # LOPA is not a second risk matrix.  Its project settings live inside the
    # normal risk-matrix template so colours, category keys, descriptions and
    # the user-defined TEL scale can be snapshotted together in a revision.
    def lopa_matrix_config(self):
        matrix = self._risk_matrix_copy(self.get_risk_matrix() or DEFAULT_MATRIX)
        return normalise_lopa_config(matrix.get('lopa'), matrix)

    def set_lopa_matrix_config(self, config):
        matrix = self._risk_matrix_copy(self.get_risk_matrix() or DEFAULT_MATRIX)
        matrix['lopa'] = normalise_lopa_config(config, matrix)
        self.set_risk_matrix(matrix)
        return matrix['lopa']

    def safeguard_types(self):
        """Project-configurable safeguard types shared by HAZOP and LOPA."""
        return list(self.lopa_matrix_config().get('safeguard_types') or
                    DEFAULT_SAFEGUARD_TYPES)

    def set_safeguard_types(self, values):
        config = self.lopa_matrix_config()
        config['safeguard_types'] = list(values or [])
        return self.set_lopa_matrix_config(config)

    def _lopa_matrix_snapshot(self):
        """Capture the full active matrix, including LOPA/TEL metadata."""
        matrix = self._risk_matrix_copy(self.get_risk_matrix() or DEFAULT_MATRIX)
        matrix['lopa'] = normalise_lopa_config(matrix.get('lopa'), matrix)
        return matrix

    @staticmethod
    def _lopa_number_sort_key(value):
        text = str(value or '').strip()
        try:
            return (0, int(text), text)
        except ValueError:
            return (1, 0, text.casefold())

    def _next_lopa_display_number(self):
        numeric = []
        for row in self.conn.execute("SELECT display_number FROM lopa_records").fetchall():
            try:
                numeric.append(int(str(row['display_number']).strip()))
            except (TypeError, ValueError):
                continue
        return f"{(max(numeric) if numeric else 0) + 1:03d}"

    def _log_lopa(self, lopa_id, revision_id, action, detail='', actor=''):
        self.conn.execute(
            "INSERT INTO lopa_change_log(lopa_id,revision_id,action,detail,actor) "
            "VALUES (?,?,?,?,?)",
            (lopa_id, revision_id, str(action or ''), str(detail or ''), str(actor or '')))

    def _lopa_revision_row(self, revision_id):
        row = self.conn.execute(
            "SELECT * FROM lopa_revisions WHERE id=?", (revision_id,)).fetchone()
        return dict(row) if row else None

    def _assert_lopa_revision_editable(self, revision_id):
        revision = self._lopa_revision_row(revision_id)
        if not revision:
            raise ValueError('LOPA-revisionen finns inte längre.')
        if revision['status'] in ('Låst', 'Godkänd', 'Arkiverad'):
            raise PermissionError('LOPA-revisionen är låst. Lås upp den innan du ändrar data.')
        return revision

    def lopa_records(self, include_archived=False):
        sql = "SELECT * FROM lopa_records"
        if not include_archived:
            sql += " WHERE archived=0"
        rows = [dict(row) for row in self.conn.execute(sql).fetchall()]
        return sorted(rows, key=lambda row: self._lopa_number_sort_key(row['display_number']))

    def get_lopa_record(self, lopa_id):
        row = self.conn.execute("SELECT * FROM lopa_records WHERE id=?", (lopa_id,)).fetchone()
        return dict(row) if row else None

    def lopa_revisions(self, lopa_id):
        return [dict(row) for row in self.conn.execute(
            "SELECT * FROM lopa_revisions WHERE lopa_id=? ORDER BY id", (lopa_id,)).fetchall()]

    def get_lopa_revision(self, revision_id):
        return self._lopa_revision_row(revision_id)

    def current_lopa_revision(self, lopa_id):
        row = self.conn.execute(
            "SELECT * FROM lopa_revisions WHERE lopa_id=? "
            "ORDER BY id DESC LIMIT 1", (lopa_id,)).fetchone()
        return dict(row) if row else None

    def lopa_revision_matrix(self, revision_id):
        revision = self._lopa_revision_row(revision_id)
        if not revision:
            return self._lopa_matrix_snapshot()
        try:
            raw = json.loads(revision.get('matrix_snapshot_json') or '{}')
        except (TypeError, ValueError, json.JSONDecodeError):
            raw = {}
        if not isinstance(raw, dict) or not raw:
            raw = self._lopa_matrix_snapshot()
        return self._risk_matrix_copy(raw)

    def create_lopa(self, display_number=None, sif_number='', sif_name='', sis_name='',
                    created_by='', notes='', allow_archived_reuse=False):
        """Create an empty, editable LOPA with revision ``00``.

        Display numbers are deliberately independent from the internal key.
        Automatic numbers never reuse an archived number; a manual archived
        reuse requires the caller to opt in after showing a warning.
        """
        manual_number = display_number is not None
        display_number = str(display_number or self._next_lopa_display_number()).strip()
        if not display_number:
            raise ValueError('Ange ett LOPA-nummer.')
        duplicate = self.conn.execute(
            "SELECT id,archived FROM lopa_records WHERE lower(display_number)=lower(?) "
            "ORDER BY archived ASC LIMIT 1", (display_number,)).fetchone()
        if duplicate and not duplicate['archived']:
            raise ValueError(f'LOPA {display_number} används redan.')
        if duplicate and duplicate['archived'] and manual_number and not allow_archived_reuse:
            raise ValueError(
                f'LOPA {display_number} är arkiverad. Bekräfta återanvändning uttryckligen.')

        snapshot = json.dumps(self._lopa_matrix_snapshot())
        with self.history_group():
            try:
                self.conn.execute('BEGIN')
                cur = self.conn.execute(
                    "INSERT INTO lopa_records(display_number,sif_number,sif_name,sis_name) VALUES (?,?,?,?)",
                    (display_number, str(sif_number or ''), str(sif_name or ''), str(sis_name or '')))
                lopa_id = cur.lastrowid
                revision_id = self.conn.execute(
                    "INSERT INTO lopa_revisions(lopa_id,label,status,created_by,matrix_snapshot_json,notes) "
                    "VALUES (?,?,?,?,?,?)",
                    (lopa_id, '00', 'Utkast', str(created_by or ''), snapshot,
                     str(notes or ''))).lastrowid
                self._log_lopa(lopa_id, revision_id, 'created', 'Ny tom LOPA-revision 00', created_by)
                self.commit()
            except Exception:
                self.conn.rollback()
                raise
        return {'lopa_id': lopa_id, 'revision_id': revision_id}

    def update_lopa_record(self, lopa_id, display_number=None, sif_number=None,
                           sif_name=None, sis_name=None):
        row = self.get_lopa_record(lopa_id)
        if not row:
            raise ValueError('LOPA:n finns inte längre.')
        values = []
        assignments = []
        if display_number is not None:
            value = str(display_number).strip()
            if not value:
                raise ValueError('LOPA-nummer får inte vara tomt.')
            duplicate = self.conn.execute(
                "SELECT id FROM lopa_records WHERE archived=0 AND lower(display_number)=lower(?) AND id<>?",
                (value, lopa_id)).fetchone()
            if duplicate:
                raise ValueError(f'LOPA {value} används redan.')
            assignments.append('display_number=?'); values.append(value)
        if sif_number is not None:
            assignments.append('sif_number=?'); values.append(str(sif_number or ''))
        if sif_name is not None:
            assignments.append('sif_name=?'); values.append(str(sif_name))
        if sis_name is not None:
            assignments.append('sis_name=?'); values.append(str(sis_name))
        if not assignments:
            return
        assignments.append("updated_at=CURRENT_TIMESTAMP")
        values.append(lopa_id)
        self.conn.execute(
            f"UPDATE lopa_records SET {', '.join(assignments)} WHERE id=?", values)
        self._log_lopa(lopa_id, None, 'updated', 'Uppdaterade LOPA-huvud')
        self.commit()

    def archive_lopa(self, lopa_id, archived=True, actor=''):
        record = self.get_lopa_record(lopa_id)
        if not record:
            return False
        self.conn.execute("UPDATE lopa_records SET archived=?,updated_at=CURRENT_TIMESTAMP WHERE id=?",
                          (int(bool(archived)), lopa_id))
        self._log_lopa(lopa_id, None, 'archived' if archived else 'restored', '', actor)
        self.commit()
        return True

    def _next_lopa_revision_label(self, lopa_id):
        numbers = []
        for row in self.lopa_revisions(lopa_id):
            try:
                numbers.append(int(str(row['label']).strip()))
            except (TypeError, ValueError):
                continue
        return f"{(max(numbers) if numbers else -1) + 1:02d}"

    def create_lopa_revision(self, lopa_id, source_revision_id=None, label=None,
                             created_by='', notes=''):
        """Deep-copy a revision, retaining evidence while starting a new draft."""
        source = (self._lopa_revision_row(source_revision_id)
                  if source_revision_id else self.current_lopa_revision(lopa_id))
        if source and source['lopa_id'] != lopa_id:
            raise ValueError('Källrevisionen hör till en annan LOPA.')
        label = str(label or self._next_lopa_revision_label(lopa_id)).strip()
        if not label:
            raise ValueError('Revisionsbeteckning saknas.')
        if self.conn.execute("SELECT 1 FROM lopa_revisions WHERE lopa_id=? AND label=?",
                             (lopa_id, label)).fetchone():
            raise ValueError(f'Revision {label} finns redan.')
        snapshot = (source.get('matrix_snapshot_json') if source else
                    json.dumps(self._lopa_matrix_snapshot()))
        with self.history_group():
            try:
                self.conn.execute('BEGIN')
                # A revision begins with the same document assumptions as its
                # predecessor.  It is still a distinct editable snapshot: a
                # later change must never rewrite the locked source revision.
                new_revision_id = self.conn.execute(
                    "INSERT INTO lopa_revisions("
                    "lopa_id,label,status,created_by,matrix_snapshot_json,notes,"
                    "performed_by_text,approved_by_text,document_date,"
                    "dimensioning_category_key,additional_actions,additional_requirements,"
                    "process_safety_time) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    (lopa_id, label, 'Utkast', str(created_by or ''), snapshot,
                     str(notes or (source or {}).get('notes') or ''),
                     str((source or {}).get('performed_by_text') or ''),
                     str((source or {}).get('approved_by_text') or ''),
                     str((source or {}).get('document_date') or ''),
                     str((source or {}).get('dimensioning_category_key') or ''),
                     str((source or {}).get('additional_actions') or ''),
                     str((source or {}).get('additional_requirements') or ''),
                     (source or {}).get('process_safety_time'))).lastrowid
                if source:
                    self._copy_lopa_revision_children(source['id'], new_revision_id)
                self._log_lopa(lopa_id, new_revision_id, 'revision-created',
                               f'Ny revision {label}', created_by)
                self.commit()
            except Exception:
                self.conn.rollback()
                raise
        return new_revision_id

    def _copy_lopa_revision_children(self, source_revision_id, target_revision_id):
        """Internal deep copy. Called inside one transaction/history group."""
        source_map = {}
        source_rows = self.lopa_sources(source_revision_id)
        for source in source_rows:
            payload = dict(source)
            old_id = payload.pop('id')
            payload.pop('equipment_tag', None)  # joined display-only value
            payload['revision_id'] = target_revision_id
            columns = list(payload)
            new_id = self.conn.execute(
                f"INSERT INTO lopa_source_scenarios({','.join(columns)}) "
                f"VALUES ({','.join('?' for _ in columns)})",
                [payload[column] for column in columns]).lastrowid
            source_map[old_id] = new_id
            for consequence in self.lopa_source_consequences(old_id):
                copy = dict(consequence); copy.pop('id'); copy['source_id'] = new_id
                cols = list(copy)
                self.conn.execute(
                    f"INSERT INTO lopa_source_consequences({','.join(cols)}) "
                    f"VALUES ({','.join('?' for _ in cols)})",
                    [copy[column] for column in cols])
            for escalation in self.lopa_escalation_rows(old_id):
                copy = dict(escalation); copy.pop('id'); copy['source_id'] = new_id
                cols = list(copy)
                self.conn.execute(
                    f"INSERT INTO lopa_escalation_rows({','.join(cols)}) "
                    f"VALUES ({','.join('?' for _ in cols)})",
                    [copy[column] for column in cols])

        group_map = {}
        for group in self.lopa_sensor_groups(source_revision_id):
            copy = dict(group); old_id = copy.pop('id'); copy['revision_id'] = target_revision_id
            cols = list(copy)
            new_id = self.conn.execute(
                f"INSERT INTO lopa_sensor_groups({','.join(cols)}) "
                f"VALUES ({','.join('?' for _ in cols)})",
                [copy[column] for column in cols]).lastrowid
            group_map[old_id] = new_id
            for member in self.lopa_sensor_members(old_id):
                copied_member = dict(member); copied_member.pop('id'); copied_member.pop('tag', None)
                copied_member['group_id'] = new_id
                member_cols = list(copied_member)
                self.conn.execute(
                    f"INSERT INTO lopa_sensor_members({','.join(member_cols)}) "
                    f"VALUES ({','.join('?' for _ in member_cols)})",
                    [copied_member[column] for column in member_cols])

        # The final-element side is independent from the sensor hierarchy,
        # but follows the identical revision-copy rule.  Keep its object
        # identity, action and voting review flag in the new draft.
        for group in self.lopa_final_groups(source_revision_id):
            copied_group = dict(group)
            old_id = copied_group.pop('id')
            copied_group['revision_id'] = target_revision_id
            group_cols = list(copied_group)
            new_id = self.conn.execute(
                f"INSERT INTO lopa_final_groups({','.join(group_cols)}) "
                f"VALUES ({','.join('?' for _ in group_cols)})",
                [copied_group[column] for column in group_cols]).lastrowid
            for member in self.lopa_final_members(old_id):
                copied_member = dict(member)
                copied_member.pop('id')
                copied_member.pop('tag', None)
                copied_member['group_id'] = new_id
                member_cols = list(copied_member)
                self.conn.execute(
                    f"INSERT INTO lopa_final_members({','.join(member_cols)}) "
                    f"VALUES ({','.join('?' for _ in member_cols)})",
                    [copied_member[column] for column in member_cols])

        for barrier in self.lopa_barriers(source_revision_id):
            copy = dict(barrier); old_id = copy.pop('id')
            copy['revision_id'] = target_revision_id
            copy['source_id'] = source_map.get(copy.get('source_id'))
            cols = list(copy)
            new_id = self.conn.execute(
                f"INSERT INTO lopa_barriers({','.join(cols)}) "
                f"VALUES ({','.join('?' for _ in cols)})",
                [copy[column] for column in cols]).lastrowid
            for category in self.lopa_barrier_categories(old_id):
                self.conn.execute(
                    "INSERT INTO lopa_barrier_categories(barrier_id,category_key,active) VALUES (?,?,?)",
                    (new_id, category['category_key'], category['active']))

    def lock_lopa_revision(self, revision_id, actor=''):
        revision = self._assert_lopa_revision_editable(revision_id)
        self.conn.execute(
            "UPDATE lopa_revisions SET status='Låst',locked_at=CURRENT_TIMESTAMP,locked_by=? WHERE id=?",
            (str(actor or ''), revision_id))
        self._log_lopa(revision['lopa_id'], revision_id, 'locked', '', actor)
        self.commit()

    def unlock_lopa_revision(self, revision_id, reason, actor=''):
        revision = self._lopa_revision_row(revision_id)
        if not revision:
            raise ValueError('LOPA-revisionen finns inte längre.')
        reason = str(reason or '').strip()
        if not reason:
            raise ValueError('Ange varför LOPA-revisionen låses upp.')
        self.conn.execute(
            "UPDATE lopa_revisions SET status='Utkast',unlock_reason=?,locked_at='',locked_by='' WHERE id=?",
            (reason, revision_id))
        self._log_lopa(revision['lopa_id'], revision_id, 'unlocked', reason, actor)
        self.commit()

    def safeguard_equipment_links(self, safeguard_id):
        return [dict(row) for row in self.conn.execute(
            "SELECT sel.*,COALESCE(ec.tag,sel.tag_snapshot,'') AS tag "
            "FROM safeguard_equipment_links sel "
            "LEFT JOIN equipment_catalog ec ON ec.id=sel.equipment_id "
            "WHERE sel.safeguard_id=? ORDER BY sel.sort_order,sel.id", (safeguard_id,)).fetchall()]

    def add_safeguard_equipment_link(self, safeguard_id, equipment_id, trigger_code='',
                                     trigger_custom=''):
        safeguard = self.get_safeguard(safeguard_id)
        equipment = self.get_equipment_by_id(equipment_id)
        if not safeguard:
            raise ValueError('Barriären finns inte längre.')
        if not equipment:
            raise ValueError('Objektet finns inte längre i objektdatabasen.')
        trigger_code = str(trigger_code or '').strip().upper()
        trigger_custom = str(trigger_custom or '').strip()
        existing = self.conn.execute(
            "SELECT id FROM safeguard_equipment_links WHERE safeguard_id=? AND equipment_id=? "
            "AND trigger_code=? AND trigger_custom=?",
            (safeguard_id, equipment_id, trigger_code, trigger_custom)).fetchone()
        if existing:
            return existing['id']
        sort_order = self.conn.execute(
            "SELECT COALESCE(MAX(sort_order),-1)+1 FROM safeguard_equipment_links WHERE safeguard_id=?",
            (safeguard_id,)).fetchone()[0]
        cur = self.conn.execute(
            "INSERT INTO safeguard_equipment_links "
            "(safeguard_id,equipment_id,trigger_code,trigger_custom,tag_snapshot,sort_order) "
            "VALUES (?,?,?,?,?,?)",
            (safeguard_id, equipment_id, trigger_code, trigger_custom,
             equipment.get('tag') or '', sort_order))
        self.commit()
        return cur.lastrowid

    def lopa_sources(self, revision_id):
        return [dict(row) for row in self.conn.execute(
            "SELECT ls.*,COALESCE(ec.tag,'') AS equipment_tag FROM lopa_source_scenarios ls "
            "LEFT JOIN equipment_catalog ec ON ec.id=ls.equipment_id "
            "WHERE ls.revision_id=? ORDER BY ls.id", (revision_id,)).fetchall()]

    def lopa_source_sync_state(self, source_id):
        """Compare one active LOPA snapshot with its originating HAZOP row.

        This method is intentionally read-only.  A LOPA analyst must see
        that HAZOP changed before deciding whether to keep a historic/local
        snapshot; a background refresh must never silently overwrite it.
        """
        source = self.conn.execute(
            "SELECT * FROM lopa_source_scenarios WHERE id=?", (source_id,)).fetchone()
        if not source:
            return {'state': 'missing', 'messages': ['LOPA-källscenariot saknas.']}
        source = dict(source)
        if not source.get('follows_hazop'):
            return {'state': 'detached', 'messages': ['Frikopplad från HAZOP lokalt.']}
        if not source.get('origin_safeguard_id'):
            return {'state': 'missing', 'messages': ['HAZOP-barriären finns inte längre.']}
        live = self.conn.execute(
            "SELECT s.id,c.description AS cause_text,c.base_frequency,co.description AS scenario_text "
            "FROM safeguards s JOIN consequences co ON co.id=s.consequence_id "
            "JOIN causes c ON c.id=co.cause_id WHERE s.id=?",
            (source['origin_safeguard_id'],)).fetchone()
        if not live:
            return {'state': 'missing', 'messages': ['HAZOP-barriären finns inte längre.']}
        live = dict(live)
        changes = []
        if (live.get('cause_text') or '') != (source.get('cause_text') or ''):
            changes.append('orsakstext')
        if (live.get('scenario_text') or '') != (source.get('scenario_text') or ''):
            changes.append('scenariotext')
        live_frequency = self.cause_base_frequency_per_year(
            self.get_cause(source.get('hazop_cause_id')))
        stored_frequency = source.get('base_frequency')
        if ((live_frequency is None) != (stored_frequency is None) or
                (live_frequency is not None and stored_frequency is not None and
                 abs(float(live_frequency) - float(stored_frequency)) > 1e-12)):
            changes.append('grundfrekvens')
        for barrier in self.lopa_barriers(source['revision_id'], source_id):
            if barrier['manual'] or not barrier['follows_hazop']:
                continue
            current = self.get_safeguard(barrier.get('source_safeguard_id'))
            if not current:
                changes.append('oberoende barriär saknas')
                continue
            current = dict(current)
            if ((current.get('description') or '') != (barrier.get('description') or '') or
                    float(current.get('rrf') or 1) != float(barrier.get('rrf') or 1) or
                    (current.get('sg_type') or '') != (barrier.get('sg_type') or '')):
                changes.append('oberoende barriär')
        if changes:
            return {'state': 'changed', 'messages': list(dict.fromkeys(changes))}
        return {'state': 'current', 'messages': []}

    def lopa_source_consequences(self, source_id):
        return [dict(row) for row in self.conn.execute(
            "SELECT * FROM lopa_source_consequences WHERE source_id=? "
            "ORDER BY category_name,category_key,id", (source_id,)).fetchall()]

    def lopa_sensor_groups(self, revision_id):
        return [dict(row) for row in self.conn.execute(
            "SELECT * FROM lopa_sensor_groups WHERE revision_id=? ORDER BY sort_order,id",
            (revision_id,)).fetchall()]

    def lopa_sensor_members(self, group_id):
        return [dict(row) for row in self.conn.execute(
            "SELECT lm.*,COALESCE(ec.tag,lm.tag_snapshot,'') AS tag "
            "FROM lopa_sensor_members lm LEFT JOIN equipment_catalog ec ON ec.id=lm.equipment_id "
            "WHERE lm.group_id=? ORDER BY lm.sort_order,lm.id", (group_id,)).fetchall()]

    def lopa_barriers(self, revision_id, source_id=None):
        values = [revision_id]
        where = 'lb.revision_id=?'
        if source_id is not None:
            where += ' AND lb.source_id=?'; values.append(source_id)
        return [dict(row) for row in self.conn.execute(
            "SELECT lb.* FROM lopa_barriers lb WHERE " + where + " ORDER BY lb.sort_order,lb.id",
            values).fetchall()]

    def lopa_barrier_categories(self, barrier_id):
        return [dict(row) for row in self.conn.execute(
            "SELECT * FROM lopa_barrier_categories WHERE barrier_id=? ORDER BY category_key",
            (barrier_id,)).fetchall()]

    def lopa_escalation_rows(self, source_id):
        return [dict(row) for row in self.conn.execute(
            "SELECT * FROM lopa_escalation_rows WHERE source_id=? ORDER BY category_key,id",
            (source_id,)).fetchall()]

    def update_lopa_revision_details(self, revision_id, performed_by_text=None,
                                     approved_by_text=None, notes=None, document_date=None,
                                     dimensioning_category_key=None, additional_actions=None,
                                     additional_requirements=None, process_safety_time=None):
        """Update the editable document fields stored on one LOPA revision."""
        revision = self._assert_lopa_revision_editable(revision_id)
        assignments, values = [], []
        if performed_by_text is not None:
            assignments.append('performed_by_text=?')
            values.append(str(performed_by_text or ''))
        if approved_by_text is not None:
            assignments.append('approved_by_text=?')
            values.append(str(approved_by_text or ''))
        if notes is not None:
            assignments.append('notes=?')
            values.append(str(notes or ''))
        if document_date is not None:
            assignments.append('document_date=?')
            values.append(str(document_date or ''))
        if dimensioning_category_key is not None:
            assignments.append('dimensioning_category_key=?')
            values.append(str(dimensioning_category_key or ''))
        if additional_actions is not None:
            assignments.append('additional_actions=?')
            values.append(str(additional_actions or ''))
        if additional_requirements is not None:
            assignments.append('additional_requirements=?')
            values.append(str(additional_requirements or ''))
        if process_safety_time is not None:
            raw_value = str(process_safety_time).strip()
            if not raw_value:
                value = None
            else:
                try:
                    value = float(raw_value.replace(',', '.'))
                except (TypeError, ValueError):
                    raise ValueError('Processäkerhetstid måste vara ett tal.')
                if value < 0:
                    raise ValueError('Processäkerhetstid får inte vara negativ.')
            assignments.append('process_safety_time=?')
            values.append(value)
        if not assignments:
            return
        values.append(revision_id)
        self.conn.execute(
            f"UPDATE lopa_revisions SET {','.join(assignments)} WHERE id=?", values)
        self._log_lopa(revision['lopa_id'], revision_id, 'revision-details-updated')
        self.commit()

    def _lopa_source_revision(self, source_id):
        row = self.conn.execute(
            "SELECT ls.*,lr.lopa_id FROM lopa_source_scenarios ls "
            "JOIN lopa_revisions lr ON lr.id=ls.revision_id WHERE ls.id=?", (source_id,)).fetchone()
        if not row:
            raise ValueError('LOPA-källscenariot finns inte längre.')
        row = dict(row)
        self._assert_lopa_revision_editable(row['revision_id'])
        return row

    def set_lopa_source_active(self, source_id, active):
        """Include or exclude one imported HAZOP source locally in the LOPA."""
        source = self._lopa_source_revision(source_id)
        self.conn.execute("UPDATE lopa_source_scenarios SET active=? WHERE id=?",
                          (int(bool(active)), source_id))
        self._log_lopa(source['lopa_id'], source['revision_id'], 'source-inclusion-changed',
                       f"Källscenario {source_id}: {'aktiv' if active else 'exkluderad'}")
        self.commit()

    def set_lopa_source_scenario_text(self, source_id, scenario_text, detached_reason=''):
        """Set a local LOPA scenario description and mark it detached from HAZOP.

        The caller is responsible for asking the user before this deliberate
        local override.  The original HAZOP ids remain stored for traceability.
        """
        source = self._lopa_source_revision(source_id)
        self.conn.execute(
            "UPDATE lopa_source_scenarios SET scenario_text=?,follows_hazop=0,detached_reason=? "
            "WHERE id=?",
            (str(scenario_text or ''), str(detached_reason or 'Redigerad lokalt i LOPA.'), source_id))
        self._log_lopa(source['lopa_id'], source['revision_id'], 'source-detached',
                       f'Källscenario {source_id}: scenariotext ändrad lokalt')
        self.commit()

    def update_lopa_source_analysis_details(self, source_id, control_frequency=None,
                                            assumption_percent=None, assumption_reason=None):
        """Store LOPA-only analysis assumptions for one source scenario.

        These fields complement, rather than overwrite, the imported HAZOP
        frequency.  ``assumption_percent`` is deliberately retained as a
        percentage: 10 means 1/10 in the calculator.
        """
        source = self._lopa_source_revision(source_id)
        assignments, values = [], []
        if control_frequency is not None:
            assignments.append('control_frequency=?')
            values.append(str(control_frequency or ''))
        if assumption_percent is not None:
            try:
                percent = float(str(assumption_percent).replace(',', '.'))
            except (TypeError, ValueError):
                raise ValueError('Förutsättning måste vara ett procenttal.')
            if percent < 0:
                raise ValueError('Förutsättning får inte vara negativ.')
            assignments.append('assumption_percent=?')
            values.append(percent)
        if assumption_reason is not None:
            assignments.append('assumption_reason=?')
            values.append(str(assumption_reason or ''))
        if not assignments:
            return
        values.append(source_id)
        self.conn.execute(
            f"UPDATE lopa_source_scenarios SET {','.join(assignments)} WHERE id=?", values)
        self._log_lopa(source['lopa_id'], source['revision_id'], 'source-analysis-updated',
                       f'Källscenario {source_id}: LOPA-underlag')
        self.commit()

    def _lopa_consequence_source(self, consequence_id):
        row = self.conn.execute(
            "SELECT lc.*,ls.revision_id,lr.lopa_id FROM lopa_source_consequences lc "
            "JOIN lopa_source_scenarios ls ON ls.id=lc.source_id "
            "JOIN lopa_revisions lr ON lr.id=ls.revision_id WHERE lc.id=?", (consequence_id,)).fetchone()
        if not row:
            raise ValueError('LOPA-konsekvensen finns inte längre.')
        row = dict(row)
        self._assert_lopa_revision_editable(row['revision_id'])
        return row

    def set_lopa_consequence_active(self, consequence_id, active):
        """Include or exclude an imported consequence only in this LOPA."""
        consequence = self._lopa_consequence_source(consequence_id)
        self.conn.execute("UPDATE lopa_source_consequences SET active=? WHERE id=?",
                          (int(bool(active)), consequence_id))
        self._log_lopa(consequence['lopa_id'], consequence['revision_id'],
                       'consequence-inclusion-changed',
                       f"Konsekvens {consequence_id}: {'aktiv' if active else 'exkluderad'}")
        self.commit()

    def update_lopa_consequence(self, consequence_id, description=None, severity=None,
                                detached_reason=''):
        """Make an explicitly local consequence override in an editable revision."""
        consequence = self._lopa_consequence_source(consequence_id)
        assignments, values = [], []
        if description is not None:
            assignments.append('description=?')
            values.append(str(description or ''))
        if severity is not None:
            try:
                severity = int(severity)
            except (TypeError, ValueError):
                raise ValueError('Konsekvensnivå måste vara ett heltal.')
            if severity < 0:
                raise ValueError('Konsekvensnivå får inte vara negativ.')
            assignments.append('severity=?')
            values.append(severity)
        if not assignments:
            return
        assignments.extend(['follows_hazop=0', 'detached_reason=?'])
        values.append(str(detached_reason or 'Redigerad lokalt i LOPA.'))
        values.append(consequence_id)
        self.conn.execute(
            f"UPDATE lopa_source_consequences SET {','.join(assignments)} WHERE id=?", values)
        self._log_lopa(consequence['lopa_id'], consequence['revision_id'], 'consequence-detached',
                       f'Konsekvens {consequence_id}: lokalt ändrad')
        self.commit()

    def add_lopa_custom_consequence(self, source_id, category_key, category_name,
                                    severity=0, description=''):
        """Add a clearly local consequence without creating anything in HAZOP."""
        source = self._lopa_source_revision(source_id)
        try:
            numeric_severity = int(severity)
        except (TypeError, ValueError):
            raise ValueError('Konsekvensnivå måste vara ett heltal.')
        if numeric_severity < 0:
            raise ValueError('Konsekvensnivå får inte vara negativ.')
        consequence_id = self.conn.execute(
            "INSERT INTO lopa_source_consequences(source_id,category_key,category_name,severity,description,"
            "follows_hazop,detached_reason) VALUES (?,?,?,?,?,0,?)",
            (source_id, str(category_key or ''), str(category_name or category_key or ''),
             numeric_severity, str(description or ''), 'Skapad lokalt i LOPA.')).lastrowid
        self._log_lopa(source['lopa_id'], source['revision_id'], 'custom-consequence-added',
                       f'Källscenario {source_id}')
        self.commit()
        return consequence_id

    def _lopa_barrier_revision(self, barrier_id):
        row = self.conn.execute(
            "SELECT lb.*,lr.lopa_id FROM lopa_barriers lb "
            "JOIN lopa_revisions lr ON lr.id=lb.revision_id WHERE lb.id=?", (barrier_id,)).fetchone()
        if not row:
            raise ValueError('LOPA-barriären finns inte längre.')
        row = dict(row)
        self._assert_lopa_revision_editable(row['revision_id'])
        return row

    def set_lopa_barrier_active(self, barrier_id, active):
        """Include or exclude a barrier only for this LOPA calculation."""
        barrier = self._lopa_barrier_revision(barrier_id)
        self.conn.execute("UPDATE lopa_barriers SET active=? WHERE id=?",
                          (int(bool(active)), barrier_id))
        self._log_lopa(barrier['lopa_id'], barrier['revision_id'], 'barrier-inclusion-changed',
                       f"Barriär {barrier_id}: {'aktiv' if active else 'exkluderad'}")
        self.commit()

    def update_lopa_barrier(self, barrier_id, description=None, rrf=None, sg_type=None,
                            detached_reason=''):
        """Update a barrier locally, retaining its HAZOP source id as evidence."""
        barrier = self._lopa_barrier_revision(barrier_id)
        assignments, values = [], []
        if description is not None:
            assignments.append('description=?')
            values.append(str(description or ''))
        if rrf is not None:
            try:
                numeric_rrf = float(rrf)
            except (TypeError, ValueError):
                raise ValueError('RRF måste vara ett tal.')
            if numeric_rrf < 1:
                raise ValueError('RRF måste vara minst 1.')
            assignments.append('rrf=?')
            values.append(numeric_rrf)
        if sg_type is not None:
            assignments.append('sg_type=?')
            values.append(str(sg_type or 'Övrigt'))
        if not assignments:
            return
        assignments.extend(['follows_hazop=0', 'detached_reason=?'])
        values.append(str(detached_reason or 'Redigerad lokalt i LOPA.'))
        values.append(barrier_id)
        self.conn.execute(f"UPDATE lopa_barriers SET {','.join(assignments)} WHERE id=?", values)
        self._log_lopa(barrier['lopa_id'], barrier['revision_id'], 'barrier-detached',
                       f'Barriär {barrier_id}: lokalt ändrad')
        self.commit()

    def set_lopa_barrier_category_keys(self, barrier_id, category_keys=None):
        """Set category applicability; ``None`` means all active categories."""
        barrier = self._lopa_barrier_revision(barrier_id)
        category_keys = None if category_keys is None else [str(key) for key in category_keys]
        with self.history_group():
            try:
                self.conn.execute('BEGIN')
                self.conn.execute("UPDATE lopa_barriers SET applies_all_categories=? WHERE id=?",
                                  (int(category_keys is None), barrier_id))
                self.conn.execute("DELETE FROM lopa_barrier_categories WHERE barrier_id=?", (barrier_id,))
                for key in category_keys or []:
                    self.conn.execute(
                        "INSERT INTO lopa_barrier_categories(barrier_id,category_key,active) VALUES (?,?,1)",
                        (barrier_id, key))
                self._log_lopa(barrier['lopa_id'], barrier['revision_id'],
                               'barrier-categories-changed', f'Barriär {barrier_id}')
                self.commit()
            except Exception:
                self.conn.rollback()
                raise

    def set_lopa_escalation_values(self, source_id, category_key, factor_values=None,
                                   reason=None, active=True):
        """Persist local escalation factors for one source/category pair."""
        source = self._lopa_source_revision(source_id)
        values = factor_values if isinstance(factor_values, dict) else {}
        self.conn.execute(
            "INSERT INTO lopa_escalation_rows(source_id,category_key,factor_values_json,reason,active) "
            "VALUES (?,?,?,?,?) ON CONFLICT(source_id,category_key) DO UPDATE SET "
            "factor_values_json=excluded.factor_values_json,reason=excluded.reason,active=excluded.active",
            (source_id, str(category_key or ''), json.dumps(values), str(reason or ''), int(bool(active))))
        self._log_lopa(source['lopa_id'], source['revision_id'], 'escalation-updated',
                       f'Källscenario {source_id}, {category_key}')
        self.commit()

    def add_lopa_sensor_group(self, revision_id, voting='1oo1'):
        """Add a separate sensor voting group to an editable LOPA revision."""
        revision = self._assert_lopa_revision_editable(revision_id)
        text = str(voting or '').strip().lower()
        if not re.fullmatch(r'\d+oo\d+', text, re.IGNORECASE):
            raise ValueError('Ange röstning som exempelvis 1oo1 eller 2oo3.')
        sort_order = self.conn.execute(
            "SELECT COALESCE(MAX(sort_order),-1)+1 FROM lopa_sensor_groups WHERE revision_id=?",
            (revision_id,)).fetchone()[0]
        group_id = self.conn.execute(
            "INSERT INTO lopa_sensor_groups(revision_id,voting,sort_order) VALUES (?,?,?)",
            (revision_id, text, sort_order)).lastrowid
        self._log_lopa(revision['lopa_id'], revision_id, 'sensor-group-added', text)
        self.commit()
        return group_id

    def add_lopa_sensor_member(self, revision_id, equipment_id, trigger_code='', trigger_custom='',
                               group_id=None, origin_safeguard_id=None):
        """Add an equipment object to a sensor voting group.

        A second active member deliberately leaves the voting at its existing
        value and raises ``needs_voting_review`` until the analyst confirms it.
        """
        revision = self._assert_lopa_revision_editable(revision_id)
        equipment = self.get_equipment_by_id(equipment_id)
        if not equipment:
            raise ValueError('Objektet finns inte längre i objektdatabasen.')
        if group_id is None:
            group_id = self._ensure_lopa_sensor_group(revision_id)
        group = self.conn.execute(
            "SELECT id FROM lopa_sensor_groups WHERE id=? AND revision_id=?", (group_id, revision_id)).fetchone()
        if not group:
            raise ValueError('Givargruppen hör inte till den här LOPA-revisionen.')
        with self.history_group():
            try:
                self.conn.execute('BEGIN')
                member_id = self._add_lopa_sensor_member(
                    revision_id, equipment_id, origin_safeguard_id,
                    trigger_code, trigger_custom, equipment.get('tag') or '', group_id=group_id)
                self._log_lopa(revision['lopa_id'], revision_id, 'sensor-added',
                               f"{equipment.get('tag') or equipment_id}")
                self.commit()
            except Exception:
                self.conn.rollback()
                raise
        return member_id

    def set_lopa_sensor_member_active(self, member_id, active):
        row = self.conn.execute(
            "SELECT lm.group_id,lg.revision_id,lr.lopa_id FROM lopa_sensor_members lm "
            "JOIN lopa_sensor_groups lg ON lg.id=lm.group_id "
            "JOIN lopa_revisions lr ON lr.id=lg.revision_id WHERE lm.id=?", (member_id,)).fetchone()
        if not row:
            raise ValueError('Givarobjektet finns inte längre.')
        row = dict(row)
        self._assert_lopa_revision_editable(row['revision_id'])
        self.conn.execute("UPDATE lopa_sensor_members SET active=? WHERE id=?",
                          (int(bool(active)), member_id))
        self._log_lopa(row['lopa_id'], row['revision_id'], 'sensor-inclusion-changed',
                       f"Givare {member_id}: {'aktiv' if active else 'exkluderad'}")
        self.commit()

    def lopa_final_groups(self, revision_id):
        return [dict(row) for row in self.conn.execute(
            "SELECT * FROM lopa_final_groups WHERE revision_id=? ORDER BY sort_order,id",
            (revision_id,)).fetchall()]

    def lopa_final_members(self, group_id):
        return [dict(row) for row in self.conn.execute(
            "SELECT fm.*,COALESCE(ec.tag,fm.name_snapshot,'') AS tag "
            "FROM lopa_final_members fm LEFT JOIN equipment_catalog ec ON ec.id=fm.equipment_id "
            "WHERE fm.group_id=? ORDER BY fm.sort_order,fm.id", (group_id,)).fetchall()]

    def add_lopa_final_group(self, revision_id, voting='1oo1'):
        revision = self._assert_lopa_revision_editable(revision_id)
        text = str(voting or '').strip().lower()
        if not re.fullmatch(r'\d+oo\d+', text, re.IGNORECASE):
            raise ValueError('Ange röstning som exempelvis 1oo1 eller 2oo3.')
        sort_order = self.conn.execute(
            "SELECT COALESCE(MAX(sort_order),-1)+1 FROM lopa_final_groups WHERE revision_id=?",
            (revision_id,)).fetchone()[0]
        group_id = self.conn.execute(
            "INSERT INTO lopa_final_groups(revision_id,voting,sort_order) VALUES (?,?,?)",
            (revision_id, text, sort_order)).lastrowid
        self._log_lopa(revision['lopa_id'], revision_id, 'final-group-added', text)
        self.commit()
        return group_id

    def set_lopa_final_group_voting(self, group_id, voting):
        row = self.conn.execute(
            "SELECT revision_id FROM lopa_final_groups WHERE id=?", (group_id,)).fetchone()
        if not row:
            raise ValueError('Manövergruppen finns inte längre.')
        self._assert_lopa_revision_editable(row['revision_id'])
        text = str(voting or '').strip().lower()
        if not re.fullmatch(r'\d+oo\d+', text, re.IGNORECASE):
            raise ValueError('Ange röstning som exempelvis 1oo1 eller 2oo3.')
        self.conn.execute(
            "UPDATE lopa_final_groups SET voting=?,needs_voting_review=0 WHERE id=?",
            (text, group_id))
        self.commit()

    def add_lopa_final_member(self, revision_id, equipment_id=None, name='', action_text='', group_id=None):
        revision = self._assert_lopa_revision_editable(revision_id)
        equipment = self.get_equipment_by_id(equipment_id) if equipment_id is not None else None
        if equipment_id is not None and not equipment:
            raise ValueError('Objektet finns inte längre i objektdatabasen.')
        if group_id is None:
            groups = self.lopa_final_groups(revision_id)
            group_id = groups[0]['id'] if groups else self.add_lopa_final_group(revision_id)
        group = self.conn.execute(
            "SELECT id FROM lopa_final_groups WHERE id=? AND revision_id=?", (group_id, revision_id)).fetchone()
        if not group:
            raise ValueError('Manövergruppen hör inte till den här LOPA-revisionen.')
        label = (equipment.get('tag') if equipment else '') or str(name or '').strip()
        if not label:
            raise ValueError('Ange ett objekt eller ett namn för manöverdelen.')
        member_count = self.conn.execute(
            "SELECT COUNT(*) FROM lopa_final_members WHERE group_id=? AND active=1", (group_id,)).fetchone()[0]
        sort_order = self.conn.execute(
            "SELECT COALESCE(MAX(sort_order),-1)+1 FROM lopa_final_members WHERE group_id=?", (group_id,)).fetchone()[0]
        cur = self.conn.execute(
            "INSERT INTO lopa_final_members(group_id,equipment_id,name_snapshot,action_text,sort_order) "
            "VALUES (?,?,?,?,?)", (group_id, equipment_id, label, str(action_text or ''), sort_order))
        if member_count >= 1:
            self.conn.execute("UPDATE lopa_final_groups SET needs_voting_review=1 WHERE id=?", (group_id,))
        self._log_lopa(revision['lopa_id'], revision_id, 'final-member-added', label)
        self.commit()
        return cur.lastrowid

    def update_lopa_final_member(self, member_id, name=None, action_text=None, active=None):
        row = self.conn.execute(
            "SELECT fm.*,fg.revision_id,lr.lopa_id FROM lopa_final_members fm "
            "JOIN lopa_final_groups fg ON fg.id=fm.group_id "
            "JOIN lopa_revisions lr ON lr.id=fg.revision_id WHERE fm.id=?", (member_id,)).fetchone()
        if not row:
            raise ValueError('Manöverobjektet finns inte längre.')
        row = dict(row)
        self._assert_lopa_revision_editable(row['revision_id'])
        assignments, values = [], []
        if name is not None:
            assignments.append('name_snapshot=?'); values.append(str(name or ''))
        if action_text is not None:
            assignments.append('action_text=?'); values.append(str(action_text or ''))
        if active is not None:
            assignments.append('active=?'); values.append(int(bool(active)))
        if not assignments:
            return
        values.append(member_id)
        self.conn.execute(f"UPDATE lopa_final_members SET {','.join(assignments)} WHERE id=?", values)
        self._log_lopa(row['lopa_id'], row['revision_id'], 'final-member-updated', str(member_id))
        self.commit()

    def lopa_comments(self, revision_id):
        return [dict(row) for row in self.conn.execute(
            "SELECT * FROM lopa_comments WHERE revision_id=? ORDER BY id", (revision_id,)).fetchall()]

    def add_lopa_comment(self, revision_id, body, author=''):
        revision = self._assert_lopa_revision_editable(revision_id)
        text = str(body or '').strip()
        if not text:
            raise ValueError('Kommentaren är tom.')
        comment_id = self.conn.execute(
            "INSERT INTO lopa_comments(revision_id,author,body) VALUES (?,?,?)",
            (revision_id, str(author or ''), text)).lastrowid
        self._log_lopa(revision['lopa_id'], revision_id, 'comment-added', text[:80])
        self.commit()
        return comment_id

    def _ensure_lopa_sensor_group(self, revision_id):
        groups = self.lopa_sensor_groups(revision_id)
        if groups:
            return groups[0]['id']
        return self.conn.execute(
            "INSERT INTO lopa_sensor_groups(revision_id,voting,sort_order) VALUES (?,'1oo1',0)",
            (revision_id,)).lastrowid

    def set_lopa_sensor_group_voting(self, group_id, voting):
        row = self.conn.execute(
            "SELECT revision_id FROM lopa_sensor_groups WHERE id=?", (group_id,)).fetchone()
        if not row:
            raise ValueError('Givargruppen finns inte längre.')
        self._assert_lopa_revision_editable(row['revision_id'])
        text = str(voting or '').strip()
        if not re.fullmatch(r'\d+oo\d+', text, re.IGNORECASE):
            raise ValueError('Ange röstning som exempelvis 1oo1 eller 2oo3.')
        self.conn.execute(
            "UPDATE lopa_sensor_groups SET voting=?,needs_voting_review=0 WHERE id=?",
            (text.lower(), group_id))
        self.commit()

    def _add_lopa_sensor_member(self, revision_id, equipment_id, origin_safeguard_id,
                                trigger_code='', trigger_custom='', tag_snapshot='', group_id=None):
        if equipment_id is None:
            return None
        group_id = group_id or self._ensure_lopa_sensor_group(revision_id)
        existing = self.conn.execute(
            "SELECT id FROM lopa_sensor_members WHERE group_id=? AND equipment_id=? "
            "AND trigger_code=? AND trigger_custom=?",
            (group_id, equipment_id, trigger_code, trigger_custom)).fetchone()
        if existing:
            return existing['id']
        member_count = self.conn.execute(
            "SELECT COUNT(*) FROM lopa_sensor_members WHERE group_id=? AND active=1",
            (group_id,)).fetchone()[0]
        sort_order = self.conn.execute(
            "SELECT COALESCE(MAX(sort_order),-1)+1 FROM lopa_sensor_members WHERE group_id=?",
            (group_id,)).fetchone()[0]
        cur = self.conn.execute(
            "INSERT INTO lopa_sensor_members(group_id,equipment_id,origin_safeguard_id,"
            "trigger_code,trigger_custom,tag_snapshot,sort_order) VALUES (?,?,?,?,?,?,?)",
            (group_id, equipment_id, origin_safeguard_id, trigger_code,
             trigger_custom, tag_snapshot, sort_order))
        # The second sensor is a deliberate review point.  The database
        # keeps 1oo1 until the user explicitly confirms/provides voting.
        if member_count >= 1:
            self.conn.execute(
                "UPDATE lopa_sensor_groups SET needs_voting_review=1 WHERE id=?",
                (group_id,))
        return cur.lastrowid

    def add_lopa_barrier(self, revision_id, source_id=None, description='', rrf=1,
                         sg_type='Övrigt', category_keys=None, manual=True):
        self._assert_lopa_revision_editable(revision_id)
        numeric_rrf = float(rrf)
        if numeric_rrf < 1:
            raise ValueError('RRF måste vara minst 1.')
        category_keys = None if category_keys is None else list(category_keys)
        sort_order = self.conn.execute(
            "SELECT COALESCE(MAX(sort_order),-1)+1 FROM lopa_barriers WHERE revision_id=?",
            (revision_id,)).fetchone()[0]
        cur = self.conn.execute(
            "INSERT INTO lopa_barriers(revision_id,source_id,sg_type,description,rrf,manual,"
            "applies_all_categories,sort_order) VALUES (?,?,?,?,?,?,?,?)",
            (revision_id, source_id, str(sg_type or 'Övrigt'), str(description or ''),
             numeric_rrf, int(bool(manual)), int(category_keys is None), sort_order))
        barrier_id = cur.lastrowid
        for key in category_keys or []:
            self.conn.execute(
                "INSERT OR REPLACE INTO lopa_barrier_categories(barrier_id,category_key,active) "
                "VALUES (?,?,1)", (barrier_id, str(key)))
        self.commit()
        return barrier_id

    def _lopa_category_key_map(self, matrix):
        result = {}
        for index, category in enumerate(matrix.get('consequence_categories') or []):
            category = category if isinstance(category, dict) else {}
            key = self._template_category_key(
                category.get('key') or category.get('name'), f'category-{index + 1}')
            result[category.get('name') or key] = key
        return result

    def add_lopa_source_from_safeguard(self, lopa_id, safeguard_id, equipment_id=None,
                                       trigger_code='', trigger_custom=''):
        """Import one HAZOP safeguard as a LOPA sensor path and source scenario.

        The selected safeguard is deliberately *not* copied as an independent
        barrier. Other safeguards under the same HAZOP cause are copied with
        their per-category applicability, while the selected safeguard's
        object+trigger becomes a sensor member.
        """
        revision = self.current_lopa_revision(lopa_id)
        if not revision:
            raise ValueError('LOPA:n saknar en aktiv revision.')
        self._assert_lopa_revision_editable(revision['id'])
        source_row = self.conn.execute(
            "SELECT s.id AS safeguard_id,s.consequence_id,c.id AS cause_id,c.node_id,"
            "c.deviation_id,c.description AS cause_text,c.likelihood,c.base_frequency,"
            "c.frequency_cleared,c.standard_cause_id,co.description AS scenario_text "
            "FROM safeguards s JOIN consequences co ON co.id=s.consequence_id "
            "JOIN causes c ON c.id=co.cause_id WHERE s.id=?", (safeguard_id,)).fetchone()
        if not source_row:
            raise ValueError('HAZOP-barriären finns inte längre.')
        source_row = dict(source_row)
        existing = self.conn.execute(
            "SELECT id FROM lopa_source_scenarios WHERE revision_id=? AND origin_safeguard_id=?",
            (revision['id'], safeguard_id)).fetchone()
        if existing:
            return {'source_id': existing['id'], 'created': False}

        links = self.safeguard_equipment_links(safeguard_id)
        if equipment_id is not None:
            links = [link for link in links if link.get('equipment_id') == equipment_id] or links
        if not links and equipment_id is not None:
            equipment = self.get_equipment_by_id(equipment_id)
            if equipment:
                links = [{
                    'equipment_id': equipment_id,
                    'trigger_code': trigger_code,
                    'trigger_custom': trigger_custom,
                    'tag': equipment.get('tag') or '',
                }]
        primary_link = links[0] if links else {}
        chosen_equipment_id = primary_link.get('equipment_id') or equipment_id
        numeric_frequency = self.cause_base_frequency_per_year(self.get_cause(source_row['cause_id']))
        origin = 'numerisk HAZOP-frekvens' if numeric_frequency is not None else 'HAZOP-frekvensnivå'
        matrix = self.lopa_revision_matrix(revision['id'])
        category_keys = self._lopa_category_key_map(matrix)

        with self.history_group():
            try:
                self.conn.execute('BEGIN')
                source_id = self.conn.execute(
                    "INSERT INTO lopa_source_scenarios(revision_id,hazop_node_id,hazop_deviation_id,"
                    "hazop_cause_id,origin_safeguard_id,equipment_id,trigger_code,trigger_custom,"
                    "cause_text,scenario_text,base_frequency,frequency_origin,assumption_percent) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    (revision['id'], source_row['node_id'], source_row['deviation_id'],
                     source_row['cause_id'], safeguard_id, chosen_equipment_id,
                     str(primary_link.get('trigger_code') or trigger_code or '').upper(),
                     str(primary_link.get('trigger_custom') or trigger_custom or ''),
                     source_row.get('cause_text') or '', source_row.get('scenario_text') or '',
                     numeric_frequency, origin,
                     (matrix.get('lopa') or {}).get('default_assumption_percent', 100.0))).lastrowid

                # A cause can produce several HAZOP consequences. Import each
                # assessed category, not just the consequence where the user
                # happened to click the safeguard.
                cause_consequences = self.consequences(source_row['cause_id'])
                category_applicability = {}
                for consequence in cause_consequences:
                    consequence = dict(consequence)
                    for severity in self.get_consequence_severities(consequence['id']):
                        severity = dict(severity)
                        category_key = category_keys.get(severity['name']) or self._template_category_key(
                            severity['name'], f"category-{severity['category_id']}")
                        self.conn.execute(
                            "INSERT INTO lopa_source_consequences(source_id,hazop_consequence_id,"
                            "hazop_category_id,category_key,category_name,severity,description) "
                            "VALUES (?,?,?,?,?,?,?)",
                            (source_id, consequence['id'], severity['category_id'], category_key,
                             severity['name'], severity['severity'], consequence.get('description') or ''))
                        # One local escalation row per active consequence
                        # category.  The factor payload starts empty so the
                        # revision snapshot's configured defaults apply until
                        # the analyst explicitly changes a percentage.
                        self.conn.execute(
                            "INSERT OR IGNORE INTO lopa_escalation_rows("
                            "source_id,category_key,factor_values_json,reason,active) VALUES (?,?, '{}','',1)",
                            (source_id, category_key))
                        category_applicability[(consequence['id'], category_key)] = severity

                # The barrier used to create the sensor remains excluded here.
                # All other barriers mirror HAZOP and keep their category-level
                # exclusions in the LOPA copy.
                for consequence in cause_consequences:
                    consequence = dict(consequence)
                    severities = self.get_consequence_severities(consequence['id'])
                    for safeguard in self.safeguards(consequence['id']):
                        safeguard = dict(safeguard)
                        if safeguard['id'] == safeguard_id:
                            continue
                        existing_barrier = self.conn.execute(
                            "SELECT id FROM lopa_barriers WHERE source_id=? AND source_safeguard_id=?",
                            (source_id, safeguard['id'])).fetchone()
                        if existing_barrier:
                            barrier_id = existing_barrier['id']
                        else:
                            barrier_id = self.conn.execute(
                                "INSERT INTO lopa_barriers(revision_id,source_id,source_safeguard_id,"
                                "sg_type,description,rrf,manual,applies_all_categories,sort_order) "
                                "VALUES (?,?,?,?,?,?,0,0,?)",
                                (revision['id'], source_id, safeguard['id'],
                                 safeguard.get('sg_type') or 'Övrigt',
                                 safeguard.get('description') or '',
                                 safeguard.get('rrf') or 1,
                                 self.conn.execute(
                                     "SELECT COALESCE(MAX(sort_order),-1)+1 FROM lopa_barriers WHERE revision_id=?",
                                     (revision['id'],)).fetchone()[0])).lastrowid
                        for severity in severities:
                            severity = dict(severity)
                            category_key = category_keys.get(severity['name']) or self._template_category_key(
                                severity['name'], f"category-{severity['category_id']}")
                            if safeguard['id'] in self.get_severity_excluded_sgs(severity['id']):
                                continue
                            self.conn.execute(
                                "INSERT OR REPLACE INTO lopa_barrier_categories(barrier_id,category_key,active) "
                                "VALUES (?,?,1)", (barrier_id, category_key))

                for link in links:
                    self._add_lopa_sensor_member(
                        revision['id'], link.get('equipment_id'), safeguard_id,
                        str(link.get('trigger_code') or trigger_code or '').upper(),
                        str(link.get('trigger_custom') or trigger_custom or ''),
                        str(link.get('tag') or ''))
                self._log_lopa(lopa_id, revision['id'], 'hazop-linked',
                               f"Barriär {safeguard_id} importerad som givardel")
                self.commit()
            except Exception:
                self.conn.rollback()
                raise
        return {'source_id': source_id, 'created': True}

    def lopa_source_calculation(self, source_id):
        source = self.conn.execute(
            "SELECT * FROM lopa_source_scenarios WHERE id=?", (source_id,)).fetchone()
        if not source:
            raise ValueError('LOPA-källscenariot finns inte längre.')
        source = dict(source)
        source['matrix'] = self.lopa_revision_matrix(source['revision_id'])
        factor_values = {}
        for row in self.lopa_escalation_rows(source_id):
            try:
                values = json.loads(row['factor_values_json'] or '{}')
            except (TypeError, ValueError, json.JSONDecodeError):
                values = {}
            factor_values[row['category_key']] = {
                'values': values if isinstance(values, dict) else {},
                'active': bool(row['active']),
            }
        source['categories'] = []
        for consequence in self.lopa_source_consequences(source_id):
            escalation = factor_values.get(consequence['category_key'], {})
            source['categories'].append({
                'category_key': consequence['category_key'],
                'category_name': consequence['category_name'],
                'severity': consequence['severity'],
                'active': bool(consequence['active']) and escalation.get('active', True),
                'factors': escalation.get('values', {}),
            })
        source['barriers'] = []
        for barrier in self.lopa_barriers(source['revision_id'], source_id):
            if not barrier['independent']:
                continue
            if barrier['applies_all_categories']:
                categories = None
            else:
                categories = {row['category_key']: bool(row['active'])
                              for row in self.lopa_barrier_categories(barrier['id'])}
            source['barriers'].append({
                'id': barrier['id'], 'description': barrier['description'],
                'rrf': barrier['rrf'], 'active': bool(barrier['active']),
                'categories': categories,
            })
        return calculate_lopa(source, source['matrix'].get('lopa') or {})

    def get_custom_risk_matrix_templates(self):
        """Return this project's named, user-created risk-matrix templates.

        Templates live alongside the project configuration rather than in the
        active matrix itself. A malformed or obsolete saved entry is ignored
        so it can never prevent the project from opening.
        """
        try:
            raw = json.loads(self.get_config('custom_risk_matrix_templates', '[]'))
        except (TypeError, ValueError, json.JSONDecodeError):
            raw = []
        templates = []
        for item in raw if isinstance(raw, list) else []:
            if not isinstance(item, dict):
                continue
            name = str(item.get('name') or '').strip()
            matrix = item.get('matrix')
            if not name or not isinstance(matrix, dict):
                continue
            templates.append({'name': name, 'matrix': self._risk_matrix_copy(matrix)})
        return templates

    def save_custom_risk_matrix_template(self, name, cfg):
        """Create or replace a named project-local risk-matrix template."""
        name = str(name or '').strip()
        if not name:
            raise ValueError("Mallnamn saknas.")
        template = {'name': name, 'matrix': self._risk_matrix_copy(cfg)}
        templates = self.get_custom_risk_matrix_templates()
        key = name.casefold()
        for index, existing in enumerate(templates):
            if existing['name'].casefold() == key:
                templates[index] = template
                break
        else:
            templates.append(template)
        self.set_config('custom_risk_matrix_templates', json.dumps(templates))
        return templates

    def delete_custom_risk_matrix_template(self, name):
        """Delete one named project-local matrix template, if it exists."""
        key = str(name or '').strip().casefold()
        templates = self.get_custom_risk_matrix_templates()
        remaining = [item for item in templates if item['name'].casefold() != key]
        if len(remaining) == len(templates):
            return False
        self.set_config('custom_risk_matrix_templates', json.dumps(remaining))
        return True

    @staticmethod
    def _template_category_key(value, fallback='category'):
        text = str(value or fallback).strip().casefold()
        text = re.sub(r'[^a-z0-9åäö]+', '-', text).strip('-')
        return text or fallback

    def _project_category_template(self, rows=5, existing_categories=None):
        """Snapshot active project categories as reusable template metadata."""
        definitions = self.get_severity_definitions()
        existing_by_key = {
            self._template_category_key(item.get('key') or item.get('name')): item
            for item in (existing_categories or []) if isinstance(item, dict)
        }
        result = []
        for index, category in enumerate(self.consequence_categories()):
            category = dict(category)
            key = self._template_category_key(category.get('name'), f'category-{index + 1}')
            existing = existing_by_key.get(key, {})
            result.append({
                'key': key,
                'name': category.get('name') or key,
                'color': existing.get('color') or '#64748b',
                'descriptions': [definitions.get(level, {}).get(category['id'], '')
                                 for level in range(1, 1 + int(rows))],
            })
        return result

    def apply_risk_matrix_template_without_assessments(self, cfg):
        """Install a complete template in an unassessed project.

        This is the non-destructive counterpart to the reviewed migration:
        it is valid only before any category assessment or description exists.
        """
        cfg = self._risk_matrix_copy(cfg)
        has_data = self.conn.execute(
            "SELECT EXISTS(SELECT 1 FROM consequence_severities) OR "
            "EXISTS(SELECT 1 FROM consequence_final_severities) OR "
            "EXISTS(SELECT 1 FROM severity_definitions)"
        ).fetchone()[0]
        if has_data:
            raise RuntimeError("Mallen innehåller bedömningar och måste migreras i granskningsdialogen.")
        try:
            self.conn.execute('BEGIN IMMEDIATE')
            self.conn.execute("DELETE FROM consequence_categories")
            for index, category in enumerate(cfg.get('consequence_categories', [])):
                cur = self.conn.execute(
                    "INSERT INTO consequence_categories(name,sort_order) VALUES (?,?)",
                    (category.get('name') or category.get('key') or 'Kategori', index))
                for level, text in enumerate(category.get('descriptions', []), start=1):
                    self.conn.execute(
                        "INSERT INTO severity_definitions(severity_level,category_id,description) VALUES (?,?,?)",
                        (level, cur.lastrowid, text or ''))
            self.conn.execute(
                "INSERT INTO app_config(key,value) VALUES('risk_matrix',?) "
                "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                (json.dumps(cfg),))
            self._commit_with_history()
        except Exception:
            self.conn.rollback()
            raise
        _risk_matrix_cache.load(self)

    # ── Risk-matrix template migration ──────────────────────────────────────
    @staticmethod
    def _risk_matrix_copy(cfg):
        """Return an independent, normalised matrix working copy."""
        return _normalise_matrix(json.loads(json.dumps(cfg or DEFAULT_MATRIX)))

    @staticmethod
    def _risk_level_label(cfg, kind, value):
        """Human-readable label for the application's stored ordinal value."""
        if kind == 'frequency':
            index = int(value) + 1       # stored F=-1 maps to matrix index 0
            labels = cfg.get('x_labels', [])
            codes = cfg.get('x_codes', [])
            prefix = f"F={value}"
        else:
            index = int(value) - 1       # stored C=1 maps to matrix index 0
            labels = cfg.get('y_labels', [])
            codes = cfg.get('y_codes', [])
            prefix = f"C={value}"
        text = labels[index] if 0 <= index < len(labels) else ''
        code = codes[index] if 0 <= index < len(codes) else prefix
        visible = f"{code} — {text}" if text else str(code)
        return f"{prefix} — {visible}"

    @staticmethod
    def _rank_level_map(source_count, target_count, source_offset, target_offset):
        """Give every source ordinal a visible, reviewable first suggestion.

        This is intentionally only a rank-based proposal.  The caller must
        let the user review/override it; labels are not trusted as evidence of
        equivalent process-safety meaning.
        """
        result = {}
        if source_count <= 0 or target_count <= 0:
            return result
        for source_index in range(source_count):
            if source_count == 1:
                target_index = 0
            else:
                target_index = round(source_index * (target_count - 1) /
                                     (source_count - 1))
            result[str(source_index + source_offset)] = target_index + target_offset
        return result

    def risk_matrix_migration_preview(self, source_cfg, target_cfg):
        """Build a pure-data, editable template-conversion plan.

        No database state is changed here.  The dialog can safely alter the
        mapping/record targets in the returned dict and later hand that exact
        reviewed plan to :meth:`apply_risk_matrix_migration`.
        """
        source = self._risk_matrix_copy(source_cfg)
        target = self._risk_matrix_copy(target_cfg)
        # The database owns the active category ids and is consequently the
        # source of truth for the source template.  Old saved matrices did
        # not include this snapshot, so build it here without changing data.
        source['consequence_categories'] = self._project_category_template(
            source['rows'], source.get('consequence_categories'))
        # Backwards compatible candidates (for example older user-created
        # files) retain the active categories.  Built-in templates always
        # provide their own category profile.
        if not target.get('consequence_categories'):
            target['consequence_categories'] = json.loads(json.dumps(
                source['consequence_categories']))

        source_categories = []
        for index, category in enumerate(self.consequence_categories()):
            category = dict(category)
            source_categories.append({
                'source_id': category['id'],
                'name': category['name'],
                'key': self._template_category_key(category['name'], f'category-{index + 1}'),
                'color': next((item.get('color', '#64748b')
                               for item in source['consequence_categories']
                               if item.get('key') == self._template_category_key(category['name'])),
                              '#64748b'),
            })
        target_categories = [dict(category) for category in target['consequence_categories']]
        category_map = {}
        used_targets = set()
        for index, category in enumerate(source_categories):
            matches = [target_category for target_category in target_categories
                       if target_category['key'] not in used_targets and
                       target_category['key'] == category['key']]
            if not matches and index < len(target_categories):
                candidate = target_categories[index]
                if candidate['key'] not in used_targets:
                    matches = [candidate]
            if matches:
                category_map[str(category['source_id'])] = matches[0]['key']
                used_targets.add(matches[0]['key'])
        source_freq_count, target_freq_count = source['cols'], target['cols']
        source_cons_count, target_cons_count = source['rows'], target['rows']
        frequency_map = self._rank_level_map(
            source_freq_count, target_freq_count, -1, -1)
        severity_map = self._rank_level_map(
            source_cons_count, target_cons_count, 1, 1)
        category_severity_maps = {
            str(category['source_id']): dict(severity_map)
            for category in source_categories
        }

        def _category_target(category_id, source_value):
            return category_severity_maps.get(str(category_id), severity_map).get(
                str(source_value), 1)

        frequency_records = []
        cause_rows = self.conn.execute("""
            SELECT c.id AS cause_id, c.description AS cause_description,
                   c.likelihood, c.base_frequency, c.frequency_cleared,
                   sc.frequency AS standard_frequency,
                   n.name AS node_name, d.description AS deviation_description
            FROM causes c
            LEFT JOIN standard_causes sc ON sc.id=c.standard_cause_id
            LEFT JOIN deviations d ON d.id=c.deviation_id
            LEFT JOIN nodes n ON n.id=c.node_id
            ORDER BY n.sort_order, n.name, d.sort_order, d.description, c.sort_order, c.id
        """).fetchall()
        for row in cause_rows:
            r = dict(row)
            if r.get('frequency_cleared'):
                continue
            numeric = (r.get('base_frequency') if r.get('base_frequency') is not None
                       else r.get('standard_frequency'))
            if numeric is not None and float(numeric) > 0:
                source_value = freq_to_f_level(float(numeric), source.get('freq_boundaries'))
                target_value = freq_to_f_level(float(numeric), target.get('freq_boundaries'))
                source_kind = 'numeric'
            else:
                source_value = int(r.get('likelihood') if r.get('likelihood') is not None else 0)
                target_value = frequency_map.get(str(source_value), -1)
                source_kind = 'manual'
            frequency_records.append({
                'key': f"cause:{r['cause_id']}", 'cause_id': r['cause_id'],
                'node': r.get('node_name') or '',
                'deviation': r.get('deviation_description') or '',
                'cause': r.get('cause_description') or '',
                'source': source_value, 'target': target_value,
                'source_kind': source_kind, 'numeric_frequency': numeric,
                'expected_likelihood': r.get('likelihood'),
            })

        severity_records = []
        base_rows = self.conn.execute("""
            SELECT co.id AS consequence_id, co.severity, co.description AS consequence_description,
                   ca.description AS cause_description, n.name AS node_name,
                   d.description AS deviation_description
            FROM consequences co
            JOIN causes ca ON ca.id=co.cause_id
            LEFT JOIN deviations d ON d.id=ca.deviation_id
            LEFT JOIN nodes n ON n.id=ca.node_id
            ORDER BY n.sort_order, n.name, d.sort_order, ca.sort_order, co.sort_order, co.id
        """).fetchall()
        for row in base_rows:
            r = dict(row); source_value = int(r['severity'])
            severity_records.append({
                'key': f"base:{r['consequence_id']}", 'kind': 'base',
                'consequence_id': r['consequence_id'], 'category_id': None,
                'node': r.get('node_name') or '', 'deviation': r.get('deviation_description') or '',
                'cause': r.get('cause_description') or '',
                'consequence': r.get('consequence_description') or '', 'category': '',
                'source': source_value, 'target': severity_map.get(str(source_value), 1),
                'expected_severity': source_value,
            })
        category_rows = self.conn.execute("""
            SELECT cs.id AS severity_id, cs.consequence_id, cs.category_id, cs.severity,
                   cc.name AS category_name, co.description AS consequence_description,
                   ca.description AS cause_description, n.name AS node_name,
                   d.description AS deviation_description
            FROM consequence_severities cs
            JOIN consequence_categories cc ON cc.id=cs.category_id
            JOIN consequences co ON co.id=cs.consequence_id
            JOIN causes ca ON ca.id=co.cause_id
            LEFT JOIN deviations d ON d.id=ca.deviation_id
            LEFT JOIN nodes n ON n.id=ca.node_id
            ORDER BY n.sort_order, n.name, d.sort_order, ca.sort_order, co.sort_order, cc.sort_order, cs.id
        """).fetchall()
        for row in category_rows:
            r = dict(row); source_value = int(r['severity'])
            severity_records.append({
                'key': f"category:{r['severity_id']}", 'kind': 'category',
                'severity_id': r['severity_id'], 'consequence_id': r['consequence_id'],
                'category_id': r['category_id'], 'node': r.get('node_name') or '',
                'deviation': r.get('deviation_description') or '',
                'cause': r.get('cause_description') or '',
                'consequence': r.get('consequence_description') or '',
                'category': r.get('category_name') or '',
                'source': source_value, 'target': _category_target(r['category_id'], source_value),
                'expected_severity': source_value,
            })
        final_rows = self.conn.execute("""
            SELECT cfs.consequence_id, cfs.category_id, cfs.severity,
                   cc.name AS category_name, co.description AS consequence_description,
                   ca.description AS cause_description, n.name AS node_name,
                   d.description AS deviation_description
            FROM consequence_final_severities cfs
            JOIN consequence_categories cc ON cc.id=cfs.category_id
            JOIN consequences co ON co.id=cfs.consequence_id
            JOIN causes ca ON ca.id=co.cause_id
            LEFT JOIN deviations d ON d.id=ca.deviation_id
            LEFT JOIN nodes n ON n.id=ca.node_id
            ORDER BY n.sort_order, n.name, d.sort_order, ca.sort_order, co.sort_order, cc.sort_order
        """).fetchall()
        for row in final_rows:
            r = dict(row); source_value = int(r['severity'])
            severity_records.append({
                'key': f"final:{r['consequence_id']}:{r['category_id']}", 'kind': 'final',
                'consequence_id': r['consequence_id'], 'category_id': r['category_id'],
                'node': r.get('node_name') or '', 'deviation': r.get('deviation_description') or '',
                'cause': r.get('cause_description') or '',
                'consequence': r.get('consequence_description') or '',
                'category': r.get('category_name') or '',
                'source': source_value, 'target': _category_target(r['category_id'], source_value),
                'expected_severity': source_value,
            })

        definition_records = []
        for row in self.conn.execute("""
            SELECT sd.id, sd.severity_level, sd.category_id, sd.description,
                   cc.name AS category_name
            FROM severity_definitions sd
            JOIN consequence_categories cc ON cc.id=sd.category_id
            ORDER BY cc.sort_order, sd.severity_level, sd.id
        """):
            r = dict(row); source_value = int(r['severity_level'])
            definition_records.append({
                'definition_id': r['id'], 'category_id': r['category_id'],
                'category': r.get('category_name') or '',
                'description': r.get('description') or '', 'source': source_value,
                'target': _category_target(r['category_id'], source_value),
                'target_category_key': category_map.get(str(r['category_id'])),
            })

        return {
            'version': 2, 'source_matrix': source, 'target_matrix': target,
            'frequency_map': frequency_map, 'severity_map': severity_map,
            'category_map': category_map,
            'category_severity_maps': category_severity_maps,
            'source_categories': source_categories,
            'target_categories': target_categories,
            'frequency_records': frequency_records,
            'severity_records': severity_records,
            'definition_records': definition_records,
        }

    def apply_risk_matrix_migration(self, plan):
        """Apply one reviewed matrix migration atomically, with hard backup.

        Raises ``ValueError`` for incomplete/conflicting plans and propagates
        database failures after rolling the complete transaction back.
        """
        if not isinstance(plan, dict) or plan.get('version') not in (1, 2):
            raise ValueError("Ogiltig migreringsplan för riskmatris.")
        source = self._risk_matrix_copy(plan.get('source_matrix'))
        target = self._risk_matrix_copy(plan.get('target_matrix'))
        current = self._risk_matrix_copy(self.get_risk_matrix())
        if plan.get('version') == 2:
            current['consequence_categories'] = self._project_category_template(
                current['rows'], current.get('consequence_categories'))
        if json.dumps(current, sort_keys=True) != json.dumps(source, sort_keys=True):
            raise RuntimeError(
                "Riskmatrisen ändrades medan migreringen granskades. Öppna förhandsgranskningen igen.")
        freq_min, freq_max = -1, target['cols'] - 2
        sev_min, sev_max = 1, target['rows']

        def _target(record, low, high, label):
            try:
                value = int(record['target'])
            except (KeyError, TypeError, ValueError):
                raise ValueError(f"{label} saknar en mål-nivå.")
            if not low <= value <= high:
                raise ValueError(f"{label} har mål-nivå utanför den nya matrisen.")
            return value

        category_map = plan.get('category_map', {})
        target_categories = {row.get('key'): row for row in plan.get(
            'target_categories', target.get('consequence_categories', [])) if row.get('key')}
        source_categories = list(plan.get('source_categories', []))
        if plan.get('version') == 2:
            if len(source_categories) != len(self.consequence_categories()):
                raise RuntimeError("Konsekvenskategorier ändrades medan migreringen granskades.")
            targets = []
            for source_category in source_categories:
                source_id = str(source_category.get('source_id'))
                target_key = category_map.get(source_id)
                if target_key not in target_categories:
                    raise ValueError("Varje befintlig konsekvenskategori måste kopplas till en mallkategori.")
                targets.append(target_key)
            if len(targets) != len(set(targets)):
                raise ValueError("Flera befintliga kategorier kan inte kopplas till samma mallkategori.")

        # The template descriptions are authoritative.  A mapped project
        # description is used only when the selected template deliberately
        # leaves that target level blank.
        fallback_definitions = {}
        for record in plan.get('definition_records', []):
            target_value = _target(record, sev_min, sev_max, "Konsekvensbeskrivning")
            target_key = record.get('target_category_key') or category_map.get(
                str(record.get('category_id')))
            if target_key not in target_categories:
                continue
            key = (target_key, target_value)
            text = (record.get('description') or '').strip()
            existing = fallback_definitions.get(key)
            if existing is not None and text and existing and text != existing:
                raise ValueError(
                    "Två olika konsekvensbeskrivningar skulle hamna på samma "
                    "nivå. Ändra kartläggningen innan du genomför bytet.")
            fallback_definitions[key] = text or existing or ''

        backup = self._write_backup(startup=True)
        if backup is None or not Path(backup).exists() or Path(backup).stat().st_size == 0:
            raise RuntimeError("Kunde inte skapa en verifierad backup före mallbytet.")

        try:
            self.conn.execute('BEGIN IMMEDIATE')
            # Existing projects normally receive this from SCHEMA at startup,
            # but keep the operation self-contained for a long-running app
            # that has just been upgraded without a restart.
            self.conn.execute("""
                CREATE TABLE IF NOT EXISTS risk_matrix_migrations (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    source_matrix TEXT NOT NULL, target_matrix TEXT NOT NULL,
                    mapping_json TEXT NOT NULL, backup_path TEXT NOT NULL DEFAULT ''
                )
            """)
            target_category_ids = {}
            if plan.get('version') == 2:
                # Retain source ids for mapped categories.  All assessments,
                # final levels and applicability exclusions therefore keep
                # their foreign keys while their category receives the
                # template's name and order.
                for source_category in source_categories:
                    source_id = int(source_category['source_id'])
                    target_key = category_map[str(source_id)]
                    target_category = target_categories[target_key]
                    self.conn.execute(
                        "UPDATE consequence_categories SET name=?, sort_order=? WHERE id=?",
                        (target_category.get('name') or target_key,
                         list(target_categories).index(target_key), source_id))
                    target_category_ids[target_key] = source_id
                for index, (target_key, target_category) in enumerate(target_categories.items()):
                    if target_key not in target_category_ids:
                        cur = self.conn.execute(
                            "INSERT INTO consequence_categories(name,sort_order) VALUES (?,?)",
                            (target_category.get('name') or target_key, index))
                        target_category_ids[target_key] = cur.lastrowid
            else:
                target_category_ids = {str(row['id']): row['id']
                                       for row in self.consequence_categories()}
            for record in plan.get('frequency_records', []):
                target_value = _target(record, freq_min, freq_max, "Frekvens")
                cur = self.conn.execute(
                    "UPDATE causes SET likelihood=? WHERE id=? AND likelihood=?",
                    (target_value, record['cause_id'], record.get('expected_likelihood')))
                if cur.rowcount != 1:
                    raise RuntimeError("En orsak ändrades medan migreringen granskades.")
            for record in plan.get('severity_records', []):
                target_value = _target(record, sev_min, sev_max, "Konsekvens")
                kind = record.get('kind')
                if kind == 'base':
                    cur = self.conn.execute(
                        "UPDATE consequences SET severity=? WHERE id=? AND severity=?",
                        (target_value, record['consequence_id'], record.get('expected_severity')))
                elif kind == 'category':
                    cur = self.conn.execute(
                        "UPDATE consequence_severities SET severity=? WHERE id=? AND severity=?",
                        (target_value, record['severity_id'], record.get('expected_severity')))
                elif kind == 'final':
                    cur = self.conn.execute(
                        "UPDATE consequence_final_severities SET severity=? "
                        "WHERE consequence_id=? AND category_id=? AND severity=?",
                        (target_value, record['consequence_id'], record['category_id'],
                         record.get('expected_severity')))
                else:
                    raise ValueError("Okänd konsekvenstyp i migreringsplanen.")
                if cur.rowcount != 1:
                    raise RuntimeError("En konsekvens ändrades medan migreringen granskades.")

            # Definition row ids are not referenced elsewhere. Rebuild from
            # the template, falling back to mapped project text only where
            # the template has intentionally left a target level blank.
            self.conn.execute("DELETE FROM severity_definitions")
            if plan.get('version') == 2:
                for target_key, target_category in target_categories.items():
                    category_id = target_category_ids[target_key]
                    descriptions = list(target_category.get('descriptions') or [])
                    for severity in range(1, sev_max + 1):
                        template_text = (descriptions[severity - 1]
                                         if severity <= len(descriptions) else '')
                        text = template_text or fallback_definitions.get(
                            (target_key, severity), '')
                        self.conn.execute(
                            "INSERT INTO severity_definitions(severity_level,category_id,description) "
                            "VALUES (?,?,?)", (severity, category_id, text))
            else:
                for record in plan.get('definition_records', []):
                    severity = _target(record, sev_min, sev_max, "Konsekvensbeskrivning")
                    self.conn.execute(
                        "INSERT INTO severity_definitions(severity_level,category_id,description) "
                        "VALUES (?,?,?)", (severity, record['category_id'],
                                             (record.get('description') or '').strip()))

            self.conn.execute(
                "INSERT INTO app_config(key,value) VALUES('risk_matrix',?) "
                "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                (json.dumps(target),))
            self.conn.execute("""
                INSERT INTO risk_matrix_migrations(source_matrix,target_matrix,mapping_json,backup_path)
                VALUES (?,?,?,?)
            """, (json.dumps(source), json.dumps(target), json.dumps(plan), str(backup)))
            self._commit_with_history()
        except Exception:
            self.conn.rollback()
            raise
        _risk_matrix_cache.load(self)
        return {'backup_path': str(backup),
                'frequency_count': len(plan.get('frequency_records', [])),
                'severity_count': len(plan.get('severity_records', []))}

    # ── Tag database ──────────────────────────────────────────────────────────
    def tag_database_entries(self, standard=None):
        if standard:
            return self.conn.execute(
                "SELECT * FROM tag_database WHERE standard=? AND active=1 ORDER BY tag_code",
                (standard,)).fetchall()
        return self.conn.execute(
            "SELECT * FROM tag_database WHERE active=1 ORDER BY tag_code").fetchall()

    def tag_database_standards(self):
        return [r[0] for r in self.conn.execute(
            "SELECT DISTINCT standard FROM tag_database WHERE standard!='' ORDER BY standard"
        ).fetchall()]

    def import_tag_database_excel(self, filepath: str):
        """Import tag codes from all relevant sheets in the Excel file."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            return 0, str(e)

        # Sheet name → standard name mapping
        SHEET_MAP = {
            'ISA-5.1':          'ISA-5.1',
            'SSG-5276':         'SSG-5276',
            'ISO-10628_14617':  'ISO-10628',
            'ISO-15519':        'ISO-15519',
            'IEC-DIN_EN_62424': 'IEC-62424',
            'DIN_19227_28000':  'DIN-19227',
            'PIP_PIC001':       'PIP-PIC001',
        }
        imported = 0
        for sheet_name, standard in SHEET_MAP.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            # Find header row (look for 'Taggkod' / 'Tag code')
            header_row = None
            for r in ws.iter_rows(max_row=10, values_only=True):
                for cell in r:
                    if cell and 'taggkod' in str(cell).lower():
                        header_row = r
                        break
                if header_row:
                    break
            if not header_row:
                continue
            # Map column indices
            cols = {str(v).strip().lower(): i
                    for i, v in enumerate(header_row) if v}
            c_code = next((i for k, i in cols.items() if 'taggkod' in k or 'tag code' in k), 0)
            c_sv   = next((i for k, i in cols.items() if 'svenska' in k or 'sv' in k or 'benom' in k), 3)
            c_en   = next((i for k, i in cols.items() if 'english' in k or 'en' in k), 4)
            c_cat  = next((i for k, i in cols.items() if 'kategori' in k or 'categ' in k), 5)

            start_row = ws.max_row  # will be overridden
            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                if row and row[c_code] and str(row[c_code]).strip().lower() == \
                        (header_row[c_code] or '').lower():
                    start_row = i + 1
                    break

            for row in ws.iter_rows(min_row=start_row, values_only=True):
                if not row or not row[c_code]:
                    continue
                code = str(row[c_code]).strip().upper()
                if not code or len(code) > 10:
                    continue
                sv  = str(row[c_sv]).strip()  if c_sv < len(row) and row[c_sv] else ''
                en  = str(row[c_en]).strip()  if c_en < len(row) and row[c_en] else ''
                cat = str(row[c_cat]).strip() if c_cat < len(row) and row[c_cat] else ''
                # Upsert
                self.conn.execute(
                    "INSERT OR REPLACE INTO tag_database "
                    "(tag_code,name_sv,name_en,category,standard,source,active) "
                    "VALUES (?,?,?,?,?,'excel',1)",
                    (code, sv, en, cat, standard))
                imported += 1

        self.commit()
        return imported, ''

    def tag_db_setting(self, key, default=None):
        r = self.conn.execute(
            "SELECT value FROM tag_database_settings WHERE key=?", (key,)).fetchone()
        return r['value'] if r else default

    def set_tag_db_setting(self, key, value):
        self.conn.execute(
            "INSERT OR REPLACE INTO tag_database_settings (key,value) VALUES (?,?)",
            (key, str(value)))
        self.commit()

    def tag_code_lookup(self, prefix: str) -> dict:
        """Look up a tag prefix in the active tag databases. Returns best match."""
        active_std = self.tag_db_setting('active_standard', '')
        if active_std:
            rows = self.conn.execute(
                "SELECT * FROM tag_database WHERE tag_code=? AND standard=? AND active=1",
                (prefix.upper(), active_std)).fetchall()
        else:
            rows = self.conn.execute(
                "SELECT * FROM tag_database WHERE tag_code=? AND active=1",
                (prefix.upper(),)).fetchall()
        return dict(rows[0]) if rows else {}

    # ── PID identified tags ───────────────────────────────────────────────────
    def pid_identified_tags(self):
        return self.conn.execute(
            "SELECT * FROM pid_identified_tags ORDER BY tag_code").fetchall()

    def upsert_pid_tag(self, tag_code, examples, name_sv, comp_type):
        """Insert or update a scanned tag entry (keeps existing confirmed status)."""
        existing = self.conn.execute(
            "SELECT confirmed FROM pid_identified_tags WHERE tag_code=?",
            (tag_code,)).fetchone()
        if existing:
            self.conn.execute(
                "UPDATE pid_identified_tags SET examples=?,name_sv=?,"
                "comp_type=CASE WHEN confirmed=0 THEN ? ELSE comp_type END "
                "WHERE tag_code=?",
                (examples, name_sv, comp_type, tag_code))
        else:
            self.conn.execute(
                "INSERT INTO pid_identified_tags "
                "(tag_code,examples,name_sv,comp_type,confirmed) VALUES (?,?,?,?,0)",
                (tag_code, examples, name_sv, comp_type))
        self.commit()

    def confirm_pid_tag(self, tag_code, comp_type, confirmed):
        self.conn.execute(
            "UPDATE pid_identified_tags SET comp_type=?,confirmed=? WHERE tag_code=?",
            (comp_type, int(confirmed), tag_code))
        self.commit()

    def confirmed_comp_for_tag(self, prefix: str) -> str:
        """Return confirmed component type for a tag prefix, or ''."""
        r = self.conn.execute(
            "SELECT comp_type FROM pid_identified_tags "
            "WHERE tag_code=? AND confirmed=1", (prefix.upper(),)).fetchone()
        return r['comp_type'] if r else ''

    def all_active_tag_codes(self) -> list:
        """Return list of all active tag codes for highlight scanning."""
        active_std = self.tag_db_setting('active_standard', '')
        if active_std:
            rows = self.conn.execute(
                "SELECT tag_code FROM tag_database WHERE standard=? AND active=1",
                (active_std,)).fetchall()
        else:
            rows = self.conn.execute(
                "SELECT tag_code FROM tag_database WHERE active=1").fetchall()
        return [r[0] for r in rows]

    # ── Equipment catalog ─────────────────────────────────────────────────────
    def equipment_items(self):
        return self.conn.execute(
            "SELECT * FROM equipment_catalog ORDER BY prefix, tag").fetchall()

    def add_equipment_item(self, tag, original_tag, prefix, page, eq_type, desc, is_ocr):
        cur = self.conn.execute(
            "INSERT INTO equipment_catalog "
            "(tag,original_tag,prefix,pid_page,equipment_type,description,is_ocr,include) "
            "VALUES (?,?,?,?,?,?,?,1)",
            (tag, original_tag, prefix, page, eq_type, desc, is_ocr))
        self.commit()
        return cur.lastrowid

    def update_equipment_item(self, id_, tag, prefix, eq_type, desc):
        self.conn.execute(
            "UPDATE equipment_catalog SET tag=?,prefix=?,equipment_type=?,description=? WHERE id=?",
            (tag, prefix, eq_type, desc, id_))
        self.commit()

    def rename_equipment_and_references(self, equipment_id, new_tag):
        """Rename one catalogue object and every visible HAZOP reference.

        Object identity is the stable ``equipment_catalog.id``.  The tag is
        also rendered inside free text in causes, consequences, safeguards
        and recommendations, so changing the catalogue row alone leaves old,
        no-longer-bold text behind.  This is the one explicit *global rename*
        operation used by the HAZOP object popup: it updates marker labels and
        whole-token HAZOP references in one SQLite transaction.

        It intentionally does not touch standard-cause templates.  Those are
        reusable library text, not occurrences in the active HAZOP study.
        """
        equipment = self.get_equipment_by_id(equipment_id)
        if not equipment:
            raise ValueError('Objektet finns inte längre i objektdatabasen.')
        old_tag = str(equipment.get('tag') or '').strip()
        new_tag = str(new_tag or '').strip().upper()
        if not new_tag:
            raise ValueError('Ange en tagg innan objektet byter namn.')
        if not old_tag or old_tag.casefold() == new_tag.casefold():
            return {'old_tag': old_tag, 'new_tag': new_tag, 'changed': 0}

        duplicate = self.conn.execute(
            "SELECT id FROM equipment_catalog WHERE UPPER(tag)=UPPER(?) AND id<>? LIMIT 1",
            (new_tag, equipment_id)).fetchone()
        if duplicate:
            raise ValueError(f'Taggen {new_tag} används redan av ett annat objekt.')

        pattern = re.compile(r'(?<![A-Za-z0-9])' + re.escape(old_tag) +
                             r'(?![A-Za-z0-9])', re.IGNORECASE)

        def replace_token(value):
            return pattern.sub(new_tag, str(value or ''))

        def replace_refs(value):
            return ','.join(
                new_tag if str(ref).strip().casefold() == old_tag.casefold()
                else str(ref).strip()
                for ref in parse_tag_refs(value or ''))

        changes = 0
        try:
            self.conn.execute('BEGIN')
            self.conn.execute("UPDATE equipment_catalog SET tag=? WHERE id=?",
                              (new_tag, equipment_id))
            self.conn.execute("UPDATE equipment_markers SET tag=? WHERE equipment_id=?",
                              (new_tag, equipment_id))

            for table, columns in (
                ('causes', ('description', 'comp_tag')),
                ('consequences', ('description', 'comp_tag', 'tagged_refs')),
                ('safeguards', ('description', 'comp_tag', 'tagged_refs')),
                ('recommendations', ('description',)),
            ):
                select_columns = list(columns)
                if table == 'causes':
                    # The normalizer needs the live group identity as well
                    # as its text fields; the ordinary rename loop only
                    # updates description/comp_tag.
                    select_columns.extend((
                        'equipment_id', 'secondary_equipment_id',
                        'group_equipment_ids'))
                rows = self.conn.execute(
                    f"SELECT id, {', '.join(select_columns)} FROM {table}").fetchall()
                for row in rows:
                    updates = {}
                    for column in columns:
                        old_value = row[column] or ''
                        new_value = (replace_refs(old_value)
                                     if column == 'tagged_refs'
                                     else replace_token(old_value))
                        if new_value != old_value:
                            updates[column] = new_value
                    if table == 'causes':
                        # A tag rename is also a natural repair point for an
                        # old one-line group.  Do this inside the same
                        # transaction: the freshly renamed tag is then
                        # visible in every member row, and subsequent inline
                        # edits cannot be trapped on the primary row merely
                        # because the legacy text was never split.
                        normalised_cause = dict(row)
                        normalised_cause.update(updates)
                        group_ids = self.group_equipment_ids_for_cause(
                            normalised_cause)
                        if len(group_ids) >= 2:
                            description = '\n'.join(
                                self.group_cause_description_lines(
                                    normalised_cause, group_ids))
                            if description != (normalised_cause.get('description') or ''):
                                updates['description'] = description
                    if updates:
                        assignments = ', '.join(f'{column}=?' for column in updates)
                        self.conn.execute(
                            f"UPDATE {table} SET {assignments} WHERE id=?",
                            (*updates.values(), row['id']))
                        changes += 1
            self.commit()
        except Exception:
            self.conn.rollback()
            raise
        return {'old_tag': old_tag, 'new_tag': new_tag, 'changed': changes}

    def replace_equipment_tag_text(self, source, target):
        """Replace text in defined equipment tags and their marker labels.

        Returns the affected catalog rows.  The operation is atomic so the
        confirmation dialog can safely apply the exact preview it showed.
        """
        import re as _re
        source, target = str(source or ''), str(target or '')
        if not source:
            return []
        rows = self.conn.execute(
            "SELECT id, tag FROM equipment_catalog "
            "WHERE LOWER(tag) LIKE LOWER(?) ORDER BY prefix, tag",
            (f'%{source}%',)).fetchall()
        changes = []
        try:
            self.conn.execute('BEGIN')
            for row in rows:
                old = row['tag'] or ''
                new = _re.sub(_re.escape(source), target, old, flags=_re.IGNORECASE)
                if new == old:
                    continue
                self.conn.execute("UPDATE equipment_catalog SET tag=? WHERE id=?",
                                  (new, row['id']))
                self.conn.execute("UPDATE equipment_markers SET tag=? WHERE equipment_id=?",
                                  (new, row['id']))
                changes.append({'id': row['id'], 'old': old, 'new': new})
            self.commit()
        except Exception:
            self.conn.rollback()
            raise
        return changes

    def update_equipment_scan_fields(self, id_, tag, prefix, page, is_ocr):
        """Update only the fields a P&ID rescan actually knows about — used
        by equipment_detection.apply_scan_result_to_equipment_catalog
        (2026-08-24, see NOTES.md) instead of a full clear+reinsert, so a
        tag that survives a rescan keeps its id (and therefore its linked
        deviations/causes) instead of orphaning them. Deliberately leaves
        equipment_type/description/include untouched — those are the
        user's own edits and a rescan shouldn't discard them."""
        self.conn.execute(
            "UPDATE equipment_catalog SET tag=?,original_tag=?,prefix=?,pid_page=?,is_ocr=? WHERE id=?",
            (tag, tag, prefix, page, is_ocr, id_))
        self.commit()

    def delete_equipment_item(self, id_):
        # deviations.equipment_id (added 2026-08-07 for "Nod → Utrustning →
        # Avvikelse", see NOTES.md) and causes.equipment_id (added 2026-08-13
        # for the "Live tag-länk mellan Orsak-cellens taggremsa och objektet
        # på P&ID", see NOTES.md) both have NO ON DELETE clause — clear
        # both references first (keeps the deviation/cause and its own
        # children, just detaches them from the deleted equipment row)
        # instead of hitting sqlite3.IntegrityError: FOREIGN KEY constraint
        # failed, the same root cause already found and fixed for
        # equipment_catalog.node_id in delete_node(). 2026-08-24: causes.
        # equipment_id was missing here entirely (only deviations.equipment_id
        # was ever cleared) — a real crash report
        # (crash_20260824_143009_IntegrityError.json) traced this exact
        # DELETE failing once a real P&ID's causes had tags linked via that
        # column. equipment_markers.equipment_id is the only one of the
        # three FKs into equipment_catalog with ON DELETE CASCADE already,
        # so it needs no manual handling here.
        # LOPA retains snapshots after an object is deleted. Flag affected
        # imports before the FK becomes NULL so the analyst sees a red/missing
        # source rather than a silently detached historical sensor.
        self.conn.execute(
            "UPDATE lopa_source_scenarios SET source_missing=1 WHERE equipment_id=?", (id_,))
        self.conn.execute("UPDATE deviations SET equipment_id=NULL WHERE equipment_id=?", (id_,))
        self.conn.execute("UPDATE causes SET equipment_id=NULL WHERE equipment_id=?", (id_,))
        self.conn.execute("DELETE FROM equipment_catalog WHERE id=?", (id_,))
        self.commit()

    def clear_equipment_catalog(self):
        # Same fix as delete_equipment_item() above, but for the full-
        # rescan-replaces-catalog path ("🔍 Skanna P&ID"/"📋 Analysera P&ID").
        self.conn.execute(
            "UPDATE lopa_source_scenarios SET source_missing=1 WHERE equipment_id IS NOT NULL")
        self.conn.execute("UPDATE deviations SET equipment_id=NULL WHERE equipment_id IS NOT NULL")
        self.conn.execute("UPDATE causes SET equipment_id=NULL WHERE equipment_id IS NOT NULL")
        self.conn.execute("DELETE FROM equipment_catalog")
        self.commit()

    # ── Nod ↔ Utrustning (2026-08-07) ────────────────────────────────────────
    def equipment_node_id(self, equipment_id):
        row = self.conn.execute(
            "SELECT node_id FROM equipment_catalog WHERE id=?", (equipment_id,)).fetchone()
        return row['node_id'] if row else None

    def set_equipment_node(self, equipment_id, node_id):
        self.conn.execute(
            "UPDATE equipment_catalog SET node_id=? WHERE id=?", (node_id, equipment_id))
        self.commit()

    def equipment_deviation_count(self, equipment_id):
        """Count cause rows that visibly reference this equipment.

        The method name is retained for compatibility with older callers,
        but the rubber-band counter represents cause rows, not distinct
        deviations. Several causes under one deviation are counted
        separately.
        """
        equipment = self.get_equipment_by_id(equipment_id)
        if not equipment:
            return 0
        tag = (equipment.get('tag') or '').strip()
        if not tag:
            return 0
        return sum(
            1 for row in self.conn.execute("SELECT * FROM causes").fetchall()
            if self._equipment_tag_matches_cause(dict(row), tag))

    @staticmethod
    def _equipment_tag_matches_row(row, tag):
        """Match an actively visible equipment tag in a HAZOP row.

        Do not use ``tagged_refs`` here: that column intentionally keeps
        historical drag references for bold rendering, even after a user
        removes the tag from the actual text. Counters must reflect current
        occurrences, not additions made in the past.
        """
        if not tag:
            return False
        tag = str(tag).strip()
        # Only text currently visible in the consequence/safeguard cell is
        # authoritative.  comp_tag/tagged_refs can be stale after editing.
        text = str(row.get('description') or '')
        return bool(re.search(
            rf'(?<![A-Za-z0-9]){re.escape(tag)}(?![A-Za-z0-9])',
            text, re.IGNORECASE))

    @staticmethod
    def _equipment_tag_matches_cause(row, tag):
        """Match the visible identity of an ordinary cause cell.

        Cause tags are rendered from ``comp_tag`` as the bold prefix, so its
        current value is part of that cell's visible text.
        """
        if not tag:
            return False
        text = ' '.join(str(row.get(key) or '')
                        for key in ('comp_tag', 'description'))
        return bool(re.search(
            rf'(?<![A-Za-z0-9]){re.escape(str(tag).strip())}(?![A-Za-z0-9])',
            text, re.IGNORECASE))

    @staticmethod
    def _equipment_tag_in_description(row, tag):
        return bool(re.search(
            rf'(?<![A-Za-z0-9]){re.escape(str(tag).strip())}(?![A-Za-z0-9])',
            str(row.get('description') or ''), re.IGNORECASE))

    def equipment_consequence_count(self, comp_tag, comp_type=''):
        """How many consequences reference this equipment's tag — the
        'förekomster i konsekvenser' counter (2026-08-11, see NOTES.md).
        consequences has no equipment_id FK (only the flat comp_tag/
        comp_type columns set by set_consequence_tag), so — unlike
        equipment_deviation_count's FK join — this matches by tag+type."""
        if not comp_tag:
            return 0
        return sum(
            1 for row in self.conn.execute("SELECT * FROM consequences").fetchall()
            if self._equipment_tag_matches_row(dict(row), comp_tag)
        )

    def equipment_safeguard_count(self, comp_tag, comp_type=''):
        """How many safeguards reference this equipment's tag — the
        'förekomster i safeguards' counter (2026-08-11, see NOTES.md).
        Mirrors equipment_consequence_count (safeguards also has no
        equipment_id FK, only comp_tag/comp_type)."""
        if not comp_tag:
            return 0
        return sum(
            1 for row in self.conn.execute("SELECT * FROM safeguards").fetchall()
            if self._equipment_tag_matches_row(dict(row), comp_tag)
        )

    def equipment_recommendation_count(self, comp_tag, comp_type=''):
        """Count recommendation rows whose visible text contains this tag.

        Recommendations are study-wide catalog rows, so a reused
        recommendation is counted once, not once per consequence link.
        """
        if not comp_tag:
            return 0
        return sum(
            1 for row in self.conn.execute("SELECT * FROM recommendations").fetchall()
            if self._equipment_tag_matches_row(dict(row), comp_tag)
        )

    def set_deviation_equipment(self, deviation_id, equipment_id):
        """Tie an EXISTING deviation to a specific equipment item — used
        when equipment is drag-and-dropped directly onto a deviation
        already sitting in the HAZOP tree (2026-08-09, see NOTES.md),
        as opposed to get_or_create_deviation()'s own equipment_id param
        (which only ever applies to a brand-new deviation it creates).
        Backs both the Nod → Ledord → Utrustning tree grouping and the
        worksheet's separate Utrustning column."""
        self.conn.execute(
            "UPDATE deviations SET equipment_id=? WHERE id=?", (equipment_id, deviation_id))
        self.commit()

    # ── Equipment types ───────────────────────────────────────────────────────
    def get_equipment_type(self, prefix: str):
        """Return saved equipment_type for this prefix, or None."""
        row = self.conn.execute(
            "SELECT equipment_type FROM equipment_types WHERE prefix=?", (prefix,)).fetchone()
        return row['equipment_type'] if row else None

    def save_equipment_type(self, prefix: str, equipment_type: str, display_name: str = ''):
        self.conn.execute(
            "INSERT OR REPLACE INTO equipment_types (prefix, equipment_type, display_name) "
            "VALUES (?,?,?)", (prefix, equipment_type, display_name))
        self.commit()

    def all_equipment_types(self):
        return self.conn.execute(
            "SELECT * FROM equipment_types ORDER BY prefix").fetchall()

    def get_equipment_by_tag(self, tag: str):
        """Return equipment_catalog row for a full tag string (case-insensitive)."""
        row = self.conn.execute(
            "SELECT * FROM equipment_catalog WHERE UPPER(tag)=UPPER(?) AND include=1 LIMIT 1",
            (tag,)).fetchone()
        return dict(row) if row else None

    def get_equipment_by_id(self, id_):
        row = self.conn.execute(
            "SELECT * FROM equipment_catalog WHERE id=?", (id_,)).fetchone()
        return dict(row) if row else None

    def get_equipment_by_marker_id(self, marker_id):
        """Resolve an equipment_markers.id (what a P&ID drag/drop mime
        carries — see NOTES.md) to its linked equipment_catalog row, or
        None if the marker has no linked equipment (untagged shape hit).
        Single shared lookup for every equipment drag-and-drop target
        (KON/SG cells, HAZOP tree deviations) instead of each repeating
        the same two-step marker->catalog join."""
        row = self.conn.execute(
            "SELECT equipment_id FROM equipment_markers WHERE id=?", (marker_id,)).fetchone()
        if not row or row['equipment_id'] is None:
            return None
        return self.get_equipment_by_id(row['equipment_id'])

    # ── Smart object recognition: study tag memory ─────────────────────────────

    def get_tag_memory(self, tag: str):
        """Look up study_tag_memory by the letter prefix of tag (active entries only)."""
        pfx = _tag_letter_prefix(tag) if tag else ''
        if not pfx:
            return None
        row = self.conn.execute(
            "SELECT * FROM study_tag_memory WHERE UPPER(tag)=UPPER(?) AND active=1 LIMIT 1",
            (pfx,)).fetchone()
        return dict(row) if row else None

    def upsert_tag_memory(self, tag: str, comp_type: str,
                          comp_tag: str = '', phash: str = ''):
        """Increment the usage counter for (prefix, comp_type).

        Each (prefix, comp_type) pair has its own counter so the same prefix
        can accumulate counts for multiple types independently.  On lookup,
        the type with the highest count wins.

        Numbers are ignored: 'PU101', 'PU102', 'E1.M1.PU103' all update 'PU'.

        Only increments the usage counter — never deactivates other types.
        The winner is determined by highest count among active entries.
        active=0 is only set manually via the Smart Recognition panel.
        """
        if not comp_type:
            return
        pfx = _tag_letter_prefix(tag) if tag else ''
        if not pfx:
            return
        now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
        existing = self.conn.execute(
            "SELECT usage_count FROM study_tag_memory "
            "WHERE UPPER(tag)=UPPER(?) AND UPPER(comp_type)=UPPER(?)",
            (pfx, comp_type)).fetchone()
        if existing:
            self.conn.execute(
                "UPDATE study_tag_memory SET comp_tag=?,phash=?,"
                "usage_count=usage_count+1,updated=? "
                "WHERE UPPER(tag)=UPPER(?) AND UPPER(comp_type)=UPPER(?)",
                (comp_tag, phash, now, pfx, comp_type))
        else:
            self.conn.execute(
                "INSERT INTO study_tag_memory (tag,comp_type,comp_tag,phash,active,updated)"
                " VALUES (?,?,?,?,1,?)",
                (pfx, comp_type, comp_tag, phash, now))
        self.commit()

    def get_prefix_memory(self, prefix: str) -> str:
        """Return the most-confirmed comp_type for a letter prefix.
        Highest usage_count among active=1 entries wins.
        Tie-broken by most recently updated (latest confirmation wins).
        """
        if not prefix:
            return ''
        try:
            row = self.conn.execute(
                "SELECT comp_type FROM study_tag_memory "
                "WHERE UPPER(tag)=UPPER(?) AND active=1 "
                "ORDER BY usage_count DESC, updated DESC LIMIT 1",
                (prefix,)).fetchone()
            return row['comp_type'] if row else ''
        except Exception:
            try:
                row = self.conn.execute(
                    "SELECT comp_type FROM study_tag_memory "
                    "WHERE UPPER(tag)=UPPER(?) "
                    "ORDER BY usage_count DESC, updated DESC LIMIT 1",
                    (prefix,)).fetchone()
                return row['comp_type'] if row else ''
            except Exception:
                return ''

    def get_tag_type_suggestions(self, tag: str, limit: int = 8):
        """Rank remembered object types for a newly detected tag.

        The alphabetic code (for example ``FT`` in ``20-FT-201``) is the
        primary key; serial digits are intentionally ignored.  Usage count
        provides the learned weighting when one code has been classified as
        more than one type.  A remembered full example tag gets a strong
        bonus, so an exact repeat wins without making other valid choices
        disappear from the dropdown.
        """
        raw = str(tag or '').strip().upper()
        prefix = _tag_letter_prefix(raw)
        if not prefix:
            return []
        try:
            rows = self.conn.execute(
                "SELECT tag, comp_type, comp_tag, usage_count, updated "
                "FROM study_tag_memory WHERE UPPER(tag)=UPPER(?) AND active=1",
                (prefix,)).fetchall()
        except Exception:
            return []
        ranked = []
        for row in rows:
            item = dict(row)
            usage = max(0, int(item.get('usage_count') or 0))
            exact = 1 if str(item.get('comp_tag') or '').strip().upper() == raw else 0
            # Prefix match is the shared evidence; count and exact examples
            # decide between competing types for the same alphabetic code.
            score = usage * 10 + exact * 1000
            ranked.append({
                'comp_type': item.get('comp_type') or '',
                'score': score,
                'usage_count': usage,
                'exact': bool(exact),
                'updated': item.get('updated') or '',
            })
        ranked.sort(key=lambda item: (item['score'], item['usage_count'], item['updated']), reverse=True)
        return ranked[:max(1, int(limit))]

    def set_tag_memory_active(self, prefix: str, comp_type: str, active: bool):
        """Enable/disable a specific (prefix, comp_type) entry."""
        self.conn.execute(
            "UPDATE study_tag_memory SET active=? "
            "WHERE UPPER(tag)=UPPER(?) AND UPPER(comp_type)=UPPER(?)",
            (1 if active else 0, prefix, comp_type))
        self.commit()

    def find_fingerprint(self, phash: str, max_distance: int = 50):
        """Return best matching symbol_fingerprints row by Hamming distance, or None."""
        if not phash:
            return None
        try:
            h1 = int(phash, 16)
        except ValueError:
            return None
        best = None; best_dist = max_distance + 1
        rows = self.conn.execute(
            "SELECT * FROM symbol_fingerprints ORDER BY usage_count DESC").fetchall()
        for row in rows:
            try:
                h2 = int(row['phash'], 16)
                dist = bin(h1 ^ h2).count('1')
                if dist < best_dist:
                    best_dist = dist
                    best = dict(row)
            except Exception:
                continue
        return best

    def store_fingerprint(self, phash: str, comp_type: str, tag_example: str = ''):
        """Save or increment usage count for a visual fingerprint."""
        if not phash or not comp_type:
            return
        existing = self.find_fingerprint(phash, max_distance=30)
        if existing:
            self.conn.execute(
                "UPDATE symbol_fingerprints SET usage_count=usage_count+1,comp_type=? WHERE id=?",
                (comp_type, existing['id']))
        else:
            self.conn.execute(
                "INSERT INTO symbol_fingerprints (phash,comp_type,tag_example) VALUES (?,?,?)",
                (phash, comp_type, tag_example))
        self.commit()

    # ── Categories ────────────────────────────────────────────────────────────
    def consequence_categories(self):
        return self.conn.execute(
            "SELECT * FROM consequence_categories ORDER BY sort_order, name").fetchall()

    def get_consequence_severities(self, consequence_id):
        """Return list of {id, category_id, name, severity} for each assessed category."""
        return self.conn.execute(
            "SELECT cs.id, cs.category_id, cc.name, cs.severity "
            "FROM consequence_severities cs "
            "JOIN consequence_categories cc ON cc.id=cs.category_id "
            "WHERE cs.consequence_id=? ORDER BY cc.sort_order, cc.name",
            (consequence_id,)).fetchall()

    def get_consequence_severities_for_consequences(self, consequence_ids):
        """Bulk version of get_consequence_severities() — same row shape
        (each row also carries consequence_id, harmless extra key for
        existing dict(row) callers) grouped by consequence_id, same
        per-group ordering (category sort_order, then name). 2026-08-24,
        see NOTES.md — fixes part of ScenarioTablePanel._build_rows()'s
        O(causes × consequences × categories) query pattern."""
        consequence_ids = list(consequence_ids)
        result = {cid: [] for cid in consequence_ids}
        if not consequence_ids:
            return result
        CHUNK = 500
        for start in range(0, len(consequence_ids), CHUNK):
            chunk = consequence_ids[start:start + CHUNK]
            placeholders = ','.join('?' * len(chunk))
            rows = self.conn.execute(
                "SELECT cs.id, cs.category_id, cc.name, cs.severity, cs.consequence_id "
                "FROM consequence_severities cs "
                "JOIN consequence_categories cc ON cc.id=cs.category_id "
                f"WHERE cs.consequence_id IN ({placeholders}) "
                "ORDER BY cs.consequence_id, cc.sort_order, cc.name", chunk).fetchall()
            for row in rows:
                result[row['consequence_id']].append(row)
        return result

    def set_consequence_severity(self, consequence_id, category_id, severity):
        """Set (or clear when severity=0) a per-category severity for a consequence."""
        if not severity:
            self.conn.execute(
                "DELETE FROM consequence_severities "
                "WHERE consequence_id=? AND category_id=?",
                (consequence_id, category_id))
        else:
            self.conn.execute(
                "INSERT OR REPLACE INTO consequence_severities "
                "(consequence_id, category_id, severity) VALUES (?,?,?)",
                (consequence_id, category_id, severity))
        self.commit()

    def get_final_consequence_severities(self, consequence_id):
        """Return explicitly selected post-barrier category severities.

        Missing categories deliberately have no row: the caller must fall
        back to the corresponding regular consequence severity.
        """
        return self.conn.execute(
            "SELECT category_id, severity FROM consequence_final_severities "
            "WHERE consequence_id=?", (consequence_id,)).fetchall()

    def get_final_consequence_severities_for_consequences(self, consequence_ids):
        """Bulk map ``{consequence_id: {category_id: severity}}`` for
        optional post-barrier severity overrides."""
        consequence_ids = list(consequence_ids)
        result = {cid: {} for cid in consequence_ids}
        if not consequence_ids:
            return result
        CHUNK = 500
        for start in range(0, len(consequence_ids), CHUNK):
            chunk = consequence_ids[start:start + CHUNK]
            placeholders = ','.join('?' * len(chunk))
            rows = self.conn.execute(
                "SELECT consequence_id, category_id, severity "
                "FROM consequence_final_severities "
                f"WHERE consequence_id IN ({placeholders})", chunk).fetchall()
            for row in rows:
                result[row['consequence_id']][row['category_id']] = row['severity']
        return result

    def set_final_consequence_severity(self, consequence_id, category_id, severity):
        """Set or clear an optional post-barrier category severity.

        A zero/false value removes the override and restores the regular
        category severity as the effective final-consequence level.
        """
        if not severity:
            self.conn.execute(
                "DELETE FROM consequence_final_severities "
                "WHERE consequence_id=? AND category_id=?",
                (consequence_id, category_id))
        else:
            self.conn.execute(
                "INSERT OR REPLACE INTO consequence_final_severities "
                "(consequence_id, category_id, severity) VALUES (?,?,?)",
                (consequence_id, category_id, severity))
        self.commit()

    def get_severity_excluded_sgs(self, severity_id):
        """Return set of safeguard_ids excluded from this category assessment."""
        rows = self.conn.execute(
            "SELECT safeguard_id FROM consequence_severity_exclusions WHERE severity_id=?",
            (severity_id,)).fetchall()
        return {r[0] for r in rows}

    def get_severity_excluded_sgs_for_severities(self, severity_ids):
        """Bulk version of get_severity_excluded_sgs() — one/few queries
        for many severity_ids instead of one per id (2026-08-24, see
        NOTES.md — fixes part of ScenarioTablePanel._build_rows()'s
        O(causes × consequences × categories) query pattern). Returns
        {severity_id: set(safeguard_ids)}, every requested id present
        (possibly with an empty set)."""
        severity_ids = list(severity_ids)
        result = {sid: set() for sid in severity_ids}
        if not severity_ids:
            return result
        CHUNK = 500
        for start in range(0, len(severity_ids), CHUNK):
            chunk = severity_ids[start:start + CHUNK]
            placeholders = ','.join('?' * len(chunk))
            rows = self.conn.execute(
                "SELECT severity_id, safeguard_id FROM consequence_severity_exclusions "
                f"WHERE severity_id IN ({placeholders})", chunk).fetchall()
            for row in rows:
                result[row[0]].add(row[1])
        return result

    def set_severity_excluded_sgs(self, severity_id, excluded_sg_ids):
        self.conn.execute(
            "DELETE FROM consequence_severity_exclusions WHERE severity_id=?",
            (severity_id,))
        for sg_id in excluded_sg_ids:
            self.conn.execute(
                "INSERT OR IGNORE INTO consequence_severity_exclusions "
                "(severity_id, safeguard_id) VALUES (?,?)", (severity_id, sg_id))
        self.commit()

    def get_severity_excluded_reduction_factors(self, severity_id):
        """Return enabler ids that do not apply to one category assessment."""
        rows = self.conn.execute(
            "SELECT reduction_factor_id FROM reduction_factor_severity_exclusions "
            "WHERE severity_id=?", (severity_id,)).fetchall()
        return {row[0] for row in rows}

    def get_severity_excluded_reduction_factors_for_severities(self, severity_ids):
        """Bulk category exclusions for enablers, mirroring safeguard lookup."""
        severity_ids = list(severity_ids)
        result = {sid: set() for sid in severity_ids}
        if not severity_ids:
            return result
        chunk_size = 500
        for start in range(0, len(severity_ids), chunk_size):
            chunk = severity_ids[start:start + chunk_size]
            placeholders = ','.join('?' * len(chunk))
            rows = self.conn.execute(
                "SELECT severity_id,reduction_factor_id "
                "FROM reduction_factor_severity_exclusions "
                f"WHERE severity_id IN ({placeholders})", chunk).fetchall()
            for row in rows:
                result[row[0]].add(row[1])
        return result

    def set_severity_excluded_reduction_factors(self, severity_id, excluded_factor_ids):
        self.conn.execute(
            "DELETE FROM reduction_factor_severity_exclusions WHERE severity_id=?",
            (severity_id,))
        for factor_id in excluded_factor_ids:
            self.conn.execute(
                "INSERT OR IGNORE INTO reduction_factor_severity_exclusions "
                "(severity_id,reduction_factor_id) VALUES (?,?)",
                (severity_id, factor_id))
        self.commit()

    def get_safeguard_excluded_causes(self, sg_id):
        """Return set of cause_ids excluded from this safeguard."""
        rows = self.conn.execute(
            "SELECT cause_id FROM safeguard_cause_exclusions WHERE safeguard_id=?",
            (sg_id,)).fetchall()
        return {r[0] for r in rows}

    def get_safeguard_excluded_causes_for_safeguards(self, sg_ids):
        """Bulk version of get_safeguard_excluded_causes() — see
        get_severity_excluded_sgs_for_severities's docstring (2026-08-24,
        NOTES.md). Returns {sg_id: set(cause_ids)}."""
        sg_ids = list(sg_ids)
        result = {sid: set() for sid in sg_ids}
        if not sg_ids:
            return result
        CHUNK = 500
        for start in range(0, len(sg_ids), CHUNK):
            chunk = sg_ids[start:start + CHUNK]
            placeholders = ','.join('?' * len(chunk))
            rows = self.conn.execute(
                "SELECT safeguard_id, cause_id FROM safeguard_cause_exclusions "
                f"WHERE safeguard_id IN ({placeholders})", chunk).fetchall()
            for row in rows:
                result[row[0]].add(row[1])
        return result

    def set_safeguard_excluded_causes(self, sg_id, cause_id_set):
        self.conn.execute(
            "DELETE FROM safeguard_cause_exclusions WHERE safeguard_id=?", (sg_id,))
        for cid in cause_id_set:
            self.conn.execute(
                "INSERT OR IGNORE INTO safeguard_cause_exclusions "
                "(safeguard_id, cause_id) VALUES (?,?)", (sg_id, cid))
        self.commit()

    def add_category(self, name):
        cur = self.conn.execute(
            "INSERT INTO consequence_categories (name) VALUES (?)", (name,))
        self.commit()
        return cur.lastrowid

    def update_category(self, id_, name):
        self.conn.execute("UPDATE consequence_categories SET name=? WHERE id=?", (name, id_))
        self.commit()

    def delete_category(self, id_):
        self.conn.execute("DELETE FROM consequence_categories WHERE id=?", (id_,))
        self.commit()

    def reorder_categories(self, ordered_ids):
        """Persist a new display order for consequence categories (2026-08-11,
        'jag vill även kunna justera ordningen') — mirrors
        reorder_standard_deviations()'s established pattern. consequence_categories()
        already ORDERs BY sort_order, so this is the only piece that was missing."""
        for i, id_ in enumerate(ordered_ids):
            self.conn.execute(
                "UPDATE consequence_categories SET sort_order=? WHERE id=?", (i, id_))
        self.commit()

    # ── Deltagarmatris: participants, analysis sessions, attendance ─────────────
    # (2026-08-11, replaces the old free-text 'project_participants' field —
    # see SettingsPanel's "Deltagare" tab and NOTES.md.)
    def list_participants(self):
        return self.conn.execute(
            "SELECT * FROM participants ORDER BY sort_order, id").fetchall()

    def add_participant(self, first_name='', last_name='', role=''):
        cur = self.conn.execute(
            "INSERT INTO participants (first_name, last_name, role) VALUES (?,?,?)",
            (first_name, last_name, role))
        self.commit()
        return cur.lastrowid

    def update_participant(self, id_, first_name=None, last_name=None, role=None):
        row = self.conn.execute("SELECT * FROM participants WHERE id=?", (id_,)).fetchone()
        if not row:
            return
        self.conn.execute(
            "UPDATE participants SET first_name=?, last_name=?, role=? WHERE id=?",
            (first_name if first_name is not None else row['first_name'],
             last_name if last_name is not None else row['last_name'],
             role if role is not None else row['role'],
             id_))
        self.commit()

    def delete_participant(self, id_):
        self.conn.execute("DELETE FROM participants WHERE id=?", (id_,))
        self.commit()

    def list_analysis_sessions(self):
        return self.conn.execute(
            "SELECT * FROM analysis_sessions ORDER BY sort_order, id").fetchall()

    def add_analysis_session(self, label='', date=None, location=None, is_digital=False):
        # New sessions normally follow the latest dated session by one day.
        # Keep the old label argument for project files created before the
        # date/location fields existed.
        if date is None:
            dates = []
            for row in self.list_analysis_sessions():
                value = row['date'] or row['label'] or ''
                try:
                    dates.append(datetime.date.fromisoformat(str(value)[:10]))
                except (TypeError, ValueError):
                    continue
            next_date = (max(dates) + datetime.timedelta(days=1)) if dates else datetime.date.today()
            date = next_date.isoformat()
        if location is None:
            previous = list(self.list_analysis_sessions())
            location = (previous[-1]['location'] or '') if previous else ''
        if not label:
            label = str(date)
        cur = self.conn.execute(
            "INSERT INTO analysis_sessions (label,date,location,is_digital,start_time,end_time) "
            "VALUES (?,?,?,?,?,?)",
            (label, date, location or '', 1 if is_digital else 0, '', ''))
        self.commit()
        return cur.lastrowid

    def update_analysis_session(self, id_, label):
        self.conn.execute("UPDATE analysis_sessions SET label=? WHERE id=?", (label, id_))
        self.commit()

    def update_analysis_session_details(self, id_, date=None, location=None, label=None,
                                        is_digital=None, start_time=None, end_time=None):
        row = self.conn.execute("SELECT * FROM analysis_sessions WHERE id=?", (id_,)).fetchone()
        if not row:
            return
        self.conn.execute(
            "UPDATE analysis_sessions SET date=?, location=?, label=?, is_digital=?, "
            "start_time=?, end_time=? WHERE id=?",
            (date if date is not None else (row['date'] or row['label'] or ''),
             location if location is not None else (row['location'] or ''),
             label if label is not None else (row['label'] or ''),
             int(is_digital) if is_digital is not None else int(row['is_digital'] or 0),
             start_time if start_time is not None else (row['start_time'] or ''),
             end_time if end_time is not None else (row['end_time'] or ''), id_))
        self.commit()

    def set_analysis_session_location(self, id_, location):
        self.conn.execute("UPDATE analysis_sessions SET location=? WHERE id=?",
                          (location or '', id_))
        self.commit()

    def delete_analysis_session(self, id_):
        self.conn.execute("DELETE FROM analysis_sessions WHERE id=?", (id_,))
        self.commit()

    def get_attendance(self, participant_id, session_id):
        row = self.conn.execute(
            "SELECT attended FROM participant_attendance WHERE participant_id=? AND session_id=?",
            (participant_id, session_id)).fetchone()
        return bool(row['attended']) if row else False

    def set_attendance(self, participant_id, session_id, attended):
        self.conn.execute(
            "INSERT INTO participant_attendance (participant_id, session_id, attended) "
            "VALUES (?,?,?) ON CONFLICT(participant_id, session_id) "
            "DO UPDATE SET attended=excluded.attended",
            (participant_id, session_id, 1 if attended else 0))
        self.commit()

    def get_attendance_matrix(self):
        """Return dict {(participant_id, session_id): bool} for all recorded attendance rows."""
        rows = self.conn.execute(
            "SELECT participant_id, session_id, attended FROM participant_attendance").fetchall()
        return {(r['participant_id'], r['session_id']): bool(r['attended']) for r in rows}

    def list_participant_columns(self):
        return [dict(r) for r in self.conn.execute(
            "SELECT * FROM participant_columns ORDER BY sort_order, id")]

    def add_participant_column(self, name):
        max_ord = (self.conn.execute(
            "SELECT COALESCE(MAX(sort_order),-1) FROM participant_columns").fetchone()[0])
        cur = self.conn.execute(
            "INSERT INTO participant_columns (name, sort_order) VALUES (?,?)",
            (name, max_ord + 1))
        self.commit()
        return cur.lastrowid

    def update_participant_column(self, id_, name):
        self.conn.execute("UPDATE participant_columns SET name=? WHERE id=?", (name, id_))
        self.commit()

    def delete_participant_column(self, id_):
        self.conn.execute("DELETE FROM participant_columns WHERE id=?", (id_,))
        self.commit()

    def get_participant_column_values(self):
        """Return dict {(participant_id, column_id): value} for all recorded values."""
        rows = self.conn.execute(
            "SELECT participant_id, column_id, value FROM participant_column_values").fetchall()
        return {(r['participant_id'], r['column_id']): r['value'] for r in rows}

    def set_participant_column_value(self, participant_id, column_id, value):
        self.conn.execute(
            "INSERT INTO participant_column_values (participant_id, column_id, value) "
            "VALUES (?,?,?) ON CONFLICT(participant_id, column_id) "
            "DO UPDATE SET value=excluded.value",
            (participant_id, column_id, value))
        self.commit()

    def get_severity_definitions(self):
        """Return dict: severity_level (1-based int) -> {category_id -> description}."""
        rows = self.conn.execute(
            "SELECT severity_level, category_id, description FROM severity_definitions"
        ).fetchall()
        result = {}
        for r in rows:
            lvl = r['severity_level']
            if lvl not in result:
                result[lvl] = {}
            result[lvl][r['category_id']] = r['description']
        return result

    def set_severity_definition(self, severity_level, category_id, description):
        self.conn.execute(
            "INSERT INTO severity_definitions (severity_level, category_id, description) "
            "VALUES (?,?,?) ON CONFLICT(severity_level,category_id) DO UPDATE SET description=excluded.description",
            (severity_level, category_id, description))
        self.commit()

    # ── Component types & failure modes ───────────────────────────────────────
    def component_types(self):
        return self.conn.execute(
            "SELECT * FROM component_types ORDER BY sort_order, name").fetchall()

    def failure_modes(self, component_id):
        return self.conn.execute(
            "SELECT * FROM failure_modes WHERE component_id=? ORDER BY sort_order, id",
            (component_id,)).fetchall()

    def add_component_type(self, name):
        cur = self.conn.execute(
            "INSERT INTO component_types (name) VALUES (?)", (name,))
        self.commit()
        return cur.lastrowid

    def update_component_type(self, id_, name):
        self.conn.execute("UPDATE component_types SET name=? WHERE id=?", (name, id_))
        self.commit()

    def delete_component_type(self, id_):
        self.conn.execute("DELETE FROM component_types WHERE id=?", (id_,))
        self.commit()

    def add_failure_mode(self, component_id, description, freq=None):
        cur = self.conn.execute(
            "INSERT INTO failure_modes (component_id, description, freq_per_year) VALUES (?,?,?)",
            (component_id, description, freq))
        self.commit()
        return cur.lastrowid

    def update_failure_mode(self, id_, description, freq=None):
        self.conn.execute(
            "UPDATE failure_modes SET description=?, freq_per_year=? WHERE id=?",
            (description, freq, id_))
        self.commit()

    def delete_failure_mode(self, id_):
        self.conn.execute("DELETE FROM failure_modes WHERE id=?", (id_,))
        self.commit()

    # ── P&ID helpers ──────────────────────────────────────────────────────────
    def get_pid_path(self):
        row = self.conn.execute("SELECT value FROM pid_config WHERE key='path'").fetchone()
        return row['value'] if row else None

    def set_pid_path(self, path):
        self.conn.execute(
            "INSERT OR REPLACE INTO pid_config (key,value) VALUES ('path',?)", (str(path),))
        self.commit()

    def get_pid_config_value(self, key):
        row = self.conn.execute(
            "SELECT value FROM pid_config WHERE key=?", (key,)).fetchone()
        return row['value'] if row else None

    def set_pid_config_value(self, key, value):
        self.conn.execute(
            "INSERT OR REPLACE INTO pid_config (key,value) VALUES (?,?)", (key, str(value)))
        self.commit()

    def clear_connector_analysis(self):
        self.conn.execute("DELETE FROM off_page_connector")
        self.conn.execute("DELETE FROM pid_connection")
        self.commit()

    def save_connectors(self, rows):
        if not rows:
            return
        for r in rows:
            r.setdefault('ref_page', None)
        self.conn.executemany(
            "INSERT INTO off_page_connector "
            "(pid_page,x_pdf,y_pdf,direction,edge,ref_text,ref_sheet,"
            "ref_line_id,media_type,weight,confidence,raw_text,ocr_used,analyzed_at,ref_page) "
            "VALUES(:pid_page,:x_pdf,:y_pdf,:direction,:edge,:ref_text,:ref_sheet,"
            ":ref_line_id,:media_type,:weight,:confidence,:raw_text,:ocr_used,:analyzed_at,:ref_page)",
            rows)
        self.commit()

    def update_connector_dot_position(self, connector_id, x, y):
        """Persist a manually dragged dot position for one off-page connector."""
        self.conn.execute(
            "UPDATE off_page_connector SET dot_scene_x=?, dot_scene_y=? WHERE id=?",
            (x, y, connector_id))
        self.commit()

    def save_pid_connections(self, rows):
        if not rows:
            return
        self.conn.executemany(
            "INSERT INTO pid_connection "
            "(from_page,to_page,from_connector,to_connector,media_type,weight,"
            "confidence,is_bidirectional,is_ghost,ghost_ref,warning) "
            "VALUES(:from_page,:to_page,:from_connector,:to_connector,:media_type,"
            ":weight,:confidence,:is_bidirectional,:is_ghost,:ghost_ref,:warning)",
            rows)
        self.commit()

    # ── Board annotations (sticky notes, feature 8) ──────────────────────────
    def get_board_annotations(self):
        return [dict(r) for r in self.conn.execute(
            "SELECT id,x,y,w,h,text,color FROM board_annotations")]

    def add_board_annotation(self, x, y, text='', color='#fff9c4', w=200, h=80):
        cur = self.conn.execute(
            "INSERT INTO board_annotations (x,y,w,h,text,color) VALUES (?,?,?,?,?,?)",
            (x, y, w, h, text, color))
        self.commit()
        return cur.lastrowid

    def update_board_annotation(self, id_, x=None, y=None, w=None, h=None,
                                 text=None, color=None):
        sets, vals = [], []
        for col, val in (('x',x),('y',y),('w',w),('h',h),('text',text),('color',color)):
            if val is not None:
                sets.append(f"{col}=?"); vals.append(val)
        if sets:
            self.conn.execute(
                f"UPDATE board_annotations SET {', '.join(sets)} WHERE id=?",
                vals + [id_])
            self.commit()

    def delete_board_annotation(self, id_):
        self.conn.execute("DELETE FROM board_annotations WHERE id=?", (id_,))
        self.commit()

    def get_pid_connections(self):
        return self.conn.execute("SELECT * FROM pid_connection").fetchall()

    def get_connectors(self):
        return self.conn.execute("SELECT * FROM off_page_connector").fetchall()

    def delete_pid_connection(self, conn_id):
        self.conn.execute("DELETE FROM pid_connection WHERE id=?", (conn_id,))
        self.commit()

    def add_manual_pid_connection(self, from_page, to_page):
        """Insert a manual (user-defined) inter-sheet link with max confidence."""
        import datetime
        self.conn.execute(
            "INSERT INTO pid_connection "
            "(from_page,to_page,from_connector,to_connector,media_type,weight,"
            "confidence,is_bidirectional,is_ghost,ghost_ref,warning) "
            "VALUES (?,?,NULL,NULL,'unknown',1.0,1.0,1,0,NULL,'manual')",
            (from_page, to_page))
        self.commit()

    # ── PID revisions & sheets ────────────────────────────────────────────────
    def add_revision(self, revision, notes, pdf_path, created_at=''):
        if not created_at:
            created_at = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
        cur = self.conn.execute(
            "INSERT INTO pid_revisions (revision,notes,created_at,pdf_path) VALUES (?,?,?,?)",
            (revision, notes, created_at, str(pdf_path)))
        self.commit()
        return cur.lastrowid

    def get_revisions(self):
        return self.conn.execute(
            "SELECT * FROM pid_revisions ORDER BY id DESC").fetchall()

    def project_revisions(self):
        return [dict(r) for r in self.conn.execute(
            "SELECT * FROM project_revisions ORDER BY sort_order, id")]

    def add_project_revision(self, label, date='', description=''):
        max_ord = (self.conn.execute(
            "SELECT COALESCE(MAX(sort_order),-1) FROM project_revisions").fetchone()[0])
        cur = self.conn.execute(
            "INSERT INTO project_revisions (label, date, description, sort_order) "
            "VALUES (?,?,?,?)", (label, date, description, max_ord + 1))
        self.commit()
        return cur.lastrowid

    def update_project_revision(self, id_, label=None, date=None, description=None):
        row = self.conn.execute(
            "SELECT * FROM project_revisions WHERE id=?", (id_,)).fetchone()
        if not row:
            return
        label = row['label'] if label is None else label
        date = row['date'] if date is None else date
        description = row['description'] if description is None else description
        self.conn.execute(
            "UPDATE project_revisions SET label=?, date=?, description=? WHERE id=?",
            (label, date, description, id_))
        self.commit()

    def delete_project_revision(self, id_):
        self.conn.execute("DELETE FROM project_revisions WHERE id=?", (id_,))
        self.commit()

    def project_custom_fields(self):
        return [dict(r) for r in self.conn.execute(
            "SELECT * FROM project_custom_fields ORDER BY sort_order, id")]

    def add_project_custom_field(self, name='', value=''):
        max_ord = (self.conn.execute(
            "SELECT COALESCE(MAX(sort_order),-1) FROM project_custom_fields").fetchone()[0])
        cur = self.conn.execute(
            "INSERT INTO project_custom_fields (name, value, sort_order) VALUES (?,?,?)",
            (name, value, max_ord + 1))
        self.commit()
        return cur.lastrowid

    def update_project_custom_field(self, id_, name=None, value=None):
        row = self.conn.execute(
            "SELECT * FROM project_custom_fields WHERE id=?", (id_,)).fetchone()
        if not row:
            return
        name = row['name'] if name is None else name
        value = row['value'] if value is None else value
        self.conn.execute(
            "UPDATE project_custom_fields SET name=?, value=? WHERE id=?", (name, value, id_))
        self.commit()

    def delete_project_custom_field(self, id_):
        self.conn.execute("DELETE FROM project_custom_fields WHERE id=?", (id_,))
        self.commit()

    def ensure_sheets_initialized(self, page_count, pdf_path=None):
        existing = self.conn.execute("SELECT COUNT(*) FROM pid_sheets").fetchone()[0]
        if existing == 0 and page_count > 0:
            # Bladnamn = "Filnamn – sida N" (2026-08-17, user-confirmed
            # format, see NOTES.md) instead of the old generic "Blad N" —
            # falls back to "Blad N" only if no path was supplied.
            stem = Path(pdf_path).stem if pdf_path else None
            for i in range(page_count):
                name = f"{stem} – sida {i + 1}" if stem else f"Blad {i + 1}"
                self.conn.execute(
                    "INSERT INTO pid_sheets (display_order,physical_page,sheet_name) VALUES (?,?,?)",
                    (i, i, name))
            self.commit()

    def get_sheets(self):
        return self.conn.execute(
            "SELECT * FROM pid_sheets ORDER BY display_order").fetchall()

    def append_sheets(self, physical_pages, sheet_names, revision_id=None):
        max_row = self.conn.execute(
            "SELECT MAX(display_order) FROM pid_sheets").fetchone()[0]
        start_order = (max_row + 1) if max_row is not None else 0
        for i, (phys, name) in enumerate(zip(physical_pages, sheet_names)):
            self.conn.execute(
                "INSERT INTO pid_sheets (display_order,physical_page,sheet_name,revision_id) "
                "VALUES (?,?,?,?)",
                (start_order + i, phys, name, revision_id))
        self.commit()

    def reorder_sheets(self, ordered_ids):
        for disp_order, sheet_id in enumerate(ordered_ids):
            self.conn.execute(
                "UPDATE pid_sheets SET display_order=? WHERE id=?",
                (disp_order, sheet_id))
        self.commit()

    def update_sheet_name(self, id_, name):
        self.conn.execute("UPDATE pid_sheets SET sheet_name=?, drawing_name=? WHERE id=?", (name, name, id_))
        self.commit()

    def get_attendance_details(self):
        """Return {(participant_id, session_id): (attended, note)}."""
        rows = self.conn.execute(
            "SELECT participant_id, session_id, attended, COALESCE(note, '') AS note "
            "FROM participant_attendance").fetchall()
        return {(r['participant_id'], r['session_id']):
                (bool(r['attended']), r['note'] or '') for r in rows}

    def set_attendance_note(self, participant_id, session_id, note):
        """Persist free text attached to one participant and analysis session."""
        self.conn.execute(
            "INSERT INTO participant_attendance (participant_id, session_id, attended, note) "
            "VALUES (?,?,COALESCE((SELECT attended FROM participant_attendance "
            "WHERE participant_id=? AND session_id=?),0),?) "
            "ON CONFLICT(participant_id, session_id) DO UPDATE SET note=excluded.note",
            (participant_id, session_id, participant_id, session_id, str(note or '')))
        self.commit()

    def update_sheet_metadata(self, id_, drawing_number=None, drawing_name=None,
                              drawing_revision=None, drawing_date=None):
        row = self.conn.execute("SELECT * FROM pid_sheets WHERE id=?", (id_,)).fetchone()
        if not row:
            return
        current = dict(row)
        name = drawing_name if drawing_name is not None else (current.get('drawing_name') or current.get('sheet_name') or '')
        self.conn.execute(
            "UPDATE pid_sheets SET drawing_number=?, drawing_name=?, sheet_name=?, "
            "drawing_revision=?, drawing_date=? WHERE id=?",
            (drawing_number if drawing_number is not None else current.get('drawing_number', ''),
             name,
             name,
             drawing_revision if drawing_revision is not None else current.get('drawing_revision', ''),
             drawing_date if drawing_date is not None else current.get('drawing_date', ''),
             id_))
        self.commit()

    def set_sheet_revision(self, id_, revision_id):
        self.conn.execute(
            "UPDATE pid_sheets SET revision_id=? WHERE id=?", (revision_id, id_))
        self.commit()

    def delete_sheets(self, ids):
        for id_ in ids:
            self.conn.execute("DELETE FROM pid_sheets WHERE id=?", (id_,))
        remaining = self.conn.execute(
            "SELECT id FROM pid_sheets ORDER BY display_order").fetchall()
        for disp_order, row in enumerate(remaining):
            self.conn.execute(
                "UPDATE pid_sheets SET display_order=? WHERE id=?",
                (disp_order, row['id']))
        self.commit()

    def delete_objects_on_pages(self, physical_pages):
        """Delete all P&ID placements (markers and node markups) on the given physical pages."""
        for page in physical_pages:
            self.conn.execute("DELETE FROM node_markups WHERE pid_page=?", (page,))
            self.conn.execute("DELETE FROM cause_markers WHERE pid_page=?", (page,))
            self.conn.execute("DELETE FROM consequence_markers WHERE pid_page=?", (page,))
            self.conn.execute("DELETE FROM safeguard_markers WHERE pid_page=?", (page,))
        self.commit()

    def objects_on_pages(self, physical_pages):
        """Return counts of HAZOP objects on the given physical page numbers.
        Returns dict: physical_page -> {markups, causes, consequences, safeguards}."""
        result = {}
        for page in physical_pages:
            markups = self.conn.execute(
                "SELECT COUNT(*) FROM node_markups WHERE pid_page=?", (page,)).fetchone()[0]
            causes = self.conn.execute(
                "SELECT COUNT(*) FROM cause_markers WHERE pid_page=?", (page,)).fetchone()[0]
            consequences = self.conn.execute(
                "SELECT COUNT(*) FROM consequence_markers WHERE pid_page=?", (page,)).fetchone()[0]
            safeguards = self.conn.execute(
                "SELECT COUNT(*) FROM safeguard_markers WHERE pid_page=?", (page,)).fetchone()[0]
            result[page] = {
                'markups': markups,
                'causes': causes,
                'consequences': consequences,
                'safeguards': safeguards,
            }
        return result

    def get_sheet_physical_page(self, display_index):
        row = self.conn.execute(
            "SELECT physical_page FROM pid_sheets ORDER BY display_order "
            "LIMIT 1 OFFSET ?", (display_index,)).fetchone()
        return row['physical_page'] if row else display_index

    def get_display_page_count(self):
        return self.conn.execute("SELECT COUNT(*) FROM pid_sheets").fetchone()[0]

    def clear_sheets(self):
        self.conn.execute("DELETE FROM pid_sheets")
        self.commit()

    def clear_all_pid_data(self):
        """Remove all P&ID revisions, sheets, placements, markups and connectors."""
        for table in (
            "pid_sheets", "pid_revisions",
            "cause_markers", "consequence_markers", "safeguard_markers",
            "node_markups", "node_red_markups",
            "off_page_connector", "pid_connection", "pid_page_rotation",
        ):
            self.conn.execute(f"DELETE FROM {table}")
        self.conn.execute("DELETE FROM pid_config WHERE key='path'")
        self.commit()

    # ── Manuell sidrotation (2026-08-12, see NOTES.md) ──────────────────────

    def get_page_rotation(self, physical_page):
        """Extra clockwise rotation (0/90/180/270) the user chose for this
        physical page, on top of the PDF's own /Rotate. 0 if never set."""
        row = self.conn.execute(
            "SELECT rotation FROM pid_page_rotation WHERE physical_page=?",
            (physical_page,)).fetchone()
        return int(row['rotation']) if row else 0

    def get_all_page_rotations(self):
        """Return {physical_page: rotation} for every page with a non-default
        override — used to repopulate PIDGraphicsView._page_rotation_override
        whenever the PDF is (re)loaded, see PIDPanel._import_pdf/try_reload_pdf."""
        rows = self.conn.execute("SELECT physical_page, rotation FROM pid_page_rotation").fetchall()
        return {int(r['physical_page']): int(r['rotation']) for r in rows}

    def set_page_rotation(self, physical_page, rotation):
        rotation = int(rotation) % 360
        self.conn.execute(
            "INSERT INTO pid_page_rotation (physical_page, rotation) VALUES (?,?) "
            "ON CONFLICT(physical_page) DO UPDATE SET rotation=excluded.rotation",
            (physical_page, rotation))
        self.commit()

    def clear_page_rotations(self):
        """Called alongside clear_sheets() whenever the working PDF is
        entirely replaced (not appended-to) — physical page numbers from the
        old file have nothing to do with the new one, so any rotation
        override left over from it would apply to the wrong page."""
        self.conn.execute("DELETE FROM pid_page_rotation")
        self.commit()

    def remap_page_rotation_positions(self, physical_page, transform_fn, angle_delta_deg=0):
        """Re-anchor every position stored for `physical_page` to the same
        physical point after PIDGraphicsView.set_page_rotation_override()
        changes what that page's PDF-space coordinate system means.

        `transform_fn(x, y) -> (new_x, new_y)` must map a point from the
        OLD rotated PDF-space to the NEW rotated PDF-space (built by the
        caller from the page's derotation_matrix/rotation_matrix before and
        after the change — see PIDPanel._rotate_page). `angle_delta_deg` is
        used only to decide whether axis-aligned rect_w/rect_h (and
        node_red_markups' symbol_w/symbol_h) need to swap: a +-90 degree
        turn swaps width and height, 0/180 do not.

        Covers cause/consequence/safeguard/equipment markers, the node
        outline (nodes.markup_points) and zone drawings (node_markups,
        node_red_markups incl. symbol_rot). Does NOT cover off_page_connector
        or board_annotations — deliberately deferred, see NOTES.md known
        limitations.
        """
        swap_wh = (int(angle_delta_deg) % 180) == 90

        def _swap(w, h):
            if swap_wh and w is not None and h is not None:
                return h, w
            return w, h

        for table, x_col, y_col in (
                ('cause_markers', 'x', 'y'),
                ('consequence_markers', 'x', 'y'),
                ('safeguard_markers', 'x', 'y'),
                ('equipment_markers', 'x', 'y')):
            has_rect = table != 'equipment_markers'
            cols = f"id, {x_col}, {y_col}" + (", rect_w, rect_h" if has_rect else "")
            rows = self.conn.execute(
                f"SELECT {cols} FROM {table} WHERE pid_page=?", (physical_page,)).fetchall()
            for r in rows:
                nx, ny = transform_fn(r[x_col], r[y_col])
                if has_rect:
                    nw, nh = _swap(r['rect_w'], r['rect_h'])
                    self.conn.execute(
                        f"UPDATE {table} SET {x_col}=?,{y_col}=?,rect_w=?,rect_h=? WHERE id=?",
                        (nx, ny, nw, nh, r['id']))
                else:
                    self.conn.execute(
                        f"UPDATE {table} SET {x_col}=?,{y_col}=? WHERE id=?",
                        (nx, ny, r['id']))

        for row in self.conn.execute(
                "SELECT id, markup_points FROM nodes WHERE pid_page=?", (physical_page,)).fetchall():
            try:
                pts = json.loads(row['markup_points'] or '[]')
            except Exception:
                continue
            if not pts:
                continue
            new_pts = [list(transform_fn(p[0], p[1])) for p in pts]
            self.conn.execute("UPDATE nodes SET markup_points=? WHERE id=?",
                              (json.dumps(new_pts), row['id']))

        for row in self.conn.execute(
                "SELECT id, points FROM node_markups WHERE pid_page=?", (physical_page,)).fetchall():
            try:
                pts = json.loads(row['points'] or '[]')
            except Exception:
                continue
            if not pts:
                continue
            new_pts = [list(transform_fn(p[0], p[1])) for p in pts]
            self.conn.execute("UPDATE node_markups SET points=? WHERE id=?",
                              (json.dumps(new_pts), row['id']))

        for row in self.conn.execute(
                "SELECT id, points, symbol_w, symbol_h, symbol_rot FROM node_red_markups "
                "WHERE pid_page=?", (physical_page,)).fetchall():
            try:
                pts = json.loads(row['points'] or '[]')
            except Exception:
                continue
            new_pts = [list(transform_fn(p[0], p[1])) for p in pts] if pts else pts
            nw, nh = _swap(row['symbol_w'], row['symbol_h'])
            new_rot = (float(row['symbol_rot'] or 0) + float(angle_delta_deg)) % 360
            self.conn.execute(
                "UPDATE node_red_markups SET points=?,symbol_w=?,symbol_h=?,symbol_rot=? WHERE id=?",
                (json.dumps(new_pts), nw, nh, new_rot, row['id']))

        self.commit()

    def add_node_with_markup(self, name, points, style, page):
        cur = self.conn.execute(
            "INSERT INTO nodes (name, markup_points, markup_style, pid_page) VALUES (?,?,?,?)",
            (name, json.dumps(points), json.dumps(style), page))
        self.commit()
        return cur.lastrowid

    def cause_markers_for_page(self, page):
        return self.conn.execute(
            "SELECT * FROM cause_markers WHERE pid_page=?", (page,)).fetchall()

    def consequence_markers_for_page(self, page):
        return self.conn.execute(
            "SELECT * FROM consequence_markers WHERE pid_page=?", (page,)).fetchall()

    def safeguard_markers_for_page(self, page):
        return self.conn.execute(
            "SELECT * FROM safeguard_markers WHERE pid_page=?", (page,)).fetchall()

    def cause_markers_for_cause(self, cause_id):
        """Return all markers for a specific cause (page, x, y, comp_type, tag)."""
        return self.conn.execute(
            "SELECT pid_page, x, y, component_type, component_tag "
            "FROM cause_markers WHERE cause_id=?",
            (cause_id,)).fetchall()

    # ── Equipment markers (auto-detected symbols, "🎯 Hitta på P&ID") ──────────
    def add_equipment_marker(self, equipment_id, tag, page, x, y, comp_type,
                             shape_outline='', confidence=0.0, link_method='',
                             detection_confidence=None, tag_reading_confidence=None,
                             tag_assignment_confidence=None, line_assignment_confidence=None,
                             line_number='', medium_code='', medium_code_verified=0,
                             nominal_size='', tag_status='tagged'):
        cur = self.conn.execute(
            "INSERT INTO equipment_markers "
            "(equipment_id,tag,pid_page,x,y,comp_type,shape_outline,confidence,link_method,"
            "detection_confidence,tag_reading_confidence,tag_assignment_confidence,"
            "line_assignment_confidence,line_number,medium_code,medium_code_verified,"
            "nominal_size,tag_status) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (equipment_id, tag, page, x, y, comp_type, shape_outline, confidence, link_method,
             detection_confidence, tag_reading_confidence, tag_assignment_confidence,
             line_assignment_confidence, line_number, medium_code, medium_code_verified,
             nominal_size, tag_status))
        self.commit()
        return cur.lastrowid

    def equipment_markers_for_page(self, page):
        return self.conn.execute(
            "SELECT * FROM equipment_markers WHERE pid_page=?", (page,)).fetchall()

    def update_equipment_marker_link(self, marker_id, equipment_id, tag):
        """Re-point a marker to a different equipment_catalog row —
        used by EquipmentPlacementPopup (2026-08-18, see NOTES.md
        "kombinerad placeringsmeny") when a tag typed/detected AFTER
        placement turns out to already belong to an existing catalog
        row: the marker created against the blank placeholder row is
        re-linked to the existing one instead of leaving a duplicate."""
        self.conn.execute(
            "UPDATE equipment_markers SET equipment_id=?, tag=? WHERE id=?",
            (equipment_id, tag, marker_id))
        self.commit()

    def update_equipment_marker_position(self, marker_id, page, x, y):
        """Move an existing P&ID equipment marker without changing its link."""
        self.conn.execute(
            "UPDATE equipment_markers SET pid_page=?, x=?, y=? WHERE id=?",
            (int(page), float(x), float(y), int(marker_id)))
        self.commit()

    def update_equipment_marker_geometry(self, marker_id, page, x, y,
                                         shape_outline):
        """Persist an equipment marker's centre and optional outline together.

        Rubber-band equipment markers store their visible contour as PDF
        points. Moving only ``x``/``y`` therefore leaves that contour behind;
        this atomic update keeps the label anchor and contour in sync. The
        geometry editor uses the same method when it rescales a marker.
        """
        self.conn.execute(
            "UPDATE equipment_markers SET pid_page=?, x=?, y=?, "
            "shape_outline=? WHERE id=?",
            (int(page), float(x), float(y), str(shape_outline or ''),
             int(marker_id)))
        self.commit()

    def update_equipment_marker_label_position(self, marker_id, label_x, label_y):
        """Persist only the visible tag/counter strip's PDF anchor.

        This is intentionally separate from the equipment geometry update:
        moving a label must never move or resize the detected P&ID object.
        """
        self.conn.execute(
            "UPDATE equipment_markers SET label_x=?, label_y=? WHERE id=?",
            (float(label_x), float(label_y), int(marker_id)))
        self.commit()

    def delete_equipment_marker(self, id_):
        self.conn.execute("DELETE FROM equipment_markers WHERE id=?", (id_,))
        self.commit()

    # ── Queries ───────────────────────────────────────────────────────────────
    def nodes(self):
        return self.conn.execute("SELECT * FROM nodes ORDER BY sort_order, id").fetchall()

    def systems(self):
        """Top-level hierarchy grouping above Nod (2026-08-24, see NOTES.md
        "Ny toppnivå System"). A node's system_id may be NULL — such nodes
        render as ungrouped top-level items in the tree, same as before
        this feature existed (e.g. any project saved before this feature,
        or a node created from a UI path that doesn't set system_id)."""
        return self.conn.execute(
            "SELECT * FROM systems ORDER BY sort_order, id").fetchall()

    def causes(self, node_id):
        return self.conn.execute(
            "SELECT * FROM causes WHERE node_id=? ORDER BY sort_order, id", (node_id,)).fetchall()

    def consequences(self, cause_id):
        return self.conn.execute(
            "SELECT * FROM consequences WHERE cause_id=? ORDER BY sort_order, id", (cause_id,)).fetchall()

    def hazop_hierarchy_reference(self, *, cause_id=None, consequence_id=None):
        """Return the visible HAZOP-tree reference for a cause/consequence.

        The value deliberately follows the *presentation* hierarchy (study,
        optional system, node, guide word, cause and consequence) instead of
        exposing mutable database IDs.  That makes a LOPA reference readable
        in a report and keeps it aligned with the ordering users see after
        reordering items in the HAZOP tree.

        The study is always level ``1``.  A system is included only when the
        node belongs to one; ungrouped legacy nodes therefore have a shorter,
        but still unambiguous, path.  ``None`` is returned when the stored
        HAZOP source no longer exists, preserving historic LOPA rows without
        inventing a misleading reference.
        """
        consequence = None
        if consequence_id is not None:
            consequence = self.get_consequence(consequence_id)
            if not consequence:
                return None
            cause_id = consequence.get('cause_id')
        if cause_id is None:
            return None
        cause = self.get_cause(cause_id)
        if not cause:
            return None
        node = self.get_node(cause.get('node_id'))
        if not node:
            return None

        reference = ['1']  # one HAZOP study in the current project document
        node_id = node['id']
        system_id = node.get('system_id')
        if system_id is not None:
            systems = [dict(row) for row in self.systems()]
            system_index = next(
                (index for index, system in enumerate(systems, start=1)
                 if system['id'] == system_id), None)
            if system_index is None:
                return None
            nodes_in_system = [dict(row) for row in self.nodes()
                               if row['system_id'] == system_id]
            node_index = next(
                (index for index, item in enumerate(nodes_in_system, start=1)
                 if item['id'] == node_id), None)
            if node_index is None:
                return None
            reference.extend((str(system_index), str(node_index)))
        else:
            ungrouped_nodes = [dict(row) for row in self.nodes()
                               if row['system_id'] is None]
            node_index = next(
                (index for index, item in enumerate(ungrouped_nodes, start=1)
                 if item['id'] == node_id), None)
            if node_index is None:
                return None
            reference.append(str(node_index))

        deviation_id = cause.get('deviation_id')
        deviations = [dict(row) for row in self.deviations(node_id)]
        groups = []
        for deviation in deviations:
            group = next((items for key, items in groups
                          if key == deviation['description']), None)
            if group is None:
                group = []
                groups.append((deviation['description'], group))
            group.append(deviation)

        cause_group = None
        deviation_index = None
        for index, (_description, members) in enumerate(groups, start=1):
            if any(member['id'] == deviation_id for member in members):
                deviation_index = index
                cause_group = members
                break
        if cause_group is None:
            # Old files can contain a cause without a retained deviation.
            # Keep its LOPA source readable while still avoiding database IDs.
            deviation_index = len(groups) + 1
            cause_group = []
        reference.append(str(deviation_index))

        ordered_causes = []
        if cause_group:
            for deviation in cause_group:
                ordered_causes.extend(
                    dict(row) for row in self.causes_for_deviation(deviation['id']))
        else:
            ordered_causes = [dict(row) for row in self.causes(node_id)
                              if row.get('deviation_id') is None]
        cause_index = next(
            (index for index, item in enumerate(ordered_causes, start=1)
             if item['id'] == cause['id']), None)
        if cause_index is None:
            return None
        reference.append(str(cause_index))

        if consequence is not None:
            consequences = [dict(row) for row in self.consequences(cause['id'])]
            consequence_index = next(
                (index for index, item in enumerate(consequences, start=1)
                 if item['id'] == consequence['id']), None)
            if consequence_index is None:
                return None
            reference.append(str(consequence_index))
        return '.'.join(reference)

    def safeguards_for_cause(self, cause_id):
        """Return all safeguards attached to any consequence of cause_id."""
        return self.conn.execute(
            "SELECT s.id FROM safeguards s "
            "JOIN consequences c ON c.id=s.consequence_id "
            "WHERE c.cause_id=?", (cause_id,)).fetchall()

    def safeguards(self, consequence_id):
        return self.conn.execute(
            "SELECT * FROM safeguards WHERE consequence_id=? ORDER BY sort_order, id", (consequence_id,)).fetchall()

    def recommendations_for_consequence(self, consequence_id):
        """Recommendations linked to one consequence, via the
        consequence_recommendations join table (2026-08-25, see
        NOTES.md "Rekommendationshantering" — replaces the old
        actions(consequence_id), same call shape/row shape so
        _add_row()/_recommendation_summary() didn't need to change
        beyond the method name)."""
        return self.conn.execute(
            "SELECT r.* FROM recommendations r "
            "JOIN consequence_recommendations cr ON cr.recommendation_id = r.id "
            "WHERE cr.consequence_id=? ORDER BY r.display_number, r.id", (consequence_id,)).fetchall()

    def recommendations_for_consequences(self, consequence_ids):
        """Bulk version of recommendations_for_consequence() — same
        {consequence_id: [rows]} shape _fetch_grouped() returns for a
        plain FK column, but recommendations are reached through the
        consequence_recommendations join table instead (a real N:M
        relation, not a FK on the consumer), so it can't reuse
        _fetch_grouped() directly. Same chunking to stay under SQLite's
        per-statement variable limit."""
        ids = list(consequence_ids)
        result = {i: [] for i in ids}
        if not ids:
            return result
        CHUNK = 500
        for start in range(0, len(ids), CHUNK):
            chunk = ids[start:start + CHUNK]
            placeholders = ','.join('?' * len(chunk))
            rows = self.conn.execute(
                f"SELECT cr.consequence_id AS _cons_id, r.* "
                f"FROM consequence_recommendations cr "
                f"JOIN recommendations r ON r.id = cr.recommendation_id "
                f"WHERE cr.consequence_id IN ({placeholders}) "
                f"ORDER BY cr.consequence_id, r.display_number, r.id", chunk).fetchall()
            for row in rows:
                result[row['_cons_id']].append(row)
        return result

    def all_recommendations(self):
        """The whole study-wide recommendation catalog, for the
        RecommendationEditorDialog's reuse/search list."""
        return self.conn.execute(
            "SELECT * FROM recommendations ORDER BY display_number, id").fetchall()

    def consequences_for_recommendation(self, recommendation_id):
        """Every consequence_id a recommendation currently links to, via
        the consequence_recommendations join table — the reverse
        direction of recommendations_for_consequence() (2026-08-26, added
        for the Rekommendationer overview page: RecommendationsPanel uses
        this to compute the studie.nod.avvikelse.orsak.konsekvens
        reference(s) a reused recommendation currently resolves to)."""
        rows = self.conn.execute(
            "SELECT consequence_id FROM consequence_recommendations "
            "WHERE recommendation_id=? ORDER BY consequence_id",
            (recommendation_id,)).fetchall()
        return [r['consequence_id'] for r in rows]

    @staticmethod
    def _clean_recommendation_text(value):
        """Remove accidental HTML document markup imported into a recommendation."""
        text = '' if value is None else str(value)
        if '<!doctype' in text.casefold() or re.search(r'<\s*(html|body|p|div)\b', text, re.I):
            text = re.sub(r'<!doctype[^>]*>', '', text, flags=re.I)
            text = re.sub(r'<[^>]+>', '', text)
            text = html.unescape(text)
        return text.strip()

    def recommendation_consequence_count(self, recommendation_id):
        """How many consequences currently link to this recommendation —
        the basis for the "used by multiple causes, update all?" prompt
        when editing a shared recommendation's text."""
        return self.conn.execute(
            "SELECT COUNT(*) FROM consequence_recommendations WHERE recommendation_id=?",
            (recommendation_id,)).fetchone()[0]

    def add_recommendation(self, description='', responsible='', due_date='', status='Öppen'):
        """Create a new, unlinked catalog row with the next display number."""
        cleaned = self._clean_recommendation_text(description)
        if cleaned:
            key = ' '.join(cleaned.split()).casefold()
            for row in self.all_recommendations():
                existing = self._clean_recommendation_text(row['description'] or '')
                if existing and ' '.join(existing.split()).casefold() == key:
                    return row['id']
        next_display = self.conn.execute(
            "SELECT COALESCE(MAX(display_number), 0) + 1 FROM recommendations").fetchone()[0]
        cur = self.conn.execute(
            "INSERT INTO recommendations "
            "(display_number,description,responsible,due_date,status) VALUES (?,?,?,?,?)",
            (next_display, cleaned, responsible, due_date, status))
        self.commit()
        return cur.lastrowid

    def add_recommendation_to_consequence(self, consequence_id, description='',
                                          responsible='', due_date='', status='Öppen'):
        """Create a new recommendation and link it to consequence_id in
        one call — same call shape the old add_action(consequence_id)
        had, and what the picker popup's "Skapa ny rekommendation"
        button calls."""
        rec_id = self.add_recommendation(description, responsible, due_date, status)
        self.link_recommendation_to_consequence(rec_id, consequence_id)
        return rec_id

    def update_recommendation(self, id_, description=None, responsible=None,
                              due_date=None, status=None):
        """Partial update — None means 'don't touch', same convention as
        update_cause(). This is the "Ja, uppdatera alla" path: same id,
        new text visible to every consequence already linked to it."""
        sets, vals = [], []
        if description is not None:
            sets.append("description=?"); vals.append(
                self._clean_recommendation_text(normalize_arrows(description)))
        if responsible is not None:
            sets.append("responsible=?"); vals.append(responsible)
        if due_date is not None:
            sets.append("due_date=?"); vals.append(due_date)
        if status is not None:
            sets.append("status=?"); vals.append(status)
        if not sets:
            return
        vals.append(id_)
        self.conn.execute(f"UPDATE recommendations SET {','.join(sets)} WHERE id=?", vals)
        self.commit()

    def delete_recommendation(self, id_):
        """Hard delete from the catalog (cascades any remaining links).
        Not called by the picker UI itself — unchecking a recommendation
        only unlinks it (see unlink_recommendation_from_consequence) so
        reusable text isn't lost — kept for completeness/tests."""
        row = self.conn.execute(
            "SELECT display_number FROM recommendations WHERE id=?", (id_,)).fetchone()
        if row is None:
            return
        self.conn.execute("DELETE FROM recommendations WHERE id=?", (id_,))
        self.conn.execute(
            "UPDATE recommendations SET display_number=display_number-1 "
            "WHERE display_number>?", (row['display_number'],))
        self.commit()

    def link_recommendation_to_consequence(self, recommendation_id, consequence_id):
        """Idempotent — checking an already-linked recommendation again
        is a no-op, never creates a duplicate link."""
        self.conn.execute(
            "INSERT OR IGNORE INTO consequence_recommendations "
            "(consequence_id,recommendation_id) VALUES (?,?)",
            (consequence_id, recommendation_id))
        self.commit()

    def unlink_recommendation_from_consequence(self, recommendation_id, consequence_id):
        """Unchecking a recommendation in the picker — removes only this
        link. The catalog row itself is left alone even if this was its
        last link, so the text stays available for reuse later."""
        self.conn.execute(
            "DELETE FROM consequence_recommendations "
            "WHERE consequence_id=? AND recommendation_id=?",
            (consequence_id, recommendation_id))
        self.commit()

    def get_node(self, id_):
        row = self.conn.execute("SELECT * FROM nodes WHERE id=?", (id_,)).fetchone()
        return dict(row) if row else None

    def get_cause(self, id_):
        row = self.conn.execute("SELECT * FROM causes WHERE id=?", (id_,)).fetchone()
        return dict(row) if row else None

    def get_consequence(self, id_):
        row = self.conn.execute("SELECT * FROM consequences WHERE id=?", (id_,)).fetchone()
        return dict(row) if row else None

    def get_safeguard(self, id_):
        row = self.conn.execute("SELECT * FROM safeguards WHERE id=?", (id_,)).fetchone()
        return dict(row) if row else None

    def get_recommendation(self, id_):
        row = self.conn.execute("SELECT * FROM recommendations WHERE id=?", (id_,)).fetchone()
        return dict(row) if row else None

    def cause_base_frequency_per_year(self, cause):
        """Return frequency in events/year from standard cause or base_frequency, or None."""
        if cause is None:
            return None
        d = dict(cause)
        if d.get('frequency_cleared'):
            return None
        std_id = d.get('standard_cause_id')
        if std_id:
            sc = self.get_standard_cause(std_id)
            if sc and sc.get('frequency') is not None:
                return sc['frequency']
        bf = d.get('base_frequency')
        return bf if bf is not None else None

    def cause_f_level(self, cause, default=3):
        """Return F-level (-1..5): standard_cause/base_frequency first, else manual likelihood."""
        base_freq_per_year = self.cause_base_frequency_per_year(cause)
        if base_freq_per_year is not None:
            return freq_to_f_level(base_freq_per_year)
        if cause is None:
            return default
        like = dict(cause).get('likelihood')
        return like if like is not None else default

    # Keep old names as aliases for backward compatibility
    cause_base_frequency = cause_base_frequency_per_year
    cause_frequency_level = cause_f_level

    # ── Add ───────────────────────────────────────────────────────────────────
    def add_system(self, name='Nytt system'):
        max_ord = (self.conn.execute(
            "SELECT COALESCE(MAX(sort_order),-1) FROM systems").fetchone()[0])
        cur = self.conn.execute(
            "INSERT INTO systems (name, sort_order) VALUES (?,?)", (name, max_ord + 1))
        self.commit()
        return cur.lastrowid

    def rename_system(self, id_, name):
        self.conn.execute("UPDATE systems SET name=? WHERE id=?", (name, id_))
        self.commit()

    def delete_system(self, id_):
        # Same "reassign, don't cascade" convention as delete_node_type():
        # a system is a pure organizational grouping — deleting it must
        # never take its nodes (and everything under them) down with it.
        self.conn.execute("UPDATE nodes SET system_id=NULL WHERE system_id=?", (id_,))
        self.conn.execute("DELETE FROM systems WHERE id=?", (id_,))
        self.commit()

    def reorder_systems(self, ordered_ids):
        for i, id_ in enumerate(ordered_ids):
            self.conn.execute("UPDATE systems SET sort_order=? WHERE id=?", (i, id_))
        self.commit()

    def set_node_system(self, node_id, system_id):
        self.conn.execute(
            "UPDATE nodes SET system_id=? WHERE id=?", (system_id, node_id))
        self.commit()

    def add_node(self, system_id=None):
        cur = self.conn.execute(
            "INSERT INTO nodes (name, system_id) VALUES ('Ny nod', ?)", (system_id,))
        node_id = cur.lastrowid
        std = [r[0] for r in self.conn.execute(
            "SELECT description FROM standard_deviations WHERE active=1 ORDER BY sort_order").fetchall()]
        for dev_type in (std or DEVIATION_TYPES):
            self.conn.execute(
                "INSERT INTO deviations (node_id, description) VALUES (?,?)",
                (node_id, dev_type))
        self.commit()
        return node_id

    def deviations(self, node_id):
        return self.conn.execute(
            "SELECT * FROM deviations WHERE node_id=? ORDER BY sort_order, id", (node_id,)).fetchall()

    def deviations_for_equipment(self, equipment_id):
        return self.conn.execute(
            "SELECT * FROM deviations WHERE equipment_id=? ORDER BY sort_order, id", (equipment_id,)).fetchall()

    def get_deviation(self, id_):
        row = self.conn.execute("SELECT * FROM deviations WHERE id=?", (id_,)).fetchone()
        return dict(row) if row else None

    def causes_for_node_all(self, node_id):
        """All causes for a node across all deviations."""
        return self.conn.execute(
            "SELECT c.* FROM causes c "
            "JOIN deviations d ON d.id=c.deviation_id "
            "WHERE d.node_id=?", (node_id,)).fetchall()

    def causes_for_equipment(self, equipment_id):
        """Every cause that 'mentions' this equipment anywhere in its own
        chain — its deviation is directly tied to the equipment
        (deviations.equipment_id), or the cause/one of its consequences/
        one of its safeguards was tagged to it (comp_tag+comp_type, e.g.
        via drag-and-drop) — used to filter the HAZOP scenario table to
        just this object's rows when its P&ID marker is clicked (2026-08-12,
        see NOTES.md: 'de orsaker som visas i hazop scenario är de där
        objektet finns med'). Mirrors the tag-matching
        equipment_consequence_count/equipment_safeguard_count already use
        for their own occurrence counters."""
        equip = self.get_equipment_by_id(equipment_id)
        tag = (equip or {}).get('tag', '') or ''
        comp_type = (equip or {}).get('equipment_type', '') or ''
        if not tag:
            rows = self.conn.execute(
                "SELECT DISTINCT c.* FROM causes c "
                "WHERE c.equipment_id=? ORDER BY c.id", (equipment_id,)).fetchall()
        else:
            rows = self.conn.execute(
                "SELECT DISTINCT c.* FROM causes c "
                "LEFT JOIN consequences k ON k.cause_id=c.id "
                "LEFT JOIN safeguards s ON s.consequence_id=k.id "
                "LEFT JOIN consequence_recommendations cr ON cr.consequence_id=k.id "
                "LEFT JOIN recommendations r ON r.id=cr.recommendation_id "
                "WHERE c.comp_tag=? "
                "   OR k.description LIKE '%' || ? || '%' "
                "   OR s.description LIKE '%' || ? || '%' "
                "   OR (r.description LIKE '%' || ? || '%') "
                "ORDER BY c.id",
                (tag, tag, tag, tag)).fetchall()
        found = {row['id'] for row in rows}
        # A grouped cause can mention this object only in its later rows;
        # those rows are not represented by the legacy comp_tag columns.
        grouped = self.conn.execute(
            "SELECT c.* FROM causes c JOIN deviations d ON d.id=c.deviation_id "
            "WHERE c.group_equipment_ids IS NOT NULL "
            "ORDER BY c.id").fetchall()
        for row in grouped:
            if row['id'] not in found:
                group_ids = self._group_equipment_ids_for_cause(row)
                if equipment_id in group_ids and self._equipment_tag_matches_cause(
                        dict(row), tag):
                    rows.append(row)
        return sorted(rows, key=lambda row: row['id'])

    def consequences_for_node(self, node_id):
        """All consequences for a node across all causes."""
        return self.conn.execute(
            "SELECT k.* FROM consequences k "
            "JOIN causes c ON c.id=k.cause_id "
            "JOIN deviations d ON d.id=c.deviation_id "
            "WHERE d.node_id=?", (node_id,)).fetchall()

    def causes_for_deviation(self, deviation_id):
        return self.conn.execute(
            "SELECT * FROM causes WHERE deviation_id=? ORDER BY sort_order, id", (deviation_id,)).fetchall()

    def _fetch_grouped(self, table, fk_column, ids):
        """Bulk-fetch every row from `table` whose `fk_column` is in `ids`
        in a small number of batched queries (chunked to stay under
        SQLite's per-statement variable limit), grouped into a
        {fk_value: [rows]} dict — same row shape/order as calling the
        single-id equivalent once per id (e.g. causes_for_deviation),
        without the per-id query cost. Every id in `ids` is present in
        the result, with an empty list if it has no matching rows.
        `table`/`fk_column` are always fixed literals from this class's
        own wrapper methods below, never external input — safe to
        interpolate directly. Added 2026-08-24 (see NOTES.md) to fix
        TreePanel.refresh()'s N+1 query pattern (one query per node, then
        per deviation, per cause, per consequence, on every tree rebuild)."""
        ids = list(ids)
        result = {i: [] for i in ids}
        if not ids:
            return result
        CHUNK = 500
        for start in range(0, len(ids), CHUNK):
            chunk = ids[start:start + CHUNK]
            placeholders = ','.join('?' * len(chunk))
            order_column = ('sort_order, ' if table in {
                'nodes', 'deviations', 'causes', 'consequences', 'safeguards'
            } else '')
            rows = self.conn.execute(
                f"SELECT * FROM {table} WHERE {fk_column} IN ({placeholders}) "
                f"ORDER BY {fk_column}, {order_column}id", chunk).fetchall()
            for row in rows:
                result[row[fk_column]].append(row)
        return result

    def deviations_for_nodes(self, node_ids):
        """Bulk version of deviations() — see _fetch_grouped."""
        return self._fetch_grouped('deviations', 'node_id', node_ids)

    def causes_for_deviations(self, deviation_ids):
        """Bulk version of causes_for_deviation() — see _fetch_grouped."""
        return self._fetch_grouped('causes', 'deviation_id', deviation_ids)

    def consequences_for_causes(self, cause_ids):
        """Bulk version of consequences() — see _fetch_grouped."""
        return self._fetch_grouped('consequences', 'cause_id', cause_ids)

    def safeguards_for_consequences(self, consequence_ids):
        """Bulk version of safeguards() — see _fetch_grouped."""
        return self._fetch_grouped('safeguards', 'consequence_id', consequence_ids)

    def _equipment_id_for_tag(self, tag, cache):
        """Resolve a free-text tag string to an equipment_catalog id via
        get_equipment_by_tag(), memoized in `cache` (a plain dict passed
        in by the caller) so a tag repeated across many rows in one
        equipment_link_types_in_scope() call only issues one query."""
        tag = (tag or '').strip()
        if not tag:
            return None
        if tag not in cache:
            eq = self.get_equipment_by_tag(tag)
            cache[tag] = eq['id'] if eq else None
        return cache[tag]

    def _tags_for_row(self, row):
        """Return only tags currently present in the cell text.

        ``tagged_refs`` is retained for migration/undo compatibility, but it
        is history rather than an active relationship.  A removed tag must
        therefore stop highlighting and stop contributing to tree scope.
        """
        description = str(row.get('description') or '')
        tags = set(parse_tag_refs(row.get('tagged_refs') or '')) & {
            tag for tag in parse_tag_refs(row.get('tagged_refs') or '')
            if re.search(rf'(?<![A-Za-z0-9]){re.escape(tag)}(?![A-Za-z0-9])',
                         description, re.IGNORECASE)
        }
        comp_tag = str(row.get('comp_tag') or '').strip()
        if comp_tag and re.search(
                rf'(?<![A-Za-z0-9]){re.escape(comp_tag)}(?![A-Za-z0-9])',
                description, re.IGNORECASE):
            tags.add(comp_tag)
        return tags

    def group_equipment_ids_for_cause(self, cause):
        """Return a grouped cause's object ids in its visual row order.

        ``group_equipment_ids`` is the canonical representation for a group.
        The first two object ids are also mirrored in the old cause columns,
        so partially migrated rows can otherwise appear as a one-row cause in
        one view and a two-row cause in another.  Repair that incomplete
        representation in memory by adding the legacy links *only when* the
        JSON list is shorter than a group.  This keeps an intentional
        one-object cause one-object, while making a damaged two-plus-object
        group behave consistently in the Scenario table, tree and P&ID scope.
        """
        if not cause:
            return []
        if not isinstance(cause, dict):
            cause = dict(cause)

        raw = cause.get('group_equipment_ids') or ''
        if isinstance(raw, str):
            try:
                raw = json.loads(raw) if raw else []
            except (TypeError, ValueError, json.JSONDecodeError):
                raw = []
        raw_ids = list(raw) if isinstance(raw, (list, tuple)) else []
        legacy_ids = [cause.get('equipment_id'),
                      cause.get('secondary_equipment_id')]
        values = list(raw_ids)
        if len(values) < 2:
            values.extend(legacy_ids)

        result = []
        for value in values:
            try:
                value = int(value)
            except (TypeError, ValueError):
                continue
            if value not in result:
                result.append(value)
        return result[:20]

    def group_cause_description_lines(self, cause, equipment_ids=None):
        """Return exactly one display/edit line per grouped-cause member.

        A group is a single cause with multiple independently editable
        object-event rows.  Older edits could leave the description as one
        arrow sentence, or with fewer rows than linked objects.  The old
        views then disagreed about which row existed; the secondary row could
        not be edited at all.  This method is deliberately pure: callers can
        render a repaired representation immediately and persist it only when
        the user changes the group.

        Existing text is never discarded.  If tags identify the old compact
        sentence, it is split at the later member tags.  Otherwise the text
        remains on the first member and the missing members receive their
        bare tag as a stable, editable anchor.
        """
        if not cause:
            return []
        if not isinstance(cause, dict):
            cause = dict(cause)
        ids = list(equipment_ids if equipment_ids is not None
                   else self.group_equipment_ids_for_cause(cause))
        if len(ids) < 2:
            return str(cause.get('description') or '').splitlines()

        fallback_tags = [part.strip() for part in re.split(
            r'\s+(?:&|OR|<>|->|\+)\s+',
            str(cause.get('comp_tag') or ''), flags=re.IGNORECASE)
            if part.strip()]
        tags = []
        for index, equipment_id in enumerate(ids):
            equipment = self.get_equipment_by_id(equipment_id)
            tag = str((equipment or {}).get('tag') or '').strip()
            tag = tag or (fallback_tags[index]
                          if index < len(fallback_tags) else '')
            tags.append(tag or f'Objekt {index + 1}')

        raw_text = str(cause.get('description') or '')
        lines = raw_text.splitlines()
        if len(lines) <= 1:
            compact = lines[0].strip() if lines else ''
            if compact in ('', 'Ny orsak'):
                lines = []
            else:
                # Split only at a recognised later member tag.  An arrow in
                # ordinary prose is not enough evidence to divide a user's
                # text between objects.
                starts = []
                search_from = 0
                for index, tag in enumerate(tags):
                    # A substring match lets e.g. ``V-1`` claim the start
                    # of ``V-10``.  Group rows are keyed by exact catalogue
                    # objects, so keep the same whole-token rule used by
                    # every other tag/reference path in the application.
                    match = re.search(
                        r'(?<![A-Za-z0-9])' + re.escape(tag) +
                        r'(?![A-Za-z0-9])', compact[search_from:],
                        re.IGNORECASE)
                    if match is None:
                        break
                    position = search_from + match.start()
                    starts.append((position, index))
                    search_from = position + len(tag)
                if len(starts) >= 2:
                    split_lines = []
                    for position_index, (start, _member_index) in enumerate(starts):
                        end = (starts[position_index + 1][0]
                               if position_index + 1 < len(starts)
                               else len(compact))
                        part = compact[start:end].strip(' ,:;\u2192->=')
                        split_lines.append(part)
                    # Text before the first tag is still user text.  Keep it
                    # attached to the first row rather than silently losing
                    # it when repairing a malformed compact group.
                    before = compact[:starts[0][0]].strip(' ,:;\u2192->=')
                    if before:
                        split_lines[0] = f'{before} {split_lines[0]}'.strip()
                    lines = split_lines
                else:
                    lines = [compact]

        if len(lines) > len(tags):
            # A group row maps to one physical table/tree line.  Keep all
            # excess user text on the final member rather than letting an
            # untagged extra line steal another object's editor hit area.
            lines = (lines[:len(tags) - 1] +
                     [' '.join(part.strip() for part in lines[len(tags) - 1:]
                               if part.strip())])
        while len(lines) < len(tags):
            lines.append('')

        # A former group-row reorder could persist the previous object's
        # tag in front of the new owner's text, for example::
        #
        #     FV-1 FI-1 felar lÃ¥gt
        #     FI-1 FV-1 Ã¶ppnar fullt
        #
        # The old normalizer only removed the *expected* tag when it was the
        # first word. It consequently prepended it again and made the stale
        # tag look like ordinary free text on every later repaint.  Consume a
        # run of known group tags at the start of a line instead. If that run
        # contains the line's actual owner, the text after the anchors is the
        # one piece of free text that belongs to that owner. References later
        # in a sentence are deliberately left intact.
        tag_patterns = [
            (tag, re.compile(r'^' + re.escape(tag) + r'(?![A-Za-z0-9])',
                             re.IGNORECASE))
            for tag in sorted(set(tags), key=len, reverse=True)
            if tag
        ]

        def leading_group_tags(value):
            remainder = str(value or '').strip()
            found = []
            # No valid group can have more leading anchors than its members;
            # the cap also protects malformed old text from a needless loop.
            for _ in range(len(tags)):
                match_tag = next((known for known, pattern in tag_patterns
                                  if pattern.match(remainder)), None)
                if not match_tag:
                    break
                found.append(match_tag)
                remainder = remainder[len(match_tag):].lstrip(' ,:;\u2192->=')
            return found, remainder

        normalised = []
        for tag, value in zip(tags, lines):
            value = str(value or '').strip()
            leading_tags, remainder = leading_group_tags(value)
            if any(known.casefold() == tag.casefold()
                   for known in leading_tags):
                value = remainder
            normalised.append(f'{tag} {value}'.strip())
        return normalised

    def _group_equipment_ids_for_cause(self, cause):
        """Compatibility wrapper for older database-internal callers."""
        return self.group_equipment_ids_for_cause(cause)

    def equipment_link_types_in_scope(self, type_, id_):
        """Every equipment_catalog id "in scope" of the given HAZOP tree
        selection, and via which link type(s) it was found there —
        {equipment_id: {'deviation'|'cause'|'consequence'|'safeguard', ...}}.

        Backs the P&ID's tree-context highlight (2026-08-27, see
        NOTES.md "Dynamisk färgmarkering av objekt på P&ID"). ONE
        recursive rule for every selectable tree level — scope(System) =
        union of scope(its Nodes), scope(Node) = union of scope(its
        Deviations), scope(Deviation) = its own direct equipment_id (if
        any) + union of scope(its Causes), scope(Cause) = its own
        equipment (equipment_id, else comp_tag) + union of scope(its
        Consequences), scope(Consequence) = its own tags (comp_tag +
        tagged_refs) + union of scope(its Safeguards), scope(Safeguard)
        = its own tags only. Selecting a Cause therefore highlights that
        cause's object AND everything tagged on its own consequences/
        safeguards, but a Consequence's own selection does NOT pull in
        its parent cause's object — the scope only ever flows downward.

        Implemented as a batched top-down traversal (same shape
        TreePanel.refresh() already uses to build the whole tree in a
        handful of queries — deviations_for_nodes/causes_for_deviations/
        consequences_for_causes/safeguards_for_consequences, all
        _fetch_grouped-based) rather than naive per-row recursion, so the
        query count stays bounded regardless of subtree size. Always
        walks causes via their deviation (never the redundant, ok-to-
        drift causes.node_id-direct / causes_for_node_all-via-deviations
        pair — see NOTES.md), matching exactly what the tree itself
        displays."""
        result: dict = {}
        if id_ is None:
            return result
        tag_cache: dict = {}

        def add_link(equipment_id, link_type):
            if equipment_id is not None:
                result.setdefault(equipment_id, set()).add(link_type)

        def collect_deviation(dev):
            eq_id = dev.get('equipment_id')
            eq = self.get_equipment_by_id(eq_id) if eq_id is not None else None
            if eq and self._equipment_tag_matches_cause(dev, eq.get('tag')):
                add_link(eq_id, 'deviation')

        def collect_cause(cause):
            group_ids = self._group_equipment_ids_for_cause(cause)
            if group_ids:
                for eq_id in group_ids:
                    eq = self.get_equipment_by_id(eq_id)
                    if eq and self._equipment_tag_matches_cause(
                            cause, eq.get('tag')):
                        add_link(eq_id, 'cause')
            else:
                eq_id = self._equipment_id_for_tag(cause.get('comp_tag'), tag_cache)
                if eq_id and self._equipment_tag_matches_cause(
                        cause, cause.get('comp_tag')):
                    add_link(eq_id, 'cause')

        def collect_consequence_or_safeguard(row, link_type):
            if not isinstance(row, dict):
                row = dict(row)
            for tag in self._tags_for_row(row):
                add_link(self._equipment_id_for_tag(tag, tag_cache), link_type)

        def collect_recommendation(row):
            if not isinstance(row, dict):
                row = dict(row)
            # Recommendations are catalogue text, not tag-bearing HAZOP
            # rows: they have no comp_tag/tagged_refs columns.  Match the
            # visible recommendation text against the equipment catalogue,
            # exactly like equipment_recommendation_count does.
            for equipment in self.conn.execute(
                    "SELECT id, tag FROM equipment_catalog "
                    "WHERE tag IS NOT NULL AND tag!=''").fetchall():
                if self._equipment_tag_in_description(row, equipment['tag']):
                    add_link(equipment['id'], 'recommendation')

        dev_rows = cause_rows = cons_rows = sg_rows = None

        if type_ == SYSTEM_T:
            node_ids = [n['id'] for n in self.nodes() if n['system_id'] == id_]
            devs_by_node = self.deviations_for_nodes(node_ids)
            dev_rows = [dict(d) for devs in devs_by_node.values() for d in devs]
        elif type_ == NODE_T:
            devs_by_node = self.deviations_for_nodes([id_])
            dev_rows = [dict(d) for d in devs_by_node.get(id_, [])]
        elif type_ == DEV_T:
            dev = self.get_deviation(id_)
            if dev:
                # TreePanel renders deviations with the same guide-word
                # text in one shared row.  Selecting that row must therefore
                # include equipment assigned to every matching deviation in
                # the same node, including an object-specific sibling row.
                dev_rows = [dict(d) for d in self.deviations(dev['node_id'])
                            if d['description'] == dev['description']]
            else:
                dev_rows = []
        elif type_ == CAUSE_T:
            cause = self.get_cause(id_)
            cause_rows = [cause] if cause else []
        elif type_ == CONS_T:
            cons = self.get_consequence(id_)
            cons_rows = [cons] if cons else []
        elif type_ == SG_T:
            sg = self.get_safeguard(id_)
            sg_rows = [sg] if sg else []
        else:
            return result

        if dev_rows is not None:
            for dev in dev_rows:
                collect_deviation(dev)
            causes_by_dev = self.causes_for_deviations([d['id'] for d in dev_rows])
            cause_rows = [dict(c) for causes in causes_by_dev.values() for c in causes]

        if cause_rows is not None:
            for cause in cause_rows:
                collect_cause(cause)
            cons_by_cause = self.consequences_for_causes([c['id'] for c in cause_rows])
            cons_rows = [dict(k) for conss in cons_by_cause.values() for k in conss]

        if cons_rows is not None:
            for cons in cons_rows:
                collect_consequence_or_safeguard(cons, 'consequence')
            sgs_by_cons = self.safeguards_for_consequences([k['id'] for k in cons_rows])
            sg_rows = [dict(s) for sgs in sgs_by_cons.values() for s in sgs]
            recommendations_by_cons = self.recommendations_for_consequences(
                [k['id'] for k in cons_rows])
            for recommendations in recommendations_by_cons.values():
                for recommendation in recommendations:
                    collect_recommendation(recommendation)

        if sg_rows is not None:
            for sg in sg_rows:
                collect_consequence_or_safeguard(sg, 'safeguard')

        return result

    def causes_for_node_excluding_deviation(self, node_id, deviation_id):
        """Return causes for the node that belong to OTHER deviations (for reuse dialog)."""
        return self.conn.execute(
            "SELECT c.id, c.description, c.comp_type, c.comp_tag, "
            "d.description AS deviation_name, d.id AS deviation_id "
            "FROM causes c "
            "JOIN deviations d ON c.deviation_id = d.id "
            "WHERE d.node_id=? AND d.id!=? "
            "ORDER BY d.id, c.id",
            (node_id, deviation_id)).fetchall()

    def add_deviation(self, node_id, description="Övrigt", equipment_id=None):
        cur = self.conn.execute(
            "INSERT OR IGNORE INTO deviations (node_id, description, equipment_id) VALUES (?,?,?)",
            (node_id, description, equipment_id))
        self.commit()
        if cur.rowcount == 0:
            row = self.conn.execute(
                "SELECT id FROM deviations WHERE node_id=? AND description=? "
                "AND equipment_id IS ? ORDER BY id LIMIT 1",
                (node_id, description, equipment_id)).fetchone()
            return row['id'] if row else None
        return cur.lastrowid

    def update_deviation(self, id_, description):
        self.conn.execute("UPDATE deviations SET description=? WHERE id=?", (description, id_))
        self.commit()

    def delete_deviation(self, id_):
        for cause in self.causes_for_deviation(id_):
            self.delete_cause(cause['id'])
        self.conn.execute("DELETE FROM deviations WHERE id=?", (id_,))
        self.commit()

    def get_or_create_deviation(self, node_id, description="Övrigt", equipment_id=None):
        row = self.conn.execute(
            "SELECT id FROM deviations WHERE node_id=? AND description=? AND equipment_id IS ? "
            "ORDER BY id LIMIT 1",
            (node_id, description, equipment_id)).fetchone()
        return row[0] if row else self.add_deviation(node_id, description, equipment_id)

    # ── Standard deviation / cause template library ───────────────────────────
    def standard_deviations(self):
        return self.conn.execute(
            "SELECT * FROM standard_deviations WHERE active=1 ORDER BY sort_order, id").fetchall()

    def add_standard_deviation(self, description, node_type_id=None):
        max_ord = self.conn.execute(
            "SELECT COALESCE(MAX(sort_order),0) FROM standard_deviations").fetchone()[0]
        cur = self.conn.execute(
            "INSERT INTO standard_deviations (description, sort_order, node_type_id) "
            "VALUES (?,?,?)",
            (description, max_ord + 1, node_type_id))
        self.commit()
        return cur.lastrowid

    def update_standard_deviation(self, id_, description):
        self.conn.execute(
            "UPDATE standard_deviations SET description=? WHERE id=?", (description, id_))
        self.commit()

    def delete_standard_deviation(self, id_):
        self.conn.execute("DELETE FROM standard_deviations WHERE id=?", (id_,))
        self.commit()

    def reorder_standard_deviations(self, ordered_ids):
        for i, id_ in enumerate(ordered_ids):
            self.conn.execute(
                "UPDATE standard_deviations SET sort_order=? WHERE id=?", (i, id_))
        self.commit()

    def node_types(self):
        """List of node types for Avvikelser & Orsaker's leftmost column
        (2026-08-17). Lazily seeds a single 'Processnod' row the first
        time this is called on a database that has none yet, rather than
        a migration-time seed — simpler, and idempotent either way."""
        rows = self.conn.execute(
            "SELECT * FROM node_types ORDER BY sort_order, id").fetchall()
        if not rows:
            self.conn.execute(
                "INSERT INTO node_types (name, sort_order) VALUES ('Processnod', 0)")
            self.commit()
            rows = self.conn.execute(
                "SELECT * FROM node_types ORDER BY sort_order, id").fetchall()
        return [dict(r) for r in rows]

    def add_node_type(self, name):
        max_ord = (self.conn.execute(
            "SELECT COALESCE(MAX(sort_order),-1) FROM node_types").fetchone()[0])
        cur = self.conn.execute(
            "INSERT INTO node_types (name, sort_order) VALUES (?,?)", (name, max_ord + 1))
        self.commit()
        return cur.lastrowid

    def rename_node_type(self, id_, name):
        self.conn.execute("UPDATE node_types SET name=? WHERE id=?", (name, id_))
        self.commit()

    def delete_node_type(self, id_):
        # Deviations left behind fall back to NULL, which the UI treats as
        # belonging to the first/default node type (normally "Processnod")
        # — never silently hidden.
        self.conn.execute(
            "UPDATE standard_deviations SET node_type_id=NULL WHERE node_type_id=?", (id_,))
        self.conn.execute("DELETE FROM node_types WHERE id=?", (id_,))
        self.commit()

    def set_deviation_node_type(self, deviation_id, node_type_id):
        self.conn.execute(
            "UPDATE standard_deviations SET node_type_id=? WHERE id=?",
            (node_type_id, deviation_id))
        self.commit()

    def copy_standard_deviation_to_node_type(self, deviation_id, node_type_id):
        """Deep, independent copy of a standard deviation AND its standard
        causes into another node type (2026-08-17, drag-and-drop between
        node types — user confirmed this must COPY, and the copy must be
        editable on its own afterward, not linked back to the original)."""
        dev = self.conn.execute(
            "SELECT * FROM standard_deviations WHERE id=?", (deviation_id,)).fetchone()
        if not dev:
            return None
        max_ord = (self.conn.execute(
            "SELECT COALESCE(MAX(sort_order),0) FROM standard_deviations").fetchone()[0])
        cur = self.conn.execute(
            "INSERT INTO standard_deviations (description, sort_order, node_type_id) "
            "VALUES (?,?,?)",
            (dev['description'], max_ord + 1, node_type_id))
        new_dev_id = cur.lastrowid
        causes = self.conn.execute(
            "SELECT * FROM standard_causes WHERE deviation_id=? ORDER BY sort_order, id",
            (deviation_id,)).fetchall()
        for c in causes:
            self.conn.execute(
                "INSERT INTO standard_causes (deviation_id, description, sort_order, "
                "object_id, comp_type, frequency, use_in_cause_form) VALUES (?,?,?,?,?,?,?)",
                (new_dev_id, c['description'], c['sort_order'], c['object_id'],
                 c['comp_type'], c['frequency'], c['use_in_cause_form']))
        self.commit()
        return new_dev_id

    def get_standard_cause(self, id_):
        row = self.conn.execute(
            "SELECT * FROM standard_causes WHERE id=?", (id_,)).fetchone()
        return dict(row) if row else None

    def standard_causes(self, deviation_id):
        return self.conn.execute(
            "SELECT * FROM standard_causes WHERE deviation_id=? AND active=1 ORDER BY sort_order, id",
            (deviation_id,)).fetchall()

    def standard_causes_for_name(self, deviation_name):
        row = self.conn.execute(
            "SELECT id FROM standard_deviations WHERE description=? LIMIT 1",
            (deviation_name,)).fetchone()
        if not row:
            return []
        return self.standard_causes(row[0])

    def add_standard_cause(self, deviation_id, description):
        max_ord = self.conn.execute(
            "SELECT COALESCE(MAX(sort_order),0) FROM standard_causes WHERE deviation_id=?",
            (deviation_id,)).fetchone()[0]
        cur = self.conn.execute(
            "INSERT INTO standard_causes (deviation_id, description, sort_order) VALUES (?,?,?)",
            (deviation_id, description, max_ord + 1))
        self.commit()
        return cur.lastrowid

    def update_standard_cause(self, id_, description=None, **kwargs):
        sets, vals = [], []
        if description is not None:
            sets.append("description=?"); vals.append(description)
        if 'frequency' in kwargs:
            sets.append("frequency=?"); vals.append(kwargs['frequency'])
        if 'use_in_cause_form' in kwargs:
            sets.append("use_in_cause_form=?"); vals.append(kwargs['use_in_cause_form'])
        if sets:
            vals.append(id_)
            self.conn.execute(f"UPDATE standard_causes SET {', '.join(sets)} WHERE id=?", vals)
            self.commit()

    def delete_standard_cause(self, id_):
        self.conn.execute("DELETE FROM standard_causes WHERE id=?", (id_,))
        self.commit()

    def reorder_standard_causes(self, ordered_ids):
        for i, id_ in enumerate(ordered_ids):
            self.conn.execute(
                "UPDATE standard_causes SET sort_order=? WHERE id=?", (i, id_))
        self.commit()

    def distinct_comp_types(self):
        """Return sorted list of all comp_type values used in standard_causes (excl. empty)."""
        rows = self.conn.execute(
            "SELECT DISTINCT comp_type FROM standard_causes "
            "WHERE comp_type != '' AND active=1 ORDER BY comp_type").fetchall()
        return [r[0] for r in rows]

    def standard_causes_for_comp_type(self, comp_type, deviation_description=None):
        """Return standard_causes for comp_type, optionally filtered to one deviation."""
        if deviation_description:
            return self.conn.execute(
                "SELECT sc.id, sc.description, sc.sort_order, sc.comp_type, sc.frequency, "
                "sd.description AS deviation_name, sd.id AS deviation_id "
                "FROM standard_causes sc "
                "JOIN standard_deviations sd ON sc.deviation_id = sd.id "
                "WHERE sc.comp_type=? AND sc.active=1 AND sd.active=1 AND sd.description=? "
                "ORDER BY sd.sort_order, sc.sort_order",
                (comp_type, deviation_description)).fetchall()
        return self.conn.execute(
            "SELECT sc.id, sc.description, sc.sort_order, sc.comp_type, sc.frequency, "
            "sd.description AS deviation_name, sd.id AS deviation_id "
            "FROM standard_causes sc "
            "JOIN standard_deviations sd ON sc.deviation_id = sd.id "
            "WHERE sc.comp_type=? AND sc.active=1 AND sd.active=1 ORDER BY sd.sort_order, sc.sort_order",
            (comp_type,)).fetchall()

    def add_standard_cause_for_comp_type(self, deviation_id, description, comp_type):
        max_ord = (self.conn.execute(
            "SELECT COALESCE(MAX(sort_order),0) FROM standard_causes WHERE deviation_id=?",
            (deviation_id,)).fetchone()[0] or 0)
        cur = self.conn.execute(
            "INSERT INTO standard_causes (deviation_id, description, sort_order, comp_type)"
            " VALUES (?,?,?,?)", (deviation_id, description, max_ord + 1, comp_type))
        self.commit()
        return cur.lastrowid

    # ── Hierarchy: deviation → object → causes ────────────────────────────────
    def objects_for_deviation(self, deviation_id):
        """Standard objects that have at least one cause for this deviation, sorted."""
        rows = self.conn.execute(
            """SELECT so.id, so.name, so.sort_order,
                      COUNT(sc.id) AS n_causes
               FROM standard_objects so
               JOIN standard_causes sc ON sc.object_id = so.id
               WHERE sc.deviation_id = ? AND sc.active=1
               GROUP BY so.id
               ORDER BY so.sort_order, so.name""",
            (deviation_id,)).fetchall()
        return [dict(r) for r in rows]

    def deviations_for_object(self, object_id):
        """Standard deviations that have at least one cause for this object,
        sorted — the mirror of objects_for_deviation(), used by
        EquipmentDeviationBar to offer the same richer, object-based
        deviation/cause set as StandardCausesPickerPopup instead of the
        narrower literal-comp_type lookup (see NOTES.md)."""
        rows = self.conn.execute(
            """SELECT sd.id, sd.description, sd.sort_order,
                      COUNT(sc.id) AS n_causes
               FROM standard_deviations sd
               JOIN standard_causes sc ON sc.deviation_id = sd.id
               WHERE sc.object_id = ? AND sc.active=1 AND sd.active=1
               GROUP BY sd.id
               ORDER BY sd.sort_order, sd.id""",
            (object_id,)).fetchall()
        return [dict(r) for r in rows]

    def all_objects_with_cause_counts(self, deviation_id):
        """All standard objects with cause count for this deviation (0 = no causes yet)."""
        rows = self.conn.execute(
            """SELECT so.id, so.name, so.sort_order,
                      COALESCE(cnt.n, 0) AS n_causes
               FROM standard_objects so
               LEFT JOIN (
                   SELECT object_id, COUNT(*) AS n
                   FROM standard_causes WHERE deviation_id=?
                   GROUP BY object_id
               ) cnt ON cnt.object_id = so.id
               ORDER BY so.sort_order, so.name""",
            (deviation_id,)).fetchall()
        return [dict(r) for r in rows]

    def standard_causes_for_object(self, deviation_id, object_id):
        """Standard causes for a specific deviation + object combination."""
        return [dict(r) for r in self.conn.execute(
            """SELECT sc.id, sc.description, sc.sort_order, sc.comp_type,
                      sc.frequency, sc.use_in_cause_form, sc.object_id
               FROM standard_causes sc
               WHERE sc.deviation_id=? AND sc.object_id=?
               ORDER BY sc.sort_order, sc.id""",
            (deviation_id, object_id))]

    def add_standard_cause_with_object(self, deviation_id, object_id, description):
        max_ord = (self.conn.execute(
            "SELECT COALESCE(MAX(sort_order),0) FROM standard_causes WHERE deviation_id=?",
            (deviation_id,)).fetchone()[0] or 0)
        # Look up comp_type from object name for backwards compat
        obj = self.conn.execute(
            "SELECT name FROM standard_objects WHERE id=?", (object_id,)).fetchone()
        comp = obj[0] if obj else ''
        cur = self.conn.execute(
            "INSERT INTO standard_causes (deviation_id, description, sort_order, comp_type, object_id)"
            " VALUES (?,?,?,?,?)",
            (deviation_id, description, max_ord + 1, comp, object_id))
        self.commit()
        return cur.lastrowid

    # ── Standard objects ──────────────────────────────────────────────────────
    def standard_objects(self):
        return [dict(r) for r in self.conn.execute(
            "SELECT id, name, sort_order FROM standard_objects ORDER BY sort_order, id")]

    def add_standard_object(self, name):
        max_ord = (self.conn.execute(
            "SELECT COALESCE(MAX(sort_order),0) FROM standard_objects").fetchone()[0] or 0)
        cur = self.conn.execute(
            "INSERT INTO standard_objects (name, sort_order) VALUES (?,?)", (name, max_ord + 1))
        self.commit()
        return cur.lastrowid

    def update_standard_object(self, id_, name):
        self.conn.execute("UPDATE standard_objects SET name=? WHERE id=?", (name, id_))
        self.commit()

    def delete_standard_object(self, id_):
        self.conn.execute("DELETE FROM standard_objects WHERE id=?", (id_,))
        self.commit()

    def reorder_standard_objects(self, ordered_ids):
        for i, id_ in enumerate(ordered_ids):
            self.conn.execute("UPDATE standard_objects SET sort_order=? WHERE id=?", (i, id_))
        self.commit()

    # ── Symbol templates ("Hitta liknande symbol" — bibliotek, 2026-08-15,
    # see NOTES.md "uppföljningsfunktioner") ────────────────────────────────
    def symbol_templates(self):
        return [dict(r) for r in self.conn.execute(
            "SELECT id, name, features_json, comp_type, created "
            "FROM symbol_templates ORDER BY name")]

    def get_symbol_template(self, id_):
        row = self.conn.execute(
            "SELECT id, name, features_json, comp_type, created "
            "FROM symbol_templates WHERE id=?", (id_,)).fetchone()
        return dict(row) if row else None

    def add_symbol_template(self, name, features_json, comp_type=''):
        cur = self.conn.execute(
            "INSERT INTO symbol_templates (name, features_json, comp_type, created) "
            "VALUES (?,?,?,datetime('now'))", (name, features_json, comp_type))
        self.commit()
        return cur.lastrowid

    def delete_symbol_template(self, id_):
        self.conn.execute("DELETE FROM symbol_templates WHERE id=?", (id_,))
        self.commit()

    def add_cause(self, deviation_id):
        dev = self.get_deviation(deviation_id)
        node_id = dev['node_id'] if dev else None
        cur = self.conn.execute(
            "INSERT INTO causes (node_id,deviation_id,description,likelihood) VALUES (?,?,'Ny orsak',1)",
            (node_id, deviation_id))
        self.commit()
        return cur.lastrowid

    def add_cause_after(self, deviation_id, after_cause_id=None):
        """Create a cause immediately after a sibling in one deviation.

        Kept separate from add_cause so existing add paths retain their
        established behaviour. Older projects may have several siblings with
        the default sort_order of zero, so normalise the sibling order first.
        """
        dev = self.get_deviation(deviation_id)
        node_id = dev['node_id'] if dev else None
        siblings = list(self.causes_for_deviation(deviation_id))
        ordered_ids = [row['id'] for row in siblings]
        insert_at = (ordered_ids.index(after_cause_id) + 1
                     if after_cause_id in ordered_ids else len(ordered_ids))
        for order, sibling_id in enumerate(ordered_ids):
            self.conn.execute(
                "UPDATE causes SET sort_order=? WHERE id=?",
                (order + (1 if order >= insert_at else 0), sibling_id))
        cur = self.conn.execute(
            "INSERT INTO causes "
            "(node_id,deviation_id,description,likelihood,sort_order) "
            "VALUES (?,?,'Ny orsak',1,?)",
            (node_id, deviation_id, insert_at))
        self.commit()
        return cur.lastrowid

    def add_consequence(self, cause_id):
        # Empty, not the literal 'Ny konsekvens' (2026-08-12, see NOTES.md)
        # — a freshly created row shows a plain "—" until actually defined
        # (ScenarioTablePanel._add_row), same convention as an empty
        # safeguard/deviation-with-no-cause row already used.
        cur = self.conn.execute(
            "INSERT INTO consequences (cause_id,description,severity) VALUES (?,'',1)", (cause_id,))
        self.commit()
        return cur.lastrowid

    def add_consequence_history(self, description):
        """Remember a consequence description that was actually typed in
        HAZOP Scenario (2026-08-26, see NOTES.md "Återanvänd tidigare
        konsekvenser") — feeds the autocomplete dropdown shown while
        editing a KON cell. INSERT OR IGNORE: a description already seen
        before is a no-op, not an error or a duplicate row."""
        description = (description or '').strip()
        if not description:
            return
        self.conn.execute(
            "INSERT OR IGNORE INTO consequence_history (description) VALUES (?)",
            (description,))
        self.commit()

    def consequence_history(self):
        return [r[0] for r in self.conn.execute(
            "SELECT description FROM consequence_history "
            "ORDER BY description COLLATE NOCASE").fetchall()]

    def add_safeguard(self, consequence_id):
        cur = self.conn.execute(
            "INSERT INTO safeguards (consequence_id,description,rrf) VALUES (?,'',1)", (consequence_id,))
        self.commit()
        return cur.lastrowid

    # ── Update ────────────────────────────────────────────────────────────────
    def set_cause_comment(self, cause_id, comment):
        self.conn.execute("UPDATE causes SET comment=? WHERE id=?", (comment, cause_id))
        self.commit()

    def get_cause_comment(self, cause_id):
        r = self.conn.execute("SELECT comment FROM causes WHERE id=?", (cause_id,)).fetchone()
        return r[0] if r else ''

    def approve_node(self, node_id, user):
        import datetime as _dt
        self.conn.execute(
            "UPDATE nodes SET study_status='approved', approved_by=?, approved_at=? WHERE id=?",
            (user, _dt.datetime.now().strftime('%Y-%m-%d %H:%M'), node_id))
        self.commit()

    def set_node_status(self, node_id, status):
        self.conn.execute("UPDATE nodes SET study_status=? WHERE id=?", (status, node_id))
        self.commit()

    # ── Backup system ─────────────────────────────────────────────────────────
    # Backups live in  <project_dir>/hazop_backups/  so they never clutter the
    # project folder itself.  Two tiers:
    #   • Hourly snapshots  kept 48 h  — cover accidental data loss within a day
    #   • Daily  snapshots  kept 30 d  — cover longer-term "I need last week"
    # The DB uses SQLite's built-in BACKUP API so the copy is always consistent
    # even while the connection is live.

    _BACKUP_DIR_NAME   = "hazop_backups"
    _HOURLY_KEEP_H     = 48      # keep hourly backups for this many hours
    _DAILY_KEEP_D      = 30      # keep daily backups for this many days
    _COMMIT_INTERVAL_S  = 120    # write a new hourly backup at most every N seconds
    _PRUNE_INTERVAL_S   = 3600   # prune at most once per hour
    _last_backup_ts: float = 0.0
    _last_prune_ts:  float = 0.0

    def _backup_dir(self) -> 'Path':
        d = self.path.parent / self._BACKUP_DIR_NAME
        d.mkdir(exist_ok=True)
        return d

    def _write_backup(self, startup: bool = False):
        """Write a timestamped backup using SQLite's online backup API.

        Throttled to at most once per _COMMIT_INTERVAL_S seconds so that
        frequent commits don't hammer the disk, but startup always writes.
        Returns the backup file's Path on success, or None (throttled or
        failed) — existing call sites all ignore the return value, so this
        is purely additive for callers (e.g. a manual "backup now" button)
        that want to report success/failure to the user.
        """
        import time, sqlite3, datetime as _dt
        now = time.monotonic()
        if not startup and (now - Database._last_backup_ts) < self._COMMIT_INTERVAL_S:
            return None
        Database._last_backup_ts = now
        try:
            d   = self._backup_dir()
            ts  = _dt.datetime.now().strftime('%Y-%m-%dT%H-%M-%S-%f')
            dst = d / f"backup_{ts}.db"
            # SQLite online backup — safe while the DB is open and being written
            bk_conn = sqlite3.connect(str(dst))
            with bk_conn:
                self.conn.backup(bk_conn)
            bk_conn.close()
            self._prune_backups(d)
            return dst
        except Exception:
            return None   # never crash the app due to backup failure

    def _prune_backups(self, d: 'Path'):
        """Remove old backups according to retention policy (rate-limited to once/hour)."""
        import time, datetime as _dt
        now_ts = time.monotonic()
        if (now_ts - Database._last_prune_ts) < self._PRUNE_INTERVAL_S:
            return
        Database._last_prune_ts = now_ts

        now   = _dt.datetime.now()
        files = sorted(d.glob("backup_*.db"), reverse=True)
        # Skip entirely if the directory is too small to prune anything
        if len(files) <= self._HOURLY_KEEP_H + self._DAILY_KEEP_D:
            return
        # Parse timestamp from filename; skip files that don't match
        def parse_ts(f):
            for fmt in ("backup_%Y-%m-%dT%H-%M-%S-%f", "backup_%Y-%m-%dT%H-%M-%S"):
                try:
                    return _dt.datetime.strptime(f.stem, fmt)
                except ValueError:
                    pass
            return None
        kept_dates = set()   # dates for which we already have a daily backup
        for f in files:
            ts = parse_ts(f)
            if ts is None:
                continue
            age_h = (now - ts).total_seconds() / 3600
            date_key = ts.date()
            # Always keep if within hourly window
            if age_h <= self._HOURLY_KEEP_H:
                continue
            # Outside hourly window — keep ONE per calendar day for the daily window
            if (now - ts).days <= self._DAILY_KEEP_D:
                if date_key not in kept_dates:
                    kept_dates.add(date_key)
                    continue   # keep this one as the day's representative
            # Otherwise delete
            try:
                f.unlink()
            except Exception:
                pass

    # ── Session undo/redo ───────────────────────────────────────────────────
    #
    # The application has a large number of write paths.  Keeping inverse
    # operations next to every INSERT/UPDATE/DELETE would be both incomplete
    # and especially fragile for cascades and hierarchy moves.  SQLite can
    # provide an exact, consistent database image after each write instead.
    # The images are kept in memory for the current session; the existing
    # on-disk backups remain the long-term recovery mechanism.

    def _capture_history_snapshot(self):
        """Return a consistent SQL dump of the committed database."""
        # serialize()/deserialize() detaches WAL-backed file connections on
        # this Python/SQLite build.  iterdump() is SQLite's own consistent
        # committed export and can be restored into both file and memory DBs.
        return '\n'.join(self.conn.iterdump())

    def _reset_history_baseline(self):
        """Start a fresh history after construction or project replacement."""
        try:
            self.conn.commit()
            self._history_snapshot = self._capture_history_snapshot()
        except Exception:
            # History must never prevent a project from opening.  A later
            # successful commit can establish the first usable baseline.
            logging.warning("Could not initialise database undo history", exc_info=True)
            self._history_snapshot = None
        self._undo_stack.clear()
        self._redo_stack.clear()
        self._history_group_depth = 0
        self._history_group_before = None
        self._history_group_changed = False
        self._history_initialized = self._history_snapshot is not None
        self._notify_history_listeners()

    def add_history_listener(self, callback):
        """Subscribe to session-history changes without coupling DB to Qt."""
        if callable(callback) and callback not in self._history_listeners:
            self._history_listeners.append(callback)

    def remove_history_listener(self, callback):
        try:
            self._history_listeners.remove(callback)
        except ValueError:
            pass

    def _notify_history_listeners(self):
        for callback in list(getattr(self, '_history_listeners', ())):
            try:
                callback()
            except Exception:
                logging.debug("Undo history listener failed", exc_info=True)

    def _push_history_entry(self, before, after):
        if before is None or before == after:
            return
        self._undo_stack.append((before, after))
        if len(self._undo_stack) > self._HISTORY_LIMIT:
            del self._undo_stack[:-self._HISTORY_LIMIT]
        # Any new edit after an undo starts a new branch.
        self._redo_stack.clear()
        self._notify_history_listeners()

    def _record_committed_state(self):
        """Record the state reached by the most recent successful commit."""
        if not self._history_initialized or self._history_restoring:
            return
        try:
            after = self._capture_history_snapshot()
        except Exception:
            logging.warning("Could not capture database undo snapshot", exc_info=True)
            return
        before = self._history_snapshot
        self._history_snapshot = after
        if before is None or before == after:
            return
        if self._history_group_depth:
            self._history_group_changed = True
            return
        self._push_history_entry(before, after)

    @contextmanager
    def history_group(self):
        """Coalesce several commits into one undo step.

        Existing write methods intentionally commit immediately.  UI code can
        wrap a multi-write user action in this context as it is brought into
        the new history model; direct callers remain safe because each normal
        commit is still independently undoable.
        """
        if not self._history_initialized:
            yield
            return
        outermost = self._history_group_depth == 0
        if outermost:
            self._history_group_before = self._history_snapshot
            self._history_group_changed = False
        self._history_group_depth += 1
        try:
            yield
        finally:
            self._history_group_depth -= 1
            if self._history_group_depth == 0:
                before = self._history_group_before
                try:
                    # This also catches a legacy raw ``conn.commit()`` inside
                    # a group which did not go through Database.commit().
                    after = self._capture_history_snapshot()
                except Exception:
                    after = self._history_snapshot
                self._history_snapshot = after
                if self._history_group_changed or before != after:
                    self._push_history_entry(before, after)
                self._history_group_before = None
                self._history_group_changed = False

    @property
    def can_undo(self):
        return bool(self._undo_stack)

    @property
    def can_redo(self):
        return bool(self._redo_stack)

    @property
    def undo_count(self):
        return len(self._undo_stack)

    @property
    def redo_count(self):
        return len(self._redo_stack)

    def clear_undo_history(self):
        """Keep the current data but forget the session's undo/redo chain."""
        self._reset_history_baseline()

    def _restore_history_snapshot(self, snapshot):
        """Restore an image without producing a second history entry."""
        self._history_restoring = True
        source = None
        try:
            # Restore into the already-open connection.  Replacing the file
            # and deleting ``-wal``/``-shm`` sidecars looks attractive, but
            # those files can be held by SQLite/OneDrive on Windows.  The old
            # implementation closed ``self.conn`` before that deletion; when
            # unlink() then raised WinError 32, every subsequent Qt repaint
            # saw a closed database and crashed with ProgrammingError.
            #
            # An in-memory source plus SQLite's backup API keeps the live
            # connection (and all panels using it) intact.  The backup is
            # transactional on the destination, so a busy/locked database
            # fails without intentionally detaching the application from its
            # database.
            source = sqlite3.connect(':memory:', timeout=30.0)
            source.row_factory = sqlite3.Row
            # sqlite3's iterdump is alphabetically ordered.  Some tables
            # therefore reference a table that appears later in the dump;
            # SQLite only accepts that schema order while foreign-key
            # enforcement is disabled.  Enable it again after the dump has
            # been materialised, matching the live connection's setting.
            source.executescript(snapshot)
            source.commit()
            source.execute("PRAGMA foreign_keys = ON")
            self.conn.commit()
            source.backup(self.conn)
            self.conn.commit()
            self.conn.execute("PRAGMA foreign_keys = ON")
        finally:
            if source is not None:
                source.close()
            self._history_restoring = False
        # Risk information is cached at module level and is otherwise stale
        # after undoing a risk-matrix/settings change.
        _risk_matrix_cache.load(self)

    def undo(self):
        """Undo the latest committed database change, if it is still current."""
        if not self._undo_stack:
            return False
        before, after = self._undo_stack[-1]
        try:
            current = self._capture_history_snapshot()
        except Exception:
            return False
        if current != after:
            # A legacy/raw writer changed the DB behind the history layer.
            # Do not restore an unsafe image over that external change.
            logging.warning("Database changed outside undo history; resetting baseline")
            self._reset_history_baseline()
            return False
        try:
            self._restore_history_snapshot(before)
        except Exception:
            # Undo is a user convenience.  A transient lock must not escape
            # into the Qt event loop and become a global crash.  Keep the
            # history entry in place so the user can try again after the lock
            # is released; the live connection is deliberately left usable by
            # _restore_history_snapshot().
            logging.warning("Could not restore undo snapshot", exc_info=True)
            return False
        self._undo_stack.pop()
        self._redo_stack.append((before, after))
        self._history_snapshot = before
        self._notify_history_listeners()
        return True

    def redo(self):
        """Redo the latest undone database change, if it is still current."""
        if not self._redo_stack:
            return False
        before, after = self._redo_stack[-1]
        try:
            current = self._capture_history_snapshot()
        except Exception:
            return False
        if current != before:
            logging.warning("Database changed outside redo history; resetting baseline")
            self._reset_history_baseline()
            return False
        try:
            self._restore_history_snapshot(after)
        except Exception:
            logging.warning("Could not restore redo snapshot", exc_info=True)
            return False
        self._redo_stack.pop()
        self._undo_stack.append((before, after))
        if len(self._undo_stack) > self._HISTORY_LIMIT:
            del self._undo_stack[:-self._HISTORY_LIMIT]
        self._history_snapshot = after
        self._notify_history_listeners()
        return True

    def _commit_with_history(self):
        """Commit and record one state transition, preserving old semantics."""
        self.conn.commit()
        self._record_committed_state()
        self._write_backup()

    def commit(self):
        """Write-through commit: history snapshot, then throttled backup."""
        try:
            self._commit_with_history()
        except Exception:
            # Preserve the established non-fatal backup/commit behaviour.  A
            # history failure must not turn an otherwise valid UI edit into a
            # crash; diagnostics are still useful during development.
            logging.debug("Database write-through commit failed", exc_info=True)

    def touch_node(self, node_id, commit=True):
        """Update updated_at/updated_by on node (feature 20)."""
        import datetime as _dt
        user = (self.get_config('user_name', '') or '').strip() or 'okänd'
        self.conn.execute(
            "UPDATE nodes SET updated_at=?,updated_by=? WHERE id=?",
            (_dt.datetime.now().strftime('%Y-%m-%d %H:%M'), user, node_id))
        if commit:
            self.commit()

    def update_node(self, id_, name, description, pid_ref,
                    media='', pressure='', temperature=''):
        self.conn.execute(
            "UPDATE nodes SET name=?,description=?,pid_ref=?,"
            "media=?,pressure=?,temperature=? WHERE id=?",
            (name, description, pid_ref, media, pressure, temperature, id_))
        # "Lägg ut nodnamn" markups (node_markups.type='text') store a
        # one-time snapshot of the node's name, not a live reference —
        # keep them in sync here, in the single shared write path every
        # rename UI (TreePanel "Döp om", PropertiesRibbon's Namn-popup,
        # NodePanel's save) already goes through, instead of requiring
        # each caller to remember its own follow-up call (2026-08-17,
        # see NOTES.md "nodnamn på P&ID uppdateras inte vid namnbyte" —
        # NodePanel already did this via an external signal connection,
        # but the other two rename paths didn't, so a rename via "Döp
        # om" silently left the on-canvas label stale).
        self.sync_node_text_markups(id_, name, commit=False)
        # Keep the visible node update and its audit metadata in one undo
        # step. ``touch_node`` remains independently usable elsewhere.
        self.touch_node(id_, commit=False)
        self.commit()

    # ── Node markup CRUD ──────────────────────────────────────────────────────
    def add_node_markup(self, node_id, type_, pts, label, color, opacity, line_width, page,
                        font_size=12):
        cur = self.conn.execute(
            "INSERT INTO node_markups (node_id,type,points,label,color,opacity,line_width,"
            "font_size,pid_page) VALUES (?,?,?,?,?,?,?,?,?)",
            (node_id, type_, json.dumps(pts), label, color, opacity, line_width,
             font_size, page))
        self.commit()
        return cur.lastrowid

    def node_markups_for_node(self, node_id):
        return self.conn.execute(
            "SELECT * FROM node_markups WHERE node_id=? ORDER BY sort_order,id",
            (node_id,)).fetchall()

    def node_markups_for_page(self, page):
        return self.conn.execute(
            "SELECT * FROM node_markups WHERE pid_page=? ORDER BY sort_order,id",
            (page,)).fetchall()

    def nodes_on_page(self, physical_page):
        """Nodes appearing on a physical page — either via the node's own
        (singular) pid_page, or via any node_markups row on that page (a
        node can have markup on multiple pages), 2026-08-17 see NOTES.md
        "Ny Noder-flik"."""
        rows = self.conn.execute(
            """SELECT DISTINCT n.* FROM nodes n
               LEFT JOIN node_markups nm ON nm.node_id = n.id
               WHERE n.pid_page = ? OR nm.pid_page = ?
               ORDER BY n.id""",
            (physical_page, physical_page)).fetchall()
        return [dict(r) for r in rows]

    def pages_for_node(self, node_id):
        """Physical pages a node appears on (own pid_page + all node_markups
        pages), deduplicated and sorted."""
        node = self.get_node(node_id)
        pages = set()
        if node and node['pid_page'] is not None:
            pages.add(node['pid_page'])
        for m in self.node_markups_for_node(node_id):
            pages.add(m['pid_page'])
        return sorted(pages)

    def analysis_pages_for_node(self, node_id):
        """Return only pages with node graphics or linked cause objects."""
        pages = set()
        node = self.get_node(node_id)
        if node and node['pid_page'] is not None:
            try:
                points = json.loads(node['markup_points'] or '[]')
            except (TypeError, ValueError):
                points = []
            if points:
                pages.add(node['pid_page'])
        for table in ('node_markups', 'node_red_markups'):
            rows = self.conn.execute(
                f"SELECT pid_page FROM {table} WHERE node_id=?", (node_id,)).fetchall()
            pages.update(r['pid_page'] for r in rows if r['pid_page'] is not None)
        rows = self.conn.execute(
            """SELECT em.pid_page
               FROM equipment_markers em
               JOIN equipment_catalog ec ON ec.id = em.equipment_id
               WHERE ec.node_id=?
               UNION
               SELECT em.pid_page
               FROM equipment_markers em
               JOIN causes c ON c.equipment_id = em.equipment_id
               JOIN deviations d ON d.id = c.deviation_id
               WHERE d.node_id=?
               UNION
               SELECT em.pid_page
               FROM equipment_markers em
               JOIN deviations d ON d.equipment_id = em.equipment_id
               WHERE d.node_id=?""",
            (node_id, node_id, node_id)).fetchall()
        pages.update(r['pid_page'] for r in rows if r['pid_page'] is not None)
        return sorted(pages)

    def analysis_objects_for_node(self, node_id):
        """Return {physical_page: [object tags]} for objects relevant to node."""
        rows = self.conn.execute(
            """SELECT em.pid_page, COALESCE(em.tag, ec.tag, '') AS tag
               FROM equipment_markers em
               JOIN equipment_catalog ec ON ec.id = em.equipment_id
               WHERE ec.node_id=?
               UNION
               SELECT em.pid_page, COALESCE(em.tag, ec.tag, '')
               FROM equipment_markers em
               JOIN equipment_catalog ec ON ec.id = em.equipment_id
               JOIN causes c ON c.equipment_id = em.equipment_id
               JOIN deviations d ON d.id = c.deviation_id
               WHERE d.node_id=?
               UNION
               SELECT em.pid_page, COALESCE(em.tag, ec.tag, '')
               FROM equipment_markers em
               JOIN equipment_catalog ec ON ec.id = em.equipment_id
               JOIN deviations d ON d.equipment_id = em.equipment_id
               WHERE d.node_id=?
               ORDER BY 1, 2""",
            (node_id, node_id, node_id)).fetchall()
        result = {}
        for row in rows:
            tag = (row['tag'] or '').strip()
            if tag:
                result.setdefault(row['pid_page'], [])
                if tag not in result[row['pid_page']]:
                    result[row['pid_page']].append(tag)
        return result

    def analysis_object_details_for_node(self, node_id):
        """Return {page: [{tag, type, deviations, count}]} for a node."""
        rows = self.conn.execute(
            """SELECT DISTINCT em.pid_page AS page, em.equipment_id,
                      COALESCE(em.tag, ec.tag, '') AS tag,
                      COALESCE(ec.equipment_type, em.comp_type, '') AS type,
                      d.description AS deviation
               FROM equipment_markers em
               JOIN equipment_catalog ec ON ec.id = em.equipment_id
               LEFT JOIN deviations d ON d.equipment_id = em.equipment_id
               WHERE ec.node_id=? OR d.node_id=?
               UNION
               SELECT DISTINCT em.pid_page, em.equipment_id,
                      COALESCE(em.tag, ec.tag, ''),
                      COALESCE(ec.equipment_type, em.comp_type, ''),
                      d.description
               FROM equipment_markers em
               JOIN equipment_catalog ec ON ec.id = em.equipment_id
               JOIN causes c ON c.equipment_id = em.equipment_id
               JOIN deviations d ON d.id = c.deviation_id
               WHERE d.node_id=?
               ORDER BY 1, 3, 5""",
            (node_id, node_id, node_id)).fetchall()
        result = {}
        by_key = {}
        for row in rows:
            page = row['page']
            key = (row['equipment_id'], row['tag'] or '')
            obj = by_key.setdefault((page, key), {
                'tag': (row['tag'] or '').strip(),
                'type': (row['type'] or '').strip(),
                'deviations': []})
            deviation = (row['deviation'] or '').strip()
            if deviation and deviation not in obj['deviations']:
                obj['deviations'].append(deviation)
        for (page, _), obj in by_key.items():
            obj['count'] = len(obj['deviations'])
            result.setdefault(page, []).append(obj)
        return result

    def get_node_markup(self, mu_id):
        row = self.conn.execute(
            "SELECT * FROM node_markups WHERE id=?", (mu_id,)).fetchone()
        return dict(row) if row else None

    def update_node_markup(self, mu_id, label=None, color=None, opacity=None,
                           line_width=None, font_size=None, visible=None, points=None):
        sets, vals = [], []
        if label      is not None: sets.append("label=?");      vals.append(label)
        if color      is not None: sets.append("color=?");      vals.append(color)
        if opacity    is not None: sets.append("opacity=?");    vals.append(opacity)
        if line_width is not None: sets.append("line_width=?"); vals.append(line_width)
        if font_size  is not None: sets.append("font_size=?");  vals.append(font_size)
        if visible    is not None: sets.append("visible=?");    vals.append(int(visible))
        if points     is not None: sets.append("points=?");     vals.append(json.dumps(points))
        if sets:
            vals.append(mu_id)
            self.conn.execute(f"UPDATE node_markups SET {','.join(sets)} WHERE id=?", vals)
            self.commit()

    def delete_node_markup(self, mu_id):
        self.conn.execute("DELETE FROM node_markups WHERE id=?", (mu_id,))
        self.commit()

    def set_all_node_markups_visible(self, node_id, visible):
        self.conn.execute("UPDATE node_markups SET visible=? WHERE node_id=?",
                          (int(visible), node_id))
        self.commit()

    def has_node_markups(self, node_id) -> bool:
        r = self.conn.execute(
            "SELECT COUNT(*) FROM node_markups WHERE node_id=?", (node_id,)).fetchone()
        return r[0] > 0

    def has_visible_node_markups(self, node_id) -> bool:
        r = self.conn.execute(
            "SELECT COUNT(*) FROM node_markups WHERE node_id=? AND visible=1",
            (node_id,)).fetchone()
        return r[0] > 0

    # ── Node red markup CRUD ──────────────────────────────────────────────────
    def add_node_red_markup(self, node_id, type_, pts, label, color, opacity,
                            line_width, page, font_size=12,
                            symbol_w=40, symbol_h=40, symbol_rot=0):
        cur = self.conn.execute(
            "INSERT INTO node_red_markups "
            "(node_id,type,points,label,color,opacity,line_width,font_size,pid_page,"
            "symbol_w,symbol_h,symbol_rot) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
            (node_id, type_, json.dumps(pts), label, color, opacity, line_width,
             font_size, page, symbol_w, symbol_h, symbol_rot))
        self.commit()
        return cur.lastrowid

    def node_red_markups_for_node(self, node_id):
        return self.conn.execute(
            "SELECT * FROM node_red_markups WHERE node_id=? ORDER BY sort_order,id",
            (node_id,)).fetchall()

    def node_red_markups_for_page(self, page):
        return self.conn.execute(
            "SELECT * FROM node_red_markups WHERE pid_page=? ORDER BY sort_order,id",
            (page,)).fetchall()

    def get_node_red_markup(self, mu_id):
        row = self.conn.execute(
            "SELECT * FROM node_red_markups WHERE id=?", (mu_id,)).fetchone()
        return dict(row) if row else None

    def update_node_red_markup(self, mu_id, label=None, color=None, opacity=None,
                               line_width=None, font_size=None, visible=None,
                               points=None, symbol_w=None, symbol_h=None, symbol_rot=None):
        sets, vals = [], []
        if label      is not None: sets.append("label=?");      vals.append(label)
        if color      is not None: sets.append("color=?");      vals.append(color)
        if opacity    is not None: sets.append("opacity=?");    vals.append(opacity)
        if line_width is not None: sets.append("line_width=?"); vals.append(line_width)
        if font_size  is not None: sets.append("font_size=?");  vals.append(font_size)
        if visible    is not None: sets.append("visible=?");    vals.append(int(visible))
        if points     is not None: sets.append("points=?");     vals.append(json.dumps(points))
        if symbol_w   is not None: sets.append("symbol_w=?");   vals.append(symbol_w)
        if symbol_h   is not None: sets.append("symbol_h=?");   vals.append(symbol_h)
        if symbol_rot is not None: sets.append("symbol_rot=?"); vals.append(symbol_rot)
        if sets:
            vals.append(mu_id)
            self.conn.execute(
                f"UPDATE node_red_markups SET {','.join(sets)} WHERE id=?", vals)
            self.commit()

    def delete_node_red_markup(self, mu_id):
        self.conn.execute("DELETE FROM node_red_markups WHERE id=?", (mu_id,))
        self.commit()

    def set_all_node_red_markups_visible(self, node_id, visible):
        self.conn.execute(
            "UPDATE node_red_markups SET visible=? WHERE node_id=?",
            (int(visible), node_id))
        self.commit()

    def has_node_red_markups(self, node_id) -> bool:
        r = self.conn.execute(
            "SELECT COUNT(*) FROM node_red_markups WHERE node_id=?", (node_id,)).fetchone()
        return r[0] > 0

    def has_visible_node_red_markups(self, node_id) -> bool:
        r = self.conn.execute(
            "SELECT COUNT(*) FROM node_red_markups WHERE node_id=? AND visible=1",
            (node_id,)).fetchone()
        return r[0] > 0

    def get_node_number(self, node_id) -> int:
        """Return 1-based position of node_id in creation order (0 if not found)."""
        rows = self.conn.execute("SELECT id FROM nodes ORDER BY id").fetchall()
        for i, r in enumerate(rows, 1):
            if r[0] == node_id:
                return i
        return 0

    def sync_node_text_markups(self, node_id, new_name, commit=True):
        """Update label of all 'text' type markups for a node to match its new name."""
        self.conn.execute(
            "UPDATE node_markups SET label=? WHERE node_id=? AND type='text'",
            (new_name, node_id))
        if commit:
            self.commit()

    _SENTINEL = object()

    def update_cause(self, id_, description=None, likelihood=None, base_frequency=_SENTINEL,
                     standard_cause_id=_SENTINEL, comp_type=_SENTINEL, comp_tag=_SENTINEL,
                     base_freq=_SENTINEL, equipment_id=_SENTINEL,
                     secondary_equipment_id=_SENTINEL,
                     group_choices_set=_SENTINEL,
                     group_equipment_ids=_SENTINEL,
                     frequency_cleared=_SENTINEL):
        # Support old parameter name for backward compatibility
        if base_freq is not Database._SENTINEL and base_frequency is Database._SENTINEL:
            base_frequency = base_freq

        sets, vals = [], []
        if description is not None:
            sets.append("description=?"); vals.append(description)
        if likelihood is not None:
            sets.append("likelihood=?"); vals.append(likelihood)
        if base_frequency is not Database._SENTINEL:
            sets.append("base_frequency=?"); vals.append(base_frequency)
        if standard_cause_id is not Database._SENTINEL:
            sets.append("standard_cause_id=?"); vals.append(standard_cause_id)
        if comp_type is not Database._SENTINEL:
            sets.append("comp_type=?"); vals.append(comp_type)
        if comp_tag is not Database._SENTINEL:
            sets.append("comp_tag=?"); vals.append(comp_tag)
        # Live link to equipment_catalog (2026-08-13, see NOTES.md) — an
        # explicit None here means "sever the link" (a custom/unmatched
        # tag), same optional-but-distinguishable-from-unset convention
        # comp_type/comp_tag already use via _SENTINEL.
        if equipment_id is not Database._SENTINEL:
            sets.append("equipment_id=?"); vals.append(equipment_id)
        if secondary_equipment_id is not Database._SENTINEL:
            sets.append("secondary_equipment_id=?"); vals.append(secondary_equipment_id)
        if group_choices_set is not Database._SENTINEL:
            sets.append("group_choices_set=?"); vals.append(group_choices_set)
        if group_equipment_ids is not Database._SENTINEL:
            sets.append("group_equipment_ids=?")
            vals.append(json.dumps(group_equipment_ids) if isinstance(group_equipment_ids, (list, tuple))
                        else group_equipment_ids)
        if frequency_cleared is not Database._SENTINEL:
            sets.append("frequency_cleared=?")
            vals.append(int(bool(frequency_cleared)))
        if sets:
            vals.append(id_)
            self.conn.execute(f"UPDATE causes SET {', '.join(sets)} WHERE id=?", vals)
            self.commit()

        # Learn from every explicit tag+type assignment, regardless of which
        # UI path triggered it.  Read back the current values if either param
        # was not part of this call so we always have both.
        effective_ct  = comp_type  if comp_type  is not Database._SENTINEL else None
        effective_tag = comp_tag   if comp_tag   is not Database._SENTINEL else None
        if effective_ct is None or effective_tag is None:
            row = self.conn.execute(
                "SELECT comp_type, comp_tag FROM causes WHERE id=?", (id_,)).fetchone()
            if row:
                if effective_ct  is None: effective_ct  = row['comp_type']  or ''
                if effective_tag is None: effective_tag = row['comp_tag']   or ''
        if effective_ct and effective_tag:
            try:
                self.upsert_tag_memory(effective_tag, effective_ct, comp_tag=effective_tag)
            except Exception:
                pass

    def update_cause_freqs_from_standard(self):
        """Overwrite base_frequency on all causes linked to a standard cause that has a frequency."""
        self.conn.execute("""
            UPDATE causes
            SET base_frequency = (
                SELECT frequency FROM standard_causes WHERE id = causes.standard_cause_id
            )
            WHERE standard_cause_id IS NOT NULL
              AND COALESCE(frequency_cleared, 0) = 0
              AND EXISTS (
                SELECT 1 FROM standard_causes
                WHERE id = causes.standard_cause_id AND frequency IS NOT NULL
              )
        """)
        n = _sync_f_levels_from_base_frequency(self.conn)
        self.commit()
        return n

    def update_consequence(self, id_, description, severity, category='',
                           consequence_chain='', comp_tag=None, comp_type=None,
                           tagged_refs=None):
        description = normalize_arrows(description)
        self.conn.execute(
            "UPDATE consequences SET description=?,severity=?,category=?,"
            "consequence_chain=? WHERE id=?",
            (description, severity, category, consequence_chain, id_))
        # comp_tag/comp_type (2026-08-07, drag-and-drop tag from P&ID —
        # see NOTES.md) — optional, None means "don't touch", same
        # backward-compatible convention update_cause already uses, so
        # every existing call site (which never passes these) is unaffected.
        # tagged_refs (2026-08-09) follows the same optional convention.
        if comp_tag is not None or comp_type is not None or tagged_refs is not None:
            parts, vals = [], []
            if comp_tag is not None:
                parts.append("comp_tag=?"); vals.append(comp_tag)
            if comp_type is not None:
                parts.append("comp_type=?"); vals.append(comp_type)
            if tagged_refs is not None:
                parts.append("tagged_refs=?"); vals.append(tagged_refs)
            vals.append(id_)
            self.conn.execute(f"UPDATE consequences SET {', '.join(parts)} WHERE id=?", vals)
        self.commit()

    def set_consequence_tag(self, id_, comp_tag, comp_type):
        """Attach an equipment tag/type to a consequence's tag-strip
        display without touching its description/severity — the
        low-level primitive append_tag_to_consequence() builds on to also
        update the free text (2026-08-09, see NOTES.md)."""
        self.conn.execute(
            "UPDATE consequences SET comp_tag=?, comp_type=? WHERE id=?",
            (comp_tag, comp_type, id_))
        self.commit()

    def set_safeguard_tag(self, id_, comp_tag, comp_type):
        """Attach an equipment tag/type to a safeguard's tag-strip display
        without touching its description/rrf — same complement-not-
        replacement rule as set_consequence_tag; append_tag_to_safeguard()
        builds on this to also update the free text (2026-08-09, see
        NOTES.md)."""
        self.conn.execute(
            "UPDATE safeguards SET comp_tag=?, comp_type=? WHERE id=?",
            (comp_tag, comp_type, id_))
        self.commit()

    def append_tag_to_consequence(self, id_, comp_tag, comp_type):
        """Drag-and-drop an equipment marker onto a KON cell (2026-08-09,
        see NOTES.md): appends the tag into the free-text description
        (building a running sentence across repeated drags, e.g. "hög
        nivå i" -> "hög nivå i TA-1" -> "... => överbreddning till TA-2"),
        instead of the tag-strip-only behavior of set_consequence_tag,
        which used to silently overwrite the PREVIOUS drop's tag on every
        new one. Still updates the tag-strip too (shows the most recently
        dropped tag) — the full history lives in the text now."""
        row = self.get_consequence(id_)
        if not row:
            return
        new_desc = append_tag_to_text(row['description'], comp_tag)
        new_refs = add_tag_ref(row.get('tagged_refs') or '', comp_tag)
        self.update_consequence(id_, new_desc, row['severity'], row['category'] or '',
                                 row['consequence_chain'] or '',
                                 comp_tag=comp_tag, comp_type=comp_type,
                                 tagged_refs=new_refs)

    def append_tag_to_safeguard(self, id_, comp_tag, comp_type):
        """Same as append_tag_to_consequence, for a safeguard cell."""
        row = self.get_safeguard(id_)
        if not row:
            return
        new_desc = append_tag_to_text(row['description'], comp_tag)
        new_refs = add_tag_ref(row.get('tagged_refs') or '', comp_tag)
        self.update_safeguard(id_, description=new_desc, tagged_refs=new_refs)
        self.set_safeguard_tag(id_, comp_tag, comp_type)

    def update_safeguard(self, id_, description=None, rrf=None, sg_type=None, tagged_refs=None):
        if description is None and rrf is None and sg_type is None and tagged_refs is None:
            return
        parts, vals = [], []
        if description is not None:
            parts.append("description=?"); vals.append(normalize_arrows(description))
        if rrf is not None:
            parts.append("rrf=?"); vals.append(rrf)
        if sg_type is not None:
            parts.append("sg_type=?"); vals.append(sg_type)
        if tagged_refs is not None:
            parts.append("tagged_refs=?"); vals.append(tagged_refs)
        vals.append(id_)
        self.conn.execute(f"UPDATE safeguards SET {', '.join(parts)} WHERE id=?", vals)
        self.commit()

    # ── Delete ────────────────────────────────────────────────────────────────
    def delete_node(self, id_):
        # Deleting a node cascades through causes -> consequences -> safeguards,
        # i.e. it can wipe out an entire branch of the HAZOP tree in one go.
        # Force an un-throttled backup right before this destructive operation
        # so a crash mid-cascade (or a bug in the cascade logic) is always
        # recoverable. Never let a backup failure block the actual delete.
        try:
            self._write_backup(startup=True)
        except Exception:
            logging.warning("Pre-delete backup failed", exc_info=True)
        # No FK cascade exists from causes(node_id) down into
        # consequence_severities / consequence_severity_exclusions / linked_consequence_id,
        # so route through delete_cause() for each direct cause (mirrors delete_deviation()).
        for cause in self.causes(id_):
            self.delete_cause(cause['id'])
        # equipment_catalog.node_id (added 2026-08-07 for "Nod → Utrustning
        # → Avvikelse", see NOTES.md) has NO ON DELETE clause — unlike
        # deviations.node_id, equipment assigned to this node must NOT be
        # deleted along with it (the assignment is optional/soft, the
        # equipment itself lives independently in the register), so clear
        # the assignment instead of cascading. Without this, deleting a
        # node with any equipment assigned to it raised sqlite3.IntegrityError:
        # FOREIGN KEY constraint failed (real crash report, 2026-08-07).
        self.conn.execute("UPDATE equipment_catalog SET node_id=NULL WHERE node_id=?", (id_,))
        self.conn.execute("DELETE FROM nodes WHERE id=?", (id_,)); self.commit()

    def delete_cause(self, id_):
        # Preserve LOPA evidence when its HAZOP source is removed. The FK is
        # intentionally SET NULL, while this flag turns that retained record
        # into an explicit missing-source warning in the LOPA view.
        self.conn.execute(
            "UPDATE lopa_source_scenarios SET source_missing=1 WHERE hazop_cause_id=?", (id_,))
        # Clean up severity/exclusion data for all consequences under this cause
        # (no FK cascade exists for these tables).
        self.conn.execute(
            "DELETE FROM reduction_factor_severity_exclusions WHERE reduction_factor_id IN ("
            "  SELECT id FROM reduction_factors WHERE consequence_id IN "
            "  (SELECT id FROM consequences WHERE cause_id=?))", (id_,))
        self.conn.execute(
            "DELETE FROM consequence_severity_exclusions "
            "WHERE severity_id IN ("
            "  SELECT cs.id FROM consequence_severities cs "
            "  JOIN consequences c ON cs.consequence_id = c.id "
            "  WHERE c.cause_id=?)",
            (id_,))
        self.conn.execute(
            "DELETE FROM consequence_severities "
            "WHERE consequence_id IN (SELECT id FROM consequences WHERE cause_id=?)",
            (id_,))
        self.conn.execute(
            "DELETE FROM consequence_final_severities "
            "WHERE consequence_id IN (SELECT id FROM consequences WHERE cause_id=?)",
            (id_,))
        self.conn.execute(
            "UPDATE causes SET linked_consequence_id=NULL "
            "WHERE linked_consequence_id IN (SELECT id FROM consequences WHERE cause_id=?)",
            (id_,))
        self.conn.execute(
            "DELETE FROM consequence_severity_exclusions WHERE safeguard_id IN ("
            "  SELECT id FROM safeguards WHERE consequence_id IN "
            "  (SELECT id FROM consequences WHERE cause_id=?))",
            (id_,))
        self.conn.execute("DELETE FROM causes WHERE id=?", (id_,)); self.commit()

    def delete_consequence(self, id_):
        self.conn.execute(
            "UPDATE lopa_source_consequences SET source_missing=1 WHERE hazop_consequence_id=?", (id_,))
        # Clean up orphaned severity data (no FK cascade exists for these tables)
        self.conn.execute(
            "DELETE FROM reduction_factor_severity_exclusions WHERE reduction_factor_id IN ("
            "SELECT id FROM reduction_factors WHERE consequence_id=?)", (id_,))
        self.conn.execute(
            "DELETE FROM consequence_severity_exclusions "
            "WHERE severity_id IN (SELECT id FROM consequence_severities WHERE consequence_id=?)",
            (id_,))
        self.conn.execute("DELETE FROM consequence_severities WHERE consequence_id=?", (id_,))
        self.conn.execute("DELETE FROM consequence_final_severities WHERE consequence_id=?", (id_,))
        # Null out any causes that chain-link to this consequence (cross-branch reference, no FK)
        self.conn.execute("UPDATE causes SET linked_consequence_id=NULL WHERE linked_consequence_id=?", (id_,))
        self.conn.execute(
            "DELETE FROM consequence_severity_exclusions WHERE safeguard_id IN "
            "(SELECT id FROM safeguards WHERE consequence_id=?)",
            (id_,))
        self.conn.execute("DELETE FROM consequences WHERE id=?", (id_,)); self.commit()

    # ── Consequence steps (Del1-Del5 escalation chain) ────────────────────────
    def get_consequence_steps(self, consequence_id):
        """Return list of dicts: {step, text, ref_tag, node_key} sorted by step."""
        rows = self.conn.execute(
            "SELECT step, text, ref_tag, node_key FROM consequence_steps "
            "WHERE consequence_id=? ORDER BY step", (consequence_id,)).fetchall()
        return [dict(r) for r in rows]

    def set_consequence_steps(self, consequence_id, steps):
        """Replace all steps for a consequence.

        steps: list of dicts with keys step(int), text(str), ref_tag(str)
        and optional node_key(str) — the consequence-graph node id, used to
        restore dependent column options when the chain is reopened.
        Empty text entries are omitted.
        """
        self.conn.execute(
            "DELETE FROM consequence_steps WHERE consequence_id=?",
            (consequence_id,))
        for s in steps:
            text = (s.get('text') or '').strip()
            if text:
                self.conn.execute(
                    "INSERT INTO consequence_steps "
                    "(consequence_id, step, text, ref_tag, node_key)"
                    " VALUES (?,?,?,?,?)",
                    (consequence_id, int(s['step']), text,
                     (s.get('ref_tag') or '').strip(),
                     (s.get('node_key') or '').strip()))
        self.commit()

    def consequence_steps_as_text(self, consequence_id):
        """Return 'Del1 → Del2 → …' string built from stored steps."""
        rows = self.get_consequence_steps(consequence_id)
        parts = [r['text'] for r in rows if r['text']]
        return ' → '.join(parts) if parts else ''

    def delete_safeguard(self, id_):
        # No FK cascade exists for consequence_severity_exclusions.safeguard_id
        self.conn.execute(
            "UPDATE lopa_source_scenarios SET source_missing=1 WHERE origin_safeguard_id=?", (id_,))
        self.conn.execute(
            "UPDATE lopa_barriers SET source_missing=1 WHERE source_safeguard_id=?", (id_,))
        self.conn.execute("DELETE FROM consequence_severity_exclusions WHERE safeguard_id=?", (id_,))
        self.conn.execute("DELETE FROM safeguards WHERE id=?", (id_,)); self.commit()

    # ── Reduction factors ─────────────────────────────────────────────────────
    def reduction_factors(self, consequence_id):
        return self.conn.execute(
            "SELECT * FROM reduction_factors WHERE consequence_id=? ORDER BY id",
            (consequence_id,)).fetchall()

    def reduction_factors_for_consequences(self, consequence_ids):
        """Bulk version of reduction_factors() — see _fetch_grouped
        (2026-08-24, NOTES.md). ScenarioTablePanel._add_row() used to call
        the single-id version once per RENDERED ROW, even though a
        consequence with several categories/safeguards produces several
        rows that all share the same consequence_id — re-fetching
        identical data every time."""
        return self._fetch_grouped('reduction_factors', 'consequence_id', consequence_ids)

    def add_reduction_factor(self, consequence_id, description='', rrf=10):
        cur = self.conn.execute(
            "INSERT INTO reduction_factors (consequence_id,description,rrf,active) VALUES (?,?,?,1)",
            (consequence_id, description, rrf))
        self.commit()
        self.remember_reduction_factor(description, rrf)
        return cur.lastrowid

    def update_reduction_factor(self, id_, description, rrf, active):
        self.conn.execute(
            "UPDATE reduction_factors SET description=?,rrf=?,active=? WHERE id=?",
            (description, rrf, int(active), id_))
        self.commit()
        self.remember_reduction_factor(description, rrf)

    def reduction_factor_catalog(self):
        return self.conn.execute(
            "SELECT * FROM reduction_factor_catalog WHERE active=1 "
            "ORDER BY description COLLATE NOCASE").fetchall()

    def retire_reduction_factor_catalog_entry(self, description):
        """Hide one user-defined enabler from future popup lists.

        Existing uses in other consequences are deliberately retained. A
        later explicit add with the same description reactivates the catalog
        row through ``remember_reduction_factor``.
        """
        description = str(description or '').strip()
        if not description:
            return
        self.conn.execute(
            "UPDATE reduction_factor_catalog SET active=0 "
            "WHERE lower(description)=lower(?)", (description,))
        self.commit()

    def remember_reduction_factor(self, description, rrf=10):
        description = str(description or '').strip()
        if not description:
            return
        self.conn.execute(
            "INSERT INTO reduction_factor_catalog(description,rrf,active) VALUES (?,?,1) "
            "ON CONFLICT(description) DO UPDATE SET rrf=excluded.rrf, active=1",
            (description, float(rrf)))
        self.commit()

    def delete_reduction_factor(self, id_):
        self.conn.execute(
            "DELETE FROM reduction_factor_severity_exclusions WHERE reduction_factor_id=?",
            (id_,))
        self.conn.execute("DELETE FROM reduction_factors WHERE id=?", (id_,))
        self.commit()

    def update_consequence_factors(self, id_, fa_active, fa_rrf, ignition_active, ignition_rrf):
        self.conn.execute(
            "UPDATE consequences SET fa_active=?,fa_rrf=?,ignition_active=?,ignition_rrf=? WHERE id=?",
            (int(fa_active), fa_rrf, int(ignition_active), ignition_rrf, id_))
        self.commit()

    # ── Copy support ──────────────────────────────────────────────────────────
    def copy_cause(self, cause_id, target_deviation_id):
        orig = self.get_cause(cause_id)
        if not orig:
            return None
        dev = self.get_deviation(target_deviation_id)
        node_id = dev['node_id'] if dev else orig['node_id']
        orig = dict(orig)
        cur = self.conn.execute(
            "INSERT INTO causes (node_id,deviation_id,description,likelihood,frequency_cleared,source_id,"
            "comp_type,comp_tag,equipment_id,secondary_equipment_id,group_equipment_ids,"
            "group_choices_set) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
            (node_id, target_deviation_id, orig['description'], orig['likelihood'],
             orig.get('frequency_cleared') or 0, cause_id,
             orig.get('comp_type') or '', orig.get('comp_tag') or '',
             orig.get('equipment_id'), orig.get('secondary_equipment_id'),
             orig.get('group_equipment_ids') or '', orig.get('group_choices_set') or 0))
        self.commit()
        return cur.lastrowid

    def copy_consequence(self, cons_id, target_cause_id):
        orig = self.get_consequence(cons_id)
        if not orig:
            return None
        cur = self.conn.execute(
            "INSERT INTO consequences (cause_id,description,severity,category,"
            "fa_active,fa_rrf,ignition_active,ignition_rrf,source_id) VALUES (?,?,?,?,?,?,?,?,?)",
            (target_cause_id, orig['description'], orig['severity'], orig['category'] or '',
             orig['fa_active'] or 0, orig['fa_rrf'] or 10,
             orig['ignition_active'] or 0, orig['ignition_rrf'] or 10, cons_id))
        self.commit()
        new_id = cur.lastrowid
        # Copy safeguards
        for sg in self.safeguards(cons_id):
            self.conn.execute(
                "INSERT INTO safeguards (consequence_id,description,rrf,sg_type,source_id) VALUES (?,?,?,?,?)",
                (new_id, sg['description'], sg['rrf'], dict(sg).get('sg_type','Övrigt'), sg['id']))
        # Copy reduction factors
        for rf in self.reduction_factors(cons_id):
            self.conn.execute(
                "INSERT INTO reduction_factors (consequence_id,description,rrf,active) VALUES (?,?,?,?)",
                (new_id, rf['description'], rf['rrf'], rf['active']))
        self.commit()
        return new_id

    def copy_safeguard(self, sg_id, target_cons_id):
        orig = self.get_safeguard(sg_id)
        if not orig:
            return None
        cur = self.conn.execute(
            "INSERT INTO safeguards (consequence_id,description,rrf,sg_type,source_id) VALUES (?,?,?,?,?)",
            (target_cons_id, orig['description'], orig['rrf'],
             dict(orig).get('sg_type', 'Övrigt'), sg_id))
        self.commit()
        return cur.lastrowid

    # ── Scoped HAZOP copying ─────────────────────────────────────────────────
    # The original copy_* helpers above predate per-category risk assessments,
    # enablers and globally shared recommendations.  Keep them for older call
    # sites, but let the worksheet/scenario use these transactional helpers so
    # a user can explicitly choose between a cell-only copy and a full branch.

    def _copy_sort_order(self, table, parent_column, parent_id):
        row = self.conn.execute(
            f"SELECT COALESCE(MAX(sort_order), -1) + 1 FROM {table} "
            f"WHERE {parent_column}=?", (parent_id,)).fetchone()
        return int(row[0] if row else 0)

    def _insert_copied_cause(self, original, target_deviation_id):
        target = self.get_deviation(target_deviation_id)
        if not target:
            return None
        original = dict(original)
        cur = self.conn.execute(
            "INSERT INTO causes "
            "(node_id,deviation_id,description,likelihood,sort_order,source_id,"
            "base_frequency,frequency_cleared,standard_cause_id,comp_type,comp_tag,comment,"
            "equipment_id,secondary_equipment_id,group_equipment_ids,group_choices_set) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (target['node_id'], target_deviation_id,
             original.get('description') or '', original.get('likelihood') or 0,
             self._copy_sort_order('causes', 'deviation_id', target_deviation_id),
             original.get('id'), original.get('base_frequency'),
             original.get('frequency_cleared') or 0,
             original.get('standard_cause_id'), original.get('comp_type') or '',
             original.get('comp_tag') or '', '', original.get('equipment_id'),
             original.get('secondary_equipment_id'),
             original.get('group_equipment_ids') or '',
             original.get('group_choices_set') or 0))
        return cur.lastrowid

    def _insert_copied_consequence(self, original, target_cause_id, cell_only=False):
        original = dict(original)
        if cell_only:
            severity, category = 1, ''
            fa_active, fa_rrf = 0, 10
            ignition_active, ignition_rrf = 0, 10
        else:
            severity = original.get('severity') or 1
            category = original.get('category') or ''
            fa_active, fa_rrf = original.get('fa_active') or 0, original.get('fa_rrf') or 10
            ignition_active = original.get('ignition_active') or 0
            ignition_rrf = original.get('ignition_rrf') or 10
        cur = self.conn.execute(
            "INSERT INTO consequences "
            "(cause_id,description,severity,category,consequence_chain,sort_order,source_id,"
            "fa_active,fa_rrf,ignition_active,ignition_rrf,comp_type,comp_tag,tagged_refs) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (target_cause_id, original.get('description') or '', severity, category,
             original.get('consequence_chain') or '',
             self._copy_sort_order('consequences', 'cause_id', target_cause_id),
             original.get('id'), fa_active, fa_rrf, ignition_active, ignition_rrf,
             original.get('comp_type') or '', original.get('comp_tag') or '',
             original.get('tagged_refs') or ''))
        return cur.lastrowid

    def _insert_copied_safeguard(self, original, target_consequence_id):
        original = dict(original)
        cur = self.conn.execute(
            "INSERT INTO safeguards "
            "(consequence_id,description,rrf,source_id,sort_order,sg_type,comp_type,comp_tag,tagged_refs) "
            "VALUES (?,?,?,?,?,?,?,?,?)",
            (target_consequence_id, original.get('description') or '',
             original.get('rrf') or 1, original.get('id'),
             self._copy_sort_order('safeguards', 'consequence_id', target_consequence_id),
             original.get('sg_type') or 'Övrigt', original.get('comp_type') or '',
             original.get('comp_tag') or '', original.get('tagged_refs') or ''))
        return cur.lastrowid

    def _copy_consequence_children(self, source_consequence_id, target_consequence_id,
                                   source_cause_id, target_cause_id):
        """Copy risk setup and descendants, translating all local IDs."""
        severity_map = {}
        for source in self.get_consequence_severities(source_consequence_id):
            source = dict(source)
            cur = self.conn.execute(
                "INSERT INTO consequence_severities(consequence_id,category_id,severity) "
                "VALUES (?,?,?)",
                (target_consequence_id, source['category_id'], source['severity']))
            severity_map[source['id']] = cur.lastrowid

        for source in self.get_final_consequence_severities(source_consequence_id):
            source = dict(source)
            self.conn.execute(
                "INSERT OR REPLACE INTO consequence_final_severities "
                "(consequence_id,category_id,severity) VALUES (?,?,?)",
                (target_consequence_id, source['category_id'], source['severity']))

        safeguard_map = {}
        for source in self.safeguards(source_consequence_id):
            source = dict(source)
            safeguard_map[source['id']] = self._insert_copied_safeguard(
                source, target_consequence_id)

        # Safeguard exclusions point to per-consequence severity records and
        # must therefore be translated only after both sets exist.
        for source_severity_id, target_severity_id in severity_map.items():
            for source_safeguard_id in self.get_severity_excluded_sgs(source_severity_id):
                target_safeguard_id = safeguard_map.get(source_safeguard_id)
                if target_safeguard_id:
                    self.conn.execute(
                        "INSERT OR IGNORE INTO consequence_severity_exclusions "
                        "(severity_id,safeguard_id) VALUES (?,?)",
                        (target_severity_id, target_safeguard_id))

        for source_safeguard_id, target_safeguard_id in safeguard_map.items():
            if source_cause_id in self.get_safeguard_excluded_causes(source_safeguard_id):
                self.conn.execute(
                    "INSERT OR IGNORE INTO safeguard_cause_exclusions "
                    "(safeguard_id,cause_id) VALUES (?,?)",
                    (target_safeguard_id, target_cause_id))

        factor_map = {}
        for source in self.reduction_factors(source_consequence_id):
            source = dict(source)
            cur = self.conn.execute(
                "INSERT INTO reduction_factors(consequence_id,description,rrf,active) "
                "VALUES (?,?,?,?)",
                (target_consequence_id, source.get('description') or '',
                 source.get('rrf') or 10, source.get('active') or 0))
            factor_map[source['id']] = cur.lastrowid

        for source_severity_id, target_severity_id in severity_map.items():
            for source_factor_id in self.get_severity_excluded_reduction_factors(
                    source_severity_id):
                target_factor_id = factor_map.get(source_factor_id)
                if target_factor_id:
                    self.conn.execute(
                        "INSERT OR IGNORE INTO reduction_factor_severity_exclusions "
                        "(severity_id,reduction_factor_id) VALUES (?,?)",
                        (target_severity_id, target_factor_id))

        # Recommendations are global numbered records. A branch copy creates
        # new links, never duplicate global recommendation numbers.
        for recommendation in self.recommendations_for_consequence(source_consequence_id):
            self.conn.execute(
                "INSERT OR IGNORE INTO consequence_recommendations "
                "(consequence_id,recommendation_id) VALUES (?,?)",
                (target_consequence_id, recommendation['id']))

    def copy_cause_scoped(self, cause_id, target_deviation_id, include_descendants=False):
        """Copy a cause to a deviation, optionally with its complete branch."""
        source = self.get_cause(cause_id)
        if not source or not self.get_deviation(target_deviation_id):
            return None
        source = dict(source)
        self.conn.execute('SAVEPOINT copy_cause_scoped')
        try:
            new_cause_id = self._insert_copied_cause(source, target_deviation_id)
            if include_descendants:
                for consequence in self.consequences(cause_id):
                    consequence = dict(consequence)
                    new_consequence_id = self._insert_copied_consequence(
                        consequence, new_cause_id)
                    self._copy_consequence_children(
                        consequence['id'], new_consequence_id,
                        cause_id, new_cause_id)
            self.conn.execute('RELEASE SAVEPOINT copy_cause_scoped')
            self.commit()
            return new_cause_id
        except Exception:
            self.conn.execute('ROLLBACK TO SAVEPOINT copy_cause_scoped')
            self.conn.execute('RELEASE SAVEPOINT copy_cause_scoped')
            raise

    def copy_consequence_scoped(self, consequence_id, target_cause_id,
                                include_descendants=False):
        """Copy a consequence to an existing cause with an explicit scope."""
        source = self.get_consequence(consequence_id)
        target_cause = self.get_cause(target_cause_id)
        if not source or not target_cause:
            return None
        source = dict(source)
        self.conn.execute('SAVEPOINT copy_consequence_scoped')
        try:
            new_consequence_id = self._insert_copied_consequence(
                source, target_cause_id, cell_only=not include_descendants)
            if include_descendants:
                self._copy_consequence_children(
                    consequence_id, new_consequence_id,
                    source['cause_id'], target_cause_id)
            self.conn.execute('RELEASE SAVEPOINT copy_consequence_scoped')
            self.commit()
            return new_consequence_id
        except Exception:
            self.conn.execute('ROLLBACK TO SAVEPOINT copy_consequence_scoped')
            self.conn.execute('RELEASE SAVEPOINT copy_consequence_scoped')
            raise

    def copy_safeguard_scoped(self, safeguard_id, target_consequence_id):
        """Copy a barrier and preserve applicable category/cause exclusions."""
        source = self.get_safeguard(safeguard_id)
        target_consequence = self.get_consequence(target_consequence_id)
        if not source or not target_consequence:
            return None
        source = dict(source)
        source_consequence_id = source['consequence_id']
        source_consequence = self.get_consequence(source_consequence_id)
        target_cause_id = target_consequence['cause_id']
        self.conn.execute('SAVEPOINT copy_safeguard_scoped')
        try:
            new_safeguard_id = self._insert_copied_safeguard(
                source, target_consequence_id)
            source_by_category = {
                row['category_id']: row['id']
                for row in self.get_consequence_severities(source_consequence_id)}
            target_by_category = {
                row['category_id']: row['id']
                for row in self.get_consequence_severities(target_consequence_id)}
            for category_id, source_severity_id in source_by_category.items():
                target_severity_id = target_by_category.get(category_id)
                if (target_severity_id and safeguard_id in
                        self.get_severity_excluded_sgs(source_severity_id)):
                    self.conn.execute(
                        "INSERT OR IGNORE INTO consequence_severity_exclusions "
                        "(severity_id,safeguard_id) VALUES (?,?)",
                        (target_severity_id, new_safeguard_id))
            source_cause_id = source_consequence['cause_id'] if source_consequence else None
            if (source_cause_id is not None and source_cause_id in
                    self.get_safeguard_excluded_causes(safeguard_id)):
                self.conn.execute(
                    "INSERT OR IGNORE INTO safeguard_cause_exclusions "
                    "(safeguard_id,cause_id) VALUES (?,?)",
                    (new_safeguard_id, target_cause_id))
            self.conn.execute('RELEASE SAVEPOINT copy_safeguard_scoped')
            self.commit()
            return new_safeguard_id
        except Exception:
            self.conn.execute('ROLLBACK TO SAVEPOINT copy_safeguard_scoped')
            self.conn.execute('RELEASE SAVEPOINT copy_safeguard_scoped')
            raise

    # ── Move support ──────────────────────────────────────────────────────────
    def move_cause(self, cause_id, target_node_id):
        self.conn.execute("UPDATE causes SET node_id=? WHERE id=?",
                          (target_node_id, cause_id))
        self.commit()

    def move_cause_to_deviation(self, cause_id, target_deviation_id):
        dev = self.get_deviation(target_deviation_id)
        if dev:
            self.conn.execute(
                "UPDATE causes SET deviation_id=?, node_id=? WHERE id=?",
                (target_deviation_id, dev['node_id'], cause_id))
            self.commit()

    def move_consequence(self, cons_id, target_cause_id):
        self.conn.execute("UPDATE consequences SET cause_id=? WHERE id=?",
                          (target_cause_id, cons_id))
        self.commit()

    def move_safeguard(self, sg_id, target_cons_id):
        self.conn.execute("UPDATE safeguards SET consequence_id=? WHERE id=?",
                          (target_cons_id, sg_id))
        self.commit()

    def set_sibling_order(self, table, parent_column, parent_id, ordered_ids):
        """Persist a user's drag-and-drop order for one tree level."""
        allowed = {
            'nodes': ('system_id', 'id'),
            'deviations': ('node_id', 'id'),
            'causes': ('deviation_id', 'id'),
            'consequences': ('cause_id', 'id'),
            'safeguards': ('consequence_id', 'id'),
        }
        expected_parent = allowed.get(table)
        if not expected_parent or expected_parent[0] != parent_column:
            raise ValueError('unsupported sibling order target')
        for order, row_id in enumerate(ordered_ids):
            self.conn.execute(
                f"UPDATE {table} SET sort_order=? WHERE id=? AND {parent_column} IS ?",
                (order, row_id, parent_id))
        self.commit()

    def stats(self):
        return {
            'nodes':        self.conn.execute("SELECT COUNT(*) FROM nodes").fetchone()[0],
            'causes':       self.conn.execute("SELECT COUNT(*) FROM causes").fetchone()[0],
            'consequences': self.conn.execute("SELECT COUNT(*) FROM consequences").fetchone()[0],
            'safeguards':   self.conn.execute("SELECT COUNT(*) FROM safeguards").fetchone()[0],
            'open_recommendations': self.conn.execute(
                "SELECT COUNT(*) FROM recommendations WHERE status='Öppen'").fetchone()[0],
        }

    def all_data(self):
        rows = []
        for node in self.nodes():
            for cause in self.causes(node['id']):
                for cons in self.consequences(cause['id']):
                    sgs = [dict(s) for s in self.safeguards(cons['id'])]
                    acts = [dict(a) for a in self.recommendations_for_consequence(cons['id'])]
                    rows.append({
                        'node_name':      node['name'],
                        'node_pid':       node['pid_ref'] or '',
                        'cause_id':       cause['id'],
                        'cause':          cause['description'],
                        'likelihood':     self.cause_frequency_level(cause),
                        'consequence_id': cons['id'],
                        'consequence':    cons['description'],
                        'severity':       cons['severity'],
                        'category':       cons['category'] or '',
                        'safeguards':     sgs,
                        'safeguards_text': '; '.join(s['description'] for s in sgs),
                        'actions':        acts,
                    })
        return rows


