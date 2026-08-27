#!/usr/bin/env python3
"""HAZOP Tool — Hazard and Operability Study Manager v2"""

import sys
import re
import json
import sqlite3
import math
import datetime
import logging
import traceback
from pathlib import Path
from functools import partial
import platform
import inspect

from constants import (
    CONFIG, SEV_LABELS, RRF_VALUES, RRF_LABELS, SG_TYPES, MARKUP_COLORS,
    RISK_ICON, NODE_T, CAUSE_T, CONS_T, SG_T, DEV_T, EQUIP_T, LEDORD_T, SYSTEM_T,
    DEVIATION_TYPES, _app_dir,
)
from database import (
    Database, DB_PATH, DEFAULT_MATRIX, _normalise_matrix, _risk_matrix_cache,
    load_matrix, get_matrix, risk_info, freq_to_f_level, DEFAULT_FREQ_BOUNDARIES,
    _STD_OBJECTS, parse_tag_refs, append_tag_to_text, add_tag_ref,
)
from pid_viewer import (
    PIDPanel, COMPONENT_TYPES, CONSEQUENCE_TEMPLATES, HAS_PYMUPDF,
    MODE_NAV, MODE_NODE, MODE_PICK_REF_TAG,
    scan_pdf_for_equipment, ocr_status, resolve_ocr_scan_choice, KNOWN_PREFIXES, invert_cause_text,
    _RED_MARKUP_SYMBOLS, _get_red_symbol_svg,
    _equip_prefix_from_tag,
    detect_equipment_symbols, EquipmentMarkerReviewDialog,
    apply_scan_result_to_equipment_catalog, upsert_identified_tags_from_scan,
    ParallelTagScanWorker, ParallelEquipmentAnalysisWorker,
    PageProgressDialog,
    FREQ_LABELS, freq_to_idx, idx_to_freq,
    _obj_type_matches,
    _mk_pm, _mk_icon, _icon, _EMOJI_ICON,
)
from ui_helpers import (
    freq_axis_label, freq_axis_label_full, cons_axis_label,
    _equipment_type_options, _lookup_comp_type_for_tag, _make_tag_completer,
    _resolve_std_deviation_id, _create_cause_from_pick, _EQ_TYPE_ITEMS,
    find_tag_bold_ranges, _draw_text_with_bold_tags,
    total_freq_reduction, CHAIN_ITEMS, build_consequence_text, parse_chain_from_json,
)
from tree_panel import (
    TreePanel, CauseObjectPopup, CauseTagPopup,
    RRFPopup, FrequencyPickerPopup,
)
from scenario_panel import (
    ScenarioTablePanel, RiskMatrixPopup, ConsequenceStepPickerDialog,
    ReductionFactorsDialog, _ScenarioDelegate, _PidDelegate, _LopaWidget,
    _CONSEQ_ENTRY, _CONSEQ_GENERIC_NEXT, _CONSEQ_NODES, _N_STEPS,
    _ORS_FIRST_LINE_H, _PID_ICON_W, _PLUS_BADGE_SIZE,
)
from equipment_panel import (
    EquipmentPanel, EquipmentTagPopup, ObjectPickerPopup, PIDAnalysisPanel,
    TagDatabasePanel, _EquipmentTableModel, _IdentifiedTagsModel, _tag_prefix,
)
from settings_panels import (
    HAZOPPreparationPanel, PIDManagementPanel, ParticipantMatrixPanel,
    SettingsPanel, StandardCausesSettingsPanel, StandardObjectsSettingsPanel,
    StudyManagementPanel,
)
from node_markup import (
    PropertiesRibbon, MarkupTablePanel, RedMarkupPanel,
)
from worksheet import HAZOPWorksheet
from recommendations_panel import RecommendationsPanel

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QSplitter, QScrollArea,
    QTreeWidget, QTreeWidgetItem, QTreeWidgetItemIterator, QStackedWidget,
    QTabWidget,
    QVBoxLayout, QHBoxLayout, QFormLayout, QGridLayout,
    QLineEdit, QTextEdit, QPlainTextEdit, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QTableView,
    QComboBox, QDialog, QDialogButtonBox, QDateEdit,
    QMessageBox, QFileDialog, QGroupBox,
    QMenu, QToolBar, QStatusBar, QSizePolicy,
    QSpinBox, QDoubleSpinBox, QSlider, QColorDialog, QFrame, QListWidget, QListWidgetItem,
    QProgressDialog, QAbstractItemView, QToolTip, QInputDialog, QCheckBox,
    QStyledItemDelegate, QStyleOptionViewItem, QStyle, QStyleOptionButton,
    QButtonGroup, QRadioButton, QToolButton,
)
from PyQt6.QtCore import (
    Qt, pyqtSignal, QSize, QPointF, QRectF, QRect, QPoint, QTimer, QMimeData, QEvent,
    QAbstractTableModel, QModelIndex, QSortFilterProxyModel, QDate,
)
from PyQt6.QtGui import QFont, QFontMetrics, QColor, QAction, QBrush, QPen, QPainter, QDrag, QPainterPath, QPixmap, QIcon, QPolygonF, QShortcut, QKeySequence, QCursor, QPalette, QTextLayout, QTextOption, QTextCharFormat

# ══════════════════════════════════════════════════════════════════════════════
# SPLASH SCREEN — STARTUP PROGRESS
# ══════════════════════════════════════════════════════════════════════════════

class SplashScreen(QWidget):
    """Modern splash screen with progress indicator during startup."""

    def __init__(self):
        super().__init__()
        self.setWindowFlags(Qt.WindowType.SplashScreen | Qt.WindowType.FramelessWindowHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setFixedSize(400, 300)

        # Center on screen
        screen = QApplication.primaryScreen()
        geometry = screen.geometry()
        x = (geometry.width() - 400) // 2
        y = (geometry.height() - 300) // 2
        self.move(x, y)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(20)
        layout.addStretch()

        # Logo/title
        title = QLabel("HAZOP Tool")
        title.setStyleSheet("font-size: 24px; font-weight: bold; color: #17191C;")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)

        # Subtitle
        subtitle = QLabel("Startar upp...")
        subtitle.setStyleSheet("font-size: 12px; color: #666666;")
        subtitle.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.subtitle = subtitle
        layout.addWidget(subtitle)

        # Spinner (simple rotating dots)
        spinner = QLabel("●  ○  ○")
        spinner.setStyleSheet("font-size: 16px; color: #17191C; letter-spacing: 8px;")
        spinner.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.spinner = spinner
        layout.addWidget(spinner)

        self._spinner_state = 0
        self._spinner_frames = [
            "●  ○  ○",
            "○  ●  ○",
            "○  ○  ●",
            "○  ●  ○",
        ]

        # Timer for spinner animation
        self._timer = QTimer()
        self._timer.timeout.connect(self._animate_spinner)
        self._timer.start(200)

        layout.addStretch()

    def _animate_spinner(self):
        self.spinner.setText(self._spinner_frames[self._spinner_state % len(self._spinner_frames)])
        self._spinner_state += 1

    def set_status(self, text):
        self.subtitle.setText(text)
        QApplication.processEvents()

    def close_splash(self):
        self._timer.stop()
        self.close()


# ══════════════════════════════════════════════════════════════════════════════
# CRASH REPORTING & DIAGNOSTICS
# ══════════════════════════════════════════════════════════════════════════════

class CrashReporter:
    """Automatic crash reporting with structured JSON output for easy analysis."""

    CRASH_DIR = _app_dir() / 'crashes'

    @classmethod
    def setup(cls):
        """Install crash handler for uncaught exceptions."""
        cls.CRASH_DIR.mkdir(exist_ok=True)
        # sys.excepthook itself is installed at module load time as
        # _global_exception_hook (defined right after this class) -- not
        # here, since this method only runs once __main__ calls it.
        logging.info(f"Crash reporting initialized: {cls.CRASH_DIR}")

    @classmethod
    def handle_exception(cls, exc_type, exc_value, exc_tb):
        """Catch uncaught exceptions and generate detailed crash report."""
        try:
            report = cls.generate_report(exc_type, exc_value, exc_tb)
            cls.save_report(report)
            cls.log_report(report)
        except Exception as e:
            # If crash reporter itself fails, fall back to stderr
            print(f"Failed to generate crash report: {e}", file=sys.stderr)
            traceback.print_exception(exc_type, exc_value, exc_tb, file=sys.stderr)

    @classmethod
    def generate_report(cls, exc_type, exc_value, exc_tb):
        """Generate structured crash report with all diagnostic data."""
        # Extract full traceback with frames
        tb_frames = []
        tb = exc_tb
        while tb is not None:
            frame = tb.tb_frame
            tb_frames.append({
                'filename': frame.f_code.co_filename,
                'function': frame.f_code.co_name,
                'lineno': tb.tb_lineno,
                'locals_preview': cls._format_locals(frame.f_locals),
            })
            tb = tb.tb_next

        report = {
            'timestamp': datetime.datetime.now().isoformat(),
            'exception': {
                'type': exc_type.__name__,
                'message': str(exc_value),
                'module': exc_type.__module__,
            },
            'traceback': {
                'frames': tb_frames,
                'full_text': ''.join(traceback.format_exception(exc_type, exc_value, exc_tb)),
            },
            'environment': {
                'python_version': platform.python_version(),
                'platform': platform.platform(),
                'machine': platform.machine(),
                'processor': platform.processor(),
            },
            'imports': cls._get_module_versions(),
        }
        return report

    @classmethod
    def _format_locals(cls, locals_dict, max_len=100):
        """Format local variables for crash report (safe truncation)."""
        result = {}
        for key, val in locals_dict.items():
            if key.startswith('_'):
                continue
            try:
                val_str = repr(val)
                if len(val_str) > max_len:
                    val_str = val_str[:max_len] + '...'
                result[key] = val_str
            except Exception:
                result[key] = f'<{type(val).__name__}>'
        return result

    @classmethod
    def _get_module_versions(cls):
        """Collect versions of key dependencies."""
        versions = {}
        for module_name in ['PyQt6', 'fitz', 'easyocr', 'rapidocr_onnxruntime', 'PIL']:
            try:
                mod = __import__(module_name.split('.')[0])
                if hasattr(mod, '__version__'):
                    versions[module_name] = mod.__version__
            except (ImportError, AttributeError):
                pass
        return versions

    @classmethod
    def save_report(cls, report):
        """Save crash report as JSON file."""
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        exc_type = report['exception']['type']
        filename = cls.CRASH_DIR / f'crash_{timestamp}_{exc_type}.json'

        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(report, f, indent=2, ensure_ascii=False)

        logging.error(f"Crash report saved to: {filename}")

    @classmethod
    def log_report(cls, report):
        """Log crash summary to console and file."""
        exc = report['exception']
        logging.error(f"=== CRASH REPORT ===")
        logging.error(f"Exception: {exc['type']}: {exc['message']}")
        logging.error(f"Location: {report['traceback']['frames'][-1]['filename']}:{report['traceback']['frames'][-1]['lineno']}")
        logging.error(f"Function: {report['traceback']['frames'][-1]['function']}")

    @classmethod
    def list_crashes(cls):
        """Return list of all crash reports, most recent first."""
        if not cls.CRASH_DIR.exists():
            return []
        crashes = sorted(cls.CRASH_DIR.glob('crash_*.json'), reverse=True)
        return crashes

    @classmethod
    def get_latest_crash_summary(cls):
        """Get summary of the most recent crash (for CLI debugging)."""
        crashes = cls.list_crashes()
        if not crashes:
            return None

        latest = crashes[0]
        try:
            with open(latest, 'r', encoding='utf-8') as f:
                report = json.load(f)

            exc = report['exception']
            frames = report['traceback']['frames']

            return {
                'file': str(latest),
                'timestamp': report['timestamp'],
                'type': exc['type'],
                'message': exc['message'],
                'location': f"{frames[-1]['filename']}:{frames[-1]['lineno']}",
                'function': frames[-1]['function'],
                'full_report': report,
            }
        except Exception as e:
            logging.error(f"Failed to read crash report {latest}: {e}")
            return None


# ══════════════════════════════════════════════════════════════════════════════
# GLOBAL EXCEPTION HOOK — keep the app alive when a Qt slot raises
# ══════════════════════════════════════════════════════════════════════════════
#
# In PyQt5.5+/PyQt6, an exception raised inside a Qt signal/slot callback
# (button click, tree selection, etc.) that is not caught anywhere propagates
# to sys.excepthook. If sys.excepthook is left at its default, PyQt prints the
# traceback and then aborts the whole process. Installing this hook means the
# offending slot call simply fails (the rest of that slot's code does not
# run), while the QApplication event loop keeps running so the user does not
# lose their whole session over one bad callback.
#
# This must be installed before app.exec() starts (module import time is
# early enough and keeps it in effect for the entire process lifetime).

def _global_exception_hook(exc_type, exc_value, exc_tb):
    """Replacement for sys.excepthook: log + report + inform the user,
    but never re-raise, sys.exit(), or os.abort() — the event loop must
    keep running afterwards."""

    # Let Ctrl+C behave normally (default handling), same as stock Python.
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_tb)
        return

    # Structured crash report (JSON under hazop/crashes/). Never allow a bug
    # in the crash reporter itself to take down the exception hook.
    try:
        CrashReporter.handle_exception(exc_type, exc_value, exc_tb)
    except Exception as e:
        logging.error(f"Failed to generate crash report: {e}")

    # Always also log the traceback via the standard logging module (this
    # matches the pattern used elsewhere in this file, e.g. the crash log
    # written to hazop_crash.log).
    msg = ''.join(traceback.format_exception(exc_type, exc_value, exc_tb))
    logging.error('UNHANDLED EXCEPTION IN SLOT\n%s', msg)

    # Show a dialog so the user knows something went wrong, but only if a
    # QApplication is actually running (otherwise there is nothing to show
    # a dialog on top of, e.g. exceptions during module import).
    try:
        from PyQt6.QtWidgets import QApplication, QMessageBox
        if QApplication.instance() is not None:
            try:
                crash_dir = CrashReporter.CRASH_DIR
                QMessageBox.critical(
                    None,
                    'Ett oväntat fel uppstod',
                    f'<b>{exc_type.__name__}:</b> {exc_value}<br><br>'
                    f'Programmet försöker fortsätta köra, men kontrollera '
                    f'att inget gick fel.<br><br>'
                    f'Detaljerad felrapport sparad i:<br>'
                    f'<code>{crash_dir}</code><br><br>'
                    f'Skicka innehållet från JSON-filen när du rapporterar felet.')
            except Exception:
                pass
    except Exception:
        pass


sys.excepthook = _global_exception_hook


def _configure_utf8_console_output():
    """Reconfigure stdout/stderr to UTF-8 so logging an emoji (the app's
    log/status messages are full of them, e.g. 🏭/🎯/🔍) never crashes.

    Windows' console/terminal streams default to the system codepage
    (cp1252 here) rather than UTF-8 unless PYTHONUTF8/PYTHONIOENCODING is
    set. The crash log's own FileHandler already passes encoding='utf-8'
    explicitly, but the console echo handler wrote straight to the
    unreconfigured sys.stderr — any emoji anywhere in a logged message
    then raised UnicodeEncodeError: 'charmap' codec can't encode character
    (real crash report, crash_20260807_115134_UnicodeEncodeError.json;
    trivially reproducible on this machine via print('\U0001f3ed')).
    errors='replace' rather than 'strict' so an exotic character can never
    crash the console echo again, even one outside cp1252 AND unexpected
    by whoever writes the next log message.
    """
    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, 'reconfigure'):
            try:
                stream.reconfigure(encoding='utf-8', errors='replace')
            except Exception:
                pass


# ══════════════════════════════════════════════════════════════════════════════
# WINDOWS 11 THEME — LJUST TEMA
# ══════════════════════════════════════════════════════════════════════════════

def _get_windows11_stylesheet():
    """Near-monochrome theme with one signal accent, matching the design mockup.

    Deliberately no `font-size` here (2026-08-11, bug report: "möjligheten
    att förstora och förminska texten" disappeared) — a universal `* {
    font-size: ... }` rule wins over ANY later QWidget.setFont() call on a
    matching widget (the same "QSS always wins over Qt::*Role" quirk
    already documented elsewhere in this file for background/foreground
    colors), which silently broke ScenarioTablePanel's "Textstorlek"
    spinbox. The app-wide default size is set via QApplication.setFont()
    in main() instead — that one DOES yield to a widget's own setFont().
    """
    return """
    * {
        font-family: "Segoe UI", -apple-system, BlinkMacSystemFont, sans-serif;
    }

    QMainWindow, QDialog, QWidget {
        background-color: #FBFBFA;
        color: #17191C;
    }

    QPushButton {
        background-color: #FFFFFF;
        color: #17191C;
        border: 1px solid #E2E3E1;
        border-radius: 4px;
        padding: 4px 10px;
        font-weight: 500;
    }
    QPushButton:hover {
        background-color: #F5F5F3;
        border-color: #CFD1CE;
    }
    QPushButton:pressed {
        background-color: #E8E9E6;
    }
    QPushButton:checked {
        background-color: #2F5FD0;
        color: #FFFFFF;
        border-color: #2F5FD0;
    }
    QPushButton:focus {
        outline: 2px solid #2F5FD0;
        outline-offset: 2px;
    }
    QPushButton:default {
        border-color: #2F5FD0;
    }

    QCheckBox, QRadioButton { color: #17191C; spacing: 6px; }
    QCheckBox::indicator, QRadioButton::indicator {
        width: 14px; height: 14px;
        border: 1px solid #CFD1CE;
        background-color: #FFFFFF;
    }
    QCheckBox::indicator { border-radius: 3px; }
    QRadioButton::indicator { border-radius: 7px; }
    QCheckBox::indicator:hover, QRadioButton::indicator:hover { border-color: #17191C; }
    QCheckBox::indicator:checked, QRadioButton::indicator:checked {
        background-color: #2F5FD0;
        border-color: #2F5FD0;
    }

    QSpinBox, QDoubleSpinBox {
        background-color: #FFFFFF;
        color: #17191C;
        border: 1px solid #CFD1CE;
        border-radius: 4px;
        padding: 3px 4px;
    }
    QSpinBox:focus, QDoubleSpinBox:focus { border: 2px solid #2F5FD0; }

    QListWidget, QListView {
        background-color: #FFFFFF;
        border: 1px solid #E2E3E1;
        border-radius: 4px;
    }
    QListWidget::item:hover, QListView::item:hover { background-color: #F5F5F3; }
    QListWidget::item:selected, QListView::item:selected {
        background-color: #E6ECFA;
        color: #17191C;
    }

    QTabWidget::pane { border: 1px solid #E2E3E1; border-radius: 4px; }
    QTabBar::tab {
        background-color: #F5F5F3;
        color: #17191C;
        padding: 5px 12px;
        border: 1px solid #E2E3E1;
        border-bottom: none;
        border-top-left-radius: 4px;
        border-top-right-radius: 4px;
    }
    QTabBar::tab:hover { background-color: #E8E9E6; }
    QTabBar::tab:selected {
        background-color: #FFFFFF;
        border-bottom: 2px solid #2F5FD0;
    }

    QLineEdit, QTextEdit, QPlainTextEdit {
        background-color: #FFFFFF;
        color: #17191C;
        border: 1px solid #CFD1CE;
        border-radius: 4px;
        padding: 4px 6px;
        selection-background-color: #2F5FD0;
        selection-color: #FFFFFF;
    }
    QLineEdit:focus, QTextEdit:focus, QPlainTextEdit:focus {
        border: 2px solid #2F5FD0;
        padding: 3px 5px;
    }

    QComboBox {
        background-color: #FFFFFF;
        color: #17191C;
        border: 1px solid #CFD1CE;
        border-radius: 4px;
        padding: 4px 8px;
    }
    QComboBox:focus { border: 2px solid #2F5FD0; }
    QComboBox::drop-down { border: none; width: 20px; }
    QComboBox::down-arrow { image: none; }

    QTableWidget, QTableView {
        background-color: #FFFFFF;
        alternate-background-color: #F5F5F3;
        gridline-color: #EEEFEC;
        border: 1px solid #E2E3E1;
        border-radius: 4px;
    }
    QTableWidget::item, QTableView::item { padding: 4px; border: none; }
    QTableWidget::item:selected, QTableView::item:selected {
        background-color: #E6ECFA;
        color: #17191C;
    }
    QHeaderView::section {
        background-color: #F5F5F3;
        color: #8D9299;
        padding: 4px;
        border: none;
        border-bottom: 1px solid #E2E3E1;
        font-weight: 600;
        font-size: 8pt;
        letter-spacing: 0.5px;
    }

    QTreeWidget, QTreeView {
        background-color: #FFFFFF;
        alternate-background-color: #F5F5F3;
        border: 1px solid #E2E3E1;
        border-radius: 4px;
    }
    QTreeWidget::item:hover, QTreeView::item:hover { background-color: #F5F5F3; }
    QTreeWidget::item:selected, QTreeView::item:selected {
        background-color: #E6ECFA;
        color: #17191C;
    }

    QFrame { background-color: #FFFFFF; border: none; }
    QGroupBox {
        color: #17191C;
        border: 1px solid #E2E3E1;
        border-radius: 6px;
        margin-top: 8px;
        padding-top: 8px;
    }
    QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 3px; color: #8D9299; }

    QScrollBar:vertical { background-color: #F5F5F3; border: none; width: 12px; }
    QScrollBar::handle:vertical { background-color: #CFD1CE; border-radius: 6px; margin: 2px; min-height: 20px; }
    QScrollBar::handle:vertical:hover { background-color: #B3B7B2; }
    QScrollBar:horizontal { background-color: #F5F5F3; border: none; height: 12px; }
    QScrollBar::handle:horizontal { background-color: #CFD1CE; border-radius: 6px; margin: 2px; min-width: 20px; }
    QScrollBar::handle:horizontal:hover { background-color: #B3B7B2; }
    QScrollBar::add-line, QScrollBar::sub-line { border: none; background: none; }

    QDialog { background-color: #FFFFFF; }
    QSplitter::handle { background-color: #E2E3E1; }
    QSplitter::handle:hover { background-color: #CFD1CE; }

    QMenuBar { background-color: #FFFFFF; color: #17191C; border-bottom: 1px solid #E2E3E1; }
    QMenuBar::item:selected { background-color: #F5F5F3; }
    QMenu { background-color: #FFFFFF; color: #17191C; border: 1px solid #CFD1CE; border-radius: 4px; }
    QMenu::item:selected { background-color: #F5F5F3; }

    QToolTip {
        background-color: #17191C;
        color: #FFFFFF;
        border: none;
        border-radius: 4px;
        padding: 4px 8px;
    }
    """


def _contrast_fg(bg_hex):
    """Return black or white text color for best contrast against bg_hex."""
    c = QColor(bg_hex)
    luminance = (0.299 * c.red() + 0.587 * c.green() + 0.114 * c.blue()) / 255
    return '#000000' if luminance > 0.55 else '#ffffff'


# Keep old alias so existing code that references LIKE_LABELS doesn't crash
LIKE_LABELS = FREQ_LABELS


def get_sev_labels():
    """Return severity labels from current matrix config (y_labels), falling back to SEV_LABELS."""
    cfg = get_matrix()
    y = cfg.get('y_labels', [])
    n = cfg.get('rows', 5)
    if y and len(y) >= n:
        return [f"C{i+1} – {y[i]}" if not y[i].startswith('C') else y[i] for i in range(n)]
    return SEV_LABELS[:n] if n <= len(SEV_LABELS) else SEV_LABELS + [f"C{i+1}" for i in range(len(SEV_LABELS), n)]


def _get_node_color(node_id):
    """Get a unique color for a node based on its ID."""
    return MARKUP_COLORS[node_id % len(MARKUP_COLORS)]


def _create_tagged_cause(db, deviation_id, comp_type, comp_tag, equipment_id=None):
    """Create a new cause under deviation_id (only its equipment tag/type
    set) plus one empty consequence — used when an equipment marker is
    dropped directly onto a deviation in the HAZOP tree (2026-08-08, see
    NOTES.md). No popup: the description defaults to the same "Ny orsak"
    placeholder _create_cause_from_pick's own fallback uses (2026-08-10 —
    was blank, unified to match every other auto-created cause/
    consequence/safeguard's placeholder-text convention), immediately
    inline-editable/overtype-able.

    `equipment_id` (2026-08-13, see NOTES.md) links the cause's tag
    strip live to the equipment_catalog row it came from, when the
    caller already knows it (as every current caller does, from the
    marker it just resolved) — no text-matching needed here.
    Returns (cause_id, consequence_id).
    """
    new_id = db.add_cause(deviation_id)
    db.update_cause(new_id, description='Ny orsak', comp_type=comp_type, comp_tag=comp_tag,
                     equipment_id=equipment_id)
    cons_id = db.add_consequence(new_id)
    return new_id, cons_id


# ══════════════════════════════════════════════════════════════════════════════
# SHARED WIDGETS
# ══════════════════════════════════════════════════════════════════════════════

def _apply_shared_recommendation_description_update(db, parent, rec_id, consequence_id, new_description):
    """Update a recommendation's description, first asking whether a
    SHARED one (linked to more than one consequence) should be updated
    everywhere or forked into a new, independent copy for just this
    consequence. Shared by _RecommendationDetailDialog._save() and the
    inline REK-cell commit path (ScenarioTablePanel, see NOTES.md
    "Redigera rekommendationer direkt i HAZOP Scenario", 2026-08-26) so
    the "shared recommendation" rule only has one implementation.

    Returns the id of the recommendation the description actually ended
    up on (== rec_id for a direct update or a solo/"update all" case, a
    NEW id if forked), or None if the user cancelled (nothing written)."""
    count = db.recommendation_consequence_count(rec_id)
    if count > 1:
        box = QMessageBox(parent)
        box.setWindowTitle("Delad rekommendation")
        box.setText(
            f"Denna rekommendation används av flera konsekvenser ({count} st). "
            "Vill du uppdatera rekommendationen för samtliga?")
        box.setStandardButtons(
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            | QMessageBox.StandardButton.Cancel)
        box.button(QMessageBox.StandardButton.Yes).setText("Ja, uppdatera alla")
        box.button(QMessageBox.StandardButton.No).setText("Nej, bara denna")
        box.button(QMessageBox.StandardButton.Cancel).setText("Avbryt")
        box.setDefaultButton(QMessageBox.StandardButton.Cancel)
        reply = box.exec()
        if reply == QMessageBox.StandardButton.Cancel:
            return None
        if reply == QMessageBox.StandardButton.No:
            new_id = db.add_recommendation(new_description)
            db.unlink_recommendation_from_consequence(rec_id, consequence_id)
            db.link_recommendation_to_consequence(new_id, consequence_id)
            return new_id
        # Yes falls through to the direct update below.
    db.update_recommendation(rec_id, description=new_description)
    return rec_id


class _RecommendationDetailDialog(QDialog):
    """Small focused editor for ONE recommendation's fields (2026-08-25,
    see NOTES.md "Rekommendationshantering — delad katalog med
    återanvändning") — replaces ActionEditor's live-typing table cells,
    since saving now has to ask "update everywhere, or just here?"
    before anything is written when the recommendation is shared, which
    doesn't fit a cell that saves on every keystroke/focus-out."""

    def __init__(self, db, recommendation_id, consequence_id, parent=None):
        super().__init__(parent)
        self.db = db
        self.recommendation_id = recommendation_id
        self.consequence_id = consequence_id
        rec = db.get_recommendation(recommendation_id) or {}
        self.setWindowTitle(f"Redigera R-{recommendation_id:03d}")
        self.setMinimumWidth(380)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("Rekommendation:"))
        self._desc = QPlainTextEdit(rec.get('description', ''))
        self._desc.setFixedHeight(70)
        layout.addWidget(self._desc)

        row = QHBoxLayout()
        row.addWidget(QLabel("Ansvarig:"))
        self._resp = QLineEdit(rec.get('responsible', '') or '')
        row.addWidget(self._resp)
        row.addWidget(QLabel("Datum:"))
        self._due = QLineEdit(rec.get('due_date', '') or '')
        row.addWidget(self._due)
        layout.addLayout(row)

        row2 = QHBoxLayout()
        row2.addWidget(QLabel("Status:"))
        self._status = QComboBox()
        self._status.addItems(['Öppen', 'Pågår', 'Klar'])
        self._status.setCurrentText(rec.get('status', 'Öppen') or 'Öppen')
        row2.addWidget(self._status)
        row2.addStretch()
        layout.addLayout(row2)

        btns = QHBoxLayout()
        btns.addStretch()
        cancel_btn = QPushButton("Avbryt")
        cancel_btn.clicked.connect(self.reject)
        btns.addWidget(cancel_btn)
        save_btn = QPushButton("Spara")
        save_btn.setDefault(True)
        save_btn.clicked.connect(self._save)
        btns.addWidget(save_btn)
        layout.addLayout(btns)

    def _save(self):
        description = self._desc.toPlainText()
        responsible = self._resp.text()
        due_date    = self._due.text()
        status      = self._status.currentText()

        target_id = _apply_shared_recommendation_description_update(
            self.db, self, self.recommendation_id, self.consequence_id, description)
        if target_id is None:
            self.reject()
            return
        # The helper's own "description" write already landed on
        # target_id (rec_id itself, or a freshly forked row) -- fill in
        # the fields it doesn't know about.
        self.db.update_recommendation(target_id, responsible=responsible,
                                      due_date=due_date, status=status)
        self.accept()


# RecommendationEditorDialog (the big modal "type new / search / check
# to link" dialog) removed 2026-08-26, see NOTES.md "Redigera
# rekommendationer direkt i HAZOP Scenario" -- replaced by inline
# editing of the REK cell plus RecommendationAssistPopup
# (scenario_panel.py), the same "small popup shown alongside an inline
# editor" pattern StandardCauseSuggestPopup already established for the
# ORS column. _RecommendationDetailDialog above (responsible/due-date/
# status + the shared-recommendation prompt) is unrelated and unchanged
# -- it's reachable from the new popup's own "✎" button.


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT
# ══════════════════════════════════════════════════════════════════════════════

def export_excel(db: Database, filepath: str):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return False, "openpyxl saknas.\nKör: pip install openpyxl"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HAZOP"

    HEADERS = ['Nod', 'P&ID', 'Orsak', 'F', 'Konsekvens', 'C', 'Risknivå',
               'Kategori', 'Safeguards', 'Åtgärder']
    COL_WIDTHS = [20, 12, 32, 6, 32, 6, 14, 12, 40, 42]
    RISK_FILLS = {'Låg': 'C6EFCE', 'Medium': 'FFEB9C', 'Hög': 'FFC7CE', 'Kritisk': 'FF0000'}

    thin   = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wrap   = Alignment(vertical='top', wrap_text=True)

    for col, (hdr, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.alignment = center; cell.border = border
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    for r, row in enumerate(db.all_data(), 2):
        level, _, _ = risk_info(row['likelihood'], row['severity'])
        acts_str = '\n'.join(
            f"• {a['description']} | {a['responsible']} | {a['due_date']} | {a['status']}"
            for a in row['actions'])
        sg_str = '; '.join(
            f"{s['description']} (RRF{s['rrf']})" if s['rrf'] > 1 else s['description']
            for s in row['safeguards'])
        values = [row['node_name'], row['node_pid'], row['cause'], row['likelihood'],
                  row['consequence'], row['severity'], level,
                  row['category'], sg_str, acts_str]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            cell.alignment = center if c in (4, 6) else wrap
            if c == 7:
                fc = RISK_FILLS.get(level, 'FFFFFF')
                cell.fill = PatternFill(start_color=fc, end_color=fc, fill_type='solid')
        ws.row_dimensions[r].height = 36

    try:
        wb.save(filepath); return True, ""
    except Exception as e:
        return False, str(e)


def export_pdf(db: Database, filepath: str):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        return False, "reportlab saknas.\nKör: pip install reportlab"

    doc = SimpleDocTemplate(filepath, pagesize=landscape(A4),
                            leftMargin=10*mm, rightMargin=10*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    cs = ParagraphStyle('c', fontSize=7, leading=9)
    hs = ParagraphStyle('h', fontSize=8, leading=10, textColor=colors.white,
                        fontName='Helvetica-Bold')

    RISK_COLORS_PDF = {
        'Låg':     colors.HexColor('#27ae60'),
        'Medium':  colors.HexColor('#f39c12'),
        'Hög':     colors.HexColor('#e67e22'),
        'Kritisk': colors.HexColor('#e74c3c'),
    }

    headers = ['Nod / P&ID', 'Orsak', 'L', 'Konsekvens', 'S', 'Risknivå', 'Safeguards', 'Åtgärder']
    table_data = [[Paragraph(h, hs) for h in headers]]
    row_styles = []

    for i, row in enumerate(db.all_data(), 1):
        level, _, _ = risk_info(row['likelihood'], row['severity'])
        acts_str = '<br/>'.join(
            f"• {a['description']} ({a['status']})" for a in row['actions']) or '—'
        sg_str = '<br/>'.join(
            f"• {s['description']}" + (f" RRF{s['rrf']}" if s['rrf'] > 1 else '')
            for s in row['safeguards']) or '—'
        table_data.append([
            Paragraph(f"{row['node_name']}<br/><font size='6'>{row['node_pid']}</font>", cs),
            Paragraph(row['cause'], cs),
            Paragraph(str(row['likelihood']), cs),
            Paragraph(row['consequence'], cs),
            Paragraph(str(row['severity']), cs),
            Paragraph(f"<b>{level}</b><br/>F={row['likelihood']} C={row['severity']}", cs),
            Paragraph(sg_str, cs),
            Paragraph(acts_str, cs),
        ])
        rc = RISK_COLORS_PDF.get(level, colors.white)
        row_styles.append(('BACKGROUND', (5, i), (5, i), rc))
        row_styles.append(('TEXTCOLOR', (5, i), (5, i), colors.white))

    col_widths = [32*mm, 40*mm, 8*mm, 45*mm, 8*mm, 22*mm, 55*mm, 55*mm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f5f5f5'), colors.white]),
    ] + row_styles))

    elements = [Paragraph("HAZOP — Rapport", styles['Title']), Spacer(1, 5*mm), t]
    try:
        doc.build(elements); return True, ""
    except Exception as e:
        return False, str(e)


# ── Feature 19: Global search dialog ──────────────────────────────────────────
class GlobalSearchDialog(QDialog):
    """Ctrl+F floating search across all nodes, causes, consequences, safeguards."""
    navigate_requested = pyqtSignal(int, int)   # (type_, id_)

    def __init__(self, db, parent=None):
        super().__init__(parent)
        self._db = db
        self.setWindowTitle("Sök i HAZOP-data")
        self.setWindowFlags(
            Qt.WindowType.Dialog | Qt.WindowType.WindowStaysOnTopHint)
        self.setMinimumWidth(520)
        self.setMinimumHeight(CONFIG['H_PANEL_MAX_ALT'])
        lay = QVBoxLayout(self)
        lay.setSpacing(6); lay.setContentsMargins(8, 8, 8, 8)

        row = QHBoxLayout()
        self._edit = QLineEdit()
        self._edit.setPlaceholderText("Sök i noder, orsaker, konsekvenser, barriärer…")
        self._edit.textChanged.connect(self._search)
        self._edit.setClearButtonEnabled(True)
        row.addWidget(self._edit)
        lay.addLayout(row)

        self._list = QListWidget()
        self._list.setAlternatingRowColors(True)
        self._list.itemDoubleClicked.connect(self._navigate)
        lay.addWidget(self._list)

        self._count = QLabel('')
        self._count.setStyleSheet("color:#888; font-size:10px;")
        lay.addWidget(self._count)

        close_btn = QPushButton("Stäng")
        close_btn.clicked.connect(self.close)
        lay.addWidget(close_btn)

        self._edit.setFocus()
        self._edit.installEventFilter(self)

    def eventFilter(self, obj, event):
        from PyQt6.QtCore import QEvent
        if obj is self._edit and event.type() == QEvent.Type.KeyPress:
            if event.key() in (Qt.Key.Key_Down, Qt.Key.Key_Return, Qt.Key.Key_Enter):
                self._list.setFocus()
                if self._list.count():
                    self._list.setCurrentRow(0)
                if event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter) and self._list.currentItem():
                    self._navigate(self._list.currentItem())
                return True
        return super().eventFilter(obj, event)

    def _search(self, text):
        self._list.clear()
        q = text.strip()
        if len(q) < 2:
            self._count.setText('')
            return
        q_low = q.lower()
        results = []
        for node in self._db.nodes():
            nd = dict(node)
            if q_low in nd['name'].lower():
                results.append((NODE_T, nd['id'], nd['name'], f"🏭 Nod: {nd['name']}"))
            for dev in self._db.deviations(nd['id']):
                for c in self._db.causes_for_deviation(dev['id']):
                    cd = dict(c)
                    if q_low in cd['description'].lower():
                        results.append((CAUSE_T, cd['id'], cd['description'],
                                        f"⚙ {nd['name']} / {dev['description'][:30]}: {cd['description']}"))
                    for cons in self._db.consequences(cd['id']):
                        kd = dict(cons)
                        if q_low in kd['description'].lower():
                            results.append((CONS_T, kd['id'], kd['description'],
                                            f"⚠ {nd['name']} / {cd['description'][:25]}: {kd['description']}"))
                        for sg in self._db.safeguards(kd['id']):
                            sd = dict(sg)
                            if q_low in sd['description'].lower():
                                results.append((SG_T, sd['id'], sd['description'],
                                                f"🛡 {nd['name']} / {cd['description'][:20]}: {sd['description']}"))
            if len(results) > 200:
                break
        for type_, id_, _, label in results[:200]:
            item = QListWidgetItem(label)
            item.setData(Qt.ItemDataRole.UserRole, type_)
            item.setData(Qt.ItemDataRole.UserRole + 1, id_)
            self._list.addItem(item)
        self._count.setText(f"{min(len(results),200)} träffar" + (" (begränsat till 200)" if len(results)>200 else ""))

    def _navigate(self, item):
        if item is None: return
        type_ = item.data(Qt.ItemDataRole.UserRole)
        id_   = item.data(Qt.ItemDataRole.UserRole + 1)
        if type_ is not None and id_ is not None:
            self.navigate_requested.emit(type_, id_)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN WINDOW
# ══════════════════════════════════════════════════════════════════════════════

class MainWindow(QMainWindow):
    def __init__(self, hzp_path: str = None):
        super().__init__()
        self._hzp_path    = None   # path to the currently open .hzp file (or None)
        self._search_dialog = None
        self.db = Database()
        load_matrix(self.db)
        self._markup_undo_stack  = []
        self.resize(1440, 900)

        # ── Menu bar ──────────────────────────────────────────────────────────
        # Former toolbar row folded into menus, grouped by purpose:
        #   Fil          — project + Skriv ut
        #   Export       — Excel / PDF / Åtgärder
        #   Analys       — Statistik / Godkänn
        #   Inställningar — Mörkt läge
        # Node/deviation/delete actions moved to a button row in TreePanel
        # instead (see TreePanel.__init__), since they act on the tree
        # selection rather than the document as a whole.
        mb = self.menuBar()
        file_menu = mb.addMenu("Fil")
        file_menu.addAction(_icon('document'), "Nytt projekt",      self._hzp_new)
        file_menu.addAction(_icon('import'), "Öppna (.hzp)…",     self._hzp_open)
        file_menu.addSeparator()
        self._act_save = file_menu.addAction(_icon('save'), "Spara",         self._hzp_save)
        file_menu.addAction(_icon('save'), "Spara som…",         self._hzp_save_as)
        file_menu.addSeparator()
        file_menu.addAction(_icon('print'), "Skriv ut",           self._print_scenario_table)
        file_menu.addSeparator()
        file_menu.addAction(_icon('close'), "Avsluta",            self.close)

        export_menu = mb.addMenu("Export")
        export_menu.addAction(_icon('chart'), "Excel",           self._export_excel)
        export_menu.addAction(_icon('document'), "PDF",             self._export_pdf)
        export_menu.addAction(_icon('clipboard'), "Åtgärder",        self._export_actions_pdf)

        analysis_menu = mb.addMenu("Analys")
        analysis_menu.addAction(_icon('trend-chart'), "Statistik",     self._show_statistics)
        analysis_menu.addAction(_icon('check'), "Godkänn",       self._approve_node)

        settings_menu = mb.addMenu("Inställningar")
        settings_menu.addAction(_icon('moon'), "Mörkt läge",    self._toggle_dark_mode)

        self._update_title()

        # ── Status bar ────────────────────────────────────────────────────────
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage(f"Databas: {self.db.path}")

        # ── Central widget ────────────────────────────────────────────────────
        # Outer vertical splitter: top pane holds [nav rail | view_stack],
        # bottom pane holds the scenario/markup tables (P&ID page only). The
        # bottom pane spans the FULL window width — it is a sibling of the
        # top pane, not nested inside it — so the nav rail only extends as
        # far down as the tree/P&ID area, not behind the scenario table too.
        self._outer_splitter = QSplitter(Qt.Orientation.Vertical)
        self.setCentralWidget(self._outer_splitter)

        top_widget = QWidget()
        root_layout = QHBoxLayout(top_widget)
        root_layout.setContentsMargins(0, 0, 0, 0)
        root_layout.setSpacing(0)

        # Nav rail — narrow vertical strip of square icon buttons, far left
        toggle_bar = QWidget()
        toggle_bar.setFixedWidth(52)
        toggle_bar.setStyleSheet("background:#17191C;")
        toggle_lay = QVBoxLayout(toggle_bar)
        toggle_lay.setContentsMargins(6, 12, 6, 12)
        toggle_lay.setSpacing(8)

        def _rail_icon(name):
            # Inverted two-state icon for the dark nav rail (2026-08-12,
            # see NOTES.md) — white on the dark unchecked background,
            # near-black on the white checked background. The opposite
            # color pairing from _mk_icon()'s own dark/white convention
            # (built for light-background toolbars), so built inline here
            # rather than reusing it.
            ic = QIcon()
            ic.addPixmap(_mk_pm(name, 18, QColor('#ffffff')), QIcon.Mode.Normal, QIcon.State.Off)
            ic.addPixmap(_mk_pm(name, 18, QColor('#17191C')), QIcon.Mode.Normal, QIcon.State.On)
            return ic

        # "HAZOP preparation" (2026-08-17, see NOTES.md) — Projekt/Deltagare/
        # Riskmatris & Kategorier/Standardorsaker, extracted out of
        # Inställningar into their own top-level page. Deliberately FIRST in
        # both the nav rail (visual order) AND view_stack (structural index
        # 0, per Anton's explicit request) — every other page's index shifts
        # +1 accordingly; see _switch_view and every other hardcoded
        # _switch_view(N) call elsewhere in this file, all renumbered
        # together with this change.
        #
        # "Rekommendationer" (2026-08-26, see NOTES.md) — inserted right
        # after Worksheet, becoming the new index 3; Utrustning/
        # Studiehantering/Inställningar shifted from 3/4/5 to 4/5/6
        # accordingly (again, _switch_view and every hardcoded
        # _switch_view(N) call site renumbered together with this change).
        self.btn_prep      = QPushButton()
        self.btn_pid       = QPushButton()
        self.btn_sheet     = QPushButton()
        self.btn_recommendations = QPushButton()
        self.btn_equip     = QPushButton()
        self.btn_admin     = QPushButton()
        self.btn_settings  = QPushButton()

        _nav_labels = {
            self.btn_prep:     "HAZOP preparation",
            self.btn_pid:      "P&ID-vy",
            self.btn_sheet:    "Worksheet",
            self.btn_recommendations: "Rekommendationer",
            self.btn_equip:    "Utrustning",
            self.btn_admin:    "Studiehantering",
            self.btn_settings: "Inställningar",
        }
        _nav_icons = {
            self.btn_prep:     'check',
            self.btn_pid:      'map',
            self.btn_sheet:    'clipboard',
            self.btn_recommendations: 'flag',
            self.btn_equip:    'bolt-nut',
            self.btn_admin:    'document',
            self.btn_settings: 'settings',
        }

        for btn in (self.btn_prep, self.btn_pid, self.btn_sheet, self.btn_recommendations,
                    self.btn_equip, self.btn_admin, self.btn_settings):
            btn.setCheckable(True)
            btn.setFixedSize(40, 40)
            btn.setToolTip(_nav_labels[btn])
            btn.setIcon(_rail_icon(_nav_icons[btn]))
            btn.setIconSize(QSize(18, 18))
            btn.setStyleSheet(
                "QPushButton{color:#fff;background:#2A2E34;border:none;"
                "border-radius:6px;font-size:16px;}"
                "QPushButton:hover{background:#383D45;}"
                "QPushButton:checked{background:#fff;color:#17191C;}")
            toggle_lay.addWidget(btn, alignment=Qt.AlignmentFlag.AlignHCenter)

        toggle_lay.addStretch()
        root_layout.addWidget(toggle_bar)

        # View stack (init BEFORE setChecked to prevent signal before object exists)
        self.view_stack = QStackedWidget()
        root_layout.addWidget(self.view_stack, 1)

        self._outer_splitter.addWidget(top_widget)

        self.btn_pid.setChecked(True)
        self.btn_prep.clicked.connect(lambda: self._switch_view(0))
        self.btn_pid.clicked.connect(lambda: self._switch_view(1))
        self.btn_sheet.clicked.connect(lambda: self._switch_view(2))
        self.btn_recommendations.clicked.connect(lambda: self._switch_view(3))
        self.btn_equip.clicked.connect(lambda: self._switch_view(4))
        self.btn_admin.clicked.connect(lambda: self._switch_view(5))
        self.btn_settings.clicked.connect(lambda: self._switch_view(6))

        # ── Page 0: HAZOP preparation ────────────────────────────────────────
        # Added FIRST so it becomes view_stack index 0 (QStackedWidget numbers
        # pages in addWidget() call order) — must precede every other
        # addWidget() call below, not just visually lead the nav rail.
        self.hazop_prep_panel = HAZOPPreparationPanel(self.db)
        self.view_stack.addWidget(self.hazop_prep_panel)

        # ── Page 1: P&ID view ─────────────────────────────────────────────────
        # No wrapper widget here — the view_stack page IS _h_splitter directly.
        # The scenario/markup tables that used to share a vertical splitter
        # with _h_splitter now live in self._v_splitter, added as the outer
        # splitter's own bottom pane (see below) so they span the full window
        # width instead of being indented by the nav rail.
        self._h_splitter = QSplitter(Qt.Orientation.Horizontal)

        self.tree_panel = TreePanel(self.db)
        self.tree_panel.setMinimumWidth(220)
        self.tree_panel.setMaximumWidth(340)
        self._h_splitter.addWidget(self.tree_panel)

        self.pid_panel = PIDPanel(self.db)
        self.pid_panel.setMinimumWidth(400)
        self._h_splitter.addWidget(self.pid_panel)

        # Narrow properties ribbon — also carries the P&ID node-markup
        # toolbar now (2026-08-19, see NOTES.md "Slå ihop nodmarkup i
        # nodinställningar"): the old, separate NodeMarkupPanel widget
        # that used to occupy its own _h_splitter slot right here is
        # gone — its buttons live inside props_ribbon's own NODE_T button
        # set, shown only while props_ribbon._markup_active.
        self.props_ribbon = PropertiesRibbon(self.db, main_window=self)
        self.props_ribbon.item_changed.connect(self._on_props_changed)
        self._h_splitter.addWidget(self.props_ribbon)

        # Never made visible (2026-08-26, see NOTES.md "Gör om Red
        # Markup-knappen") — kept only as the non-visual state/signal
        # object _on_place_symbol_requested/_on_red_markup_draw_finished
        # still need (open_symbol_picker()/get_current_style()/
        # get_symbol_dims(), tool_changed/symbol_selected signals). Still
        # added to the splitter (at its existing zero-width slot) rather
        # than left parentless, simplest way to keep it a normal,
        # cleanly-parented QWidget without extra plumbing.
        self.red_markup_panel = RedMarkupPanel(self.db)
        self.red_markup_panel.setVisible(False)
        self._h_splitter.addWidget(self.red_markup_panel)

        self._h_splitter.setSizes([260, 650, 62, 0])
        self.view_stack.addWidget(self._h_splitter)

        # Bottom pane of the OUTER splitter (full window width, below the
        # nav rail) — scenario table + the two markup-edit tables that swap
        # in during edit modes. Hidden entirely on non-P&ID pages (see
        # _switch_view) so those pages use the whole window height.
        self._v_splitter = QSplitter(Qt.Orientation.Vertical)

        self.scenario_panel = ScenarioTablePanel(self.db)
        # Same opt-in HAZOPWorksheet already uses (see always_show_deviation_column
        # docstring) — needed so load_node() from the equipment bar shows WHICH
        # deviation/equipment each row belongs to, not just the causes themselves.
        self.scenario_panel.always_show_deviation_column()
        self._v_splitter.addWidget(self.scenario_panel)

        self.markup_table_panel = MarkupTablePanel(self.db)
        self.markup_table_panel.setVisible(False)
        self._v_splitter.addWidget(self.markup_table_panel)

        # RedMarkupTablePanel (the "existing red markups" table this splitter
        # slot used to hold) was deleted outright 2026-08-26 (see NOTES.md
        # "Gör om Red Markup-knappen") along with the rest of the old Red
        # Markup view — see MainWindow._on_place_symbol_requested's docstring
        # for the resulting capability loss. Only two widgets share this
        # splitter now, not three.
        self._v_splitter.setSizes([220, 0])
        self._outer_splitter.addWidget(self._v_splitter)
        self._outer_splitter.setSizes([640, 220])

        # ── Page 2: Worksheet ─────────────────────────────────────────────────
        self.worksheet = HAZOPWorksheet(self.db)
        self.view_stack.addWidget(self.worksheet)

        # ── Page 3: Recommendations ("Rekommendationer", 2026-08-26) ──────────
        # Must be added right here — QStackedWidget numbers pages in
        # addWidget() call order, and this page is index 3 (right after
        # Worksheet, before Equipment), per the nav-rail renumbering above.
        self.recommendations_panel = RecommendationsPanel(self.db)
        self.view_stack.addWidget(self.recommendations_panel)

        # ── Page 4: Equipment ─────────────────────────────────────────────────
        self.equipment_panel = EquipmentPanel(self.db)
        self.equipment_panel.markers_saved.connect(self.pid_panel.reload_overlays)
        # An inline tag/type edit in the Utrustningsregister also reaches
        # equipment_catalog directly (2026-08-18, see NOTES.md "Objektets
        # identitet ...") — the tree's EQUIP_T rows and the scenario
        # table's ORS tag strip both resolve an object's identity LIVE
        # from that same table, so both must refresh here too, not just
        # the P&ID markers.
        self.equipment_panel.markers_saved.connect(self.scenario_panel.schedule_rebuild)
        self.equipment_panel.markers_saved.connect(self.tree_panel.refresh)
        self.view_stack.addWidget(self.equipment_panel)

        # ── Page 5: Study management ──────────────────────────────────────────
        self.admin_panel = StudyManagementPanel(self.db)
        self.view_stack.addWidget(self.admin_panel)

        # ── Page 6: Settings ──────────────────────────────────────────────────
        self.settings_panel = SettingsPanel(self.db)
        self.settings_panel.matrix_changed.connect(self._on_matrix_changed)
        self.hazop_prep_panel.matrix_changed.connect(self._on_matrix_changed)
        self.view_stack.addWidget(self.settings_panel)

        # ── Undo shortcut (Ctrl+Z) — only active during markup editing ────────
        self._undo_shortcut = QShortcut(QKeySequence("Ctrl+Z"), self)
        self._undo_shortcut.setEnabled(False)
        self._undo_shortcut.activated.connect(self._undo_last_markup)

        # ── Global search Ctrl+F (feature 19) ────────────────────────────────
        _search_sc = QShortcut(QKeySequence("Ctrl+F"), self)
        _search_sc.activated.connect(self._open_global_search)
        self._search_dialog = None

        # ── Wire signals ──────────────────────────────────────────────────────
        self.tree_panel.item_selected.connect(self._on_selected)
        self.tree_panel.structure_changed.connect(self._on_structure_changed)
        # Noder tab (HAZOPPreparationPanel, 2026-08-17) must reflect node
        # add/rename/delete that originated in the tree, not just its own edits.
        self.tree_panel.structure_changed.connect(self.hazop_prep_panel.refresh_nodes)
        # Inline tree editing (2026-08-17, see NOTES.md) — mirrors
        # scenario_panel.item_edited's own connection (_on_scenario_item_edited)
        # but in the opposite direction: an edit made IN THE TREE must also
        # refresh the scenario table + P&ID overlays (bold tag references,
        # markers/tooltips) since they all read from the same DB rows.
        self.tree_panel.item_edited_inline.connect(
            lambda type_, id_: (self.scenario_panel.refresh(),
                                 self.pid_panel.reload_overlays()))
        self.tree_panel.item_edited_inline.connect(self.hazop_prep_panel.refresh_nodes)
        self.tree_panel.visibility_changed.connect(
            lambda t, v: self.pid_panel.viewer.set_marker_visibility(t, v))

        self.scenario_panel.item_selected.connect(self._on_scenario_item_selected)
        self.scenario_panel.new_item_created.connect(
            lambda type_, id_: (
                # A scenario-created row belongs in the tree immediately, but
                # must NOT become the tree's current item. Selecting a hidden
                # consequence/safeguard made auto-collapse expand its complete
                # ancestor path, so a compact tree showing only "Ny nod"
                # suddenly unfolded after Enter. Rebuild without a target;
                # refresh() still preserves whatever the user manually had
                # open, while the scenario table handles its own cursor below.
                self.tree_panel.refresh(),
                self.scenario_panel.refresh(),
                # Land the editing cursor on the new item instead of leaving
                # selection wherever the rebuild happened to reset it to --
                # otherwise the user's view visibly "jumps away" from the row
                # they were just working on, making it hard to keep typing
                # the next cause/consequence in one flow.
                self.scenario_panel.select_item(type_, id_)))
        self.scenario_panel.item_edited.connect(self._on_scenario_item_edited)
        self.scenario_panel.structure_changed.connect(
            lambda: (self.tree_panel.refresh(), self.pid_panel.reload_overlays()))

        self.tree_panel.equipment_dropped_on_deviation.connect(
            self._on_equipment_dropped_on_deviation)
        self.tree_panel.edit_node_markup_requested.connect(self._on_edit_node_markup)
        self.tree_panel.node_markup_vis_requested.connect(self._on_node_markup_vis)
        self.tree_panel.node_jump_to_markup.connect(self._on_jump_to_node_markup)

        # Node markup toolbar signals (2026-08-19: merged into
        # PropertiesRibbon — see NOTES.md "Slå ihop nodmarkup i
        # nodinställningar"; same signal names/payloads as the old,
        # separate NodeMarkupPanel, just a different sender)
        self.props_ribbon.markup_mode_toggled.connect(self._on_markup_mode_toggled)
        self.props_ribbon.tool_changed.connect(
            lambda t: self.pid_panel.set_markup_tool(
                t, *self.props_ribbon.get_current_style()[:3]))
        self.props_ribbon.tool_changed.connect(self._on_node_markup_tool_activated)
        self.props_ribbon.all_vis_toggled.connect(
            lambda _: self.pid_panel.refresh_markup_overlays())
        self.props_ribbon.style_changed.connect(
            lambda color, opacity, width: self.pid_panel.viewer.set_pen_style(
                color, width, int(opacity * 210)))
        self.props_ribbon.snap_changed.connect(
            self.pid_panel.viewer.set_snap)
        self.props_ribbon.navigate_node_requested.connect(self._on_markup_navigate_node_requested)
        self.props_ribbon.bottom_panel_toggled.connect(self._on_toggle_bottom_panel)
        self.props_ribbon.place_symbol_requested.connect(self._on_place_symbol_requested)
        self._return_to_node_markup_node_id = None
        # Red markup ribbon signals
        self.red_markup_panel.closed.connect(self._on_close_red_markup)
        self.red_markup_panel.tool_changed.connect(
            lambda t: self.pid_panel.set_red_markup_tool(
                t, *self.red_markup_panel.get_current_style()[:3],
                symbol_id=self.red_markup_panel.get_selected_symbol()))
        self.red_markup_panel.symbol_selected.connect(
            lambda sid: self.pid_panel.set_red_markup_tool(
                'symbol', *self.red_markup_panel.get_current_style()[:3],
                symbol_id=sid))
        # Markup table panel signals
        self.markup_table_panel.item_deleted.connect(
            lambda _: self.pid_panel.refresh_markup_overlays())
        self.markup_table_panel.item_vis_toggled.connect(
            lambda mu_id, vis: self.pid_panel.viewer.set_markup_item_visible(mu_id, vis))
        self.markup_table_panel.item_selected.connect(
            lambda mu_id: self.pid_panel.viewer.highlight_markup(mu_id))
        # RedMarkupTablePanel and its signal wiring were deleted outright
        # 2026-08-26 (see NOTES.md "Gör om Red Markup-knappen") along with
        # the rest of the old Red Markup view.

        self.pid_panel.markup_label_edited.connect(self._on_markup_label_edited)
        self.pid_panel.markup_duplicate_requested.connect(self._on_duplicate_markup)
        self.markup_table_panel.item_style_changed.connect(
            lambda _: self.pid_panel.refresh_markup_overlays())
        self.markup_table_panel.item_duplicated.connect(self._on_duplicate_markup)
        self.pid_panel.markup_draw_finished.connect(self._on_markup_draw_finished)
        self.pid_panel.markup_moved.connect(self._on_markup_moved)
        self.pid_panel.markup_item_selected.connect(
            self.markup_table_panel.select_markup)
        self.pid_panel.red_markup_draw_finished.connect(self._on_red_markup_draw_finished)
        self.pid_panel.red_markup_moved.connect(self._on_red_markup_moved)
        # pid_panel.red_markup_item_selected used to drive
        # red_markup_table_panel.select_markup — that panel is gone
        # (2026-08-26, see NOTES.md "Gör om Red Markup-knappen"); the
        # signal itself is left defined/emitted in pid_panel_mod.py
        # (harmless with no listener) rather than touching that
        # heavily-tested module for this.
        self.pid_panel.markup_symbol_dims_changed.connect(self._on_markup_symbol_dims_changed)
        self.pid_panel.board_layout_changed.connect(self._on_board_layout_changed)
        self.tree_panel.exit_pid_mode_requested.connect(
            lambda: self.pid_panel._set_mode(MODE_NAV))

        self.pid_panel.node_created.connect(
            # emit_selection=False: _on_selected() is called explicitly right
            # after, so refresh()'s setCurrentItem must not also cascade via
            # currentItemChanged -> _on_select -> item_selected ->
            # _on_selected (same anti-pattern fixed in _on_marker_navigate,
            # commit 84c8b7c) — that would double-fire scenario_panel.load_node()
            # and schedule zoom_to_node() twice.
            lambda nid: (self.tree_panel.refresh(NODE_T, nid, emit_selection=False),
                         self._on_selected(NODE_T, nid)))
        self.pid_panel.ref_tag_picked.connect(self._on_ref_tag_picked)
        def _on_cause_template_created(cid):
            # emit_selection=False + explicit load_node (2026-08-07
            # follow-up — same "kan inte lägga till konsekvens" bug class
            # as _on_equipment_deviation_created, but triggered by CAUSE
            # creation this time, not deviation creation): refresh(CAUSE_T,
            # cid) used to run WITHOUT emit_selection=False, cascading into
            # _on_selected(CAUSE_T, cid) -> scenario_panel.load_deviation(...)
            # right as the user's very next move (picking a cause from
            # EquipmentDeviationBar's chip/dropdown, then immediately
            # clicking that row's KON cell to type) lands — narrowing/
            # rebuilding the worksheet mid-interaction. load_node() instead
            # keeps every deviation under the node visible, same as
            # _on_equipment_deviation_created already does.
            self.tree_panel.refresh(CAUSE_T, cid, emit_selection=False)
            cause = self.db.get_cause(cid)
            node_id = cause.get('node_id') if cause else None
            if node_id is not None:
                self.pid_panel.set_active_cause(cid)
                # Stay on the equipment filter if one is active (2026-08-12,
                # see NOTES.md: 'en mindre fix är att det skall se ut såhär
                # även när jag klickar på en rödmarkerad och lägger till
                # exempelvis lågt och högt flöde') — this handler also fires
                # for the auto-created cause right after
                # _on_equipment_deviation_created's own load_equipment()
                # call (checking a deviation box immediately suggests a
                # cause), so an unconditional load_node() here would
                # silently widen the just-filtered worksheet back out on
                # every single checkbox toggle.
                if self.scenario_panel.get_equipment_filter() is not None:
                    self.scenario_panel.refresh()
                else:
                    self.scenario_panel.load_node(node_id)
            self.scenario_panel.refresh_placed()
            # select_cause() itself refuses to steal the current cell from
            # an active edit / a row the user already navigated to (see
            # ScenarioTablePanel.select_cause) — this deferral just lets
            # the rebuild above settle before it scans _row_meta.
            QTimer.singleShot(0, lambda c=cid: self.scenario_panel.select_cause(c))
            # Refresh Smart Recognition panel so new learning is immediately visible
            try:
                self.settings_panel.refresh_tag_memory()
            except Exception:
                pass
        self.pid_panel.cause_template_created.connect(_on_cause_template_created)
        self.pid_panel.equipment_placement_requested.connect(self._on_equipment_placement_requested)
        self.pid_panel.equipment_edit_requested.connect(self._on_equipment_edit_requested)
        self.pid_panel.equipment_updated.connect(self._on_equipment_changed_from_marker)
        self.pid_panel.equipment_deleted.connect(self._on_equipment_changed_from_marker)
        self.pid_panel.marker_navigated.connect(self._on_marker_navigate)
        # Shift+click a marker while an ORS/KON/SG cell is being edited
        # inserts the tag into the open editor instead of navigating
        # away and destroying it (2026-08-13, see NOTES.md).
        self.pid_panel._active_edit_query_fn = self.scenario_panel.active_edit_target
        # Renaming an equipment via the ORS tag strip (2026-08-13, see
        # NOTES.md) reaches into equipment_catalog directly from
        # ScenarioTablePanel — keep the P&ID markers' own overlay text
        # AND the tree's EQUIP_T rows in sync right away too (2026-08-18,
        # see NOTES.md "Objektets identitet ..." — the tree used to only
        # pick this up on its next unrelated rebuild), not just P&ID.
        self.scenario_panel.equipment_renamed.connect(self.pid_panel.reload_overlays)
        self.scenario_panel.equipment_renamed.connect(self.tree_panel.refresh)
        self.pid_panel.equipment_deviation_created.connect(self._on_equipment_deviation_created)
        self.pid_panel.pid_analysis_done.connect(self._on_pid_analysis_done)
        self.admin_panel._pid_mgmt.sheets_changed.connect(self._on_sheets_changed)
        # Blad moved to HAZOPPreparationPanel (2026-08-17, see NOTES.md) — it
        # needs its own sheets_changed wired to the same reload, AND to
        # refresh when "Rensa samtliga P&ID" (still in Revisioner) wipes
        # sheets out from under it.
        self.hazop_prep_panel.sheets_changed.connect(self._on_sheets_changed)
        self.admin_panel._pid_mgmt.sheets_changed.connect(self.hazop_prep_panel.refresh_sheets)
        self.hazop_prep_panel.structure_changed.connect(self._on_hazop_prep_structure_changed)

        self._cur_type = None
        self._cur_id   = None

        self.tree_panel.refresh()
        self.pid_panel.try_reload_pdf()

        # Open a project passed on the command line (e.g. double-clicking a
        # .hzp file once it's registered as a Windows file association, see
        # NOTES.md "Paketera HAZOP-appen som en installationsfil") — this
        # parameter existed before but was never actually used; the window
        # always started on the default, empty hazop_project.db regardless
        # of what was passed in.
        if hzp_path:
            self._load_hzp(hzp_path)

    def _switch_view(self, page):
        prev = self.view_stack.currentIndex()
        self.view_stack.setCurrentIndex(page)
        # Bottom pane (scenario/markup tables) only makes sense on the P&ID
        # page — hidden elsewhere so those pages use the full window height
        # instead of leaving an empty strip below the nav rail's height.
        # Index 1, not 0, since HAZOP preparation (2026-08-17, see NOTES.md)
        # is now index 0 — every branch below shifted +1 accordingly.
        # 2026-08-26: Rekommendationer inserted as the new index 3, shifting
        # Utrustning/Studiehantering/Inställningar from 3/4/5 to 4/5/6.
        self._v_splitter.setVisible(page == 1)
        self.btn_prep.setChecked(page == 0)
        self.btn_pid.setChecked(page == 1)
        self.btn_sheet.setChecked(page == 2)
        self.btn_recommendations.setChecked(page == 3)
        self.btn_equip.setChecked(page == 4)
        self.btn_admin.setChecked(page == 5)
        self.btn_settings.setChecked(page == 6)
        if page == 1 and prev != 1:
            self.pid_panel.reload_overlays()
        if page == 2: self.worksheet.refresh()
        if page == 3: self.recommendations_panel.refresh()
        if page == 4: self.equipment_panel.refresh()
        if page == 5:
            self.admin_panel.refresh()
            self.admin_panel.refresh_pid()
        if page == 6:
            # Guard against settings_panel not being initialized yet
            if hasattr(self, 'settings_panel') and self.settings_panel:
                self.settings_panel.refresh_tag_memory()

    def _on_props_changed(self):
        """PropertiesRibbon saved a field — refresh tree + scenario."""
        if self._cur_type is not None and self._cur_id is not None:
            # emit_selection=False: the explicit scenario_panel._rebuild()
            # below already rebuilds the table for the current item, so
            # letting refresh()'s setCurrentItem cascade via
            # currentItemChanged -> _on_select -> item_selected ->
            # _on_selected would trigger a second, redundant
            # scenario_panel._rebuild() (same anti-pattern fixed in
            # _on_marker_navigate, commit 84c8b7c). This handler fires on
            # every properties-field save, so it is a frequent path.
            self.tree_panel.refresh(self._cur_type, self._cur_id, emit_selection=False)
            # A node rename here (PropertiesRibbon's Namn-popup) already
            # updated node_markups via Database.update_node() (2026-08-17,
            # see NOTES.md), but nothing else on this path re-draws the
            # P&ID — without this the on-canvas "Lägg ut nodnamn" label
            # stays visibly stale until some unrelated action happens to
            # trigger reload_overlays().
            if self._cur_type == NODE_T:
                self.pid_panel.refresh_markup_overlays()
        self.scenario_panel.rebuild()

    def _on_scenario_item_selected(self, type_, id_):
        """Scenario table cell click — update ribbon only; do NOT change scenario filter."""
        self._cur_type = type_
        self._cur_id   = id_
        self.props_ribbon.set_item(type_, id_)
        if type_ == CAUSE_T:
            self.pid_panel.set_active_cause(id_)
        elif type_ == CONS_T:
            self.pid_panel.set_active_consequence(id_)

    def _on_selected(self, type_, id_):
        self._cur_type = type_
        self._cur_id   = id_
        self.props_ribbon.set_item(type_, id_)
        # Markup-läget ska INTE längre öppnas automatiskt bara för att man
        # klickar på en nod i trädet (2026-08-25, se NOTES.md — reverterar
        # 2026-08-18-beteendet: "Om man klickar på en nod i trädet idag
        # försvinner hazop scenario och man kommer direkt in i
        # ritningläget på P&ID... jag behöver aktivt trycka på pennan till
        # höger"). Väljer man en nod medan markup-läget REDAN är aktivt
        # (pennan intryckt) rebinds det ändå till den nya noden — annars
        # skulle "ritar jag i noden skall det vara kopplat till noden jag
        # står på" (2026-08-18) sluta fungera vid navigering mellan noder
        # medan man faktiskt ritar. Stängs längst ned i denna metod så
        # fort valet lämnar Nod-nivån (Avvikelse, Orsak, Konsekvens,
        # Safeguard, eller något oväntat läge).
        if type_ == NODE_T:
            if self.props_ribbon._markup_active:
                self._on_edit_node_markup(id_)
            self.pid_panel.set_active_node(id_)
            self.scenario_panel.load_node(id_)
            if self.view_stack.currentIndex() == 0:
                QTimer.singleShot(80, lambda nid=id_: self.zoom_to_node(nid))
        elif type_ == DEV_T:
            self.pid_panel.set_active_deviation(id_)
            self.scenario_panel.load_deviation(id_)
        elif type_ == CAUSE_T:
            self.pid_panel.set_active_cause(id_)
            _row = self.db.get_cause(id_); cause = dict(_row) if _row else None
            if cause and cause.get('deviation_id'):
                self.scenario_panel.load_deviation(cause['deviation_id'])
            else:
                self.scenario_panel.load_cause(id_)
        elif type_ == CONS_T:
            self.pid_panel.set_active_consequence(id_)
            _cons = self.db.get_consequence(id_)
            cons = dict(_cons) if _cons else None
            if cons:
                _cause = self.db.get_cause(cons['cause_id']) if cons.get('cause_id') else None
                cause = dict(_cause) if _cause else None
                if cause and cause.get('deviation_id'):
                    self.scenario_panel.load_deviation(cause['deviation_id'])
                else:
                    self.scenario_panel.load_consequence(id_)
            else:
                self.scenario_panel.load_consequence(id_)
        elif type_ == SG_T:
            sg = self.db.get_safeguard(id_)
            if sg:
                cons = self.db.get_consequence(sg['consequence_id'])
                if cons:
                    self.pid_panel.set_active_consequence(cons['id'])
                    _cr = self.db.get_cause(cons['cause_id']) if cons.get('cause_id') else None
                    cause = dict(_cr) if _cr else None
                    if cause and cause.get('deviation_id'):
                        self.scenario_panel.load_deviation(cause['deviation_id'])
                    else:
                        self.scenario_panel.load_consequence(cons['id'])
        elif type_ == SYSTEM_T:
            # SYSTEM_T had no branch at all before 2026-08-27 (see
            # NOTES.md "Dynamisk färgmarkering av objekt på P&ID") —
            # selecting a System still isn't a valid cause/consequence/
            # safeguard PLACEMENT target (no scenario-table/zoom behavior
            # added here, out of scope for that feature), but it must
            # still clear whatever active_node/cause/consequence a
            # previous NODE_T/CAUSE_T click left set, same as every
            # other branch above keeps that state coherent via its own
            # set_active_*() call.
            self.pid_panel.clear_active_selection()
        # Tree-context P&ID highlight (2026-08-27, see NOTES.md) —
        # covers every branch above, including the new SYSTEM_T one.
        self.pid_panel.set_tree_context(type_, id_)
        if type_ != NODE_T and self.props_ribbon._markup_active:
            self._on_close_node_markup()

    def _on_scenario_item_edited(self, type_, id_):
        """Scenario table committed an edit — sync tree and P&ID labels.

        emit_selection=False: without it, tree_panel.refresh()'s
        setCurrentItem cascades via currentItemChanged -> _on_select ->
        item_selected -> _on_selected, which reloads the scenario panel and
        triggers a SECOND, redundant _rebuild() on every single cell edit
        (same anti-pattern already fixed for _on_marker_navigate,
        _on_safeguard_created, _on_props_changed and node_created). Besides
        the redundant work, this is what made the table visibly "jump away"
        after committing an edit -- the extra rebuild reset the current
        cell/selection a second time on top of whatever the first rebuild
        already did.
        """
        self.tree_panel.refresh(type_, id_, emit_selection=False)
        self.pid_panel.reload_overlays()
        # Tree-context highlight (2026-08-27, see NOTES.md) — the edit
        # itself may have changed the very tag data the highlight scope
        # depends on (e.g. dragging a new object tag onto this
        # consequence/safeguard), so force the expensive recompute
        # (set_tree_context), not just reload_overlays()'s own cheap
        # remap, even though the tree selection (self._cur_type/_cur_id)
        # itself hasn't changed.
        self.pid_panel.set_tree_context(self._cur_type, self._cur_id)

    def _on_structure_changed(self):
        self._cur_type = None
        self._cur_id   = None
        self.scenario_panel.clear()
        self.pid_panel.clear_active_selection()
        self.pid_panel.set_tree_context(None, None)
        self.pid_panel.reload_overlays()
        if self.props_ribbon._markup_active:
            self._on_close_node_markup()

    def _on_sheets_changed(self):
        """Reload the study board after sheets are added or deleted."""
        self.pid_panel.try_reload_pdf()

    def _on_hazop_prep_structure_changed(self):
        """A node was added/renamed from HAZOPPreparationPanel's Noder tab
        (2026-08-17) — the tree doesn't know about it on its own since the
        edit didn't originate there, unlike TreePanel's own structure_changed."""
        self.tree_panel.refresh()
        self._on_structure_changed()

    def _on_pid_analysis_done(self):
        """Switch to Settings → Identifierade objekt after P&ID analysis,
        then offer to chain straight into '🎯 Hitta objekt på P&ID' (2026-08-11):
        'Efter jag klickat på analysera P&ID vill jag få upp samma popupruta
        som innan, sedan vill jag att en popupfråga om jag vill hitta objekt
        på P&ID ska komma upp. Då ska samma körning som "hitta objekt på
        P&ID" knappen köras.' The 'Analys klar' popup itself is unchanged
        (shown earlier, in PIDPanel._analyze_pid, before this signal fires);
        this only adds the follow-up confirm + chained run."""
        self._switch_view(6)   # Settings page
        self.settings_panel.analysis_panel.refresh()
        # Switch to the "Identifierade objekt" tab inside settings
        tabs = self.settings_panel.findChild(QTabWidget)
        if tabs:
            for i in range(tabs.count()):
                if "Identifierade" in tabs.tabText(i):
                    tabs.setCurrentIndex(i)
                    break

        reply = QMessageBox.question(
            self, "Hitta objekt på P&ID",
            "Vill du köra 🎯 Hitta objekt på P&ID nu, med de tagg-prefix "
            "som just hittades?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes)
        if reply == QMessageBox.StandardButton.Yes:
            # Same run as clicking the button itself — refresh the register
            # model first so it reflects the tags the analysis just wrote.
            self.equipment_panel.refresh()
            self.equipment_panel.autodetect()

    def _on_marker_navigate(self, item_type: str, item_id: int):
        """Navigate tree and detail panel when a P&ID marker is clicked."""
        if item_type == 'equipment':
            # item_id here is equipment_markers.id (the marker row), not
            # equipment_catalog.id — look up which register row it links to.
            self._on_equipment_marker_navigate(item_id)
            return
        if item_type == 'equipment_register':
            # From EquipmentDeviationBar's "Visa i register" link — item_id
            # here IS already equipment_catalog.id (the bar knows it
            # directly, no marker lookup needed).
            self._switch_view(4)
            self.equipment_panel.select_row_by_equipment_id(item_id)
            return
        type_map = {'cause': CAUSE_T, 'consequence': CONS_T, 'safeguard': SG_T}
        t = type_map.get(item_type)
        if t is None:
            return
        # emit_selection=False: avoid double-firing _on_selected (and its
        # downstream scenario_panel._rebuild()) — setCurrentItem's cascade via
        # currentItemChanged is suppressed here since we call _on_selected
        # explicitly right below.
        self.tree_panel.refresh(t, item_id, emit_selection=False)
        self._on_selected(t, item_id)

    def _on_equipment_marker_navigate(self, marker_id: int):
        """Clicking an already-placed (red or green) equipment marker on the
        P&ID should surface ONLY the HAZOP scenario rows that mention it —
        causes, consequences AND safeguards (2026-08-12, see NOTES.md: 'de
        orsaker som visas i hazop scenario är de där objektet finns med').
        Filtered via scenario_panel.load_equipment(), not the whole node —
        an earlier version of this feature (2026-08-11) showed the entire
        node instead, which the user clarified was too broad.

        No _switch_view() call is needed: the marker can only be clicked
        while already on the P&ID page, and the scenario table is the
        bottom pane of that same page.
        """
        row = self.db.conn.execute(
            "SELECT equipment_id FROM equipment_markers WHERE id=?", (marker_id,)).fetchone()
        if not row or row['equipment_id'] is None:
            return
        equipment_id = row['equipment_id']
        node_id = self.db.equipment_node_id(equipment_id)
        if node_id is not None:
            self.tree_panel.refresh(NODE_T, node_id, emit_selection=False)
        self.scenario_panel.load_equipment(equipment_id)
        self.scenario_panel.refresh_placed()
        self.equipment_panel.select_row_by_equipment_id(equipment_id)

    def _on_equipment_deviation_created(self, deviation_id, equipment_id):
        """A deviation was checked (or unchecked) on in EquipmentDeviationBar
        — refresh the tree (new LEDORD_T/EQUIP_T/DEV_T items) and worksheet.

        emit_selection=False (2026-08-07 fix, real crash-adjacent bug
        report: "kan inte lägga till konsekvens"): the original version
        called tree_panel.refresh(DEV_T, deviation_id) with the default
        emit_selection=True, which cascades into _on_selected(DEV_T, ...)
        -> scenario_panel.load_deviation(...) — narrowing the worksheet to
        just THIS ONE deviation on every single checkbox toggle. That hid
        sibling deviations for the same node/equipment (the user's "I want
        to see BOTH deviations" complaint) and left keyboard focus inside
        the bar's comboboxes, so the worksheet's Enter-to-add-consequence
        shortcut never fired (the user's "I can't add a consequence"
        complaint) — same emit_selection anti-pattern already fixed
        elsewhere per commit 84c8b7c (see _on_props_changed,
        _on_safeguard_created, _on_marker_navigate).

        scenario_panel.load_equipment(), not load_node() (2026-08-12, see
        NOTES.md): if the worksheet was already filtered to this equipment
        (the user just clicked its red/green marker), checking a deviation
        box in the bar used to silently widen it back to the whole node —
        'en mindre fix är att det skall se ut såhär även när jag klickar på
        en rödmarkerad och lägger till exempelvis lågt och högt flöde. Då
        skall det enbart vara kopplat till det objektet.' load_equipment()
        still shows every deviation under THIS equipment together (same
        "not narrowed to one deviation" property load_node() had), just
        without pulling in the rest of the node's unrelated causes too.
        """
        self.tree_panel.refresh(DEV_T, deviation_id, emit_selection=False)
        self.scenario_panel.load_equipment(equipment_id)
        self.scenario_panel.refresh_placed()

    def _on_equipment_placement_requested(self, suggested_tag, scene_pos, page, pdf_rect=None):
        """P&ID right-click OR right-drag-rubber-band menu -> "🔧 Objekt"
        (2026-08-07/2026-08-09, see NOTES.md). pdf_rect (rubber-band case
        only) is threaded straight through to place_equipment_marker so
        the new marker gets a real outline shape.

        `suggested_tag` is always '' by the time this fires (2026-08-18,
        see NOTES.md "kombinerad placeringsmeny") — pid_panel_mod.py's
        _on_zone_drawn/_on_context_action no longer resolve a tag
        synchronously before emitting. place_equipment_marker() itself
        now shows the combined tag+typ+avvikelser popup (Equipment
        PlacementPopup) and starts the background tag search, so this
        handler just forwards straight through — it used to ALSO show
        EquipmentTagPopup first, which produced two dialogs in a row for
        the same placement (2026-08-18 bug report: "dubbla dialogfönster
        ... först en med objekt och sedan en med utrustning + objekttyp +
        avvikelser")."""
        self.pid_panel.place_equipment_marker(
            suggested_tag, '', scene_pos, page, pdf_rect=pdf_rect)

    def _on_equipment_edit_requested(self, marker_id):
        """Right-click "✏️ Redigera objekt" on an existing equipment marker
        (2026-08-12, see NOTES.md — reported feedback: "jag vill kunna
        högerklicka på ett objekt och kunna editera det, både tagnummer
        och vad det är för typ av utrustning"). Same EquipmentTagPopup as
        placing a brand new object, just pre-filled from and writing back
        to the existing equipment_catalog row instead of creating a new
        marker/placement."""
        equip = self.db.get_equipment_by_marker_id(marker_id)
        if not equip:
            QMessageBox.information(self, "Inget objekt",
                "Den här markören är inte kopplad till något registrerat objekt.")
            return
        popup = EquipmentTagPopup(self.db, suggested_tag=equip.get('tag') or '',
                                  suggested_type=equip.get('equipment_type') or '',
                                  parent=self)

        def _on_picked(tag, comp_type):
            tag = tag.strip().upper()
            pfx = _tag_prefix(tag) if tag else (equip.get('prefix') or '')
            self.db.update_equipment_item(
                equip['id'], tag, pfx, comp_type, equip.get('description') or '')
            self.pid_panel.reload_overlays()
            # Live tag link (2026-08-13, see NOTES.md) — any ORS tag
            # strip linked to this equipment resolves its display live
            # via causes.equipment_id, so a rebuild here shows the new
            # name immediately instead of on the next unrelated redraw.
            self.scenario_panel.schedule_rebuild()
            # The tree's EQUIP_T rows read equipment_catalog live too
            # (2026-08-18, see NOTES.md "Objektets identitet ...") — must
            # refresh here for the same reason, not just scenario/P&ID.
            self.tree_panel.refresh()

        popup.committed.connect(_on_picked)
        popup.exec()

    def _on_equipment_changed_from_marker(self, equipment_id):
        """EquipmentDeviationBar's own tag/typ edit or "Ta bort" (left-
        click an equipment marker, 2026-08-25, see NOTES.md) — PIDPanel
        already redrew its own overlays before emitting either signal
        (equipment_updated/equipment_deleted), so this only needs to
        refresh the other two places that resolve an object's identity
        live from equipment_catalog, same as _on_equipment_edit_requested's
        own refresh triplet above.

        Passes the tree's OWN currently selected (type, id) back into
        refresh() (2026-08-26, see NOTES.md "Behåll HAZOP-vyn när P&ID
        ändras") — refresh() already preserves every expanded node
        unconditionally (see its own `expanded` set at the top), but only
        re-selects/scrolls to an item when told which one via its
        select_type/select_id params; called bare (as this was) it always
        left the tree with nothing selected after "Ta bort" on a P&ID
        marker, even though the item the user was actually looking at
        (anything other than the just-deleted object itself) was still
        right there in the rebuilt tree.

        Also guards against the HAZOP Scenario view going blank: if the
        table is currently filtered to just this object's rows
        (ScenarioTablePanel.load_equipment(), triggered by clicking its
        P&ID marker) and the object was just deleted rather than merely
        edited, a plain rebuild would silently resolve to zero rows —
        load_equipment()'s own tag/type match has nothing left to match
        once the row is gone — instead of falling back to something
        useful."""
        if (self.scenario_panel.get_equipment_filter() == equipment_id
                and self.db.get_equipment_by_id(equipment_id) is None):
            self.scenario_panel.load_all()
        else:
            self.scenario_panel.schedule_rebuild()
        cur_type, cur_id = self.tree_panel._current()
        self.tree_panel.refresh(cur_type, cur_id)

    def _on_equipment_dropped_on_deviation(self, dev_id, marker_ids):
        """One or more equipment markers dragged from the P&ID onto a
        deviation in the HAZOP tree (2026-08-08, see NOTES.md) — creates
        one empty, tagged cause per marker directly (no popup), same
        immediate-inline-editable spirit as the worksheet/tree "+"
        auto-consequence feature. If the equipment has no node yet, it's
        assigned to the deviation's own node — same "slipp välja nod varje
        gång" convenience rule EquipmentDeviationBar.load() already uses.

        Also ties the FIRST dropped marker's equipment to the deviation
        itself, if it doesn't already have one (2026-08-09, see NOTES.md:
        "drar jag ett eller flera objekt till trädet skall även kolumnen
        utrustning fyllas i så det blir stringent, inte bara under
        orsak") — previously only the created CAUSE got tagged
        (comp_tag/comp_type, shown in the ORS column), leaving the
        worksheet's separate Utrustning column (driven by the
        deviation's own equipment_id, not the cause's tag) empty and
        inconsistent with the EquipmentDeviationBar checkbox flow, which
        always sets both."""
        dev = self.db.get_deviation(dev_id)
        if not dev:
            return
        node_id = dev['node_id']
        dev_equipment_id = dev['equipment_id']
        last_cause_id = None
        for marker_id in marker_ids:
            equip = self.db.get_equipment_by_marker_id(marker_id)
            if not equip:
                continue
            if equip.get('node_id') is None and node_id is not None:
                self.db.set_equipment_node(equip['id'], node_id)
            if dev_equipment_id is None:
                self.db.set_deviation_equipment(dev_id, equip['id'])
                dev_equipment_id = equip['id']
            cause_id, _cons_id = _create_tagged_cause(
                self.db, dev_id, equip.get('equipment_type', ''), equip.get('tag', ''),
                equipment_id=equip['id'])
            last_cause_id = cause_id
        if last_cause_id is not None:
            self.tree_panel.refresh(CAUSE_T, last_cause_id)
            self.scenario_panel.refresh_placed()
            # Scope the scenario view to just the cause the drop created
            # (2026-08-26, see NOTES.md "Filtrera orsaker i trädet") —
            # this used to call load_node(node_id), showing every
            # deviation/cause under the WHOLE node the object landed in;
            # reported feedback: "jag ser flera objekt [men] jag vill
            # bara se det objektet som precis dragits". load_cause()
            # narrows the table to just this one cause (and its own
            # consequences), same as clicking that cause anywhere else
            # in the tree already does.
            self.scenario_panel.load_cause(last_cause_id)

    def _open_consequence_step_picker(self, cons_id: int):
        """Open ConsequenceStepPickerDialog after a new consequence is created on P&ID."""
        logging.info('_open_consequence_step_picker: cons_id=%s', cons_id)
        try:
            cons = self.db.get_consequence(cons_id)
            if not cons:
                logging.warning('_open_consequence_step_picker: consequence %s not found in DB', cons_id)
                return
            cause = self.db.get_cause(cons['cause_id'])
            dev_desc = comp = cause_tx = ''
            if cause:
                cause_d = dict(cause)
                comp    = cause_d.get('comp_type', '') or ''
                cause_tx = cause_d.get('description', '') or ''
                dev_id  = cause_d.get('deviation_id')
                if dev_id:
                    dev = self.db.get_deviation(dev_id)
                    if dev:
                        dev_desc = dev['description'] or ''

            initial_tag = ''
            # If the consequence tag is known, look up the object type via
            # smart recognition so the dialog can pre-select the right category.
            if initial_tag and not comp:
                comp = _lookup_comp_type_for_tag(initial_tag, self.db)
            logging.info('_open_consequence_step_picker: creating dialog (dev=%r comp=%r tag=%r)', dev_desc, comp, initial_tag)

            dlg = ConsequenceStepPickerDialog(
                self.db, cons_id,
                deviation=dev_desc, comp_type=comp, cause_text=cause_tx,
                initial_ref_tag=initial_tag,
                parent=self)
            # Open next to the HAZOP scenario table's row for this consequence
            # (falls back to the current cursor position if the row isn't
            # visible in the table's current node/deviation/cause scope)
            # instead of the OS's default centered placement.
            dlg.move(self.scenario_panel.position_near_row(cons_id, dlg.sizeHint()))
            self._active_step_picker = dlg
            logging.info('_open_consequence_step_picker: calling dlg.exec()')
            result = dlg.exec()
            logging.info('_open_consequence_step_picker: dlg.exec() returned %s', result)
            if result == QDialog.DialogCode.Accepted:
                self._active_step_picker = None
                logging.info('_open_consequence_step_picker: accepted — calling scenario_panel.rebuild()')
                try:
                    self.scenario_panel.rebuild()
                    logging.info('_open_consequence_step_picker: _rebuild() done')
                except Exception:
                    logging.exception('_open_consequence_step_picker: CRASH in _rebuild()')
            else:
                logging.info('_open_consequence_step_picker: cancelled/rejected')
                self._active_step_picker = None
        except Exception:
            logging.exception('_open_consequence_step_picker: CRASH')

    def _on_ref_tag_picked(self, tag: str):
        """Called when user clicks P&ID in MODE_PICK_REF_TAG — fill the waiting column."""
        dlg = getattr(self, '_active_step_picker', None)
        if dlg is None or not dlg.isVisible():
            return
        col_idx = getattr(dlg, '_waiting_col_idx', None)
        if col_idx is not None and 0 <= col_idx < len(dlg._cols):
            dlg._cols[col_idx]['ref_edit'].setText(tag)
        dlg._waiting_col_idx = None
        dlg.show()
        dlg.raise_()
        dlg.activateWindow()

    def _on_edit_node_markup(self, node_id):
        """Entered explicitly only — tree right-click → 'Editera
        nodmarkup', the ✏️ toggle in props_ribbon, prev/next-node
        navigation, or returning from a red-markup symbol-placement
        detour. Also called from _on_selected, but ONLY while markup mode
        is already active, to rebind it to whatever node is now selected
        (2026-08-25, see NOTES.md — see that method's own comment); a
        plain node click while NOT already editing no longer reaches this
        method at all. Rebinding is idempotent — calling this again for a
        different node while already editing just re-targets everything
        at the new node, matching "ritar jag in något i noden skall detta
        vara kopplat till den noden jag står på" (2026-08-18, see
        NOTES.md).

        2026-08-17 (see NOTES.md "nodmarkup dockas till höger"): this used
        to hide tree_panel/props_ribbon/scenario_panel entirely, replacing
        almost the whole window with just the P&ID canvas + a narrow
        ribbon — Anton wanted the panel to feel docked alongside the rest
        of the app instead. tree_panel and props_ribbon now stay visible.

        2026-08-18 (see NOTES.md): the bottom strip does not force-swap to
        markup_table_panel here; HAZOP scenario stays visible until the
        user actually starts drawing (_on_node_markup_tool_activated),
        matching "HAZOP scenario fönstret ska fortsatt vara öppet om jag
        inte börjar använda någon av ritverktygen". The toggle button
        still lets the user flip between the two manually at any time.
        (2026-08-18's OTHER change — reaching this automatically from a
        plain node click with no explicit action — was reverted
        2026-08-25, see NOTES.md: "Om man klickar på en nod i trädet idag
        försvinner hazop scenario och man kommer direkt in i ritningläget
        på P&ID... jag behöver aktivt trycka på pennan till höger".)

        2026-08-19 (see NOTES.md "Slå ihop nodmarkup i nodinställningar"):
        the separate NodeMarkupPanel widget is gone — its toolbar now
        lives inside props_ribbon's own NODE_T button set. This method
        now also syncs props_ribbon's displayed item to node_id
        (set_item) — needed so the prev/next-node navigation buttons
        (which call straight into this method, bypassing _on_selected)
        keep the ribbon's plain node-settings fields pointed at the
        RIGHT node too, not just the markup toolbar; a real, pre-existing
        gap the old two-panel split had, invisible before since the two
        panels didn't share a container.

        Deliberately does NOT also sync tree_panel's own highlighted item
        here (a first attempt did, via tree_panel.refresh(...)) — this
        method can run REENTRANTLY from inside an in-progress
        tree_panel.refresh() call (e.g. _rename_node() refreshing →
        emitting item_selected → _on_selected → here), and a nested
        refresh() call while the outer one is still mid-rebuild deletes
        the very QTreeWidgetItem the outer call is still holding a
        reference to (RuntimeError: wrapped C/C++ object ... has been
        deleted — a real crash caught by test_regression.py's existing
        HAZOPPreparationBladNoderTests). See
        _on_markup_navigate_node_requested for where prev/next
        navigation's OWN tree-sync happens instead — safely, since a
        plain button click is never nested inside a refresh() call."""
        self._switch_view(1)
        self._cur_type = NODE_T
        self._cur_id   = node_id
        self.props_ribbon.set_item(NODE_T, node_id)
        self.markup_table_panel.load(node_id)
        self.props_ribbon.enter_markup_mode(node_id)
        self._h_splitter.setSizes([260, 650, 62, 0])
        self.pid_panel.enter_markup_edit(node_id)
        self._markup_undo_stack.clear()
        self._undo_shortcut.setEnabled(True)

    def _on_markup_navigate_node_requested(self, node_id):
        """props_ribbon's prev/next-node (⬆/⬇) buttons — syncs the tree's
        own highlighted item to node_id (safe here: always reached from a
        plain button click, never nested inside an in-progress
        tree_panel.refresh() call — see _on_edit_node_markup's own
        docstring for why that distinction matters) before rebinding the
        ribbon/markup toolbar to it."""
        self.tree_panel.refresh(NODE_T, node_id, emit_selection=False)
        self._on_edit_node_markup(node_id)

    def _on_node_markup_tool_activated(self, tool):
        """A real drawing tool (not the neutral 'select' tool, which also
        fires this signal on load/init) was picked in the node markup
        toolbar (2026-08-18, see NOTES.md) — swap the bottom strip to the
        Nodmarkeringar list so the user sees what they're drawing/have
        already drawn, matching "om jag klickar på exempelvis polygon, då
        ska jag istället se listan med nod markups". The ⇄ toggle can
        still bring HAZOP scenario back without leaving markup-edit mode."""
        if tool == 'select':
            return
        self.markup_table_panel.setVisible(True)
        self.scenario_panel.setVisible(False)
        self.props_ribbon.set_bottom_toggle_checked(False)
        self._v_splitter.setSizes([0, 200])
        self._outer_splitter.setSizes([560, 200])

    def _on_toggle_bottom_panel(self, checked):
        """Switch inside props_ribbon's markup toolbar (2026-08-17) —
        flips the bottom strip between Nodmarkeringar (default while
        editing) and HAZOP scenario, without leaving markup-edit mode."""
        self.markup_table_panel.setVisible(not checked)
        self.scenario_panel.setVisible(checked)
        self._v_splitter.setSizes([220, 0] if checked else [0, 200])

    def _on_markup_mode_toggled(self, checked):
        """props_ribbon's ✏️ toggle button (2026-08-19, replaces the old
        NodeMarkupPanel's one-shot "✕ Avsluta" button — see NOTES.md "Slå
        ihop nodmarkup i nodinställningar"): lets the user temporarily
        leave markup-edit mode — so P&ID marker clicks and the rubber-
        band placement gesture work again (both are disabled while markup
        mode has the canvas, see PIDPanel.enter_markup_edit) — without
        deselecting the node in the tree, and flip back into it again for
        the SAME node. The old close button had no way back in for the
        same node short of reselecting it."""
        if checked:
            node_id = self.props_ribbon.node_id
            if node_id is not None:
                self._on_edit_node_markup(node_id)
        else:
            self._on_close_node_markup()

    def _on_close_node_markup(self):
        """Markup-toggle switched off, or the tree selection moved off
        the Node level (see _on_selected) — leave markup edit mode."""
        self.pid_panel.exit_markup_mode()
        self.pid_panel.reload_overlays()
        self.props_ribbon.exit_markup_mode()
        self.scenario_panel.setVisible(True)
        self.markup_table_panel.setVisible(False)
        self._h_splitter.setSizes([260, 650, 62, 0])
        self._v_splitter.setSizes([220, 0])
        self._outer_splitter.setSizes([640, 220])
        self._markup_undo_stack.clear()
        self._undo_shortcut.setEnabled(False)

    def _on_place_symbol_requested(self):
        """PropertiesRibbon's "Lägg ut P&ID-symbol" button (2026-08-19,
        see NOTES.md "Slå ihop nodmarkup i nodinställningar").

        Reworked 2026-08-26 ("Gör om Red Markup-knappen", see NOTES.md):
        this used to route through _on_edit_red_markup, which opened the
        full old "Red Markup view" — a visible RedMarkupPanel ribbon plus
        a RedMarkupTablePanel list of existing red markups, revealed by
        resizing _h_splitter/_v_splitter/_outer_splitter and hiding
        scenario_panel. That whole view is torn down now:
        _on_edit_red_markup is deleted, RedMarkupTablePanel is deleted,
        and this method changes NO visible chrome at all — tree_panel,
        props_ribbon, scenario_panel and every splitter are left exactly
        as they were. The only visible effect of clicking this button is
        the small _SymbolSelectorPopup opening (via
        red_markup_panel.open_symbol_picker()).

        Underneath, pid_panel still needs to be told this placement is a
        red-markup (not node-markup) one — the two stayed separate state
        machines by design (see NOTES.md "Red markup konsolideras", lower
        regression risk than merging them) — so enter_red_markup_edit()
        must still run: it binds the active node
        (pid_panel.set_active_node, needed so the eventual DB save in
        _on_red_markup_draw_finished knows which node the symbol belongs
        to) and connects viewer.markup_draw_finished/markup_item_clicked
        (UniqueConnection-guarded) so a click on the canvas actually
        reaches _on_red_markup_draw_finished at all. Skipping this call
        would make drawing a symbol silently do nothing.

        _return_to_node_markup_node_id records which node to snap back to
        — the snap-back itself (_on_close_red_markup) is triggered from
        _on_red_markup_draw_finished right after a symbol is actually
        placed/saved, not from here (placing one symbol is the entire
        point of this brief detour).

        Judgement call on the OTHER way this could end — the user opens
        the popup and picks no symbol at all (dismisses it via outside
        click/Escape): deliberately left unhandled here rather than
        adding new popup-cancel signal plumbing to the heavily-tested
        picker/drawing code, because it already has a safe, natural way
        out — reselecting any node in the tree, the ⬆/⬇ node navigation,
        or the ✏️ toggle all force pid_panel back into a known state
        (_on_edit_node_markup / _on_close_node_markup), regardless of
        whether it was still sitting in the red-markup detour. The only
        practical effect of not picking a symbol is that the P&ID
        canvas's right-drag marker-placement gesture stays unavailable
        until one of those happens — the same trade-off already accepted
        for node-markup edit mode generally (see PIDPanel.enter_markup_edit),
        not a new regression introduced by this rework."""
        node_id = self.props_ribbon.node_id
        if node_id is None:
            return
        self._return_to_node_markup_node_id = node_id
        self.red_markup_panel.load(node_id)
        self.pid_panel.enter_red_markup_edit(node_id)
        self.red_markup_panel.open_symbol_picker()

    def _on_close_red_markup(self):
        """Leave red-markup edit mode. Reached automatically right after a
        symbol is placed (see _on_red_markup_draw_finished) — always
        returns to node markup editing for the same node now (see
        _on_place_symbol_requested) rather than closing everything, since
        the user only ever asked to place a symbol, not to leave node
        markup editing. RedMarkupPanel's own ✕ close button (the `closed`
        signal) still wires here too, defensively — it's permanently
        unreachable now that the panel is never shown (2026-08-26), but
        leaving the connection costs nothing.

        2026-08-26: no longer touches red_markup_table_panel (deleted
        along with the rest of the old Red Markup view — see NOTES.md
        "Gör om Red Markup-knappen"). The fallback branch below (no
        return target) still restores tree_panel/props_ribbon/
        scenario_panel visibility and the old splitter sizes for the one
        remaining caller with no return target
        (test_closing_red_markup_without_place_symbol_flow_goes_to_welcome) —
        harmless/idempotent since _on_place_symbol_requested no longer
        changes any of that chrome in the first place."""
        self.pid_panel.exit_red_markup_mode()
        self.pid_panel.reload_overlays()
        self.red_markup_panel.setVisible(False)
        return_node_id = self._return_to_node_markup_node_id
        self._return_to_node_markup_node_id = None
        if return_node_id is not None:
            self._on_edit_node_markup(return_node_id)
            # Placing a symbol is itself a drawing action — show the
            # Nodmarkeringar list with the new symbol on return, same as
            # picking any other drawing tool would (2026-08-18).
            self._on_node_markup_tool_activated('symbol')
            return
        self.tree_panel.setVisible(True)
        self.props_ribbon.setVisible(True)
        self.scenario_panel.setVisible(True)
        self._h_splitter.setSizes([260, 650, 62, 0])
        self._v_splitter.setSizes([220, 0])
        self._outer_splitter.setSizes([640, 220])

    def _on_red_markup_draw_finished(self, type_, node_id, pts, page, label):
        """New red markup drawn on P&ID — save to DB and refresh.

        2026-08-26: only 'symbol' can actually reach here now (see NOTES.md
        "Gör om Red Markup-knappen") — _on_place_symbol_requested is the
        sole entry point into red-markup mode and always leaves
        RedMarkupPanel's tool on 'symbol' (open_symbol_picker); its other
        drawing tools (polygon/polyline/comment/smart) were already
        removed 2026-08-17 (see NOTES.md "Red markup konsolideras"). The
        `else` branch is kept only as a defensive fallback for a code path
        that no longer exists, not a live one.

        Right after saving a symbol, immediately leaves red-markup mode
        and returns to node-markup editing (_on_close_red_markup) — this
        is the trigger the old RedMarkupPanel ✕ close button used to
        provide before that button became permanently unreachable (the
        panel is never shown anymore); placing one symbol is the entire
        point of this brief detour, so there is no "Red Markup view" left
        open to require a separate close action."""
        color, opacity, line_width, font_size = self.red_markup_panel.get_current_style()
        if type_ == 'symbol':
            symbol_id = label  # label holds the symbol_id for symbol type
            sw, sh, sr = self.red_markup_panel.get_symbol_dims()
            mu_id = self.db.add_node_red_markup(
                node_id, type_, pts, symbol_id, color, opacity, line_width,
                page, font_size, sw, sh, sr)
        else:
            mu_id = self.db.add_node_red_markup(
                node_id, type_, pts, label, color, opacity, line_width, page, font_size)
        self.pid_panel.viewer._pending_path_item = None
        self.pid_panel.refresh_red_markup_overlays()
        if type_ == 'symbol':
            self._on_close_red_markup()

    def _on_red_markup_moved(self, mu_id, new_pts):
        """Red markup item dragged to new position — save to DB."""
        self.db.update_node_red_markup(mu_id, points=new_pts)

    def _on_markup_symbol_dims_changed(self, mu_id, w, h, rot):
        """Symbol resized or rotated — save new dims to DB and re-render."""
        self.db.update_node_red_markup(mu_id, symbol_w=w, symbol_h=h, symbol_rot=rot)
        self.pid_panel.refresh_red_markup_overlays()

    def _on_board_layout_changed(self, layout_json):
        self.db.set_pid_config_value('board_layout', layout_json)

    def _on_node_markup_vis(self, node_id, visible):
        """Tree context menu hide/show all markups for a node."""
        self.db.set_all_node_markups_visible(node_id, visible)
        self.pid_panel.refresh_markup_overlays()
        if self.markup_table_panel.isVisible() and self.markup_table_panel.node_id == node_id:
            self.markup_table_panel.refresh()

    def _on_markup_draw_finished(self, type_, node_id, pts, page, label):
        """New markup drawn on P&ID — save to DB and refresh."""
        color, opacity, line_width, font_size = self.props_ribbon.get_current_style()
        mu_id = self.db.add_node_markup(
            node_id, type_, pts, label, color, opacity, line_width, page, font_size)
        self._markup_undo_stack.append({'op': 'draw', 'mu_id': mu_id})
        self.pid_panel.viewer._pending_path_item = None
        self.pid_panel.refresh_markup_overlays()
        self.markup_table_panel.refresh()
        self.markup_table_panel.select_markup(mu_id)

    def _on_markup_moved(self, mu_id, new_pts):
        """Markup item dragged to new position — save to DB, push undo entry."""
        red_row = self.db.get_node_red_markup(mu_id)
        if red_row:
            # Red markup item — route to red markup handler
            self._on_red_markup_moved(mu_id, new_pts)
            return
        old_row = self.db.get_node_markup(mu_id)
        if old_row:
            old_pts = json.loads(dict(old_row).get('points', '[]') or '[]')
            self._markup_undo_stack.append({'op': 'move', 'mu_id': mu_id, 'old_pts': old_pts})
        self.db.update_node_markup(mu_id, points=new_pts)
        self.markup_table_panel.refresh()

    def _on_markup_label_edited(self, mu_id, new_label):
        self.db.update_node_markup(mu_id, label=new_label)
        self.pid_panel.refresh_markup_overlays()
        self.markup_table_panel.refresh()

    def _on_jump_to_node_markup(self, node_id):
        """Double-click node with markups in tree → enter markup edit mode and zoom to items."""
        self._on_edit_node_markup(node_id)
        markups = self.db.node_markups_for_node(node_id)
        if not markups:
            return
        markups = [dict(m) for m in markups]
        mu_ids = [m['id'] for m in markups]

        # Navigate to the physical page of the first markup
        phys_page = markups[0].get('pid_page', 0)
        sheet_map = self.pid_panel._sheet_map  # display_index → physical_page
        if sheet_map:
            display_n = next((k for k, v in sheet_map.items() if v == phys_page), 0)
        else:
            display_n = phys_page
        self.pid_panel._goto_page(display_n)

        # Zoom to markup bounding box (overlays already loaded by enter_markup_edit)
        self.pid_panel.viewer.zoom_to_markup_items(mu_ids)

    def _on_duplicate_markup(self, mu_id):
        mu = self.db.get_node_markup(mu_id)
        if not mu:
            return
        mu = dict(mu)
        pts = json.loads(mu.get('points', '[]') or '[]')
        offset_pts = [[p[0] + 20, p[1] + 20] if len(p) >= 2 else p for p in pts]
        new_id = self.db.add_node_markup(
            node_id=mu['node_id'],
            type_=mu['type'],
            pts=offset_pts,
            label=mu['label'] + ' (kopia)',
            color=mu['color'],
            opacity=float(mu['opacity']),
            line_width=int(mu['line_width']),
            page=mu['pid_page'],
            font_size=int(mu['font_size']))
        self._markup_undo_stack.append({'op': 'draw', 'mu_id': new_id})
        self.pid_panel.refresh_markup_overlays()
        self.markup_table_panel.refresh()
        self.markup_table_panel.select_markup(new_id)

    def _undo_last_markup(self):
        if not self._markup_undo_stack:
            return
        entry = self._markup_undo_stack.pop()
        if entry['op'] == 'draw':
            self.db.delete_node_markup(entry['mu_id'])
            self.pid_panel.refresh_markup_overlays()
            self.markup_table_panel.refresh()
        elif entry['op'] == 'move':
            self.db.update_node_markup(entry['mu_id'], points=entry['old_pts'])
            self.pid_panel.refresh_markup_overlays()
            self.markup_table_panel.refresh()

    def _on_matrix_changed(self):
        self.tree_panel.refresh()
        if self._cur_type == CAUSE_T and self._cur_id:
            self.scenario_panel.load_cause(self._cur_id)

    def _export_excel(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Exportera Excel", "hazop_rapport.xlsx", "Excel (*.xlsx)")
        if not path: return
        ok, err = export_excel(self.db, path)
        if ok:
            self.status_bar.showMessage(f"Excel sparad: {path}", 6000)
            QMessageBox.information(self, "Klar", f"Exporterad till:\n{path}")
        else:
            QMessageBox.critical(self, "Fel vid export", err)

    def _export_pdf(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Exportera PDF", "hazop_rapport.pdf", "PDF (*.pdf)")
        if not path: return
        ok, err = export_pdf(self.db, path)
        if ok:
            self.status_bar.showMessage(f"PDF sparad: {path}", 6000)
            QMessageBox.information(self, "Klar", f"Exporterad till:\n{path}")
        else:
            QMessageBox.critical(self, "Fel vid export", err)

    # ── Dark mode ─────────────────────────────────────────────────────────────
    _dark_mode = False

    def _toggle_dark_mode(self):
        self._dark_mode = not self._dark_mode
        if self._dark_mode:
            p = QApplication.instance().palette()
            dark = QColor(30, 30, 30); mid = QColor(50, 50, 50)
            text = QColor(220, 220, 220); accent = QColor(99, 155, 255)
            p.setColor(p.ColorRole.Window,      dark)
            p.setColor(p.ColorRole.WindowText,  text)
            p.setColor(p.ColorRole.Base,        mid)
            p.setColor(p.ColorRole.AlternateBase, QColor(40, 40, 40))
            p.setColor(p.ColorRole.Text,        text)
            p.setColor(p.ColorRole.Button,      mid)
            p.setColor(p.ColorRole.ButtonText,  text)
            p.setColor(p.ColorRole.Highlight,   accent)
            p.setColor(p.ColorRole.HighlightedText, QColor(0,0,0))
            p.setColor(p.ColorRole.ToolTipBase, dark)
            p.setColor(p.ColorRole.ToolTipText, text)
            QApplication.instance().setPalette(p)
        else:
            QApplication.instance().setPalette(QApplication.style().standardPalette())

    # ── Excel export (IEC 61511 layout) ──────────────────────────────────────
    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.critical(self, "Saknar beroende",
                "openpyxl krävs: pip install openpyxl")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Exportera Excel", "hazop_rapport.xlsx", "Excel (*.xlsx)")
        if not path: return

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_fill = PatternFill("solid", fgColor="1F4E79")
        hdr_font = Font(color="FFFFFF", bold=True, size=10)
        hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        risk_colors = {'Låg': 'C6EFCE', 'Mellan': 'FFEB9C',
                       'Hög': 'FFC7CE', 'Kritisk': 'FF0000'}

        def make_sheet(wb, node):
            title = node['name'][:25].replace('/', '-').replace('\\', '-')
            ws = wb.create_sheet(title=title)
            ws.sheet_view.showGridLines = True
            cols = ['Nod', 'Avvikelse', 'Orsak', 'Konsekvens',
                    'Risk före', 'Barriär', 'RRF', 'Risk efter', 'Åtgärd']
            for ci, col in enumerate(cols, 1):
                c = ws.cell(row=1, column=ci, value=col)
                c.fill = hdr_fill; c.font = hdr_font
                c.alignment = hdr_align; c.border = border
            ws.row_dimensions[1].height = 28
            ws.column_dimensions['A'].width = 14
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 35
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 28
            ws.column_dimensions['G'].width = 8
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 30
            return ws

        for node in self.db.nodes():
            nd = dict(node)
            ws = make_sheet(wb, nd)
            r = 2
            for dev in self.db.deviations(nd['id']):
                for cause in self.db.causes_for_deviation(dev['id']):
                    cd = dict(cause)
                    cons_list = list(self.db.consequences(cd['id']))
                    if not cons_list:
                        ws.cell(r, 1, nd['name']); ws.cell(r, 2, dev['description'])
                        ws.cell(r, 3, cd['description']); r += 1
                        continue
                    for cons in cons_list:
                        kd = dict(cons)
                        sgs = list(self.db.safeguards(kd['id']))
                        if not sgs:
                            ws.cell(r, 1, nd['name']); ws.cell(r, 2, dev['description'])
                            ws.cell(r, 3, cd['description']); ws.cell(r, 4, kd['description'])
                            r += 1; continue
                        for sg in sgs:
                            sd = dict(sg)
                            ws.cell(r, 1, nd['name']); ws.cell(r, 2, dev['description'])
                            ws.cell(r, 3, cd['description']); ws.cell(r, 4, kd['description'])
                            ws.cell(r, 6, sd['description']); ws.cell(r, 7, sd.get('rrf', 1))
                            for c in range(1, 10):
                                ws.cell(r, c).border = border
                                ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical='top')
                            r += 1
            ws.freeze_panes = 'A2'
        wb.save(path)
        self.status_bar.showMessage(f"Excel sparad: {path}", 6000)
        QMessageBox.information(self, "Klar", f"Exporterad till:\n{path}")

    # ── Action report PDF ─────────────────────────────────────────────────────
    def _export_actions_pdf(self):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
        except ImportError:
            QMessageBox.critical(self, "Saknar beroende",
                "reportlab krävs: pip install reportlab")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Åtgärdsrapport", "hazop_atgarder.pdf", "PDF (*.pdf)")
        if not path: return

        styles = getSampleStyleSheet()
        doc = SimpleDocTemplate(path, pagesize=A4)
        elements = []
        elements.append(Paragraph("HAZOP Åtgärdsrapport", styles['Title']))
        elements.append(Spacer(1, 12))
        data = [['Nod', 'Orsak/Konsekvens', 'Åtgärd', 'Ansvarig', 'Deadline', 'Status']]
        for node in self.db.nodes():
            nd = dict(node)
            for dev in self.db.deviations(nd['id']):
                for cause in self.db.causes_for_deviation(dev['id']):
                    for cons in self.db.consequences(cause['id']):
                        for act in self.db.recommendations_for_consequence(cons['id']):
                            ad = dict(act)
                            data.append([
                                nd['name'][:20],
                                f"{cause['description'][:25]} → {cons['description'][:25]}",
                                ad.get('description', '')[:40],
                                ad.get('responsible', ''),
                                ad.get('due_date', ''),
                                ad.get('status', ''),
                            ])
        if len(data) == 1:
            elements.append(Paragraph("Inga åtgärder registrerade.", styles['Normal']))
        else:
            t = Table(data, colWidths=[80, 130, 140, 70, 60, 55])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E79')),
                ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
                ('FONTSIZE',   (0,0), (-1,-1), 7),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f0f4f8')]),
                ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ]))
            elements.append(t)
        doc.build(elements)
        self.status_bar.showMessage(f"Åtgärdsrapport sparad: {path}", 6000)
        QMessageBox.information(self, "Klar", f"Sparad till:\n{path}")

    # ── Print scenario table ──────────────────────────────────────────────────
    def _print_scenario_table(self):
        from PyQt6.QtPrintSupport import QPrinter, QPrintPreviewDialog
        from PyQt6.QtGui import QTextDocument, QPageSize, QPageLayout

        html = ['<html><body style="font-family:Arial;font-size:9pt;">',
                '<h2>HAZOP Scenariotabell</h2>',
                '<table border="1" cellspacing="0" cellpadding="3" width="100%">',
                '<tr style="background:#1F4E79;color:white;font-weight:bold;">',
                '<th>Nod</th><th>Avvikelse</th><th>Orsak</th>',
                '<th>Konsekvens</th><th>Barriär</th><th>Risk</th></tr>']
        alt = False
        for node in self.db.nodes():
            nd = dict(node)
            for dev in self.db.deviations(nd['id']):
                for cause in self.db.causes_for_deviation(dev['id']):
                    cons_list = list(self.db.consequences(cause['id']))
                    for cons in (cons_list or [None]):
                        kd = dict(cons) if cons else {}
                        sgs = list(self.db.safeguards(kd['id'])) if kd else []
                        bg = '#f8f9fa' if alt else '#ffffff'
                        html.append(f'<tr style="background:{bg}">')
                        html.append(f'<td>{nd["name"]}</td>')
                        html.append(f'<td>{dev["description"]}</td>')
                        html.append(f'<td>{cause["description"]}</td>')
                        html.append(f'<td>{kd.get("description","")}</td>')
                        html.append(f'<td>{"<br>".join(s["description"] for s in sgs)}</td>')
                        html.append(f'<td></td></tr>')
                        alt = not alt
        html.append('</table></body></html>')

        doc = QTextDocument()
        doc.setHtml(''.join(html))
        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        printer.setPageSize(QPageSize(QPageSize.PageSizeId.A4))
        printer.setPageOrientation(QPageLayout.Orientation.Landscape)
        preview = QPrintPreviewDialog(printer, self)
        preview.paintRequested.connect(doc.print)
        preview.exec()

    # ── Statistics ────────────────────────────────────────────────────────────
    def _show_statistics(self):
        db = self.db
        nodes  = list(db.nodes())
        n_devs = sum(len(list(db.deviations(n['id']))) for n in nodes)
        n_caus = db.conn.execute("SELECT COUNT(*) FROM causes").fetchone()[0]
        n_cons = db.conn.execute("SELECT COUNT(*) FROM consequences").fetchone()[0]
        n_sg   = db.conn.execute("SELECT COUNT(*) FROM safeguards").fetchone()[0]
        n_act  = db.conn.execute("SELECT COUNT(*) FROM recommendations").fetchone()[0]
        # Completeness: causes with ≥1 consequence with severity
        n_complete = db.conn.execute(
            "SELECT COUNT(DISTINCT c.id) FROM causes c "
            "JOIN consequences k ON k.cause_id=c.id "
            "WHERE k.severity > 0").fetchone()[0]
        pct = round(100 * n_complete / max(n_caus, 1))
        high_risk = db.conn.execute(
            "SELECT COUNT(*) FROM consequences WHERE severity >= 4").fetchone()[0]

        msg = (f"<h3>Projektstatistik</h3>"
               f"<table cellspacing='4'>"
               f"<tr><td><b>Noder:</b></td><td>{len(nodes)}</td></tr>"
               f"<tr><td><b>Avvikelser:</b></td><td>{n_devs}</td></tr>"
               f"<tr><td><b>Orsaker:</b></td><td>{n_caus}</td></tr>"
               f"<tr><td><b>Konsekvenser:</b></td><td>{n_cons}</td></tr>"
               f"<tr><td><b>Barriärer:</b></td><td>{n_sg}</td></tr>"
               f"<tr><td><b>Åtgärder:</b></td><td>{n_act}</td></tr>"
               f"<tr><td><b>Täckningsgrad:</b></td><td>{pct}% ({n_complete}/{n_caus})</td></tr>"
               f"<tr><td><b>Hög/Kritisk risk:</b></td>"
               f"<td style='color:#c0392b;font-weight:bold'>{high_risk}</td></tr>"
               f"</table>")
        box = QMessageBox(self)
        box.setWindowTitle("Statistik")
        box.setText(msg)
        box.setTextFormat(Qt.TextFormat.RichText)
        box.exec()

    # ── Sign-off / approval ───────────────────────────────────────────────────
    def _approve_node(self, node_id: int = None):
        if node_id is None:
            node_id = self._cur_id if self._cur_type == NODE_T else None
        if node_id is None:
            QMessageBox.information(self, "Välj nod",
                "Välj en nod i trädet innan du godkänner den.")
            return
        node = self.db.get_node(node_id)
        if not node:
            QMessageBox.information(self, "Nod saknas",
                "Noden finns inte längre (kan ha tagits bort).")
            return
        status = node['study_status'] if 'study_status' in node.keys() else 'draft'
        statuses = ['draft', 'under_review', 'approved']
        labels   = ['Utkast', 'Under granskning', 'Godkänd']
        choice, ok = QInputDialog.getItem(
            self, 'Status', f'Nod: {node["name"]}\nSätt status:',
            labels, statuses.index(status) if status in statuses else 0, False)
        if not ok: return
        new_status = statuses[labels.index(choice)]
        user = self.db.get_config('user_name', '') or 'okänd'
        self.db.set_node_status(node_id, new_status)
        if new_status == 'approved':
            self.db.approve_node(node_id, user)
        self.tree_panel.refresh(NODE_T, node_id)
        self.status_bar.showMessage(f"Status satt till: {choice}", 4000)

    # ── Feature 6: zoom to active node ────────────────────────────────────────
    def zoom_to_node(self, node_id):
        """Collect all markers for node_id and fit the board view to them."""
        if self.pid_panel.viewer.pdf_doc is None:
            return
        rs = self.pid_panel.viewer.render_scale
        offsets = self.pid_panel.viewer._page_offsets
        pts = []
        for c in self.db.causes_for_node_all(node_id):
            for m in self.db.cause_markers_for_cause(c['id']):
                ox, oy = offsets.get(m['pid_page'], (0.0, 0.0))
                pts.append((ox + m['x'] * rs, oy + m['y'] * rs))
        for cons in self.db.consequences_for_node(node_id):
            for m in self.db.conn.execute(
                    "SELECT pid_page,x,y FROM consequence_markers WHERE consequence_id=?",
                    (cons['id'],)).fetchall():
                ox, oy = offsets.get(m['pid_page'], (0.0, 0.0))
                pts.append((ox + m['x'] * rs, oy + m['y'] * rs))
        if not pts:
            return
        PAD = 150
        min_x = min(p[0] for p in pts) - PAD
        min_y = min(p[1] for p in pts) - PAD
        max_x = max(p[0] for p in pts) + PAD
        max_y = max(p[1] for p in pts) + PAD
        from PyQt6.QtCore import QRectF
        self.pid_panel.viewer.fitInView(
            QRectF(min_x, min_y, max_x - min_x, max_y - min_y),
            Qt.AspectRatioMode.KeepAspectRatio)
        self.pid_panel.viewer._apply_lod(self.pid_panel.viewer.transform().m11())
        self.pid_panel.viewer._schedule_lod_update()

    # ── Feature 19: global search ─────────────────────────────────────────────
    def _open_global_search(self):
        if self._search_dialog is not None:
            self._search_dialog.raise_()
            self._search_dialog.activateWindow()
            return
        dlg = GlobalSearchDialog(self.db, self)
        dlg.navigate_requested.connect(self._on_search_navigate)
        dlg.finished.connect(lambda _: setattr(self, '_search_dialog', None))
        self._search_dialog = dlg
        dlg.show()

    def _on_search_navigate(self, type_, id_):
        self._on_selected(type_, id_)
        self._switch_view(1)
        if type_ == CAUSE_T:
            c = self.db.get_cause(id_)
            if c:
                markers = self.db.cause_markers_for_cause(id_)
                if markers:
                    m = markers[0]
                    self.pid_panel.navigate_to_marker(m['pid_page'], m['x'], m['y'])


    # ══════════════════════════════════════════════════════════════════════════
    # HZP file format  (ZIP archive renamed to .hzp)
    # Contents:
    #   hazop_project.db   — the SQLite database
    #   pid/<filename>.pdf — the P&ID PDF (if one is loaded)
    #   meta.json          — {"hzp_version":1, "created":..., "pdf_name":...}
    # ══════════════════════════════════════════════════════════════════════════

    def _update_title(self):
        name = Path(self._hzp_path).name if self._hzp_path else "Osparad"
        self.setWindowTitle(f"HAZOP Tool  —  {name}")

    # Every project-scoped table, cleared by _wipe_project_tables() below.
    # Deliberately NOT listed (preserved across "New Project" on purpose):
    # standard_causes/standard_deviations/standard_objects (the reusable
    # template library), tag_database/tag_database_settings (taught
    # cross-project tag vocabulary), component_types/node_types/
    # failure_modes (static reference lookups), symbol_templates and
    # equipment_types (no clear per-project ownership, left alone to avoid
    # unrelated scope creep).
    _PROJECT_TABLES = [
        'nodes', 'deviations', 'causes', 'consequences', 'safeguards',
        'actions', 'cause_markers', 'consequence_markers', 'safeguard_markers',
        'study_tag_memory', 'symbol_fingerprints',
        'equipment_catalog', 'equipment_markers', 'pid_identified_tags',
        'pid_config',
        'off_page_connector', 'board_annotations', 'pid_connection',
        'node_markups', 'node_red_markups',
        # Deltagarmatris (2026-08-11) — replaces the old
        # 'project_participants' app_config field, see NOTES.md.
        'participants', 'analysis_sessions', 'participant_attendance',
        'participant_columns', 'participant_column_values',
        # Manuell sidrotation (2026-08-12), see NOTES.md.
        'pid_page_rotation',
        # Projekt-flikens revisionstabell + fria fält (2026-08-17), se NOTES.md.
        'project_revisions', 'project_custom_fields',
        # Konsekvenskategorier + severity-nivåer per kategori — per-
        # projekt, inte en delad mall (till skillnad från standard_*).
        'consequence_categories', 'severity_definitions',
        'consequence_severities', 'consequence_severity_exclusions',
        'safeguard_cause_exclusions', 'reduction_factors',
        # Rekommendationskatalog (2026-08-25/26), se NOTES.md.
        'recommendations', 'consequence_recommendations',
        # Konsekvenskedjan (event-tree-stegen per konsekvens).
        'consequence_steps',
        # Sidordning/revisionshistorik för P&ID:t.
        'pid_revisions', 'pid_sheets',
        # Toppnivå-grupperingen "System" ovanför Nod (2026-08-24).
        'systems',
    ]
    _PROJECT_APP_CONFIG_KEYS = (
        'project_name', 'project_date', 'project_revision',
        # 'project_date' kept above for legacy DBs created before
        # the 2026-08-11 date-range change; it is no longer
        # written or read by SettingsPanel, only cleaned up here.
        # 'project_hazop_leader' kept above for legacy DBs created
        # before the field was removed (2026-08-17) in favor of a
        # free Deltagare column, see NOTES.md.
        'project_date_start', 'project_date_end', 'project_number',
        'project_client',
        'project_facility', 'project_hazop_leader', 'project_participants',
        'ocr_default_engine', 'pid_page_orientation_hint',
        'pid_path', 'pid_layout', 'fill_screen',
    )

    def _wipe_project_tables(self):
        """Clear every project-scoped table using the CURRENT self.db
        connection — guarantees a clean slate even if the DB file itself
        can't be deleted afterward (OneDrive/lock — see _hzp_new's Step 2).
        Factored out of _hzp_new (2026-08-27, see NOTES.md "Nytt projekt
        rensar inte P&ID-objekt") so this, the part that actually
        determines what "New Project" clears, is testable directly against
        a plain self.db without touching DB_PATH or the filesystem at all.

        'nodes' and 'equipment_catalog' used to get deleted here while
        foreign-key enforcement (PRAGMA foreign_keys = ON, set once at
        connection time — see Database.__init__) was still ACTIVE. Two
        columns added later via ALTER TABLE reference them WITHOUT
        ON DELETE CASCADE (equipment_catalog.node_id -> nodes.id,
        deviations.equipment_id / causes.equipment_id -> equipment_
        catalog.id) — any project that had actually linked an object to a
        node/deviation/cause (the normal, everyday "dra objekt till
        trädet" workflow) made those DELETEs fail with a silent FK
        constraint violation (caught by the bare except below), leaving
        nodes AND equipment_catalog completely untouched. equipment_
        markers (the actual P&ID marker positions/tags) was never even in
        _PROJECT_TABLES to begin with, and every one of these tables' many
        per-consequence/per-cause children (severities, recommendations,
        reduction factors, safeguard exclusions, revisions, sheets, the
        System grouping...) was relying entirely on ON DELETE CASCADE to
        get cleared — cascade only actually fires while foreign_keys
        enforcement is on, so once THAT silently blocked the parent
        delete, none of its children got touched either. Wrapping this
        whole bulk wipe in foreign_keys=OFF sidesteps the ordering problem
        entirely (there is no "wrong order" once every row in every listed
        table is simply going to be deleted regardless) — but means every
        project-scoped table now needs to be listed explicitly rather than
        leaning on cascade, hence _PROJECT_TABLES' expansion alongside
        this fix.

        This bug was effectively masked whenever Step 2's physical file
        delete+recreate succeeded (a truly fresh file has nothing left to
        survive, regardless of whether this loop's deletes succeeded) —
        it only became visible in practice when that file delete failed
        (e.g. a OneDrive sync lock, or another running instance still
        holding the file open), which is exactly the scenario the original
        2026-06-18 "clear tables first" fix was written to guard against."""
        try:
            self.db.conn.execute("PRAGMA foreign_keys = OFF")
        except Exception:
            pass
        for tbl in self._PROJECT_TABLES:
            try:
                self.db.conn.execute(f"DELETE FROM {tbl}")
            except Exception:
                pass
        for key in self._PROJECT_APP_CONFIG_KEYS:
            try:
                self.db.conn.execute("DELETE FROM app_config WHERE key=?", (key,))
            except Exception:
                pass
        try:
            self.db.commit()
        except Exception:
            pass
        # SQLite only allows toggling foreign_keys between transactions —
        # attempting this WHILE the deletes above are still part of an
        # open transaction is a silent no-op (not an error, so the bare
        # except above wouldn't have caught it either) — verified
        # empirically after this exact ordering mistake shipped once
        # already. Must run strictly after the commit above.
        try:
            self.db.conn.execute("PRAGMA foreign_keys = ON")
        except Exception:
            pass

    def _hzp_new(self):
        if not self._confirm_discard():
            return

        # Step 1: clear all project-specific tables using the existing
        # connection — see _wipe_project_tables() for what/why.
        self._wipe_project_tables()

        # Step 2: flush WAL, close, then try to delete the file for a clean
        # file (optional — data is already gone from step 1).
        try:
            self.db.conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
        except Exception:
            pass
        try:
            self.db.conn.close()
        except Exception:
            pass

        import time as _time
        for ext in ('', '-shm', '-wal'):
            p = Path(str(DB_PATH) + ext)
            for _ in range(5):
                try:
                    p.unlink(missing_ok=True)
                    break
                except PermissionError:
                    _time.sleep(0.1)

        # Step 3: open (or create) a fresh Database object
        try:
            self.db = Database(DB_PATH)
        except Exception as e:
            QMessageBox.critical(self, "Nytt projekt",
                                 f"Kunde inte skapa ny databas:\n{e}")
            return

        self._hzp_path = None
        self._update_title()
        self._reload_all_panels(pdf_path=None)

    def _hzp_open(self):
        if not self._confirm_discard():
            return
        path, _ = QFileDialog.getOpenFileName(
            self, "Öppna HAZOP-projekt", "", "HAZOP-filer (*.hzp);;Alla filer (*)")
        if not path:
            return
        self._load_hzp(path)

    def _hzp_save(self):
        if not self._hzp_path:
            self._hzp_save_as()
        else:
            self._write_hzp(self._hzp_path)

    def _hzp_save_as(self):
        default = (Path(self._hzp_path).stem if self._hzp_path
                   else "nytt_projekt") + ".hzp"
        path, _ = QFileDialog.getSaveFileName(
            self, "Spara HAZOP-projekt", default, "HAZOP-filer (*.hzp)")
        if not path:
            return
        if not path.lower().endswith('.hzp'):
            path += '.hzp'
        self._write_hzp(path)
        self._hzp_path = path
        self._update_title()

    # ── Internal helpers ──────────────────────────────────────────────────────

    def _confirm_discard(self):
        if self._hzp_path:
            r = QMessageBox.question(
                self, "Osparade ändringar",
                "Vill du spara det nuvarande projektet innan du fortsätter?",
                QMessageBox.StandardButton.Save |
                QMessageBox.StandardButton.Discard |
                QMessageBox.StandardButton.Cancel,
                QMessageBox.StandardButton.Save)
            if r == QMessageBox.StandardButton.Cancel:
                return False
            if r == QMessageBox.StandardButton.Save:
                self._hzp_save()
        return True

    def _write_hzp(self, path: str):
        import zipfile, json, tempfile, datetime, sqlite3 as _sql
        self.db.conn.commit()   # flush all pending writes
        self.db._write_backup(startup=True)   # force immediate backup snapshot

        # Use SQLite online backup API (safe with WAL mode; consistent snapshot)
        fd, tmp_path = tempfile.mkstemp(suffix='.db')
        import os; os.close(fd)
        tmp_db = Path(tmp_path)
        try:
            bk_conn = _sql.connect(str(tmp_db))
            try:
                with bk_conn:
                    self.db.conn.backup(bk_conn)
            finally:
                bk_conn.close()

            # Collect PDF path from pid_config
            pdf_src = None
            try:
                pdf_str = self.db.get_pid_path()
                if pdf_str and Path(pdf_str).exists():
                    pdf_src = Path(pdf_str)
            except Exception:
                pass

            meta = {
                "hzp_version": 1,
                "created": datetime.datetime.now().isoformat(timespec='seconds'),
                "app": "HAZOP Tool",
                "pdf_name": pdf_src.name if pdf_src else None,
            }

            with zipfile.ZipFile(path, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
                zf.write(tmp_db, "hazop_project.db")
                if pdf_src:
                    zf.write(pdf_src, f"pid/{pdf_src.name}")
                zf.writestr("meta.json", json.dumps(meta, ensure_ascii=False, indent=2))
        finally:
            try:
                tmp_db.unlink(missing_ok=True)   # always clean up even on exception
            except Exception:
                pass   # don't let cleanup failure mask the real error

        self.status_bar.showMessage(f"Sparat: {path}", 5000)
        self._hzp_path = path
        self._update_title()

    def _load_hzp(self, path: str):
        import zipfile, json, shutil, tempfile

        # Extract to a persistent work dir next to the hzp file
        work_dir = Path(path).parent / (Path(path).stem + "_files")
        work_dir.mkdir(exist_ok=True)

        try:
            with zipfile.ZipFile(path, 'r') as zf:
                zf.extractall(work_dir)
        except zipfile.BadZipFile:
            QMessageBox.critical(self, "Fel", f"Filen är inte en giltig .hzp-fil:\n{path}")
            return

        # Load meta
        pdf_name = None
        try:
            meta = json.loads((work_dir / "meta.json").read_text(encoding='utf-8'))
            pdf_name = meta.get("pdf_name")
        except Exception:
            pass

        # Copy extracted DB over the active DB_PATH.
        #
        # The old connection is closed FIRST, before the copy — 2026-08-21
        # fix (see NOTES.md "Paketera HAZOP-appen som en installationsfil"):
        # closing a still-open WAL-mode SQLite connection performs a
        # checkpoint (Database.__del__ does this explicitly; a plain
        # .close() does one implicitly too) that flushes that connection's
        # OWN buffered WAL frames back onto whatever file DB_PATH now
        # names. Closing AFTER the copy (the original order here) meant
        # that checkpoint flushed the OLD project's pre-copy WAL state
        # onto the newly-copied file, silently clobbering the just-loaded
        # project back to whatever the old one looked like — reproduced
        # directly against the real Database class, not just theorised.
        # Closing first means the checkpoint lands on the OLD file (which
        # we're about to discard anyway) before the new one ever exists at
        # that path, so there's nothing left to clobber it with.
        #
        # This does give up a little of the original recovery guarantee
        # ("a failed copy leaves self.db working") since the old
        # connection is already gone by the time a copy could fail — but
        # DB_PATH itself is never touched by a FAILED copy (shutil.copy2
        # either fully succeeds or leaves the destination as it was), so
        # simply reopening DB_PATH on failure below restores the exact
        # same working state the old connection pointed at.
        extracted_db = work_dir / "hazop_project.db"
        if not extracted_db.exists():
            QMessageBox.critical(self, "Fel", "Projektet saknar databasfil.")
            return

        self.db.conn.close()
        try:
            shutil.copy2(extracted_db, DB_PATH)
        except Exception as e:
            self.db = Database(DB_PATH)   # recover: DB_PATH itself is untouched
            QMessageBox.critical(self, "Fel vid inläsning",
                                 f"Kunde inte kopiera databas:\n{e}")
            return

        # Reopen DB at same path
        new_db = Database(DB_PATH)
        self.db = new_db

        # Resolve PDF
        pdf_path = None
        if pdf_name:
            candidate = work_dir / "pid" / pdf_name
            if candidate.exists():
                # Keep PDF in work_dir and update pid_config to point there
                pdf_path = str(candidate)
                try:
                    new_db.set_pid_config_value('path', pdf_path)
                except Exception:
                    pass

        self._hzp_path = path
        self._update_title()
        self._reload_all_panels(pdf_path=pdf_path)
        # Immediately create a startup backup of the freshly loaded project
        self.db._write_backup(startup=True)
        self.status_bar.showMessage(f"Öppnat: {path}", 5000)

    def _reload_all_panels(self, pdf_path=None):
        """Swap self.db on every panel and refresh all content."""
        db = self.db
        load_matrix(db)

        # Update db reference on every panel (props_ribbon included)
        for panel in [self.tree_panel,
                      self.scenario_panel, self.equipment_panel,
                      self.admin_panel, self.settings_panel, self.hazop_prep_panel,
                      self.markup_table_panel,
                      self.red_markup_panel,
                      self.worksheet, self.props_ribbon, self.recommendations_panel]:
            try:
                panel.db = db
            except Exception:
                pass

        # EquipmentPanel's QTableView is backed by _EquipmentTableModel,
        # which keeps its OWN db reference (needed for setData()/delete_row()
        # to write through directly) separate from EquipmentPanel.db above —
        # found via a real crash: "Cannot operate on a closed database" when
        # switching to the Utrustning page after a project reload, because
        # the model was still holding the OLD (by-then-closed) connection.
        try:
            self.equipment_panel.set_db(db)
        except Exception:
            pass

        # Also update db on settings sub-panels (they have their own db reference)
        sp = self.settings_panel
        for attr in ('_std_objects_panel', '_tag_memory_panel', '_tag_db_panel',
                     'analysis_panel'):
            try:
                sub = getattr(sp, attr, None)
                if sub is not None:
                    sub.db = db
            except Exception:
                pass

        # Same nested-sub-panel-with-its-own-db-reference pattern, for the
        # sub-panels that moved from SettingsPanel to HAZOPPreparationPanel
        # (2026-08-17, see NOTES.md) — same fix, new host object.
        hp = self.hazop_prep_panel
        for attr in ('_std_causes_panel', '_participant_matrix_panel'):
            try:
                sub = getattr(hp, attr, None)
                if sub is not None:
                    sub.db = db
            except Exception:
                pass
        # analysis_panel (PIDAnalysisPanel) is likewise backed by
        # _IdentifiedTagsModel with its own db reference — same fix as
        # equipment_panel above.
        try:
            sp.analysis_panel._model.db = db
        except Exception:
            pass

        # HAZOPWorksheet embeds its own ScenarioTablePanel instance (distinct
        # from self.scenario_panel above) — it has its own stale db reference.
        try:
            self.worksheet._table_panel.db = db
        except Exception:
            pass

        # StudyManagementPanel (admin_panel) embeds its own PIDManagementPanel
        # (revision history + sheet reordering) — same nested-sub-panel-with-
        # its-own-db-reference pattern as equipment_panel._model/worksheet.
        # _table_panel above, missed here until a real crash: "Cannot operate
        # on a closed database" in Database.get_revisions() when switching to
        # the Administration tab after a project reload (2026-08-11, see
        # NOTES.md).
        try:
            self.admin_panel._pid_mgmt.db = db
        except Exception:
            pass

        # PIDPanel wires differently — it holds db on itself and the viewer
        try:
            self.pid_panel.db = db
            self.pid_panel.viewer.db = db
        except Exception:
            pass

        # EquipmentDeviationBar (the bottom-of-P&ID bar shown when an
        # equipment marker is clicked) also keeps its own db reference —
        # same "Cannot operate on a closed database" crash class as
        # equipment_panel._model/worksheet._table_panel above, found via a
        # real crash report the first time a user clicked an equipment
        # marker after a project reload.
        try:
            self.pid_panel._equipment_bar.db = db
        except Exception:
            pass

        # Reload tree + scenario
        self.tree_panel.refresh()
        self.scenario_panel.clear()

        # Reload P&ID
        if pdf_path:
            self.pid_panel.try_reload_pdf(pdf_path)
        else:
            self.pid_panel.try_reload_pdf()


if __name__ == '__main__':
    # MUST be the very first thing that runs in __main__, before any other
    # setup (2026-08-24, see NOTES.md "Analysera P&ID kraschar och startar
    # om appen vid flera sidor"). ParallelTagScanWorker/
    # ParallelEquipmentAnalysisWorker/ImageSymbolSearchWorker (pid_viewer.py)
    # use concurrent.futures.ProcessPoolExecutor for documents with 4+ pages
    # (see _should_parallelize) -- on Windows, spawning a new process
    # re-executes THIS SAME frozen .exe from scratch when packaged with
    # PyInstaller, unless multiprocessing.freeze_support() has already run.
    # Without it, every worker process re-entered this __main__ block as if
    # freshly launched, opening ANOTHER full MainWindow -- which is exactly
    # what "crashes and restarts" looked like from the outside: not a crash
    # at all, but the app relaunching itself once per worker process. Never
    # triggered unpackaged (`python hazop.py`), or on documents under 4
    # pages, which is why it went unnoticed until the app was actually
    # packaged and tested on a real multi-page P&ID.
    import multiprocessing
    multiprocessing.freeze_support()

    import logging

    _configure_utf8_console_output()

    # ── Crash logger ───────────────────────────────────────────────────────────
    # Structured crash reporting: saves detailed diagnostic info to JSON files
    # in hazop/crashes/ directory for automatic analysis. Also maintains legacy
    # hazop_crash.log for backward compatibility.
    _LOG = _app_dir() / 'hazop_crash.log'
    logging.basicConfig(
        filename=str(_LOG),
        level=logging.DEBUG,
        format='%(asctime)s %(levelname)s %(message)s',
        encoding='utf-8',
    )
    # Also echo to stderr (visible in the console window)
    # During startup, suppress DEBUG output to console; keep to file
    console_handler = logging.StreamHandler(sys.stderr)
    console_handler.setLevel(logging.INFO)
    logging.getLogger().addHandler(console_handler)

    # Setup structured crash reporting
    CrashReporter.setup()

    # sys.excepthook was already installed at module load time (see
    # _global_exception_hook, defined right after the CrashReporter class,
    # near the top of this file) — it logs, writes a structured crash
    # report, and shows a QMessageBox, without calling sys.exit()/os.abort()
    # or re-raising, so a single bad slot cannot take down the whole app.
    # Nothing here should reassign sys.excepthook; doing so would just
    # shadow that hook for the rest of the process.
    assert sys.excepthook is _global_exception_hook, (
        "sys.excepthook was reassigned somewhere between module import and "
        "__main__ -- _global_exception_hook must remain installed")

    # Catch exceptions raised inside Qt signal handlers (they bypass sys.excepthook)
    def _qt_message_handler(mode, context, message):
        if mode in (QtMsgType.QtWarningMsg,
                    QtMsgType.QtCriticalMsg,
                    QtMsgType.QtFatalMsg):
            logging.warning('Qt [%s] %s', mode.name, message)
        else:
            logging.debug('Qt [%s] %s', mode.name, message)

    try:
        from PyQt6.QtCore import qInstallMessageHandler, QtMsgType
        qInstallMessageHandler(_qt_message_handler)
    except ImportError:
        pass

    # Route all Python warnings through the log as well
    import warnings
    logging.captureWarnings(True)

    logging.info('=== HAZOP Tool started ===')

    app = QApplication(sys.argv)
    app.setStyle('Fusion')

    # Override Fusion's default OS/theme-blue Highlight role with the app's
    # own accent — otherwise native-rendered selections (custom delegate
    # paints via option.palette.highlight(), combo box popups, etc.) show
    # a different blue than the one used throughout the QSS below.
    _palette = app.palette()
    _palette.setColor(QPalette.ColorRole.Highlight, QColor('#2F5FD0'))
    _palette.setColor(QPalette.ColorRole.HighlightedText, QColor('#FFFFFF'))
    _palette.setColor(QPalette.ColorRole.Window, QColor('#FBFBFA'))
    _palette.setColor(QPalette.ColorRole.Base, QColor('#FFFFFF'))
    _palette.setColor(QPalette.ColorRole.Text, QColor('#17191C'))
    _palette.setColor(QPalette.ColorRole.WindowText, QColor('#17191C'))
    # Tooltips are top-level windows.  Set their contrast on the application
    # palette so every tooltip stays readable even when its source widget has
    # a local stylesheet of its own.
    _palette.setColor(QPalette.ColorRole.ToolTipBase, QColor('#17191C'))
    _palette.setColor(QPalette.ColorRole.ToolTipText, QColor('#FFFFFF'))
    app.setPalette(_palette)

    # Apply Windows 11 light theme
    app.setStyleSheet(_get_windows11_stylesheet())
    # Default font size lives here, not in the QSS above — see that
    # function's own docstring for why a QSS font-size rule would block
    # widgets (like ScenarioTablePanel's "Textstorlek" spinbox) from ever
    # changing their own font size again.
    app.setFont(QFont("Segoe UI", 9))

    # Show splash screen during startup
    splash = SplashScreen()
    splash.show()
    app.processEvents()

    # Catch exceptions in Qt event loop (slots/signals)
    _original_notify = app.notify
    def _safe_notify(receiver, event):
        try:
            return _original_notify(receiver, event)
        except Exception:
            logging.exception('Exception in Qt event loop (notify)')
            return False
    app.notify = _safe_notify

    try:
        splash.set_status("Laddar databas...")
        # A .hzp file double-clicked once registered as a Windows file
        # association (see NOTES.md "Paketera HAZOP-appen som en
        # installationsfil") arrives as argv[1] — only honour it if it
        # actually looks like a real .hzp project, so a stray/unexpected
        # argument can't be mistaken for one.
        hzp_arg = sys.argv[1] if len(sys.argv) > 1 else None
        if hzp_arg and not (Path(hzp_arg).suffix.lower() == '.hzp' and Path(hzp_arg).is_file()):
            hzp_arg = None
        win = MainWindow(hzp_arg)
        splash.set_status("Initialiserar gränssnitt...")
        app.processEvents()
        splash.close_splash()
        win.show()
        code = app.exec()
        logging.info('=== HAZOP Tool exited (code %d) ===', code)

        # Clean up global resources before exit (OCR models, DB connections)
        try:
            from pid_viewer import cleanup_ocr_resources
            cleanup_ocr_resources()
        except Exception:
            pass

        sys.exit(code)
    except Exception:
        logging.exception('Fatal exception during startup or main loop')
        raise
