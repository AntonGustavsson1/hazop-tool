"""Excel export for revisioned LOPA analyses.

The exporter deliberately reads only the stored LOPA revision and its matrix
snapshot.  It never re-reads live HAZOP fields while generating a report, so
a locked revision remains a faithful historic document.
"""

from __future__ import annotations

import re


def _number(value, digits=6):
    return '—' if value is None else f'{float(value):.{digits}g}'


def _sheet_name(base, used):
    text = re.sub(r'[\\/*?:\[\]]', ' ', base).strip() or 'LOPA'
    text = text[:31]
    candidate = text
    index = 2
    while candidate in used:
        suffix = f' {index}'
        candidate = f'{text[:31 - len(suffix)]}{suffix}'
        index += 1
    used.add(candidate)
    return candidate


def _write_table(ws, row, headers, rows, *, widths=None):
    """Write a compact styled table and return the first free row below it."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    header_fill = PatternFill('solid', fgColor='3F4A54')
    header_font = Font(color='FFFFFF', bold=True)
    border = Border(bottom=Side(style='thin', color='AAB2B9'))
    for column, header in enumerate(headers, 1):
        cell = ws.cell(row, column, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    for row_offset, values in enumerate(rows, 1):
        for column, value in enumerate(values, 1):
            cell = ws.cell(row + row_offset, column, value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = border
    if widths:
        for column, width in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + column)].width = width
    return row + len(rows) + 2


def _source_summary(db, revision_id, source):
    result = db.lopa_source_calculation(source['id'])
    state = 'Följer HAZOP' if source['follows_hazop'] else 'Frikopplad från HAZOP'
    if source['source_missing']:
        state = 'Källa saknas i HAZOP'
    return result, state


def _selected_revisions(db, selections):
    if selections is not None:
        result = []
        for lopa_id, revision_id in selections:
            record = db.get_lopa_record(lopa_id)
            revision = db.get_lopa_revision(revision_id)
            if record and revision and revision['lopa_id'] == lopa_id:
                result.append((record, revision))
        return result
    result = []
    for record in db.lopa_records(include_archived=True):
        revision = db.current_lopa_revision(record['id'])
        if revision:
            result.append((record, revision))
    return result


def export_lopa_excel(db, filepath, selections=None):
    """Export selected ``(lopa_id, revision_id)`` pairs to a workbook.

    A summary sheet is always generated.  Every LOPA revision becomes a
    separate, print-friendly sheet containing the same analysis sections as
    the LOPA workspace: source scenario, sensor/final element, barriers,
    escalation, criterion and document comments.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return False, 'openpyxl saknas.\nKör: pip install openpyxl'

    selected = _selected_revisions(db, selections)
    workbook = openpyxl.Workbook()
    summary = workbook.active
    summary.title = 'LOPA sammanfattning'
    summary['A1'] = 'Skyddsbarriäranalys (LOPA) – sammanfattning'
    summary['A1'].font = Font(bold=True, size=14)
    summary.merge_cells('A1:I1')
    summary.freeze_panes = 'A4'
    summary_row = 3
    summary_rows = []
    for record, revision in selected:
        calculations = [db.lopa_source_calculation(source['id'])
                        for source in db.lopa_sources(revision['id'])]
        governing = max((item for item in calculations if item['required_rrf'] is not None),
                         key=lambda item: item['required_rrf'], default=None)
        summary_rows.append([
            record['display_number'], record.get('sif_number') or '', record['sif_name'],
            record['sis_name'], revision['label'], revision['status'],
            governing['governing_category_name'] if governing else '—',
            _number(governing['required_rrf'] if governing else None),
            governing['sil'] if governing else 'Ofullständig',
        ])
    _write_table(summary, summary_row,
                 ['LOPA-nr', 'SIF-nr', 'SIF-namn', 'SIS', 'Rev.', 'Status',
                  'Dimensionerande kategori', 'RRF', 'Beräknad SIL'], summary_rows,
                 widths=[12, 14, 28, 20, 8, 13, 24, 14, 18])

    used_names = {summary.title}
    for record, revision in selected:
        ws = workbook.create_sheet(_sheet_name(
            f"LOPA {record['display_number']} R{revision['label']}", used_names))
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'A5'
        ws['A1'] = 'Skyddsbarriäranalys (LOPA)'
        ws['A1'].font = Font(bold=True, size=15)
        ws.merge_cells('A1:H1')
        metadata = [
            ('LOPA-nr', record['display_number']), ('SIF-nr', record.get('sif_number') or ''),
            ('SIF-namn', record['sif_name']), ('SIS', record['sis_name']),
            ('Revision', revision['label']), ('Status', revision['status']),
            ('Datum', revision.get('document_date') or ''),
            ('Utförd av', revision.get('performed_by_text') or ''),
            ('Godkänd av', revision.get('approved_by_text') or ''),
        ]
        for index, (label, value) in enumerate(metadata):
            row = 3 + index // 3
            column = 1 + (index % 3) * 3
            ws.cell(row, column, label).font = Font(bold=True)
            ws.cell(row, column + 1, value)
            ws.merge_cells(start_row=row, start_column=column + 1, end_row=row, end_column=column + 2)
        row = 7
        ws.cell(row, 1, 'Källscenarier från HAZOP').font = Font(bold=True, size=12)
        row += 1
        source_rows = []
        sources = db.lopa_sources(revision['id'])
        for source in sources:
            result, status = _source_summary(db, revision['id'], source)
            trigger = ' '.join(part for part in (
                source.get('equipment_tag') or '', source.get('trigger_code') or '',
                source.get('trigger_custom') or '') if part)
            source_rows.append([
                'Ja' if source['active'] else 'Nej', trigger or '—', source.get('cause_text') or '',
                source.get('scenario_text') or '', _number(source.get('base_frequency')),
                f"{source.get('assumption_percent') or 0:.6g} %", source.get('control_frequency') or '',
                result['governing_category_name'] or '—', _number(result['required_rrf']),
                result['sil'] or 'Ofullständig', status,
            ])
        row = _write_table(ws, row,
                           ['Aktiv', 'Objekt / anrop', 'Orsak', 'Scenario', 'Felfrekvens /år',
                            'Förutsättning', 'Kontrollfrekvens', 'Dim. kategori', 'RRF', 'SIL', 'Synkstatus'],
                           source_rows,
                           widths=[9, 22, 28, 33, 16, 14, 20, 20, 12, 13, 24])

        ws.cell(row, 1, 'Givardel').font = Font(bold=True, size=12)
        row += 1
        sensor_rows = []
        for group_index, group in enumerate(db.lopa_sensor_groups(revision['id']), 1):
            members = db.lopa_sensor_members(group['id'])
            if not members:
                sensor_rows.append([f'Givardel {group_index}', group['voting'], '—', '—', '—'])
            for member in members:
                trip = ' '.join(part for part in (member.get('trigger_code') or '',
                                                   member.get('trigger_custom') or '') if part)
                sensor_rows.append([
                    f'Givardel {group_index}', group['voting'], member.get('tag') or '—', trip or '—',
                    'Ja' if member['active'] else 'Nej',
                ])
        row = _write_table(ws, row, ['Grupp', 'Voting', 'Objekt', 'Anrop', 'Aktiv'], sensor_rows,
                           widths=[18, 12, 24, 18, 10])

        ws.cell(row, 1, 'Manöverdel').font = Font(bold=True, size=12)
        row += 1
        final_rows = []
        for group_index, group in enumerate(db.lopa_final_groups(revision['id']), 1):
            members = db.lopa_final_members(group['id'])
            if not members:
                final_rows.append([f'Manöverdel {group_index}', group['voting'], '—', '—', '—'])
            for member in members:
                final_rows.append([
                    f'Manöverdel {group_index}', group['voting'], member.get('tag') or '—',
                    member.get('action_text') or '—', 'Ja' if member['active'] else 'Nej',
                ])
        row = _write_table(ws, row, ['Grupp', 'Voting', 'Objekt', 'Åtgärd', 'Aktiv'], final_rows,
                           widths=[18, 12, 24, 30, 10])

        ws.cell(row, 1, 'Oberoende barriärer').font = Font(bold=True, size=12)
        row += 1
        barrier_rows = []
        for source in sources:
            for barrier in db.lopa_barriers(revision['id'], source['id']):
                categories = ('Alla' if barrier['applies_all_categories'] else
                              ', '.join(item['category_key'] for item in
                                        db.lopa_barrier_categories(barrier['id']) if item['active']) or 'Ingen')
                state = ('Manuell LOPA-barriär' if barrier['manual'] else
                         ('Följer HAZOP' if barrier['follows_hazop'] else 'Frikopplad från HAZOP'))
                if barrier['source_missing']:
                    state = 'Källa saknas i HAZOP'
                barrier_rows.append([
                    source.get('cause_text') or '', 'Ja' if barrier['active'] else 'Nej',
                    barrier.get('sg_type') or '', barrier.get('description') or '', _number(barrier.get('rrf')),
                    categories, state,
                ])
        row = _write_table(ws, row,
                           ['Källscenario', 'Aktiv', 'Typ', 'Beskrivning', 'RRF', 'Kategorier', 'Status'],
                           barrier_rows, widths=[28, 9, 18, 34, 12, 22, 24])

        ws.cell(row, 1, 'Eskalering och beräkningsspår').font = Font(bold=True, size=12)
        row += 1
        escalation_rows = []
        for source in sources:
            calculation = db.lopa_source_calculation(source['id'])
            escalation = {item['category_key']: item for item in db.lopa_escalation_rows(source['id'])}
            for item in calculation['categories']:
                values = escalation.get(item['category_key'], {}).get('factor_values_json') or '{}'
                escalation_rows.append([
                    source.get('cause_text') or '', item['category_name'], item['severity'],
                    _number(item['tel']), values, _number(item['barrier_rrf']),
                    _number(item['escalation_factor']), _number(item['remaining_frequency']),
                    _number(item['accident_frequency']), _number(item['required_rrf']), item['sil'] or '—',
                ])
        row = _write_table(ws, row,
                           ['Källscenario', 'Kategori', 'Nivå', 'TEL /år', 'Faktorer (%)',
                            'Barriär-RRF', 'Eskalering', 'Kvarvarande /år', 'Olycka /år', 'Krävd RRF', 'SIL'],
                           escalation_rows,
                           widths=[28, 18, 9, 14, 30, 14, 14, 18, 16, 14, 13])

        ws.cell(row, 1, 'Övriga uppgifter').font = Font(bold=True, size=12)
        row += 1
        extra_rows = [
            ['Ytterligare åtgärder', revision.get('additional_actions') or ''],
            ['Ytterligare säkerhetskrav', revision.get('additional_requirements') or ''],
            ['Processäkerhetstid (s)', _number(revision.get('process_safety_time'))],
        ]
        row = _write_table(ws, row, ['Fält', 'Innehåll'], extra_rows, widths=[30, 75])
        comments = db.lopa_comments(revision['id'])
        if comments:
            ws.cell(row, 1, 'Kommentarer').font = Font(bold=True, size=12)
            row += 1
            _write_table(ws, row, ['Datum', 'Namn', 'Kommentar'],
                         [[item.get('created_at') or '', item.get('author') or '', item.get('body') or '']
                          for item in comments], widths=[22, 24, 80])
        for current_row in range(1, ws.max_row + 1):
            ws.row_dimensions[current_row].height = max(ws.row_dimensions[current_row].height or 15, 18)
        for cells in ws.iter_rows():
            for cell in cells:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

    if not selected:
        summary['A3'] = 'Inga LOPA-revisioner valdes för export.'
    try:
        workbook.save(filepath)
    except Exception as exc:
        return False, str(exc)
    return True, ''
