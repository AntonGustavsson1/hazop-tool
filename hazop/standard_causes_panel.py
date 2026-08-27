#!/usr/bin/env python3
"""StandardCausesSettingsPanel -- split out of settings_panels.py 2026-08-21, see NOTES.md "Dela upp settings_panels.py"."""

import re
import json
from pathlib import Path
from functools import partial

from PyQt6.QtWidgets import (
    QAbstractItemView, QCheckBox, QColorDialog, QComboBox, QDateEdit,
    QDoubleSpinBox, QFileDialog, QFormLayout, QGridLayout,
    QGroupBox, QHBoxLayout, QHeaderView, QInputDialog, QLabel, QLineEdit,
    QListWidget, QListWidgetItem, QMenu, QMessageBox, QPushButton,
    QScrollArea, QSizePolicy, QSpinBox, QSplitter, QTableWidget,
    QTableWidgetItem, QTabWidget, QTextEdit, QToolButton, QVBoxLayout,
    QWidget,
)
from PyQt6.QtCore import Qt, pyqtSignal, QDate, QEvent, QMimeData
from PyQt6.QtGui import QBrush, QColor, QDrag, QFont, QFontMetrics

from constants import CONFIG, SEV_LABELS
from database import (
    Database, DEFAULT_MATRIX, DEFAULT_FREQ_BOUNDARIES, _STD_OBJECTS,
    _normalise_matrix, _risk_matrix_cache, get_matrix, freq_to_f_level,
    risk_info,
)
from pid_viewer import _icon, FREQ_LABELS, ocr_status
from ui_helpers import freq_axis_label
from equipment_panel import TagDatabasePanel, PIDAnalysisPanel


class StandardCausesSettingsPanel(QWidget):
    """4-level editable hierarchy: Nodtyp → Avvikelse → Objekt → Orsaker."""

    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self._loading = False
        self._loading_nt = False
        self._node_type_ids = []

        layout = QHBoxLayout(self)

        # ── Col 0: Nodtyp (2026-08-17 user request) ──────────────────────────
        # Filters _dev_list to deviations belonging to the selected node
        # type; drag a deviation from _dev_list onto a node type here to
        # COPY it (deep, independent copy incl. its causes — user confirmed
        # via AskUserQuestion, not a move/link) into that type.
        c0 = QVBoxLayout()
        c0.addWidget(QLabel("<b>Nodtyp</b>"))
        self._nodetype_list = QListWidget()
        self._nodetype_list.currentRowChanged.connect(lambda _row: self._load_deviations())
        self._nodetype_list.itemChanged.connect(self._on_nodetype_item_changed)
        self._nodetype_list.setAcceptDrops(True)
        self._nodetype_list.viewport().setAcceptDrops(True)
        self._nodetype_list.installEventFilter(self)
        self._nodetype_list.viewport().installEventFilter(self)
        c0.addWidget(self._nodetype_list)
        c0b = QHBoxLayout()
        for icon, slot in (('+', self._add_node_type), ('−', self._del_node_type)):
            b = QPushButton(icon); b.setFixedWidth(28); b.clicked.connect(slot); c0b.addWidget(b)
        c0b.addStretch(); c0.addLayout(c0b)
        # _load_node_types() is deferred to the end of __init__ (below,
        # after _dev_list exists) — its setCurrentRow() call fires
        # currentRowChanged immediately, which cascades into
        # _load_deviations(), so calling it here (before _dev_list is
        # built) would crash with AttributeError.

        # ── Col 1: Avvikelse ──────────────────────────────────────────────────
        c1 = QVBoxLayout()
        c1.addWidget(QLabel("<b>Avvikelse</b>"))
        self._dev_list = QListWidget()
        self._dev_list.currentRowChanged.connect(self._on_dev_sel)
        self._dev_list.setDragEnabled(True)
        # Instance-level override (same monkeypatch pattern already used
        # elsewhere in this file, e.g. ParticipantMatrixPanel's Enter-key
        # handling) — carries the deviation's DB id as custom mime text
        # instead of Qt's default internal model-index payload, so the
        # Nodtyp column's drop handler can read it directly.
        def _dev_list_mime_data(items, _list=self._dev_list):
            md = QMimeData()
            if items:
                dev_id = items[0].data(Qt.ItemDataRole.UserRole)
                md.setText(f'hzp:stddev:{dev_id}')
            return md
        self._dev_list.mimeData = _dev_list_mime_data
        c1.addWidget(self._dev_list)
        c1b = QHBoxLayout()
        for icon, slot in (('+', self._add_dev), ('−', self._del_dev),
                           ('↑', lambda: self._move_dev(-1)), ('↓', lambda: self._move_dev(1))):
            b = QPushButton(icon); b.setFixedWidth(28); b.clicked.connect(slot); c1b.addWidget(b)
        c1b.addStretch(); c1.addLayout(c1b)

        # ── Col 2: Objekt ─────────────────────────────────────────────────────
        c2 = QVBoxLayout()
        self._obj_lbl = QLabel("<b>Objekt</b>")
        c2.addWidget(self._obj_lbl)
        self._obj_list = QListWidget()
        self._obj_list.currentRowChanged.connect(self._on_obj_sel)
        c2.addWidget(self._obj_list)
        # "implementera de funktioner som finns i standardobjekt även i
        # standard orsaker så man kan lägga till nya objekt under
        # standardorsaker" (2026-08-17, see NOTES.md) — same add/delete/
        # reorder/rename CRUD as StandardObjectsSettingsPanel's own
        # _list, over the exact same standard_objects table, so a new
        # object type no longer requires switching tabs.
        c2b = QHBoxLayout()
        for icon, slot in (('+', self._add_obj), ('−', self._del_obj),
                           ('↑', lambda: self._move_obj(-1)), ('↓', lambda: self._move_obj(1))):
            b = QPushButton(icon); b.setFixedWidth(28); b.clicked.connect(slot); c2b.addWidget(b)
        c2b.addStretch(); c2.addLayout(c2b)
        # Show all objects; objects with causes are highlighted
        self._show_all_obj_chk = QCheckBox("Visa alla objekt")
        self._show_all_obj_chk.setChecked(True)
        self._show_all_obj_chk.stateChanged.connect(lambda _: self._load_objects())
        c2.addWidget(self._show_all_obj_chk)

        # ── Col 3: Orsaker ────────────────────────────────────────────────────
        c3 = QVBoxLayout()
        self._cause_lbl = QLabel("<b>Orsaker</b>")
        c3.addWidget(self._cause_lbl)
        self._cause_list = QListWidget()
        self._cause_list.currentRowChanged.connect(self._on_cause_sel)
        c3.addWidget(self._cause_list)

        # Frequency field for selected cause
        freq_row = QHBoxLayout()
        freq_lbl = QLabel("Frekvens (/år):")
        freq_lbl.setStyleSheet("font-size:10px; color:#555;")
        freq_row.addWidget(freq_lbl)
        self._freq_edit = QLineEdit()
        self._freq_edit.setPlaceholderText("t.ex. 0.01")
        self._freq_edit.setMaximumWidth(90)
        self._freq_edit.setToolTip("Basfrekvens för vald orsak (händelser/år). Lämna tomt om okänd.")
        self._freq_edit.editingFinished.connect(self._save_freq)
        freq_row.addWidget(self._freq_edit)
        self._freq_level_lbl = QLabel("")
        self._freq_level_lbl.setStyleSheet("color:#8D9299; font-size:10px;")
        freq_row.addWidget(self._freq_level_lbl)
        freq_row.addStretch()
        c3.addLayout(freq_row)

        c3b = QHBoxLayout()
        for icon, slot in (('+', self._add_cause), ('−', self._del_cause),
                           ('↑', lambda: self._move_cause(-1)), ('↓', lambda: self._move_cause(1))):
            b = QPushButton(icon); b.setFixedWidth(28); b.clicked.connect(slot); c3b.addWidget(b)
        c3b.addStretch(); c3.addLayout(c3b)
        btn_sync = QPushButton("Synka frekvenser →")
        btn_sync.setToolTip("Uppdaterar frekvensen på alla orsaker kopplade till standardorsaker.")
        btn_sync.clicked.connect(self._sync_freqs)
        c3.addWidget(btn_sync)
        # Feature 16: export/import buttons
        io_row = QHBoxLayout()
        btn_exp = QPushButton("↑ Exportera")
        btn_exp.setToolTip("Exportera hela standardbiblioteket till JSON")
        btn_exp.clicked.connect(self._export_library)
        btn_imp = QPushButton("↓ Importera")
        btn_imp.setToolTip("Importera standardbibliotek från JSON (lägger till, skriver ej över)")
        btn_imp.clicked.connect(self._import_library)
        io_row.addWidget(btn_exp); io_row.addWidget(btn_imp)
        c3.addLayout(io_row)
        btn_exp_xlsx = QPushButton("↑ Exportera Excel")
        btn_exp_xlsx.setToolTip(
            "Exportera samtliga standardavvikelser (grupperade per objekttyp) "
            "till en redigerbar Excel-fil")
        btn_exp_xlsx.clicked.connect(self._export_library_excel)
        c3.addWidget(btn_exp_xlsx)

        layout.addLayout(c0, 1)
        layout.addLayout(c1, 1)
        layout.addLayout(c2, 1)
        layout.addLayout(c3, 1)
        self._load_node_types()   # cascades into _load_deviations() via currentRowChanged

    # ── Load helpers ──────────────────────────────────────────────────────────
    # ── Node type CRUD (2026-08-17, see NOTES.md) ────────────────────────────
    def _load_node_types(self):
        self._loading_nt = True
        cur = self._nodetype_list.currentRow()
        self._nodetype_list.clear()
        types = self.db.node_types()
        self._node_type_ids = [t['id'] for t in types]
        for t in types:
            item = QListWidgetItem(t['name'])
            item.setData(Qt.ItemDataRole.UserRole, t['id'])
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
            self._nodetype_list.addItem(item)
        self._loading_nt = False
        self._nodetype_list.setCurrentRow(max(0, min(cur, self._nodetype_list.count() - 1)))

    def _current_node_type_id(self):
        item = self._nodetype_list.currentItem()
        return item.data(Qt.ItemDataRole.UserRole) if item else None

    def _on_nodetype_item_changed(self, item):
        if self._loading_nt:
            return
        id_ = item.data(Qt.ItemDataRole.UserRole)
        name = item.text().strip()
        if id_ is not None and name:
            self.db.rename_node_type(id_, name)

    def _add_node_type(self):
        new_id = self.db.add_node_type('Ny nodtyp')
        item = QListWidgetItem('Ny nodtyp')
        item.setData(Qt.ItemDataRole.UserRole, new_id)
        item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
        self._nodetype_list.addItem(item)
        self._node_type_ids.append(new_id)
        self._nodetype_list.setCurrentItem(item)
        self._nodetype_list.editItem(item)

    def _del_node_type(self):
        item = self._nodetype_list.currentItem()
        if not item:
            return
        id_ = item.data(Qt.ItemDataRole.UserRole)
        if len(self._node_type_ids) <= 1:
            QMessageBox.information(self, 'Kan inte ta bort',
                                     'Minst en nodtyp måste finnas kvar.')
            return
        if QMessageBox.question(
                self, 'Ta bort nodtyp',
                'Ta bort nodtypen? Avvikelser under den flyttas till standardtypen.',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                ) == QMessageBox.StandardButton.Yes:
            self.db.delete_node_type(id_)
            self._load_node_types()

    def eventFilter(self, obj, event):
        from PyQt6.QtCore import QEvent
        _drop_targets = (self._nodetype_list, self._nodetype_list.viewport())
        if obj in _drop_targets and event.type() == QEvent.Type.DragEnter:
            if event.mimeData().hasText() and event.mimeData().text().startswith('hzp:stddev:'):
                event.acceptProposedAction()
                return True
        if obj in _drop_targets and event.type() == QEvent.Type.DragMove:
            if event.mimeData().hasText() and event.mimeData().text().startswith('hzp:stddev:'):
                event.acceptProposedAction()
                return True
        if obj in _drop_targets and event.type() == QEvent.Type.Drop:
            text = event.mimeData().text() if event.mimeData().hasText() else ''
            if text.startswith('hzp:stddev:'):
                self._handle_deviation_drop(event, obj)
                return True
        return super().eventFilter(obj, event)

    def _handle_deviation_drop(self, event, source_obj):
        text = event.mimeData().text()
        try:
            dev_id = int(text.split(':')[2])
        except (IndexError, ValueError):
            event.ignore()
            return
        pos = event.position().toPoint() if hasattr(event, 'position') else event.pos()
        # Qt delivers drop events to either the outer QListWidget or its
        # viewport depending on version/setup (same lesson as TreePanel's
        # equipment-drop handling) — itemAt() always expects viewport
        # coordinates, so remap only when the event landed on the outer
        # widget instead of the viewport directly.
        if source_obj is self._nodetype_list:
            pos = self._nodetype_list.viewport().mapFrom(self._nodetype_list, pos)
        item = self._nodetype_list.itemAt(pos)
        if item is None:
            event.ignore()
            return
        node_type_id = item.data(Qt.ItemDataRole.UserRole)
        self.db.copy_standard_deviation_to_node_type(dev_id, node_type_id)
        event.acceptProposedAction()
        if node_type_id == self._current_node_type_id():
            self._load_deviations()

    def _load_deviations(self):
        self._loading = True
        cur = self._dev_list.currentRow()
        self._dev_list.clear()
        nt_id = self._current_node_type_id()
        default_nt_id = self._node_type_ids[0] if self._node_type_ids else None
        for d in self.db.standard_deviations():
            d_nt = d['node_type_id']
            belongs = (d_nt == nt_id) or (d_nt is None and nt_id == default_nt_id)
            if not belongs:
                continue
            item = QListWidgetItem(d['description'])
            item.setData(Qt.ItemDataRole.UserRole, d['id'])
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
            self._dev_list.addItem(item)
        self._loading = False
        self._dev_list.setCurrentRow(max(0, min(cur, self._dev_list.count()-1)))

    def _current_dev_id(self):
        item = self._dev_list.currentItem()
        return item.data(Qt.ItemDataRole.UserRole) if item else None

    def _load_objects(self, dev_id=None):
        if dev_id is None:
            dev_id = self._current_dev_id()
        self._loading = True
        cur = self._obj_list.currentRow()
        self._obj_list.clear()
        if dev_id is None:
            self._loading = False; return
        show_all = self._show_all_obj_chk.isChecked()
        if show_all:
            rows = self.db.all_objects_with_cause_counts(dev_id)
        else:
            rows = self.db.objects_for_deviation(dev_id)
        for r in rows:
            # Database queries return sqlite3.Row objects; normalize before
            # optional-field access so one malformed `.get()` cannot leave
            # the entire object list empty after it was cleared.
            r = dict(r)
            label = r['name']
            n = r.get('n_causes', 0)
            if n:
                label = f"{r['name']}  ({n})"
            item = QListWidgetItem(label)
            item.setData(Qt.ItemDataRole.UserRole, r['id'])
            item.setData(Qt.ItemDataRole.UserRole + 1, r['name'])
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
            if n:
                item.setForeground(QColor('#17191C'))
            self._obj_list.addItem(item)
        try:
            self._obj_list.itemChanged.disconnect(self._on_obj_changed)
        except TypeError:
            pass   # wasn't connected yet (first call)
        self._obj_list.itemChanged.connect(self._on_obj_changed)
        self._loading = False
        self._obj_list.setCurrentRow(max(0, min(cur, self._obj_list.count()-1)))

    def _current_obj_id(self):
        item = self._obj_list.currentItem()
        return item.data(Qt.ItemDataRole.UserRole) if item else None

    def _load_causes(self, dev_id=None, obj_id=None):
        if dev_id is None: dev_id = self._current_dev_id()
        if obj_id is None: obj_id = self._current_obj_id()
        self._loading = True
        cur = self._cause_list.currentRow()
        self._cause_list.clear()
        if dev_id is None or obj_id is None:
            self._loading = False; return
        dev_item = self._dev_list.currentItem()
        obj_item = self._obj_list.currentItem()
        dev_name = dev_item.text() if dev_item else ''
        obj_name = obj_item.data(Qt.ItemDataRole.UserRole + 1) if obj_item else ''
        self._cause_lbl.setText(f"<b>Orsaker</b> — {dev_name} / {obj_name}")
        for c in self.db.standard_causes_for_object(dev_id, obj_id):
            c = dict(c)
            freq = c.get('frequency')
            label = c['description']
            if freq is not None:
                label += f"  [{freq:g}/år]"
            item = QListWidgetItem(label)
            item.setData(Qt.ItemDataRole.UserRole,     c['id'])
            item.setData(Qt.ItemDataRole.UserRole + 1, c['description'])
            item.setData(Qt.ItemDataRole.UserRole + 2, freq)
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
            self._cause_list.addItem(item)
        self._loading = False
        self._cause_list.setCurrentRow(max(0, min(cur, self._cause_list.count()-1)))
        # Clear freq field if no cause selected after reload
        if self._cause_list.currentRow() < 0:
            self._freq_edit.clear()
            self._freq_level_lbl.setText('')

    # ── Slot chains ───────────────────────────────────────────────────────────
    def _on_dev_sel(self, row):
        if self._loading: return
        dev_item = self._dev_list.item(row)
        if dev_item:
            self._obj_lbl.setText(f"<b>Objekt</b> — {dev_item.text()}")
        self._load_objects()

    def _on_obj_sel(self, row):
        if self._loading: return
        self._load_causes()

    def _on_cause_sel(self, row):
        if self._loading: return
        item = self._cause_list.item(row)
        # Populate freq field
        freq = item.data(Qt.ItemDataRole.UserRole + 2) if item else None
        self._freq_edit.blockSignals(True)
        self._freq_edit.setText(f"{freq:g}" if freq is not None else '')
        self._freq_edit.blockSignals(False)
        self._freq_level_lbl.setText(
            freq_axis_label(freq_to_f_level(freq)) if freq is not None else '')

    def _save_freq(self):
        """Save the edited frequency for the currently selected standard cause."""
        item = self._cause_list.currentItem()
        if not item: return
        cid = item.data(Qt.ItemDataRole.UserRole)
        if cid is None: return
        text = self._freq_edit.text().strip()
        if not text:
            freq = None
            self._freq_level_lbl.setText('')
        else:
            try:
                freq = float(text)
                self._freq_level_lbl.setText(freq_axis_label(freq_to_f_level(freq)))
            except ValueError:
                self._freq_level_lbl.setText('Ogiltigt')
                return
        self.db.update_standard_cause(cid, frequency=freq)
        # Update display label in list
        item.setData(Qt.ItemDataRole.UserRole + 2, freq)
        desc = item.data(Qt.ItemDataRole.UserRole + 1) or item.text()
        if freq is not None:
            item.setText(f"{desc}  [{freq:g}/år]")
        else:
            item.setText(desc)

    # ── Deviation CRUD ────────────────────────────────────────────────────────
    def _add_dev(self):
        new_id = self.db.add_standard_deviation('Ny avvikelse', self._current_node_type_id())
        item = QListWidgetItem('Ny avvikelse')
        item.setData(Qt.ItemDataRole.UserRole, new_id)
        item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
        self._dev_list.addItem(item)
        self._dev_list.editItem(item)

    def _del_dev(self):
        item = self._dev_list.currentItem()
        if not item: return
        id_ = item.data(Qt.ItemDataRole.UserRole)
        if id_ and QMessageBox.question(self, 'Ta bort', 'Ta bort avvikelse och alla dess orsaker?',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                ) == QMessageBox.StandardButton.Yes:
            self.db.delete_standard_deviation(id_)
            self._load_deviations()

    def _move_dev(self, d):
        row = self._dev_list.currentRow()
        new_row = row + d
        if not (0 <= new_row < self._dev_list.count()): return
        a = self._dev_list.takeItem(row)
        self._dev_list.insertItem(new_row, a)
        self._dev_list.setCurrentRow(new_row)
        ids = [self._dev_list.item(i).data(Qt.ItemDataRole.UserRole)
               for i in range(self._dev_list.count())]
        self.db.reorder_standard_deviations(ids)

    # ── Object CRUD (2026-08-17, see NOTES.md — same standard_objects
    # table/methods as StandardObjectsSettingsPanel's own _list) ─────────────────
    def _on_obj_changed(self, item):
        if self._loading:
            return
        id_ = item.data(Qt.ItemDataRole.UserRole)
        if id_ is None:
            return
        # Strip the "  (n)" cause-count suffix _load_objects() appends
        # for display — an edit must only ever change the object's own
        # name, never bake the count into it.
        name = re.sub(r'\s*\(\d+\)$', '', item.text()).strip()
        if name:
            self.db.update_standard_object(id_, name)
        self._load_objects()

    def _add_obj(self):
        new_id = self.db.add_standard_object('Nytt objekt')
        if not self._show_all_obj_chk.isChecked():
            # A brand-new object has zero causes yet, so it wouldn't
            # appear in the "only objects with causes" view at all —
            # switch to "Visa alla objekt" so it's actually visible to
            # rename/edit right away. Its own stateChanged already
            # triggers _load_objects().
            self._show_all_obj_chk.setChecked(True)
        else:
            self._load_objects()
        for i in range(self._obj_list.count()):
            if self._obj_list.item(i).data(Qt.ItemDataRole.UserRole) == new_id:
                self._obj_list.setCurrentRow(i)
                self._obj_list.editItem(self._obj_list.item(i))
                break

    def _del_obj(self):
        item = self._obj_list.currentItem()
        if not item:
            return
        id_ = item.data(Qt.ItemDataRole.UserRole)
        if id_ is None:
            return
        self.db.delete_standard_object(id_)
        self._load_objects()

    def _move_obj(self, direction):
        row = self._obj_list.currentRow()
        new_row = row + direction
        if not (0 <= new_row < self._obj_list.count()):
            return
        a = self._obj_list.takeItem(row)
        self._obj_list.insertItem(new_row, a)
        self._obj_list.setCurrentRow(new_row)
        ids = [self._obj_list.item(i).data(Qt.ItemDataRole.UserRole)
               for i in range(self._obj_list.count())]
        self.db.reorder_standard_objects(ids)

    # ── Cause CRUD ────────────────────────────────────────────────────────────
    def _add_cause(self):
        dev_id = self._current_dev_id()
        obj_id = self._current_obj_id()
        if dev_id is None or obj_id is None: return
        new_id = self.db.add_standard_cause_with_object(dev_id, obj_id, 'Ny orsak')
        item = QListWidgetItem('Ny orsak')
        item.setData(Qt.ItemDataRole.UserRole, new_id)
        item.setData(Qt.ItemDataRole.UserRole + 1, 'Ny orsak')
        item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
        self._cause_list.addItem(item)
        self._cause_list.editItem(item)
        self._load_objects()   # refresh object cause counts

    def _del_cause(self):
        item = self._cause_list.currentItem()
        if not item: return
        id_ = item.data(Qt.ItemDataRole.UserRole)
        if id_:
            self.db.delete_standard_cause(id_)
            row = self._cause_list.row(item)
            self._cause_list.takeItem(row)
            self._load_objects()

    def _move_cause(self, d):
        row = self._cause_list.currentRow()
        new_row = row + d
        if not (0 <= new_row < self._cause_list.count()): return
        a = self._cause_list.takeItem(row)
        self._cause_list.insertItem(new_row, a)
        self._cause_list.setCurrentRow(new_row)
        ids = [self._cause_list.item(i).data(Qt.ItemDataRole.UserRole)
               for i in range(self._cause_list.count())]
        self.db.reorder_standard_causes(ids)

    def _on_cause_changed(self, item):
        if self._loading: return
        id_ = item.data(Qt.ItemDataRole.UserRole)
        if id_:
            self.db.update_standard_cause(id_, description=item.text().strip())

    # ── Sync ──────────────────────────────────────────────────────────────────
    def _sync_freqs(self):
        ret = QMessageBox.question(self, 'Synka frekvenser',
            'Uppdatera frekvenser på alla kopplade orsaker?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No)
        if ret == QMessageBox.StandardButton.Yes:
            n = self.db.update_cause_freqs_from_standard()
            QMessageBox.information(self, 'Klart', f'{n} orsak(er) uppdaterades.')

    # ── Feature 16: Export/import standard library ────────────────────────────
    def _export_library(self):
        path, _ = QFileDialog.getSaveFileName(
            self, 'Exportera standardbibliotek', '', 'JSON (*.json)')
        if not path: return
        data = {'deviations': [], 'objects': []}
        for dev in self.db.standard_deviations():
            dd = {'description': dev['description'], 'causes': []}
            for c in self.db.standard_causes(dev['id']):
                cd = dict(c)
                dd['causes'].append({k: cd.get(k) for k in
                    ['description', 'comp_type', 'frequency', 'object_id']})
            data['deviations'].append(dd)
        for obj in self.db.standard_objects():
            data['objects'].append(obj['name'])
        import json as _json
        with open(path, 'w', encoding='utf-8') as f:
            f.write(_json.dumps(data, ensure_ascii=False, indent=2))
        QMessageBox.information(self, 'Exporterat', f'Sparat till:\n{path}')

    # ── Excel export (2026-08-26): editable, re-importable spreadsheet ────────
    def _export_library_excel(self):
        """Export every standard cause, grouped by object type (Objekttyp),
        to a plain flat .xlsx table -- one data row per standard cause, no
        merged cells -- so it stays trivially sortable/filterable in Excel
        AND re-importable later (a future importer only needs to match rows
        by their own (Objekttyp, Avvikelse, Orsak) text, the same identity
        JSON import already matches on in _import_library above -- no
        hidden id columns needed)."""
        path, _ = QFileDialog.getSaveFileName(
            self, 'Exportera standardavvikelser till Excel',
            'standardavvikelser.xlsx', 'Excel-filer (*.xlsx)')
        if not path:
            return
        if not path.lower().endswith('.xlsx'):
            path += '.xlsx'
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = 'Standardavvikelser'
        headers = ['Objekttyp', 'Avvikelse', 'Orsak', 'Frekvens (/år)']
        ws.append(headers)
        header_fill = PatternFill('solid', fgColor='1F4E79')
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='left')

        deviations = self.db.standard_deviations()
        band_fill = PatternFill('solid', fgColor='EEF2F7')
        band = False
        row = 2
        for obj in self.db.standard_objects():
            obj_rows = []
            for dev in deviations:
                for c in self.db.standard_causes_for_object(dev['id'], obj['id']):
                    obj_rows.append((dev, c))
            if not obj_rows:
                # No causes at all yet for this object type -- skip it
                # rather than emitting an empty group; nothing to edit.
                continue
            band = not band
            for dev, c in obj_rows:
                freq = c.get('frequency')
                ws.append([obj['name'], dev['description'], c['description'],
                           freq if freq is not None else None])
                if band:
                    for cell in ws[row]:
                        cell.fill = band_fill
                row += 1

        widths = {1: 28, 2: 22, 3: 60, 4: 16}
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        info = wb.create_sheet('Läs mig')
        info.column_dimensions['A'].width = 100
        for text in (
            'Denna flik listar programmets samtliga standardavvikelser, grupperade per objekttyp.',
            '',
            'Kolumner:',
            '  Objekttyp -- måste stavas exakt som i programmets objekttypslista '
            '(Inställningar -- Standardobjekt) för att kunna matchas vid en framtida import.',
            '  Avvikelse -- t.ex. "Högt flöde", "Lågt tryck".',
            '  Orsak -- fritext, en rad per orsak.',
            '  Frekvens (/år) -- lämna tom om okänd.',
            '',
            'Radera eller redigera rader fritt, eller lägg till nya längst ned i valfri grupp.',
            'Filen är ett rent tabellformat (inga sammanslagna celler) för att gå att läsa in igen.',
        ):
            info.append([text])
        info['A1'].font = Font(bold=True)
        info['A3'].font = Font(bold=True)

        wb.save(path)
        QMessageBox.information(self, 'Exporterat', f'Sparat till:\n{path}')

    def _import_library(self):
        path, _ = QFileDialog.getOpenFileName(
            self, 'Importera standardbibliotek', '', 'JSON (*.json)')
        if not path: return
        import json as _json
        try:
            with open(path, encoding='utf-8') as f:
                data = _json.loads(f.read())
        except Exception as e:
            QMessageBox.critical(self, 'Fel', str(e)); return
        added_devs = added_causes = added_objs = 0
        for obj_name in data.get('objects', []):
            if not self.db.conn.execute(
                    "SELECT id FROM standard_objects WHERE name=?", (obj_name,)).fetchone():
                self.db.add_standard_object(obj_name); added_objs += 1
        for dev_d in data.get('deviations', []):
            dev_row = self.db.conn.execute(
                "SELECT id FROM standard_deviations WHERE description=? AND active=1",
                (dev_d['description'],)).fetchone()
            if not dev_row:
                dev_id = self.db.add_standard_deviation(dev_d['description'])
                added_devs += 1
            else:
                dev_id = dev_row[0]
            for c in dev_d.get('causes', []):
                obj_id = c.get('object_id')
                if not self.db.conn.execute(
                        "SELECT id FROM standard_causes WHERE deviation_id=? AND description=? AND active=1",
                        (dev_id, c['description'])).fetchone():
                    self.db.add_standard_cause_with_object(dev_id, obj_id or 0, c['description'])
                    added_causes += 1
        self.db.conn.commit()
        self._load_deviations()
        QMessageBox.information(self, 'Importerat',
            f'Lagt till: {added_devs} avvikelser, {added_causes} orsaker, {added_objs} objekt.')
