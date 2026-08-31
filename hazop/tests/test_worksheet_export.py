import os
import sys
import tempfile
import unittest
from pathlib import Path

os.environ.setdefault('QT_QPA_PLATFORM', 'offscreen')
_TEST_DIR = Path(__file__).resolve().parent
_HAZOP_DIR = _TEST_DIR.parent
for _path in (_HAZOP_DIR, _TEST_DIR):
    if str(_path) not in sys.path:
        sys.path.insert(0, str(_path))

# Import hazop first.  Its existing module graph initialises ui_helpers before
# the Qt-backed pid_viewer imports it; worksheet_export then reuses those
# already-loaded helpers without introducing a new circular import path.
import hazop  # noqa: E402,F401
from database import Database  # noqa: E402
from worksheet_export import export_worksheet_excel  # noqa: E402


class WorksheetExportTests(unittest.TestCase):
    def setUp(self):
        self.db = Database(':memory:')
        self.system_id = self.db.conn.execute(
            'SELECT id FROM systems ORDER BY id LIMIT 1').fetchone()[0]
        # Database(':memory:') contains one seeded demo node.  It is not part
        # of this fixture and would otherwise precede the test node.
        self.db.conn.execute('DELETE FROM nodes')
        self.db.conn.commit()

    def tearDown(self):
        self.db.conn.close()

    def _node(self, name):
        return self.db.conn.execute(
            'INSERT INTO nodes(name, system_id, sort_order) VALUES (?,?,?)',
            (name, self.system_id, 1)).lastrowid

    def _deviation(self, node_id, description, sort_order):
        return self.db.conn.execute(
            'INSERT INTO deviations(node_id, description, sort_order) VALUES (?,?,?)',
            (node_id, description, sort_order)).lastrowid

    def test_export_follows_worksheet_rows_and_keeps_empty_deviations(self):
        node_id = self._node('Node A')
        # Insert in the opposite order to prove sort_order, not insertion id,
        # controls the workbook order.
        empty_deviation = self._deviation(node_id, 'Empty deviation', 1)
        populated_deviation = self._deviation(node_id, 'Populated deviation', 2)
        self.db.conn.commit()

        cause_id = self.db.add_cause(populated_deviation)
        self.db.update_cause(cause_id, description='Cause A', likelihood=2)
        consequence_id = self.db.add_consequence(cause_id)
        self.db.update_consequence(
            consequence_id, description='Consequence A', severity=3,
            category='Process')
        safeguard_1 = self.db.add_safeguard(consequence_id)
        safeguard_2 = self.db.add_safeguard(consequence_id)
        self.db.update_safeguard(safeguard_1, description='Barrier A', rrf=10)
        self.db.update_safeguard(safeguard_2, description='Barrier B', rrf=100)
        rec_1 = self.db.add_recommendation('Recommendation A')
        rec_2 = self.db.add_recommendation('Recommendation B')
        self.db.link_recommendation_to_consequence(rec_1, consequence_id)
        self.db.link_recommendation_to_consequence(rec_2, consequence_id)

        fd, path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        try:
            ok, error = export_worksheet_excel(
                self.db, path, merge_identical=False)
            self.assertTrue(ok, error)

            import openpyxl
            workbook = openpyxl.load_workbook(path)
            worksheet = workbook['HAZOP Scenario']
            self.assertEqual(worksheet['A1'].value,
                             'HAZOP Scenario – redigerbar sammanställning')
            self.assertEqual(worksheet['A3'].value, 'Nod')
            self.assertEqual(worksheet['B3'].value, 'Avvikelse')
            self.assertEqual(worksheet['K3'].value, 'Recommendation')

            # The first data row is the empty deviation.  The next two rows
            # are the shared consequence grid: one row per barrier and one
            # recommendation per row, in worksheet order.
            self.assertEqual(worksheet['B5'].value, '1. Empty deviation')
            self.assertEqual(worksheet['B6'].value, '2. Populated deviation')
            self.assertEqual(worksheet['G6'].value, '1. Barrier A')
            self.assertEqual(worksheet['G7'].value, '2. Barrier B')
            self.assertEqual(worksheet['K6'].value, '001. Recommendation A')
            self.assertEqual(worksheet['K7'].value, '002. Recommendation B')
            self.assertEqual(worksheet.max_column, 11)
            self.assertEqual(workbook.sheetnames,
                             ['HAZOP Scenario', 'Rekommendationer'])
        finally:
            os.unlink(path)


if __name__ == '__main__':
    unittest.main()
