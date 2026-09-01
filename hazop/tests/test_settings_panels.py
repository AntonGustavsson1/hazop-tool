#!/usr/bin/env python3
"""Split out of test_regression.py 2026-08-20 (see NOTES.md
"Dela upp test_regression.py i per-modul testfiler") — tests
primarily covering settings_panels.py, plus any cross-module glue they
directly depend on. Test bodies are unchanged from the
original file, only their file location moved."""
"""Regression test suite for the HAZOP PyQt6 application.

Covers crash patterns that have been found and fixed in this codebase over
recent sessions:

  1. Orphaned-data crashes: deleting a "cause" can leave orphaned
     "consequence"/"safeguard" records; P&ID overlay code that draws
     connection lines between markers used to crash with KeyError /
     AttributeError when it hit an orphaned record's missing parent
     reference.
  2. sqlite3.Row objects do not support `.get()` — several code paths used
     to call `.get()` directly on a raw Row instead of converting to a dict
     first, causing AttributeError.
  3. ComboBox `currentIndex()` returning -1 (uninitialized/empty widget)
     used to cause IndexError when used to index into arrays such as
     RRF_VALUES / SG_TYPES.
  4. A settings panel referenced `self._sev_def_panel`, which was never
     actually instantiated, causing AttributeError when deleting a
     consequence category.

Run with:
    python -m pytest hazop/test_regression.py -v
or:
    python -m unittest hazop.test_regression -v

Requires QT_QPA_PLATFORM=offscreen for headless CI environments — this is
set automatically at the top of this file, before PyQt6/hazop is imported,
so the suite runs without a display (CI, SSH, etc.).
"""

import gc
import io
import os
import sys
import shutil
import sqlite3
import tempfile
import unittest
import unittest.mock
from pathlib import Path

# ── Headless Qt setup — MUST happen before importing PyQt6 or hazop ────────
os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

# hazop.py / pid_viewer.py are large standalone scripts (not a package) that
# import each other via plain `from pid_viewer import ...`, so the hazop/
# directory must be on sys.path for those imports to resolve regardless of
# the current working directory the tests are launched from.
_TEST_DIR = Path(__file__).resolve().parent
_HAZOP_DIR = _TEST_DIR.parent
for _p in (_HAZOP_DIR, _TEST_DIR):
    if str(_p) not in sys.path:
        sys.path.insert(0, str(_p))

import hazop  # noqa: E402  (import after sys.path setup, by design)
from hazop import (  # noqa: E402
    Database, TreePanel, MainWindow,
    NODE_T, DEV_T, CAUSE_T, CONS_T, SG_T, EQUIP_T, LEDORD_T,
    freq_to_idx,
)
from PyQt6.QtWidgets import (  # noqa: E402
    QApplication, QGraphicsPixmapItem, QTreeWidgetItemIterator, QCheckBox,
    QComboBox, QPushButton, QMessageBox, QInputDialog, QLineEdit,
)
from PyQt6.QtGui import QPixmap, QFocusEvent, QKeyEvent  # noqa: E402
from PyQt6.QtCore import Qt, QPoint, QDate, QEvent, QThread, pyqtSignal  # noqa: E402
from equipment_detection import COMPONENT_TYPES  # noqa: E402


# ══════════════════════════════════════════════════════════════════════════
# Shared QApplication — Qt only allows one per process.
# ══════════════════════════════════════════════════════════════════════════


from test_helpers import (
    _ensure_qapp, _menu_action_labels, _fake_pdf_loaded,
    _TempDbMainWindow, _find_tree_item,
)

class ManualBackupButtonTests(unittest.TestCase):
    """StudyManagementPanel's "💾 Skapa säkerhetskopia nu" button forces an
    immediate, unthrottled backup and reports success/failure to the user
    (2026-08-09, see NOTES.md) — previously the only way to get a backup
    was to wait for the automatic throttled/startup ones."""

    @classmethod
    def setUpClass(cls):
        cls.app = _ensure_qapp()

    def setUp(self):
        self._tmpdir = tempfile.mkdtemp(prefix="hazop_manualbackup_test_")
        self.db = Database(path=os.path.join(self._tmpdir, "test_project.db"))
        Database._last_backup_ts = 0.0

    def tearDown(self):
        try:
            del self.db
        except Exception:
            pass
        shutil.rmtree(self._tmpdir, ignore_errors=True)

    def test_backup_now_shows_success_message_with_path(self):
        from hazop import StudyManagementPanel
        panel = StudyManagementPanel(self.db)
        try:
            with unittest.mock.patch.object(QMessageBox, 'information') as mock_info:
                panel._backup_now()
            self.assertEqual(mock_info.call_count, 1)
        finally:
            panel.deleteLater()

    def test_backup_now_shows_warning_on_failure(self):
        from hazop import StudyManagementPanel
        panel = StudyManagementPanel(self.db)
        try:
            with unittest.mock.patch.object(Database, '_write_backup', return_value=None), \
                 unittest.mock.patch.object(QMessageBox, 'warning') as mock_warn:
                panel._backup_now()
            self.assertEqual(mock_warn.call_count, 1)
        finally:
            panel.deleteLater()


class StandardCausesObjectCrudTests(unittest.TestCase):
    """"implementera de funktioner som finns i standardobjekt även i
    standard orsaker så man kan lägga till nya objekt under
    standardorsaker" (2026-08-17, see NOTES.md) — the Objekt column in
    Inställningar → Standardorsaker gets the same add/delete/reorder/
    rename CRUD StandardObjectsSettingsPanel's own list already has,
    over the identical standard_objects table."""

    @classmethod
    def setUpClass(cls):
        cls.app = _ensure_qapp()

    def setUp(self):
        self._tmpdir = tempfile.mkdtemp(prefix="hazop_stdcauses_test_")
        self.db = Database(path=os.path.join(self._tmpdir, "test_project.db"))
        self.dev_id = self.db.add_standard_deviation('Test-avvikelse')

    def tearDown(self):
        try:
            del self.db
        except Exception:
            pass
        shutil.rmtree(self._tmpdir, ignore_errors=True)

    def _make_panel(self):
        # Database() seeds a real default standard_deviations/standard_objects
        # library on construction — self.dev_id is NOT necessarily row 0.
        from hazop import StandardCausesSettingsPanel
        panel = StandardCausesSettingsPanel(self.db)
        row = next(i for i in range(panel._dev_list.count())
                   if panel._dev_list.item(i).data(Qt.ItemDataRole.UserRole) == self.dev_id)
        panel._dev_list.setCurrentRow(row)   # loads _obj_list for self.dev_id
        return panel

    def test_add_obj_creates_a_new_standard_object_and_shows_it(self):
        panel = self._make_panel()
        try:
            before = {o['id'] for o in self.db.standard_objects()}
            panel._add_obj()
            after = {o['id'] for o in self.db.standard_objects()}
            new_ids = after - before
            self.assertEqual(len(new_ids), 1)
            new_id = new_ids.pop()
            shown_ids = {panel._obj_list.item(i).data(Qt.ItemDataRole.UserRole)
                         for i in range(panel._obj_list.count())}
            self.assertIn(new_id, shown_ids,
                "a freshly-added object must be visible immediately, even with 0 causes yet")
        finally:
            panel.deleteLater()

    def test_renaming_obj_item_persists_and_strips_the_count_suffix(self):
        panel = self._make_panel()
        try:
            obj_id = self.db.add_standard_object('Gammalt namn')
            self.db.add_standard_cause_with_object(self.dev_id, obj_id, 'En orsak')
            panel._load_objects()
            row = next(i for i in range(panel._obj_list.count())
                       if panel._obj_list.item(i).data(Qt.ItemDataRole.UserRole) == obj_id)
            item = panel._obj_list.item(row)
            self.assertIn('(1)', item.text(), "sanity check: cause count suffix must be shown")

            # setText() already fires the connected itemChanged signal for
            # real (== _on_obj_changed), same as a user committing an
            # in-place edit — no separate manual call needed or wanted.
            item.setText('Nytt namn  (1)')

            updated = next(o for o in self.db.standard_objects() if o['id'] == obj_id)
            self.assertEqual(updated['name'], 'Nytt namn',
                "the count suffix must never be saved as part of the object's own name")
        finally:
            panel.deleteLater()

    def test_del_obj_removes_the_standard_object(self):
        panel = self._make_panel()
        try:
            obj_id = self.db.add_standard_object('Att ta bort')
            panel._load_objects()
            row = next(i for i in range(panel._obj_list.count())
                       if panel._obj_list.item(i).data(Qt.ItemDataRole.UserRole) == obj_id)
            panel._obj_list.setCurrentRow(row)

            panel._del_obj()

            remaining_ids = {o['id'] for o in self.db.standard_objects()}
            self.assertNotIn(obj_id, remaining_ids)
        finally:
            panel.deleteLater()

    def test_move_obj_persists_new_sort_order(self):
        panel = self._make_panel()
        try:
            id_a = self.db.add_standard_object('A-objekt')
            id_b = self.db.add_standard_object('B-objekt')
            panel._load_objects()
            row_b = next(i for i in range(panel._obj_list.count())
                         if panel._obj_list.item(i).data(Qt.ItemDataRole.UserRole) == id_b)
            panel._obj_list.setCurrentRow(row_b)

            panel._move_obj(-1)   # move B up, above A

            ordered_ids = [o['id'] for o in self.db.standard_objects()]
            self.assertLess(ordered_ids.index(id_b), ordered_ids.index(id_a))
        finally:
            panel.deleteLater()


class StandardCausesExcelExportTests(unittest.TestCase):
    """"Exportera standardavvikelser till Excel" (2026-08-26): a plain,
    re-importable-by-design .xlsx export of every standard cause, grouped
    by object type (Objekttyp) -- no merged cells, one data row per
    standard cause, so it stays trivially sortable/filterable to edit by
    hand and later re-parseable by a future importer."""

    @classmethod
    def setUpClass(cls):
        cls.app = _ensure_qapp()

    def setUp(self):
        self._tmpdir = tempfile.mkdtemp(prefix="hazop_stdcauses_xlsx_test_")
        self.db = Database(path=os.path.join(self._tmpdir, "test_project.db"))
        self.out_path = os.path.join(self._tmpdir, "export.xlsx")

    def tearDown(self):
        try:
            del self.db
        except Exception:
            pass
        shutil.rmtree(self._tmpdir, ignore_errors=True)

    def _export(self):
        from hazop import StandardCausesSettingsPanel
        panel = StandardCausesSettingsPanel(self.db)
        try:
            with unittest.mock.patch(
                    'standard_causes_panel.QFileDialog.getSaveFileName',
                    return_value=(self.out_path, 'Excel-filer (*.xlsx)')), \
                 unittest.mock.patch('standard_causes_panel.QMessageBox.information'):
                panel._export_library_excel()
        finally:
            panel.deleteLater()

    def test_export_writes_a_flat_table_grouped_by_object_type(self):
        self._export()
        self.assertTrue(os.path.exists(self.out_path))

        from openpyxl import load_workbook
        wb = load_workbook(self.out_path)
        self.assertIn('Standardavvikelser', wb.sheetnames)
        self.assertIn('Läs mig', wb.sheetnames)
        ws = wb['Standardavvikelser']

        header = [c.value for c in ws[1]]
        self.assertEqual(header, ['Objekttyp', 'Avvikelse', 'Orsak', 'Frekvens (/år)'])

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        expected_causes = self.db.conn.execute(
            "SELECT COUNT(*) FROM standard_causes").fetchone()[0]
        self.assertEqual(len(rows), expected_causes,
            "every real standard_causes row (the DB's own default seed "
            "data) must appear exactly once, no merged/grouped duplicates")

        # No merged cells anywhere -- every row must be fully self-contained
        # (that's what makes the file re-importable later).
        self.assertEqual(len(ws.merged_cells.ranges), 0)

        # Grouped by object type: once a given Objekttyp value's run of rows
        # ends, it must never reappear further down the sheet.
        seen_and_closed = set()
        current = None
        for r in rows:
            obj_name = r[0]
            if obj_name != current:
                self.assertNotIn(obj_name, seen_and_closed,
                    f"Objekttyp {obj_name!r} appeared in two separate, "
                    "non-contiguous blocks -- rows are not grouped by object type")
                if current is not None:
                    seen_and_closed.add(current)
                current = obj_name

    def test_export_includes_a_known_real_seeded_cause_with_its_frequency(self):
        self._export()
        from openpyxl import load_workbook
        wb = load_workbook(self.out_path)
        ws = wb['Standardavvikelser']
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        match = [r for r in rows if r[0] == 'Pump' and r[2] == 'Pump stopp']
        self.assertEqual(len(match), 1)
        self.assertEqual(match[0][1], 'Lågt flöde',
            "Pump stopp is seeded under the Lågt flöde deviation")
        self.assertAlmostEqual(match[0][3], 0.02)

    def test_object_types_with_zero_causes_are_omitted(self):
        empty_obj_id = self.db.add_standard_object('Helt tom objekttyp')
        self._export()
        from openpyxl import load_workbook
        wb = load_workbook(self.out_path)
        ws = wb['Standardavvikelser']
        obj_names = {r[0] for r in ws.iter_rows(min_row=2, values_only=True)}
        self.assertNotIn('Helt tom objekttyp', obj_names)

    def test_cancelling_the_save_dialog_writes_nothing(self):
        from hazop import StandardCausesSettingsPanel
        panel = StandardCausesSettingsPanel(self.db)
        try:
            with unittest.mock.patch(
                    'standard_causes_panel.QFileDialog.getSaveFileName',
                    return_value=('', '')):
                panel._export_library_excel()
            self.assertFalse(os.path.exists(self.out_path))
        finally:
            panel.deleteLater()


class StandardCausesNodeTypeTests(unittest.TestCase):
    """"Ny kolumn till vänster om Avvikelse för nodtyper (standard: en typ,
    'Processnod', men användaren ska kunna skapa fler) — med drag-and-drop
    mellan nodtyper (kopiera avvikelser)" (2026-08-17). Drag-and-drop
    COPIES (deep, independent — user confirmed via AskUserQuestion, not a
    move/link). Also covers the tab rename ("Standardorsaker" →
    "Avvikelser & Orsaker") and the full removal of the Orsaksbeskrivningar
    column."""

    @classmethod
    def setUpClass(cls):
        cls.app = _ensure_qapp()

    def setUp(self):
        self._tmpdir = tempfile.mkdtemp(prefix="hazop_nodetype_test_")
        self.db = Database(path=os.path.join(self._tmpdir, "test_project.db"))

    def tearDown(self):
        try:
            del self.db
        except Exception:
            pass
        shutil.rmtree(self._tmpdir, ignore_errors=True)

    def _make_drop_event(self, text, pos):
        from PyQt6.QtCore import QEvent, QMimeData, QPointF
        mime = QMimeData()
        mime.setText(text)
        event = unittest.mock.MagicMock()
        event.type.return_value = QEvent.Type.Drop
        event.mimeData.return_value = mime
        event.position.return_value = QPointF(pos)
        return event

    def test_tab_renamed_to_avvikelser_och_orsaker(self):
        from hazop import HAZOPPreparationPanel
        panel = HAZOPPreparationPanel(self.db)
        try:
            titles = [panel._tabs.tabText(i) for i in range(panel._tabs.count())]
            self.assertIn("Avvikelser & Orsaker", titles)
            self.assertNotIn("Standardorsaker", titles)
        finally:
            panel.deleteLater()

    def test_orsaksbeskrivningar_column_fully_removed(self):
        from hazop import StandardCausesSettingsPanel
        panel = StandardCausesSettingsPanel(self.db)
        try:
            self.assertFalse(hasattr(panel, '_desc_list'))
            self.assertFalse(hasattr(panel, '_add_desc'))
            self.assertFalse(hasattr(self.db, 'cause_descriptions'))
        finally:
            panel.deleteLater()

    def test_default_processnod_type_is_seeded(self):
        types = self.db.node_types()
        self.assertEqual(len(types), 1)
        self.assertEqual(types[0]['name'], 'Processnod')

    def test_pre_existing_deviation_with_no_type_shows_under_default(self):
        """standard_deviations.node_type_id is NULL for every pre-migration
        row — it must fall back to the default (first) node type, never be
        silently hidden."""
        dev_id = self.db.add_standard_deviation('Legacy-avvikelse')   # node_type_id=None
        from hazop import StandardCausesSettingsPanel
        panel = StandardCausesSettingsPanel(self.db)
        try:
            shown_ids = {panel._dev_list.item(i).data(Qt.ItemDataRole.UserRole)
                         for i in range(panel._dev_list.count())}
            self.assertIn(dev_id, shown_ids)
        finally:
            panel.deleteLater()

    def test_add_node_type_and_filter_deviations(self):
        from hazop import StandardCausesSettingsPanel
        panel = StandardCausesSettingsPanel(self.db)
        try:
            default_dev_id = panel._dev_list.item(0).data(Qt.ItemDataRole.UserRole)
            panel._add_node_type()
            new_type_row = panel._nodetype_list.currentRow()
            self.assertEqual(panel._nodetype_list.count(), 2)
            # The new (empty) node type shows no deviations yet.
            self.assertEqual(panel._dev_list.count(), 0)
            # Switching back to the default type shows the original deviation again.
            panel._nodetype_list.setCurrentRow(0)
            shown_ids = {panel._dev_list.item(i).data(Qt.ItemDataRole.UserRole)
                         for i in range(panel._dev_list.count())}
            self.assertIn(default_dev_id, shown_ids)
        finally:
            panel.deleteLater()

    def test_delete_last_node_type_is_blocked(self):
        from hazop import StandardCausesSettingsPanel
        panel = StandardCausesSettingsPanel(self.db)
        try:
            with unittest.mock.patch.object(QMessageBox, 'information', staticmethod(lambda *a, **k: None)):
                panel._del_node_type()
            self.assertEqual(len(self.db.node_types()), 1,
                              "the only remaining node type must not be deletable")
        finally:
            panel.deleteLater()

    def test_deleting_node_type_reassigns_its_deviations_to_default(self):
        from hazop import StandardCausesSettingsPanel
        panel = StandardCausesSettingsPanel(self.db)
        try:
            panel._add_node_type()
            new_type_id = panel._nodetype_list.currentItem().data(Qt.ItemDataRole.UserRole)
            dev_id = self.db.add_standard_deviation('I nya typen', new_type_id)
            panel._load_deviations()

            with unittest.mock.patch.object(QMessageBox, 'question',
                                             return_value=QMessageBox.StandardButton.Yes):
                panel._del_node_type()

            default_id = self.db.node_types()[0]['id']
            dev = self.db.conn.execute(
                "SELECT node_type_id FROM standard_deviations WHERE id=?", (dev_id,)).fetchone()
            self.assertIsNone(dev['node_type_id'])   # falls back to default via NULL, not re-pointed
            panel._nodetype_list.setCurrentRow(0)
            self.assertEqual(panel._nodetype_list.item(0).data(Qt.ItemDataRole.UserRole), default_id)
            shown_ids = {panel._dev_list.item(i).data(Qt.ItemDataRole.UserRole)
                         for i in range(panel._dev_list.count())}
            self.assertIn(dev_id, shown_ids)
        finally:
            panel.deleteLater()

    def test_drag_deviation_onto_node_type_creates_independent_deep_copy(self):
        """User-confirmed semantics (AskUserQuestion): dragging COPIES the
        deviation and its causes as a brand-new, fully independent row —
        not a move, not a link — editable on its own afterward."""
        src_dev_id = self.db.add_standard_deviation('Test-avvikelse för kopiering')
        obj_id = self.db.add_standard_object('Ventil-X')
        cause_id = self.db.add_standard_cause_with_object(src_dev_id, obj_id, 'Orsak A')
        self.db.update_standard_cause(cause_id, frequency=0.05)
        src_description = 'Test-avvikelse för kopiering'

        from hazop import StandardCausesSettingsPanel
        panel = StandardCausesSettingsPanel(self.db)
        try:
            panel._add_node_type()
            target_item = panel._nodetype_list.currentItem()
            target_id = target_item.data(Qt.ItemDataRole.UserRole)
            pos = panel._nodetype_list.visualItemRect(target_item).center()

            event = self._make_drop_event(f'hzp:stddev:{src_dev_id}', pos)
            handled = panel.eventFilter(panel._nodetype_list.viewport(), event)
            self.assertTrue(handled)

            devs_in_target = [d for d in self.db.standard_deviations()
                               if d['node_type_id'] == target_id]
            self.assertEqual(len(devs_in_target), 1)
            new_dev_id = devs_in_target[0]['id']
            self.assertNotEqual(new_dev_id, src_dev_id)
            self.assertEqual(devs_in_target[0]['description'], src_description)

            new_causes = self.db.standard_causes(new_dev_id)
            self.assertEqual(len(new_causes), 1)
            self.assertNotEqual(new_causes[0]['id'], cause_id)
            self.assertEqual(new_causes[0]['description'], 'Orsak A')
            self.assertEqual(new_causes[0]['frequency'], 0.05)

            # Independent afterward: editing the copy must not touch the original.
            self.db.update_standard_cause(new_causes[0]['id'], description='Orsak A (ändrad)')
            original_causes = self.db.standard_causes(src_dev_id)
            self.assertEqual(original_causes[0]['description'], 'Orsak A')
        finally:
            panel.deleteLater()


class HAZOPPreparationBladNoderTests(unittest.TestCase):
    """Fas C (2026-08-17, see NOTES.md "Blad flyttas till HAZOP preparation
    + ny Noder-flik"): "Blad" moved from Studiehantering → PID-hantering
    into HAZOPPreparationPanel, sheet names now come from the PDF filename,
    P&ID-revision is settable per blad, and a new "Noder" tab mirrors the
    HAZOP tree both ways."""

    @classmethod
    def setUpClass(cls):
        cls.app = _ensure_qapp()

    def setUp(self):
        self._tmpdir = tempfile.mkdtemp(prefix="hazop_blad_noder_test_")
        self.db = Database(path=os.path.join(self._tmpdir, "test_project.db"))

    def tearDown(self):
        try:
            del self.db
        except Exception:
            pass
        shutil.rmtree(self._tmpdir, ignore_errors=True)

    def test_blad_tab_lives_in_hazop_preparation_not_pid_management(self):
        from hazop import HAZOPPreparationPanel, PIDManagementPanel
        prep = HAZOPPreparationPanel(self.db)
        mgmt = PIDManagementPanel(self.db)
        try:
            prep_titles = [prep._tabs.tabText(i) for i in range(prep._tabs.count())]
            self.assertIn("Blad", prep_titles)
            self.assertTrue(hasattr(prep, '_sheet_list'))
            self.assertFalse(hasattr(mgmt, '_sheet_list'),
                              "Blad's sheet list must no longer live on PIDManagementPanel")
        finally:
            prep.deleteLater(); mgmt.deleteLater()

    def test_sheet_name_derived_from_pdf_filename(self):
        self.db.ensure_sheets_initialized(2, '/some/path/Process P&ID Rev A.pdf')
        sheets = self.db.get_sheets()
        self.assertEqual(sheets[0]['sheet_name'], 'Process P&ID Rev A – sida 1')
        self.assertEqual(sheets[1]['sheet_name'], 'Process P&ID Rev A – sida 2')

    def test_sheet_name_falls_back_to_generic_when_no_path_given(self):
        self.db.ensure_sheets_initialized(1)
        self.assertEqual(self.db.get_sheets()[0]['sheet_name'], 'Blad 1')

    def test_appended_sheets_also_use_filename_format(self):
        from pid_viewer import PIDPanel
        panel = PIDPanel(self.db)
        try:
            self.db.ensure_sheets_initialized(1, '/x/Original.pdf')
            rev_id = self.db.add_revision('B', '', '/x/Merged File.pdf', '2026-08-17')
            self.db.append_sheets([1, 2], ['Merged File – sida 2', 'Merged File – sida 3'], rev_id)
            names = [s['sheet_name'] for s in self.db.get_sheets()]
            self.assertIn('Merged File – sida 2', names)
            self.assertIn('Merged File – sida 3', names)
        finally:
            panel.deleteLater()

    def test_set_sheet_revision_round_trips(self):
        self.db.ensure_sheets_initialized(1, '/x/Fil.pdf')
        sheet_id = self.db.get_sheets()[0]['id']
        rev_id = self.db.add_revision('A', 'notes', '/x/Fil.pdf', '2026-08-17')
        self.db.set_sheet_revision(sheet_id, rev_id)
        self.assertEqual(self.db.get_sheets()[0]['revision_id'], rev_id)

    def test_sheet_revision_combo_reflects_selection_and_writes_on_change(self):
        from hazop import HAZOPPreparationPanel
        self.db.ensure_sheets_initialized(2, '/x/Fil.pdf')
        rev_id = self.db.add_revision('A', '', '/x/Fil.pdf', '2026-08-17')
        panel = HAZOPPreparationPanel(self.db)
        try:
            panel._sheet_list.setCurrentRow(0)
            idx = panel._sheet_rev_combo.findData(rev_id)
            panel._sheet_rev_combo.setCurrentIndex(idx)
            sheet_id = panel._sheet_list.item(0).data(Qt.ItemDataRole.UserRole)
            self.assertEqual(
                next(s for s in self.db.get_sheets() if s['id'] == sheet_id)['revision_id'],
                rev_id)
        finally:
            panel.deleteLater()

    def test_nodes_on_page_finds_nodes_via_own_pid_page_and_via_markup(self):
        node_a = self.db.add_node()
        self.db.conn.execute("UPDATE nodes SET pid_page=0 WHERE id=?", (node_a,))
        node_b = self.db.add_node()
        self.db.add_node_markup(node_b, 'polygon', [[0, 0], [1, 1]], '', '#000', 0.5, 4, 0)
        self.db.commit()
        found_ids = {n['id'] for n in self.db.nodes_on_page(0)}
        self.assertIn(node_a, found_ids)
        self.assertIn(node_b, found_ids)

    def test_noder_tab_lists_all_nodes_with_sheet_names(self):
        from hazop import HAZOPPreparationPanel
        self.db.ensure_sheets_initialized(1, '/x/MinFil.pdf')
        node_id = self.db.add_node()
        self.db.conn.execute("UPDATE nodes SET pid_page=0, name=? WHERE id=?",
                              ('Nod Alpha', node_id))
        self.db.commit()
        panel = HAZOPPreparationPanel(self.db)
        try:
            panel.refresh_nodes()
            rows = {panel._nodes_table.item(r, 1).text(): panel._nodes_table.item(r, 2).text()
                    for r in range(panel._nodes_table.rowCount())}
            self.assertIn('Nod Alpha', rows)
            self.assertIn('MinFil – sida 1', rows['Nod Alpha'])
        finally:
            panel.deleteLater()

    def test_add_node_from_noder_tab_emits_structure_changed_and_syncs_to_tree(self):
        with _TempDbMainWindow() as win:
            before = len(win.db.nodes())
            win.hazop_prep_panel._add_node_from_noder_tab()
            self.assertEqual(len(win.db.nodes()), before + 1)
            # structure_changed -> MainWindow._on_hazop_prep_structure_changed -> tree_panel.refresh()
            new_id = win.db.nodes()[-1]['id']
            item = _find_tree_item(win.tree_panel.tree, NODE_T, new_id)
            self.assertIsNotNone(item, "a node added from the Noder tab must appear in the tree")

    def test_renaming_node_from_noder_tab_syncs_to_tree(self):
        with _TempDbMainWindow() as win:
            node_id = win.db.add_node()
            win.tree_panel.refresh()
            win.hazop_prep_panel.refresh_nodes()
            row = next(r for r in range(win.hazop_prep_panel._nodes_table.rowCount())
                       if win.hazop_prep_panel._nodes_table.item(r, 1).data(
                           Qt.ItemDataRole.UserRole) == node_id)
            with unittest.mock.patch.object(
                    QInputDialog, 'getText', return_value=("Nytt namn", True)):
                win.hazop_prep_panel._on_nodes_table_double_clicked(row, 1)
            self.assertEqual(win.db.get_node(node_id)['name'], "Nytt namn")
            item = _find_tree_item(win.tree_panel.tree, NODE_T, node_id)
            self.assertIsNotNone(item)
            self.assertIn("Nytt namn", item.text(0))

    def test_renaming_node_from_tree_syncs_to_noder_tab(self):
        with _TempDbMainWindow() as win:
            node_id = win.db.add_node()
            win.tree_panel.refresh()
            with unittest.mock.patch.object(
                    QInputDialog, 'getText', return_value=("Från trädet", True)):
                win.tree_panel._rename_node(node_id)
            win.hazop_prep_panel.refresh_nodes()
            names = [win.hazop_prep_panel._nodes_table.item(r, 1).text()
                     for r in range(win.hazop_prep_panel._nodes_table.rowCount())]
            self.assertIn("Från trädet", names)


class PIDManagementPanelRevisionRefreshTests(unittest.TestCase):
    """Real crash found via the actual app (2026-08-18, not caught by the
    settings_panels.py extraction's own test run since every existing test
    constructed PIDManagementPanel against a fresh, revision-less DB):
    PIDManagementPanel.refresh() calls Path(rev['pdf_path']) for every
    revision row that has one set, but the settings_panels.py module split
    (see NOTES.md "Förenkla koden + dela upp hazop.py i fler filer") missed
    importing pathlib.Path into that file — NameError the moment a project
    with real revision history (a revision row whose pdf_path is set) was
    opened."""

    @classmethod
    def setUpClass(cls):
        cls.app = _ensure_qapp()

    def setUp(self):
        self._tmpdir = tempfile.mkdtemp(prefix="hazop_pidmgmt_revision_test_")
        self.db = Database(path=os.path.join(self._tmpdir, "test_project.db"))

    def tearDown(self):
        try:
            del self.db
        except Exception:
            pass
        shutil.rmtree(self._tmpdir, ignore_errors=True)

    def test_refresh_does_not_crash_with_a_revision_that_has_a_pdf_path(self):
        from hazop import PIDManagementPanel
        self.db.add_revision('A', '', pdf_path='C:/some/real/path.pdf')
        panel = PIDManagementPanel(self.db)
        try:
            panel.refresh()
        except NameError as e:
            self.fail(f"PIDManagementPanel.refresh() raised {e!r}")
        finally:
            panel.deleteLater()

    def test_revision_row_shows_filename_not_full_path(self):
        from hazop import PIDManagementPanel
        self.db.add_revision('A', '', pdf_path='C:/some/real/path.pdf')
        panel = PIDManagementPanel(self.db)
        try:
            panel.refresh()
            # Column 3 ("PDF-fil") holds Path(rev['pdf_path']).name, not the full path.
            self.assertEqual(panel._rev_table.item(0, 3).text(), 'path.pdf')
        finally:
            panel.deleteLater()


# ══════════════════════════════════════════════════════════════════════════
# SettingsPanel: three bundled UI changes (2026-08-11)
#   A. Riskmatris + Kategorier merged into one tab (QSplitter)
#   B. Projekt tab expanded (facility, leader, participants, date RANGE)
#   C. "P&ID" tab renamed to "P&ID-inställningar" + OCR default / page
#      orientation settings added
# ══════════════════════════════════════════════════════════════════════════

class SettingsPanelMergedRiskmatrisKategorierTests(unittest.TestCase):
    """"'riskmatris' och 'kategorier' borde gå att slå ihop till en sida.
    Testa detta." / "Låt Claude välja bästa GUI-lösningen" (2026-08-11) —
    SettingsPanel now builds a single "Riskmatris & Kategorier" tab holding
    both former tabs side-by-side in a QSplitter (categories narrow/left,
    matrix main/right — see the design-choice comment in
    SettingsPanel.__init__). Verifies the merge dropped no functionality:
    the old tab names are gone, the new combined name is present, and both
    category CRUD and matrix grid editing/save still work unchanged."""

    @classmethod
    def setUpClass(cls):
        cls.app = _ensure_qapp()

    def setUp(self):
        self._tmpdir = tempfile.mkdtemp(prefix="hazop_settings_merge_test_")
        self.db_path = os.path.join(self._tmpdir, "test_project.db")
        self.db = Database(path=self.db_path)

    def tearDown(self):
        try:
            del self.db
        except Exception:
            pass
        shutil.rmtree(self._tmpdir, ignore_errors=True)

    def _tab_titles(self, panel):
        return [panel._tabs.tabText(i) for i in range(panel._tabs.count())]

    def test_tabs_merged_into_single_named_tab(self):
        from hazop import HAZOPPreparationPanel
        panel = HAZOPPreparationPanel(self.db)
        try:
            titles = self._tab_titles(panel)
            # Renamed from "Riskmatris & Kategorier" to "Riskmatris" (2026-08-17
            # user request) — still a single merged tab, not split back apart.
            self.assertIn("Riskmatris", titles)
            self.assertNotIn("Riskmatris & Kategorier", titles)
            self.assertNotIn("Kategorier", titles)
        finally:
            panel.deleteLater()

    def test_category_add_rename_delete_still_works_from_merged_tab(self):
        from hazop import HAZOPPreparationPanel
        panel = HAZOPPreparationPanel(self.db)
        try:
            panel.db.add_category("Miljö")
            panel._load_categories()
            self.assertTrue(any(panel._cat_list.item(i).text() == "Miljö"
                                 for i in range(panel._cat_list.count())))

            for i in range(panel._cat_list.count()):
                if panel._cat_list.item(i).text() == "Miljö":
                    panel._cat_list.setCurrentRow(i)
                    break

            original_get_text = QInputDialog.getText
            QInputDialog.getText = staticmethod(lambda *a, **k: ("Miljöpåverkan", True))
            try:
                panel._cat_rename()
            finally:
                QInputDialog.getText = original_get_text
            self.assertTrue(any(panel._cat_list.item(i).text() == "Miljöpåverkan"
                                 for i in range(panel._cat_list.count())))

            for i in range(panel._cat_list.count()):
                if panel._cat_list.item(i).text() == "Miljöpåverkan":
                    panel._cat_list.setCurrentRow(i)
                    break
            panel._cat_delete()
            self.assertFalse(any(panel._cat_list.item(i).text() == "Miljöpåverkan"
                                  for i in range(panel._cat_list.count())))
        finally:
            panel.deleteLater()

    def test_matrix_editing_and_save_still_works_from_merged_tab(self):
        from hazop import HAZOPPreparationPanel
        panel = HAZOPPreparationPanel(self.db)
        try:
            saved = []
            panel.matrix_changed.connect(lambda: saved.append(True))
            panel._rows_spin.setValue(4)
            panel._cols_spin.setValue(5)
            panel._apply_size()
            self.assertEqual(len(panel._cell_buttons), 4)
            # _save_matrix() shows a blocking QMessageBox.information("Sparat", ...)
            # confirmation -- headless offscreen Qt still runs a real modal event
            # loop for exec(), so it must be stubbed out or the test hangs forever
            # waiting for a click that never comes.
            with unittest.mock.patch.object(
                    panel, '_ask_custom_matrix_template_name',
                    return_value=('Testmall', True)):
                original_information = QMessageBox.information
                QMessageBox.information = staticmethod(lambda *a, **k: None)
                try:
                    panel._save_matrix()
                finally:
                    QMessageBox.information = original_information
            self.assertTrue(saved, "matrix_changed signal should still fire on save")
            cfg = self.db.get_risk_matrix()
            self.assertEqual(cfg['rows'], 4)
            self.assertEqual(cfg['cols'], 5)
        finally:
            panel.deleteLater()

    def test_unified_matrix_save_creates_a_named_template_button(self):
        """One save action persists both the active matrix and a reusable name."""
        from hazop import HAZOPPreparationPanel
        panel = HAZOPPreparationPanel(self.db)
        try:
            self.assertFalse(hasattr(panel, '_axes_save_btn'))
            panel._rows_spin.setValue(4)
            with unittest.mock.patch.object(
                    panel, '_ask_custom_matrix_template_name',
                    return_value=('Min processmall', True)), \
                    unittest.mock.patch.object(QMessageBox, 'information'):
                panel._save_matrix()
            templates = self.db.get_custom_risk_matrix_templates()
            self.assertEqual([item['name'] for item in templates], ['Min processmall'])
            self.assertEqual(templates[0]['matrix']['rows'], 4)
            panel._rows_spin.setValue(5)
            with unittest.mock.patch.object(
                    panel, '_ask_custom_matrix_template_name',
                    return_value=('Min processmall', True)), \
                    unittest.mock.patch.object(QMessageBox, 'information'):
                panel._save_matrix()
            templates = self.db.get_custom_risk_matrix_templates()
            self.assertEqual(len(templates), 1, "a repeated name updates the existing template")
            self.assertEqual(templates[0]['matrix']['rows'], 5)
            buttons = [button for button in panel.findChildren(QPushButton)
                       if button.text() == 'Min processmall']
            self.assertTrue(buttons, "the custom template must be shown below standard templates")
            with unittest.mock.patch.object(
                    QMessageBox, 'question', return_value=QMessageBox.StandardButton.Yes):
                panel._delete_custom_matrix_template('Min processmall')
            self.assertEqual(self.db.get_custom_risk_matrix_templates(), [])
            self.assertFalse(panel._custom_matrix_templates_widget.isVisible())
        finally:
            panel.deleteLater()

    def test_left_click_matrix_cell_offers_separate_colour_and_text_edits(self):
        from hazop import HAZOPPreparationPanel
        from PyQt6.QtGui import QColor
        from PyQt6.QtWidgets import QColorDialog

        panel = HAZOPPreparationPanel(self.db)
        try:
            button = panel._cell_buttons[0][1][0]
            menu = panel._cell_edit_menu(button)
            self.assertEqual([action.text() for action in menu.actions()],
                             ['Ändra färg…', 'Ändra text…'])
            self.assertEqual([action.data() for action in menu.actions()],
                             ['color', 'text'])

            with unittest.mock.patch.object(
                    QColorDialog, 'getColor', return_value=QColor('#fef3c7')):
                panel._edit_cell_color(button)
            self.assertEqual(button.color(), '#fef3c7')
            self.assertEqual(button.fg_color(), '#000000')

            with unittest.mock.patch.object(
                    QInputDialog, 'getText', return_value=('Ny risktext', True)):
                panel._edit_cell_text(button)
            self.assertEqual(button.label(), 'Ny risktext')

            with unittest.mock.patch.object(
                    QInputDialog, 'getText', return_value=('', True)):
                panel._edit_cell_text(button)
            self.assertEqual(button.label(), '')

            source = panel._cell_buttons[0][1][1]
            target = panel._cell_buttons[0][1][2]
            source.set_cell('#123456', 'Kopierad text', '#ffffff')
            target.set_cell('#abcdef', 'Gammal text', '#000000')
            target._apply_matrix_cell_payload(source._matrix_cell_payload())
            self.assertEqual(target.color(), '#123456')
            self.assertEqual(target.label(), 'Kopierad text')
            self.assertEqual(target.fg_color(), '#ffffff')

            # Exercise the QAction emitted by a real left-click menu choice,
            # then save and verify the Scenario popup reads that text.
            menu = panel._cell_edit_menu(button)
            with unittest.mock.patch.object(
                    QInputDialog, 'getText', return_value=('Text från meny', True)):
                menu.actions()[1].trigger()
            self.assertEqual(button.label(), 'Text från meny')
            with unittest.mock.patch.object(
                    panel, '_ask_custom_matrix_template_name',
                    return_value=('Testmall', True)), \
                    unittest.mock.patch.object(QMessageBox, 'information'):
                panel._save_matrix()
            from hazop import RiskMatrixPopup
            popup = RiskMatrixPopup(button.col - 1, button.row + 1)
            try:
                self.assertEqual(
                    popup._grid_buttons[(button.col - 1, button.row + 1)][1],
                    'Text från meny')
            finally:
                popup.close()
        finally:
            panel.deleteLater()

    def test_saved_axis_codes_and_popup_cache_follow_current_matrix(self):
        """Edited consequence/frequency codes must reach the risk popup.

        Exercise both reversed display directions and deliberately prime the
        global cache with another project first.  This is the exact route
        that previously let the popup retain an older matrix after Save.
        """
        from hazop import HAZOPPreparationPanel
        from database import get_matrix, load_matrix

        other_db = Database(path=os.path.join(self._tmpdir, 'other_project.db'))
        other_db.set_risk_matrix({
            'rows': 2, 'cols': 2,
            'x_labels': ['old F0', 'old F1'],
            'y_labels': ['old C1', 'old C2'],
        })
        load_matrix(other_db)
        panel = HAZOPPreparationPanel(self.db)
        try:
            # Frequency on X, with both axes reversed.  The visible left/top
            # labels are respectively the semantic highest/lowest entries.
            panel._x_rev_chk.setChecked(True)
            panel._y_rev_chk.setChecked(True)
            for i, edit in enumerate(panel._x_label_edits):
                edit.setText(f'F-visible-{i}')
            for i, edit in enumerate(panel._y_label_edits):
                edit.setText(f'C-visible-{i}')

            with unittest.mock.patch.object(
                    panel, '_ask_custom_matrix_template_name',
                    return_value=('Testmall 1', True)), \
                    unittest.mock.patch.object(QMessageBox, 'information'):
                panel._save_matrix()

            cfg = self.db.get_risk_matrix()
            self.assertEqual(cfg['x_codes'], [
                f'F-visible-{i}' for i in reversed(range(len(panel._x_label_edits)))])
            self.assertEqual(cfg['y_codes'], [
                f'C-visible-{i}' for i in range(len(panel._y_label_edits))])
            self.assertEqual(get_matrix()['x_codes'], cfg['x_codes'])
            self.assertEqual(get_matrix()['y_codes'], cfg['y_codes'])

            # Swap axes too: consequence now uses the X mapping and
            # frequency uses the Y mapping, so both label families remain
            # correct through the same Save action.
            panel._axis_combo.setCurrentIndex(
                panel._axis_combo.findData('consequence'))
            panel._x_rev_chk.setChecked(True)
            panel._y_rev_chk.setChecked(False)
            for i, edit in enumerate(panel._x_label_edits):
                edit.setText(f'C2-visible-{i}')
            for i, edit in enumerate(panel._y_label_edits):
                edit.setText(f'F2-visible-{i}')
            with unittest.mock.patch.object(
                    panel, '_ask_custom_matrix_template_name',
                    return_value=('Testmall 2', True)), \
                    unittest.mock.patch.object(QMessageBox, 'information'):
                panel._save_matrix()

            cfg = self.db.get_risk_matrix()
            self.assertEqual(cfg['y_codes'], [
                f'C2-visible-{i}' for i in reversed(range(len(panel._x_label_edits)))])
            self.assertEqual(cfg['x_codes'], [
                f'F2-visible-{i}' for i in reversed(range(len(panel._y_label_edits)))])
            self.assertEqual(get_matrix()['x_codes'], cfg['x_codes'])
            self.assertEqual(get_matrix()['y_codes'], cfg['y_codes'])
        finally:
            panel.deleteLater()
            other_db.conn.close()

    def test_st1_matrix_preset_loads_without_saving_immediately(self):
        from hazop import HAZOPPreparationPanel
        panel = HAZOPPreparationPanel(self.db)
        try:
            before = self.db.get_risk_matrix()
            panel._apply_st1_preset()
            self.assertEqual(panel._rows_spin.value(), 6)
            self.assertEqual(panel._cols_spin.value(), 5)
            self.assertTrue(panel._y_rev_chk.isChecked())
            self.assertEqual(len(panel._cell_buttons), 6)
            self.assertEqual(len(panel._cell_buttons[0][1]), 5)
            self.assertEqual(panel._last_built_cfg['x_codes'], ['A', 'B', 'C', 'D', 'E'])
            self.assertEqual(panel._last_built_cfg['y_codes'], ['0', '1', '2', '3', '4', '5'])
            self.assertEqual([edit.text() for edit in panel._x_label_edits],
                             ['A', 'B', 'C', 'D', 'E'])
            self.assertTrue(all(panel._last_built_cfg['x_labels']))
            self.assertTrue(all(panel._last_built_cfg['y_labels']))
            self.assertEqual(len(panel._last_built_cfg['cell_colors']), 6)
            self.assertTrue(all(len(row) == 5
                                for row in panel._last_built_cfg['cell_colors']))
            self.assertEqual(self.db.get_risk_matrix(), before,
                             'choosing a preset must not save before Spara riskmatris')
        finally:
            panel.deleteLater()

    def test_axis_reverse_buttons_are_checkable_toolbuttons(self):
        """"Ersätt kryssrutorna 'Vänd X'/'Vänd Y' med klickbara pilar"
        (2026-08-17 user request) — QToolButton in checkable mode replaces
        QCheckBox, but every downstream call site only ever used
        isChecked()/setChecked()/toggled, so the matrix-rebuild behavior
        must be unchanged."""
        from hazop import HAZOPPreparationPanel
        from PyQt6.QtWidgets import QToolButton, QCheckBox
        panel = HAZOPPreparationPanel(self.db)
        try:
            self.assertIsInstance(panel._x_rev_chk, QToolButton)
            self.assertIsInstance(panel._y_rev_chk, QToolButton)
            self.assertNotIsInstance(panel._x_rev_chk, QCheckBox)
            self.assertTrue(panel._x_rev_chk.isCheckable())
            self.assertTrue(panel._y_rev_chk.isCheckable())

            self.assertFalse(panel._x_rev_chk.isChecked())
            panel._x_rev_chk.setChecked(True)
            self.assertTrue(panel._x_rev_chk.isChecked())
            # toggled -> _apply_size() rebuilds the in-memory grid immediately;
            # persisting to DB only happens on explicit "Spara riskmatris".
            self.assertTrue(panel._last_built_cfg.get('x_reversed'))
        finally:
            panel.deleteLater()

    def test_boundary_edit_does_not_overwrite_frequency_code(self):
        """"Visa '<'-tecknet korrekt i gränsvärden (t.ex. '< 0.1')"
        (2026-08-17). Root cause: the axis-label QLineEdit is a fixed 80px
        wide, and QLineEdit.setText() leaves the cursor at the END of the
        text — a label like "< 0.1/år" is wider than the field at the 8px
        header font, so the widget auto-scrolled to keep the cursor
        visible, hiding the leading "<". Fix: setCursorPosition(0) after
        every setText() on these label edits, both at initial build and
        whenever _sync_freq_label_from_boundary regenerates one."""
        from hazop import HAZOPPreparationPanel
        panel = HAZOPPreparationPanel(self.db)
        try:
            # Freshly built grid — every column header must show from the start.
            for e in panel._x_label_edits:
                self.assertEqual(e.cursorPosition(), 0)

            # Editing a boundary regenerates the two adjacent labels — must
            # also reset to the start, even though the new text is longer
            # than the 80px field.
            panel._x_label_edits[0].setText("A")
            panel._freq_boundary_edits[0].setText("0.1")
            panel._sync_freq_label_from_boundary(panel._freq_boundary_edits[0], 0)
            self.assertEqual(panel._x_label_edits[0].text(), "A")
        finally:
            panel.deleteLater()

    def test_blank_cell_and_space_description_survive_axis_swap(self):
        """Changing presentation must preserve deliberately blank data."""
        from hazop import HAZOPPreparationPanel
        panel = HAZOPPreparationPanel(self.db)
        try:
            blank = panel._cell_buttons[0][1][0]
            blank.set_cell(blank.color(), '', blank.fg_color())
            blank_row, blank_col = blank.row, blank.col
            panel._last_built_cfg['x_labels'][0] = '   '
            panel._axis_combo.setCurrentIndex(
                panel._axis_combo.findData('consequence'))
            rebuilt = next(
                btn for _row, buttons in panel._cell_buttons for btn in buttons
                if btn.row == blank_row and btn.col == blank_col)
            self.assertEqual(rebuilt.label(), '')
            self.assertEqual(panel._last_built_cfg['x_labels'][0], '   ')
        finally:
            panel.deleteLater()

    def test_axes_save_editable_codes_used_by_matrix_popup(self):
        """Codes and descriptions are separate editable fields in Axlar."""
        from hazop import HAZOPPreparationPanel, RiskMatrixPopup
        from PyQt6.QtWidgets import QLabel
        panel = HAZOPPreparationPanel(self.db)
        try:
            panel._set_risk_subview(1)
            panel._frequency_axis_table.item(0, 0).setText('A')
            panel._frequency_axis_table.item(0, 1).setText('Aldrig')
            panel._consequence_axis_table.item(0, 0).setText('0')
            panel._consequence_axis_table.item(0, 1).setText('Ingen skada')
            with unittest.mock.patch.object(
                    panel, '_ask_custom_matrix_template_name',
                    return_value=('Testmall', True)), \
                    unittest.mock.patch.object(QMessageBox, 'information'):
                panel._save_axes_and_categories()
            cfg = self.db.get_risk_matrix()
            self.assertEqual(cfg['x_codes'][0], 'A')
            self.assertEqual(cfg['x_labels'][0], 'Aldrig')
            self.assertEqual(cfg['y_codes'][0], '0')
            self.assertEqual(cfg['y_labels'][0], 'Ingen skada')
            popup = RiskMatrixPopup(-1, 1)
            try:
                visible = [label.text() for label in popup.findChildren(QLabel)]
                self.assertIn('A', visible)
                self.assertIn('0', visible)
            finally:
                popup.close()
        finally:
            panel.deleteLater()

    def test_axes_view_accepts_a_multirow_excel_paste_and_saves_descriptions(self):
        """Five copied Excel cells map to five consequence rows in one category.

        Description editing deliberately lives outside the visual matrix, so
        wrapped text cannot change the size or placement of risk cells.
        """
        from hazop import HAZOPPreparationPanel
        panel = HAZOPPreparationPanel(self.db)
        try:
            cat_id = self.db.add_category("Miljö")
            panel._load_categories()
            panel._set_risk_subview(1)
            table = panel._category_definition_table
            category_column = next(
                col for col in range(1, table.columnCount())
                if table.item(0, col).data(Qt.ItemDataRole.UserRole) == cat_id)
            table.setCurrentCell(0, category_column)
            QApplication.clipboard().setText(
                "Nivå 1\nNivå 2\nNivå 3\nNivå 4\nNivå 5")
            paste = QKeyEvent(QEvent.Type.KeyPress, Qt.Key.Key_V,
                              Qt.KeyboardModifier.ControlModifier)
            self.assertTrue(panel.eventFilter(table, paste))
            self.assertEqual(
                [table.item(row, category_column).text() for row in range(5)],
                ["Nivå 1", "Nivå 2", "Nivå 3", "Nivå 4", "Nivå 5"])
            self.assertEqual(self.db.get_severity_definitions().get(1, {}).get(cat_id, ''),
                             '', "paste is a working copy until Spara")

            with unittest.mock.patch.object(
                    panel, '_ask_custom_matrix_template_name',
                    return_value=('Testmall', True)), \
                    unittest.mock.patch.object(QMessageBox, 'information'):
                panel._save_axes_and_categories()
            definitions = self.db.get_severity_definitions()
            self.assertEqual(
                [definitions.get(row, {}).get(cat_id, '') for row in range(1, 6)],
                ["Nivå 1", "Nivå 2", "Nivå 3", "Nivå 4", "Nivå 5"])
            saved_category = next(category for category in self.db.get_risk_matrix()
                                  ['consequence_categories']
                                  if category['name'] == 'Miljö')
            self.assertEqual(saved_category['descriptions'],
                             ["Nivå 1", "Nivå 2", "Nivå 3", "Nivå 4", "Nivå 5"])
        finally:
            panel.deleteLater()

    def test_axes_is_a_local_view_and_matrix_columns_keep_a_common_width(self):
        from hazop import HAZOPPreparationPanel
        panel = HAZOPPreparationPanel(self.db)
        try:
            self.assertEqual(panel._risk_substack.count(), 2)
            self.assertEqual(panel._risk_substack.currentIndex(), 0)
            panel._set_risk_subview(1)
            self.assertEqual(panel._risk_substack.currentIndex(), 1)
            self.assertTrue(panel._category_definition_table.isVisible() or
                            panel._category_definition_table.parentWidget() is not None)

            panel._on_matrix_cell_width_changed(118)
            row_buttons = panel._cell_buttons[0][1]
            self.assertTrue(all(btn.maximumWidth() == 118 for btn in row_buttons))
            self.assertNotIn('margin:-1px', row_buttons[0].styleSheet())
        finally:
            panel.deleteLater()

    def test_template_migration_dialog_keeps_mapping_state_and_row_override_separate(self):
        from hazop_preparation_panel import RiskMatrixMigrationDialog
        node_id = self.db.add_node()
        dev_id = self.db.deviations(node_id)[0]['id']
        cause_id = self.db.add_cause(dev_id)
        cons_id = self.db.add_consequence(cause_id)
        self.db.update_cause(cause_id, description='P-101 felar', likelihood=0)
        self.db.update_consequence(cons_id, 'Trycket ökar', 2)
        source = {'rows': 3, 'cols': 3,
                  'x_labels': ['A', 'B', 'C'], 'y_labels': ['Låg', 'Mellan', 'Hög'],
                  'freq_boundaries': [0.01, 1.0]}
        target = {'rows': 4, 'cols': 2,
                  'x_labels': ['1', '2'], 'y_labels': ['1', '2', '3', '4'],
                  'freq_boundaries': [0.1]}
        self.db.set_risk_matrix(source)
        dialog = RiskMatrixMigrationDialog(self.db, source, target, 'Testmall')
        try:
            # This is the same entry point used by a chip drag/drop.  It
            # changes the shared visual mapping state, not data in DB.
            dialog._on_level_dropped('frequency', 0, -1)
            self.assertEqual(dialog.plan['frequency_map']['0'], -1)
            self.assertEqual(dialog._mapping[('frequency', 0)], -1)
            record = next(row for row in dialog.plan['frequency_records']
                          if row['cause_id'] == cause_id)
            self.assertEqual(record['target'], -1)

            # A user can split the same source level through a row-specific
            # decision without changing the drag/drop default for others.
            row_index = dialog.plan['frequency_records'].index(record)
            dialog._set_record_mapping('frequency_records', row_index, 0)
            self.assertEqual(record['target'], 0)
            self.assertTrue(record['override'])
            self.assertEqual(dialog.plan['frequency_map']['0'], -1)

            # Both tabs read the same state. A mapped source chip is visible
            # in both collections, without copying an independent map.
            self.assertIn(('frequency', 0), dialog._link_field.old_chips)
            self.assertIn(('frequency', 0), dialog._matrix_against_matrix.old_chips)

            # Changing the display axis deliberately starts a fresh mapping
            # session, per the migration handoff specification.
            dialog._swap_target_axes()
            self.assertEqual(dialog.plan['frequency_map'], {})
            self.assertEqual(dialog._mapping, {})
            self.assertEqual(dialog.display_x_axis, 'severity')
        finally:
            dialog.deleteLater()

    def test_template_migration_first_page_is_compact_and_second_page_keeps_room(self):
        from hazop_preparation_panel import RiskMatrixMigrationDialog
        source = {
            'rows': 2, 'cols': 2, 'x_labels': ['A', 'B'],
            'y_labels': ['1', '2'], 'freq_boundaries': [0.1],
        }
        target = {
            'rows': 2, 'cols': 2, 'x_labels': ['1', '2'],
            'y_labels': ['1', '2'], 'freq_boundaries': [0.1],
        }
        self.db.set_risk_matrix(source)
        dialog = RiskMatrixMigrationDialog(self.db, source, target, 'Testmall')
        try:
            self.assertEqual(dialog._pages.currentIndex(), 0)
            self.assertLess(dialog.height(), 600)

            dialog._show_axis_page()
            self.assertEqual(dialog.height(), dialog._axis_page_height)

            dialog._show_category_page()
            self.assertEqual(dialog.height(), dialog._category_page_height)
        finally:
            dialog.deleteLater()

    def test_template_migration_exposes_category_and_matrix_mapping_views(self):
        """Category matching uses the same interactive mapping canvas as axes."""
        from hazop_preparation_panel import RiskMatrixMigrationDialog
        source = {
            'rows': 2, 'cols': 2, 'x_axis': 'frequency', 'y_reversed': True,
            'x_codes': ['A', 'B'], 'y_codes': ['0', '1'],
            'x_labels': ['Sällsynt händelse', 'Har inträffat'],
            'y_labels': ['Ingen skada', 'Mindre skada'],
            'cell_colors': [['#123456', '#234567'], ['#345678', '#456789']],
            'cell_labels': [['Låg', 'Mellan'], ['Mellan', 'Hög']],
            'freq_boundaries': [0.1],
        }
        target = {
            'rows': 2, 'cols': 2, 'x_axis': 'frequency', 'y_reversed': True,
            'x_codes': ['1', '2'], 'y_codes': ['1', '2'],
            'x_labels': ['Låg sannolikhet', 'Hög sannolikhet'],
            'y_labels': ['Liten konsekvens', 'Stor konsekvens'],
            'cell_colors': [['#abcdef', '#bcdef0'], ['#cdef01', '#def012']],
            'cell_labels': [['Låg', 'Mellan'], ['Mellan', 'Hög']],
            'freq_boundaries': [0.1],
        }
        self.db.set_risk_matrix(source)
        dialog = RiskMatrixMigrationDialog(self.db, source, target, 'Testmall')
        try:
            self.assertEqual([dialog._tabs.tabText(i) for i in range(dialog._tabs.count())],
                             ['1. Kopplingsfält', '2. Matris mot matris'])
            self.assertEqual(dialog._pages.currentIndex(), 0)
            self.assertTrue(hasattr(dialog, '_category_panel'))
            source_chip = dialog._link_field.old_chips[('frequency', -1)]
            target_chip = dialog._link_field.target_chips[('frequency', -1)]
            self.assertEqual(source_chip.code, 'A')
            self.assertEqual(source_chip.toolTip(), 'Sällsynt händelse')
            self.assertEqual(target_chip.code, '1')
            self.assertEqual(
                dialog._matrix_against_matrix.old_chips[('severity', 1)].code, '0')

            dialog.clear_mappings()
            dialog.activate_old_step('frequency', -1)
            dialog.activate_target_step('frequency', 0)
            self.assertEqual(dialog._mapping[('frequency', -1)], 0)
            self.assertTrue(dialog._link_field.old_chips[('frequency', -1)].isEnabled())
            self.assertTrue(dialog._matrix_against_matrix.old_chips[('frequency', -1)].isEnabled())
            self.assertIn('1 av 4 gamla steg mappade.', dialog._progress.text())
            self.assertIn('konsekvenskategorier mappade.', dialog._progress.text())
        finally:
            dialog.deleteLater()

    def test_template_migration_allows_category_mapping_in_its_own_tab(self):
        from hazop_preparation_panel import RiskMatrixMigrationDialog
        source = {'rows': 3, 'cols': 2, 'x_labels': ['A', 'B'],
                  'y_labels': ['1', '2', '3'], 'freq_boundaries': [0.1]}
        target = {
            'rows': 4, 'cols': 2, 'x_labels': ['L', 'H'],
            'y_labels': ['1', '2', '3', '4'], 'freq_boundaries': [0.1],
            'consequence_categories': [
                {'key': key, 'name': name, 'color': color,
                 'descriptions': [f'{name} {level}' for level in range(1, 5)]}
                for key, name, color in [
                    ('person', 'Människor', '#2563eb'), ('miljo', 'Miljö', '#16a34a'),
                    ('ekonomi', 'Ekonomi', '#d97706'), ('assets', 'Tillgångar', '#7c3aed'),
                    ('rykte', 'Rykte', '#475569')]
            ],
        }
        self.db.set_risk_matrix(source)
        dialog = RiskMatrixMigrationDialog(self.db, source, target, 'Testmall')
        try:
            self.assertEqual(dialog._tabs.count(), 2)
            self.assertEqual(dialog._tabs.tabText(0), '1. Kopplingsfält')
            self.assertEqual(dialog._pages.currentIndex(), 0)
            source_id = str(dialog.plan['source_categories'][0]['source_id'])
            self.assertEqual(dialog.plan['category_map'][source_id], 'person')
            canvas = dialog._category_panel._mapping_canvas
            self.assertIn(source_id, dialog._category_panel.source_chips)
            self.assertIn('person', dialog._category_panel.target_chips)
            self.assertIs(canvas.old_chips[('category', source_id)],
                          dialog._category_panel.source_chips[source_id])
            self.assertLess(dialog._category_panel._mapping_canvas.geometry().y(), 50)
            self.assertIn('konsekvenskategorier mappade.', dialog._progress.text())

            # Category links are a required first page.  The next page is
            # available only after every source category has a destination.
            dialog._show_axis_page()
            self.assertEqual(dialog._pages.currentIndex(), 1)
            self.assertTrue(dialog._global_scope_button.isChecked())

            dialog.clear_mappings()
            dialog.activate_source_category(source_id)
            dialog.activate_target_category('person')
            self.assertEqual(dialog.plan['category_map'][source_id], 'person')
            self.assertEqual(dialog.category_target_count('person'), 1)
        finally:
            dialog.deleteLater()

    def test_category_mapping_can_be_changed_without_touching_axis_mapping(self):
        from hazop_preparation_panel import RiskMatrixMigrationDialog
        source = {
            'rows': 2, 'cols': 2, 'x_labels': ['A', 'B'],
            'y_labels': ['1', '2'], 'freq_boundaries': [0.1],
            'consequence_categories': [
                {'key': 'old-person', 'name': 'Person', 'color': '#2563eb'},
                {'key': 'old-asset', 'name': 'Tillgångar', 'color': '#d97706'},
            ],
        }
        target = {
            'rows': 2, 'cols': 2, 'x_labels': ['1', '2'],
            'y_labels': ['1', '2'], 'freq_boundaries': [0.1],
            'consequence_categories': [
                {'key': 'new-person', 'name': 'Människor', 'color': '#16a34a'},
                {'key': 'new-assets', 'name': 'Assets', 'color': '#7c3aed'},
            ],
        }
        self.db.set_risk_matrix(source)
        dialog = RiskMatrixMigrationDialog(self.db, source, target, 'Testmall')
        try:
            self.assertTrue(hasattr(dialog, '_category_panel'))
            self.assertEqual(dialog._pages.currentIndex(), 0)
            source_id = str(dialog.plan['source_categories'][0]['source_id'])
            dialog.clear_mappings()
            dialog.activate_source_category(source_id)
            dialog.activate_target_category('new-person')
            self.assertEqual(dialog.plan['category_map'][source_id], 'new-person')
            self.assertEqual(dialog._mapping, {})
            self.assertEqual(dialog._category_panel._mapping_canvas.iter_mappings(),
                             [(('category', source_id), 'new-person')])
        finally:
            dialog.deleteLater()

    def test_category_severity_calibration_inherits_global_mapping_and_allows_override(self):
        """Global severity mapping is shared, with optional per-category links."""
        from hazop_preparation_panel import RiskMatrixMigrationDialog
        source = {
            'rows': 3, 'cols': 2, 'x_labels': ['A', 'B'],
            'y_labels': ['1', '2', '3'], 'freq_boundaries': [0.1],
            'consequence_categories': [
                {'key': 'old-person', 'name': 'Person', 'color': '#2563eb'},
                {'key': 'old-asset', 'name': 'Tillgångar', 'color': '#d97706'},
            ],
        }
        target = {
            'rows': 4, 'cols': 2, 'x_labels': ['1', '2'],
            'y_labels': ['1', '2', '3', '4'], 'freq_boundaries': [0.1],
            'consequence_categories': [
                {'key': 'new-person', 'name': 'Människor', 'color': '#16a34a'},
                {'key': 'new-assets', 'name': 'Assets', 'color': '#7c3aed'},
            ],
        }
        # Match the active category count so the real Next validation can be
        # exercised without opening its warning dialog.
        for category in list(self.db.consequence_categories()):
            self.db.delete_category(category['id'])
        self.db.add_category('Person')
        self.db.add_category('Tillgångar')
        self.db.set_risk_matrix(source)
        dialog = RiskMatrixMigrationDialog(self.db, source, target, 'Testmall')
        try:
            source_id = str(dialog.plan['source_categories'][0]['source_id'])
            dialog._show_axis_page()
            dialog._select_category_scope(source_id)
            self.assertFalse(dialog._category_scope_stack.isHidden())
            self.assertTrue(dialog._tabs.isHidden())
            dialog._select_global_scope()
            self.assertFalse(dialog._tabs.isHidden())

            dialog.set_axis_mapping('severity', 1, 4)
            self.assertEqual(dialog.category_level_target(source_id, 1), 4)
            dialog.remove_axis_mapping('severity', 1)
            self.assertEqual(dialog.category_level_target(source_id, 1), 1)
            dialog.set_axis_mapping('severity', 1, 4)

            dialog.set_category_severity_mapping(source_id, 1, 2)
            self.assertEqual(dialog.category_level_target(source_id, 1), 2)
            self.assertEqual(dialog.plan['category_severity_maps'][source_id]['1'], 2)

            # Later global changes affect categories without an override, but
            # must not overwrite the explicitly calibrated category.
            dialog.set_axis_mapping('severity', 2, 3)
            self.assertEqual(dialog.category_level_target(source_id, 1), 2)
            self.assertEqual(dialog.category_level_target(source_id, 2), 3)

            # The page-two Rensa action is deliberately axis-only.  Category
            # links and a category-specific override must survive it.
            dialog.set_category_mapping(source_id, 'new-person')
            dialog.clear_axis_mappings()
            self.assertEqual(dialog.plan['category_map'][source_id], 'new-person')
            self.assertEqual(dialog.category_level_target(source_id, 1), 2)

            dialog.reset_category_severity_mapping(source_id)
            self.assertEqual(dialog.category_level_target(source_id, 1), 1)
            self.assertNotIn('1', dialog._category_severity_overrides[source_id])
        finally:
            dialog.deleteLater()

    def test_deleting_category_refreshes_matrix_cell_buttons(self):
        """'När jag lägger till eller tar bort en konsekvenskategori skall
        detta synas i riskmatrisen direkt.' (2026-08-11) — _cat_add already
        called _apply_size() to rebuild the matrix grid; _cat_delete did
        not, so the matrix silently kept its old grid after a delete. Use
        _cell_buttons' identity (not just its length, which a fixed-size
        matrix wouldn't change) to prove a real rebuild happened."""
        from hazop import HAZOPPreparationPanel
        panel = HAZOPPreparationPanel(self.db)
        try:
            panel.db.add_category("Miljö")
            panel._load_categories()
            for i in range(panel._cat_list.count()):
                if panel._cat_list.item(i).text() == "Miljö":
                    panel._cat_list.setCurrentRow(i)
                    break

            buttons_before = panel._cell_buttons
            panel._cat_delete()
            self.assertIsNot(panel._cell_buttons, buttons_before,
                "_cell_buttons must be a freshly rebuilt list after deleting "
                "a category — same object identity means _apply_size() never ran")
        finally:
            panel.deleteLater()

    def test_reorder_categories_persists_new_order(self):
        """'Jag vill även kunna justera ordningen, exempelvis genom vilken
        ordning de dyker upp.' (2026-08-11) — up/down buttons move the
        selected category and persist the new order via
        Database.reorder_categories(), which consequence_categories()
        (ORDER BY sort_order, name) then reflects."""
        from hazop import HAZOPPreparationPanel
        panel = HAZOPPreparationPanel(self.db)
        try:
            # Database() seeds default categories (Person/Miljö/Ekonomi/...)
            # on creation -- clear them so this test only has to reason
            # about its own two categories' relative order.
            for cat in list(panel.db.consequence_categories()):
                panel.db.delete_category(cat['id'])
            a_id = panel.db.add_category("Alfa")
            b_id = panel.db.add_category("Beta")
            panel.db.reorder_categories([a_id, b_id])
            panel._load_categories()
            self.assertEqual(panel._cat_list.item(0).text(), "Alfa")
            self.assertEqual(panel._cat_list.item(1).text(), "Beta")

            panel._cat_list.setCurrentRow(1)   # select "Beta"
            panel._cat_move(-1)                # move it up

            names_in_ui = [panel._cat_list.item(i).text() for i in range(panel._cat_list.count())]
            self.assertEqual(names_in_ui, ["Beta", "Alfa"])
            names_in_db = [dict(c)['name'] for c in panel.db.consequence_categories()]
            self.assertEqual(names_in_db, ["Beta", "Alfa"],
                "new order must be persisted, not just reflected in the UI list")
        finally:
            panel.deleteLater()

    def test_move_up_at_top_and_move_down_at_bottom_are_no_ops(self):
        from hazop import HAZOPPreparationPanel
        panel = HAZOPPreparationPanel(self.db)
        try:
            for cat in list(panel.db.consequence_categories()):
                panel.db.delete_category(cat['id'])
            a_id = panel.db.add_category("Alfa")
            b_id = panel.db.add_category("Beta")
            panel.db.reorder_categories([a_id, b_id])
            panel._load_categories()

            panel._cat_list.setCurrentRow(0)
            panel._cat_move(-1)   # already at top -- must not raise or reorder
            names = [panel._cat_list.item(i).text() for i in range(panel._cat_list.count())]
            self.assertEqual(names, ["Alfa", "Beta"])

            panel._cat_list.setCurrentRow(1)
            panel._cat_move(1)    # already at bottom -- must not raise or reorder
            names = [panel._cat_list.item(i).text() for i in range(panel._cat_list.count())]
            self.assertEqual(names, ["Alfa", "Beta"])
        finally:
            panel.deleteLater()


class SettingsPanelProjektExpansionTests(unittest.TestCase):
    """"Fliken projekt innehåller bara Projektnamn, datum och revision,
    utveckla detta. Gör så att datum kan väljas inom ett intervall osv." /
    "Även Anläggning, HAZOP-ledare, Deltagare" (2026-08-11) — three new
    fields (Anläggning, HAZOP-ledare, Deltagare) plus a start/end
    QDateEdit date range replacing the old single free-text 'Datum' field.
    Verifies every field round-trips through db.get_config/set_config,
    including the new date-range keys ('project_date_start' /
    'project_date_end')."""

    @classmethod
    def setUpClass(cls):
        cls.app = _ensure_qapp()

    def setUp(self):
        self._tmpdir = tempfile.mkdtemp(prefix="hazop_settings_projekt_test_")
        self.db_path = os.path.join(self._tmpdir, "test_project.db")
        self.db = Database(path=self.db_path)

    def tearDown(self):
        try:
            del self.db
        except Exception:
            pass
        shutil.rmtree(self._tmpdir, ignore_errors=True)

    def test_facility_leader_round_trip(self):
        from hazop import HAZOPPreparationPanel
        panel = HAZOPPreparationPanel(self.db)
        try:
            panel._proj_facility.setText("Gävle Depå")
            panel._proj_facility.editingFinished.emit()
            panel._proj_number.setText("P-2026-042")
            panel._proj_number.editingFinished.emit()
            panel._proj_client.setText("Hybrit AB")
            panel._proj_client.editingFinished.emit()

            self.assertEqual(self.db.get_config('project_facility'), "Gävle Depå")
            self.assertEqual(self.db.get_config('project_number'), "P-2026-042")
            self.assertEqual(self.db.get_config('project_client'), "Hybrit AB")
            # HAZOP-ledare removed (2026-08-17) — the role now lives as a free
            # Deltagare column instead, see NOTES.md.
            self.assertFalse(hasattr(panel, '_proj_leader'))
        finally:
            panel.deleteLater()

    def test_date_range_round_trips_through_config(self):
        from hazop import HAZOPPreparationPanel
        panel = HAZOPPreparationPanel(self.db)
        try:
            panel._proj_date_start.setDate(QDate(2026, 9, 1))
            panel._proj_date_end.setDate(QDate(2026, 9, 3))
            self.assertEqual(self.db.get_config('project_date_start'), "2026-09-01")
            self.assertEqual(self.db.get_config('project_date_end'), "2026-09-03")
        finally:
            panel.deleteLater()

    def test_date_range_reloads_from_config_on_new_panel(self):
        from hazop import HAZOPPreparationPanel
        self.db.set_config('project_date_start', '2027-01-10')
        self.db.set_config('project_date_end', '2027-01-12')
        panel = HAZOPPreparationPanel(self.db)
        try:
            self.assertEqual(panel._proj_date_start.date().toString('yyyy-MM-dd'), '2027-01-10')
            self.assertEqual(panel._proj_date_end.date().toString('yyyy-MM-dd'), '2027-01-12')
        finally:
            panel.deleteLater()

    def test_legacy_project_date_key_is_included_in_reset_cleanup(self):
        """The old single-value 'project_date' key must not be silently
        orphaned: MainWindow's project-reset cleanup list must still clear
        it (for pre-existing databases) alongside all the new keys."""
        import inspect
        import hazop as hazop_mod
        src = inspect.getsource(hazop_mod.MainWindow)
        for key in ('project_date', 'project_date_start', 'project_date_end',
                    'project_facility', 'project_hazop_leader', 'project_participants'):
            self.assertIn(f"'{key}'", src,
                           f"Project-reset cleanup list should still mention {key!r}")


class SettingsPanelDateWidgetsAndTodayButtonTests(unittest.TestCase):
    """"Inställningarna under projekt ser konstig ut. Datumväljren tar upp
    jättemycket plats. skulle även gilla om knappen today fanns."
    (2026-08-11) — the date-range row's container (date_row_w) used to be
    stretched to the tab's full width by QFormLayout's default
    field-growth policy even though the two QDateEdit widgets inside it
    only need a small fraction of that space; fixed by capping both the
    QDateEdit widths and the container's own size policy. Also verifies
    the newly added "Idag" (Today) buttons."""

    @classmethod
    def setUpClass(cls):
        cls.app = _ensure_qapp()

    def setUp(self):
        self._tmpdir = tempfile.mkdtemp(prefix="hazop_settings_datewidgets_test_")
        self.db_path = os.path.join(self._tmpdir, "test_project.db")
        self.db = Database(path=self.db_path)

    def tearDown(self):
        try:
            del self.db
        except Exception:
            pass
        shutil.rmtree(self._tmpdir, ignore_errors=True)

    def test_date_edits_have_a_real_width_constraint(self):
        from hazop import HAZOPPreparationPanel
        panel = HAZOPPreparationPanel(self.db)
        try:
            for edit in (panel._proj_date_start, panel._proj_date_end):
                max_w = edit.maximumWidth()
                # Qt's "no maximum set" sentinel is 16777215; anything even
                # remotely close to that means the widget was left
                # unconstrained. A real cap should comfortably fit
                # "yyyy-MM-dd" plus the calendar-popup arrow, i.e. well
                # under 250px, and definitely nowhere near the full
                # ~1100px+ a stretched QFormLayout field reaches.
                self.assertLess(max_w, 250,
                                 "QDateEdit should have a compact maximum width")
                self.assertGreater(max_w, 0)
        finally:
            panel.deleteLater()

    def test_date_row_container_does_not_stretch_to_full_tab_width(self):
        """Renders the actual Projekt tab and confirms the date row's
        container widget no longer stretches to the tab's full width the
        way the other QLineEdit form rows still do (by design — this test
        is what would have caught the original bug)."""
        from hazop import HAZOPPreparationPanel
        from PyQt6.QtWidgets import QFormLayout
        panel = HAZOPPreparationPanel(self.db)
        try:
            panel.resize(900, 700)
            panel.show()
            self.app.processEvents()
            tabs = panel._tabs
            proj_idx = next(i for i in range(tabs.count())
                             if tabs.tabText(i) == "Projekt")
            tabs.setCurrentIndex(proj_idx)
            self.app.processEvents()
            proj_tab = tabs.widget(proj_idx)
            # proj_tab's own layout is an outer QVBoxLayout (2026-08-17,
            # holds the form PLUS the revision table and custom-fields box
            # below it) — the QFormLayout lives on the first child widget.
            fl = proj_tab.layout().itemAt(0).widget().layout()
            self.assertIsInstance(fl, QFormLayout)

            name_field_w = panel._proj_name.width()
            date_row_w = None
            for r in range(fl.rowCount()):
                item = fl.itemAt(r, QFormLayout.ItemRole.FieldRole)
                if item and item.widget() is not None and item.widget() not in (
                        panel._proj_name, panel._proj_number, panel._proj_client,
                        panel._proj_facility):
                    date_row_w = item.widget()
                    break
            self.assertIsNotNone(date_row_w, "Could not find the date row's container widget")
            self.assertLess(date_row_w.width(), name_field_w * 0.8,
                             "Date row container should be visibly narrower than a "
                             "full-width text field row, not stretched to match it")
        finally:
            panel.deleteLater()

    def test_today_button_sets_start_date_to_today(self):
        from hazop import HAZOPPreparationPanel
        panel = HAZOPPreparationPanel(self.db)
        try:
            panel._proj_date_start.setDate(QDate(2020, 1, 1))
            panel._proj_date_start_today_btn.click()
            self.assertEqual(panel._proj_date_start.date(), QDate.currentDate())
        finally:
            panel.deleteLater()

    def test_today_button_sets_end_date_to_today(self):
        from hazop import HAZOPPreparationPanel
        panel = HAZOPPreparationPanel(self.db)
        try:
            panel._proj_date_end.setDate(QDate(2020, 1, 1))
            panel._proj_date_end_today_btn.click()
            self.assertEqual(panel._proj_date_end.date(), QDate.currentDate())
        finally:
            panel.deleteLater()

    def test_start_today_button_does_not_touch_end_date(self):
        from hazop import HAZOPPreparationPanel
        panel = HAZOPPreparationPanel(self.db)
        try:
            panel._proj_date_end.setDate(QDate(2020, 1, 1))
            panel._proj_date_start_today_btn.click()
            self.assertEqual(panel._proj_date_end.date(), QDate(2020, 1, 1))
        finally:
            panel.deleteLater()


class ParticipantMatrixTests(unittest.TestCase):
    """"Jag tror det vore bra m du byggde en till flik med deltagare
    istället där man definerar förnamn, efternamn, roll på y axel och
    analystillfälen på x axeln så det blir en matris." (2026-08-11) —
    replaces the old free-text "Deltagare" field/tab-row with a dedicated
    "Deltagare" tab holding a QTableWidget: participants as rows
    (Förnamn/Efternamn/Roll), analysis sessions as columns, attendance as
    a checkbox per cell. Covers both the raw Database CRUD methods and the
    ParticipantMatrixPanel UI wiring."""

    @classmethod
    def setUpClass(cls):
        cls.app = _ensure_qapp()

    def setUp(self):
        self._tmpdir = tempfile.mkdtemp(prefix="hazop_participant_matrix_test_")
        self.db_path = os.path.join(self._tmpdir, "test_project.db")
        self.db = Database(path=self.db_path)

    def tearDown(self):
        try:
            del self.db
        except Exception:
            pass
        shutil.rmtree(self._tmpdir, ignore_errors=True)

    # ── Database layer ───────────────────────────────────────────────────
    def test_participant_crud_round_trips(self):
        pid = self.db.add_participant("Anna", "Andersson", "Processägare")
        rows = self.db.list_participants()
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['first_name'], "Anna")
        self.assertEqual(rows[0]['last_name'], "Andersson")
        self.assertEqual(rows[0]['role'], "Processägare")

        self.db.update_participant(pid, role="HAZOP-ledare")
        rows = self.db.list_participants()
        self.assertEqual(rows[0]['role'], "HAZOP-ledare")
        self.assertEqual(rows[0]['first_name'], "Anna",
                          "update_participant should leave other fields untouched "
                          "when only one keyword is passed")

        self.db.delete_participant(pid)
        self.assertEqual(self.db.list_participants(), [])

    def test_analysis_session_crud_round_trips(self):
        sid = self.db.add_analysis_session("Session 1 (2026-09-01)")
        rows = self.db.list_analysis_sessions()
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['label'], "Session 1 (2026-09-01)")

        self.db.update_analysis_session(sid, "Session 1 (omschemalagd)")
        rows = self.db.list_analysis_sessions()
        self.assertEqual(rows[0]['label'], "Session 1 (omschemalagd)")

        self.db.delete_analysis_session(sid)
        self.assertEqual(self.db.list_analysis_sessions(), [])

    def test_attendance_round_trips_and_toggles(self):
        pid = self.db.add_participant("Bengt", "Bengtsson", "Drift")
        sid = self.db.add_analysis_session("Session 1")

        self.assertFalse(self.db.get_attendance(pid, sid),
                          "Attendance should default to False for an unrecorded pair")

        self.db.set_attendance(pid, sid, True)
        self.assertTrue(self.db.get_attendance(pid, sid))

        # Toggling twice must not create duplicate rows (composite PK / upsert).
        self.db.set_attendance(pid, sid, False)
        self.db.set_attendance(pid, sid, True)
        count = self.db.conn.execute(
            "SELECT COUNT(*) FROM participant_attendance WHERE participant_id=? AND session_id=?",
            (pid, sid)).fetchone()[0]
        self.assertEqual(count, 1)
        self.assertTrue(self.db.get_attendance(pid, sid))

    def test_attendance_matrix_reflects_all_recorded_pairs(self):
        p1 = self.db.add_participant("Anna", "Andersson", "")
        p2 = self.db.add_participant("Bengt", "Bengtsson", "")
        s1 = self.db.add_analysis_session("Session 1")
        s2 = self.db.add_analysis_session("Session 2")
        self.db.set_attendance(p1, s1, True)
        self.db.set_attendance(p1, s2, False)
        self.db.set_attendance(p2, s2, True)

        matrix = self.db.get_attendance_matrix()
        self.assertTrue(matrix.get((p1, s1)))
        self.assertFalse(matrix.get((p1, s2), False))
        self.assertTrue(matrix.get((p2, s2)))
        self.assertNotIn((p2, s1), matrix,
                          "Never-toggled pairs should not have a stored row at all")

    def test_deleting_participant_cascades_attendance(self):
        pid = self.db.add_participant("Anna", "Andersson", "")
        sid = self.db.add_analysis_session("Session 1")
        self.db.set_attendance(pid, sid, True)
        self.db.delete_participant(pid)
        count = self.db.conn.execute(
            "SELECT COUNT(*) FROM participant_attendance WHERE participant_id=?",
            (pid,)).fetchone()[0]
        self.assertEqual(count, 0, "Deleting a participant should cascade-delete attendance rows")

    def test_deleting_session_cascades_attendance(self):
        pid = self.db.add_participant("Anna", "Andersson", "")
        sid = self.db.add_analysis_session("Session 1")
        self.db.set_attendance(pid, sid, True)
        self.db.delete_analysis_session(sid)
        count = self.db.conn.execute(
            "SELECT COUNT(*) FROM participant_attendance WHERE session_id=?",
            (sid,)).fetchone()[0]
        self.assertEqual(count, 0, "Deleting a session should cascade-delete attendance rows")

    # ── UI layer (ParticipantMatrixPanel) ────────────────────────────────
    def test_panel_add_participant_creates_row_and_db_record(self):
        from hazop import ParticipantMatrixPanel
        panel = ParticipantMatrixPanel(self.db)
        try:
            panel._add_participant()
            self.assertEqual(panel._table.rowCount(), 1)
            self.assertEqual(len(self.db.list_participants()), 1)
        finally:
            panel.deleteLater()

    def test_panel_add_session_creates_column(self):
        """No popup (2026-08-18) — adds the column directly with today's
        date as the default label."""
        from hazop import ParticipantMatrixPanel
        panel = ParticipantMatrixPanel(self.db)
        try:
            panel._add_session()
            self.assertEqual(panel._table.columnCount(), len(panel._FIXED_COLS) + 1)
            self.assertEqual(len(self.db.list_analysis_sessions()), 1)
            header_text = panel._table.horizontalHeaderItem(len(panel._FIXED_COLS)).text()
            self.assertRegex(header_text, r'^\d{4}-\d{2}-\d{2}$')
        finally:
            panel.deleteLater()

    def test_panel_add_session_starts_inline_header_edit(self):
        """The new column drops straight into inline header editing so the
        user can adjust the date/label without a popup (2026-08-18)."""
        from hazop import ParticipantMatrixPanel
        panel = ParticipantMatrixPanel(self.db)
        try:
            panel._add_session()
            col = len(panel._FIXED_COLS)
            header = panel._table.horizontalHeader()
            editors = [c for c in header.findChildren(QLineEdit)]
            self.assertEqual(len(editors), 1)
            self.assertFalse(editors[0].isHidden())
        finally:
            panel.deleteLater()

    def test_inline_header_edit_renames_session(self):
        from hazop import ParticipantMatrixPanel
        panel = ParticipantMatrixPanel(self.db)
        try:
            panel._add_session()
            col = len(panel._FIXED_COLS)
            sid = self.db.list_analysis_sessions()[0]['id']
            panel._edit_header_label(col)
            header = panel._table.horizontalHeader()
            editor = header.findChildren(QLineEdit)[0]
            editor.setText("Kickoff 2026-09-01")
            editor.editingFinished.emit()
            self.assertEqual(
                self.db.list_analysis_sessions()[0]['label'], "Kickoff 2026-09-01")
            self.assertEqual(
                panel._table.horizontalHeaderItem(col).text(), "Kickoff 2026-09-01")
        finally:
            panel.deleteLater()

    def test_inline_header_edit_renames_custom_column(self):
        from hazop import ParticipantMatrixPanel
        panel = ParticipantMatrixPanel(self.db)
        try:
            with unittest.mock.patch.object(
                    QInputDialog, 'getText', return_value=("Roll", True)):
                panel._add_column()
            col = len(panel._FIXED_COLS)
            panel._edit_header_label(col)
            header = panel._table.horizontalHeader()
            editor = header.findChildren(QLineEdit)[0]
            editor.setText("Företag")
            editor.editingFinished.emit()
            self.assertEqual(self.db.list_participant_columns()[0]['name'], "Företag")
            self.assertEqual(panel._table.horizontalHeaderItem(col).text(), "Företag")
        finally:
            panel.deleteLater()

    def test_inline_header_edit_escape_cancels_without_saving(self):
        from hazop import ParticipantMatrixPanel
        from PyQt6.QtCore import QEvent
        from PyQt6.QtGui import QKeyEvent
        panel = ParticipantMatrixPanel(self.db)
        try:
            with unittest.mock.patch.object(
                    QInputDialog, 'getText', return_value=("Roll", True)):
                panel._add_column()
            col = len(panel._FIXED_COLS)
            panel._edit_header_label(col)
            header = panel._table.horizontalHeader()
            editor = header.findChildren(QLineEdit)[0]
            editor.setText("Företag")
            ev = QKeyEvent(QEvent.Type.KeyPress, Qt.Key.Key_Escape, Qt.KeyboardModifier.NoModifier)
            editor.keyPressEvent(ev)
            self.assertEqual(self.db.list_participant_columns()[0]['name'], "Roll")
        finally:
            panel.deleteLater()

    def test_header_double_click_on_fixed_column_does_nothing(self):
        from hazop import ParticipantMatrixPanel
        panel = ParticipantMatrixPanel(self.db)
        try:
            panel._edit_header_label(0)
            header = panel._table.horizontalHeader()
            self.assertEqual(len(header.findChildren(QLineEdit)), 0)
        finally:
            panel.deleteLater()

    def test_panel_editing_name_cells_persists_to_db(self):
        from hazop import ParticipantMatrixPanel
        panel = ParticipantMatrixPanel(self.db)
        try:
            panel._add_participant()
            panel._table.item(0, 0).setText("Anna")
            panel._table.item(0, 1).setText("Andersson")
            rows = self.db.list_participants()
            self.assertEqual(rows[0]['first_name'], "Anna")
            self.assertEqual(rows[0]['last_name'], "Andersson")
        finally:
            panel.deleteLater()

    def test_panel_editing_custom_column_cell_persists_to_db(self):
        """Roll is no longer a hardcoded column (2026-08-17) — a free,
        user-named column (e.g. "Roll") between Efternamn and the session
        columns takes its place, see NOTES.md."""
        from hazop import ParticipantMatrixPanel
        panel = ParticipantMatrixPanel(self.db)
        try:
            panel._add_participant()
            self.db.add_participant_column("Roll")
            panel.refresh()
            col = len(panel._FIXED_COLS)
            panel._table.item(0, col).setText("Processägare")
            pid = self.db.list_participants()[0]['id']
            col_id = panel._column_ids[0]
            values = self.db.get_participant_column_values()
            self.assertEqual(values[(pid, col_id)], "Processägare")
        finally:
            panel.deleteLater()

    def test_panel_toggling_attendance_checkbox_persists_to_db(self):
        from hazop import ParticipantMatrixPanel
        panel = ParticipantMatrixPanel(self.db)
        try:
            panel._add_participant()
            panel._add_session()
            pid = self.db.list_participants()[0]['id']
            sid = self.db.list_analysis_sessions()[0]['id']

            cell = panel._table.item(0, len(panel._FIXED_COLS))
            self.assertFalse(cell.data(Qt.ItemDataRole.UserRole + 20))
            self.db.set_attendance(pid, sid, True)
            self.assertTrue(self.db.get_attendance(pid, sid))

            self.db.set_attendance(pid, sid, False)
            self.assertFalse(self.db.get_attendance(pid, sid))
        finally:
            panel.deleteLater()

    def test_panel_delete_participant_removes_row_and_db_record(self):
        from hazop import ParticipantMatrixPanel
        panel = ParticipantMatrixPanel(self.db)
        try:
            panel._add_participant()
            panel._table.setCurrentCell(0, 0)
            panel._delete_participant()
            self.assertEqual(panel._table.rowCount(), 0)
            self.assertEqual(self.db.list_participants(), [])
        finally:
            panel.deleteLater()

    def test_panel_delete_session_removes_column_and_db_record(self):
        from hazop import ParticipantMatrixPanel
        panel = ParticipantMatrixPanel(self.db)
        try:
            panel._add_participant()   # ensures there's a row to select a cell in
            panel._add_session()
            panel._table.setCurrentCell(0, len(panel._FIXED_COLS))
            panel._delete_session()
            self.assertEqual(panel._table.columnCount(), len(panel._FIXED_COLS))
            self.assertEqual(self.db.list_analysis_sessions(), [])
        finally:
            panel.deleteLater()

    def test_add_column_button_creates_column(self):
        from hazop import ParticipantMatrixPanel
        panel = ParticipantMatrixPanel(self.db)
        try:
            with unittest.mock.patch.object(
                    QInputDialog, 'getText', return_value=("E-post", True)):
                panel._add_column()
            self.assertEqual(panel._table.columnCount(), len(panel._FIXED_COLS) + 1)
            self.assertEqual(len(self.db.list_participant_columns()), 1)
            self.assertEqual(
                panel._table.horizontalHeaderItem(len(panel._FIXED_COLS)).text(), "E-post")
        finally:
            panel.deleteLater()

    def test_delete_column_removes_column_and_db_record(self):
        from hazop import ParticipantMatrixPanel
        panel = ParticipantMatrixPanel(self.db)
        try:
            panel._add_participant()
            with unittest.mock.patch.object(
                    QInputDialog, 'getText', return_value=("E-post", True)):
                panel._add_column()
            panel._table.setCurrentCell(0, len(panel._FIXED_COLS))
            panel._delete_column()
            self.assertEqual(panel._table.columnCount(), len(panel._FIXED_COLS))
            self.assertEqual(self.db.list_participant_columns(), [])
        finally:
            panel.deleteLater()

    def test_custom_column_sits_before_session_columns(self):
        """Egna kolumner ska ligga mellan Efternamn och analystillfällena,
        inte efter dem (2026-08-17 user request)."""
        from hazop import ParticipantMatrixPanel
        panel = ParticipantMatrixPanel(self.db)
        try:
            panel._add_session()
            sid = self.db.list_analysis_sessions()[0]['id']
            self.db.update_analysis_session(sid, "2026-09-01")
            with unittest.mock.patch.object(
                    QInputDialog, 'getText', return_value=("E-post", True)):
                panel._add_column()
            panel.refresh()
            headers = [panel._table.horizontalHeaderItem(c).text()
                       for c in range(panel._table.columnCount())]
            self.assertEqual(headers, ["Förnamn", "Efternamn", "E-post", "2026-09-01"])
        finally:
            panel.deleteLater()

    def test_enter_key_on_selected_cell_adds_participant(self):
        """Enter på en enkelklickad (icke-redigerande) rad ska lägga till en
        ny deltagare, precis som "+"-knappen (2026-08-17 user request)."""
        from PyQt6.QtCore import QEvent
        from PyQt6.QtGui import QKeyEvent
        from hazop import ParticipantMatrixPanel
        panel = ParticipantMatrixPanel(self.db)
        try:
            panel._add_participant()
            panel._table.setCurrentCell(0, 0)
            ev = QKeyEvent(QEvent.Type.KeyPress, Qt.Key.Key_Return, Qt.KeyboardModifier.NoModifier)
            panel._table.keyPressEvent(ev)
            self.assertEqual(len(self.db.list_participants()), 2)
        finally:
            panel.deleteLater()

    def test_panel_loads_existing_data_on_construction(self):
        from hazop import ParticipantMatrixPanel
        pid = self.db.add_participant("Anna", "Andersson", "Processägare")
        sid = self.db.add_analysis_session("Session 1")
        self.db.set_attendance(pid, sid, True)
        col_id = self.db.add_participant_column("Roll")
        self.db.set_participant_column_value(pid, col_id, "Processägare")

        panel = ParticipantMatrixPanel(self.db)
        try:
            self.assertEqual(panel._table.rowCount(), 1)
            self.assertEqual(panel._table.item(0, 0).text(), "Anna")
            self.assertEqual(panel._table.item(0, 1).text(), "Andersson")
            self.assertEqual(panel._table.item(0, len(panel._FIXED_COLS)).text(), "Processägare")
            self.assertTrue(panel._table.item(0, len(panel._FIXED_COLS) + 1)
                            .data(Qt.ItemDataRole.UserRole + 20))
        finally:
            panel.deleteLater()

    def test_deltagare_tab_exists_in_settings_panel(self):
        from hazop import HAZOPPreparationPanel
        panel = HAZOPPreparationPanel(self.db)
        try:
            titles = [panel._tabs.tabText(i) for i in range(panel._tabs.count())]
            self.assertIn("Deltagare", titles)
        finally:
            panel.deleteLater()

    def test_old_freetext_participants_field_is_gone(self):
        """"istället" (instead) in the user's request means the new matrix
        REPLACES the old free-text field — it must not still exist
        alongside it."""
        from hazop import HAZOPPreparationPanel
        panel = HAZOPPreparationPanel(self.db)
        try:
            self.assertFalse(hasattr(panel, '_proj_participants'),
                              "Old free-text Deltagare QPlainTextEdit should be removed")
            titles = [panel._tabs.tabText(i) for i in range(panel._tabs.count())]
            proj_idx = titles.index("Projekt")
            proj_tab = panel._tabs.widget(proj_idx)
            from PyQt6.QtWidgets import QFormLayout, QPlainTextEdit
            # proj_tab's own layout is an outer QVBoxLayout (2026-08-17,
            # holds the form PLUS the revision table and custom-fields box
            # below it) — the QFormLayout lives on the first child widget.
            fl = proj_tab.layout().itemAt(0).widget().layout()
            self.assertIsInstance(fl, QFormLayout)
            for r in range(fl.rowCount()):
                item = fl.itemAt(r, QFormLayout.ItemRole.FieldRole)
                if item and item.widget() is not None:
                    self.assertNotIsInstance(item.widget(), QPlainTextEdit,
                                              "Projekt tab should no longer contain the "
                                              "free-text participants QPlainTextEdit")
        finally:
            panel.deleteLater()


class SettingsPanelPidTabRenameAndNewSettingsTests(unittest.TestCase):
    """"Fliken PID borde kunna ändras till något mer generiskt för
    inställning. Detta borde även kunna utvecklas med fler inställningar."
    / "Byt namn + lägg till OCR/sid-inställningar" (2026-08-11) — the old
    "P&ID" tab is renamed "P&ID-inställningar" and gains two new setting
    groups (OCR-standardval, Sid-orientering) while the pre-existing
    Tagg-identifiering checkbox (tag_strip_spaces) keeps working exactly
    as before."""

    @classmethod
    def setUpClass(cls):
        cls.app = _ensure_qapp()

    def setUp(self):
        self._tmpdir = tempfile.mkdtemp(prefix="hazop_settings_pid_test_")
        self.db_path = os.path.join(self._tmpdir, "test_project.db")
        self.db = Database(path=self.db_path)

    def tearDown(self):
        try:
            del self.db
        except Exception:
            pass
        shutil.rmtree(self._tmpdir, ignore_errors=True)

    def test_tab_renamed(self):
        from hazop import SettingsPanel
        panel = SettingsPanel(self.db)
        try:
            titles = [panel._tabs.tabText(i) for i in range(panel._tabs.count())]
            self.assertIn("P&ID-inställningar", titles)
            self.assertNotIn("P&ID", titles)
        finally:
            panel.deleteLater()

    def test_min_pid_line_width_setting_persists_and_can_be_disabled(self):
        from hazop import SettingsPanel
        panel = SettingsPanel(self.db)
        try:
            self.assertTrue(panel._min_pid_lines_chk.isChecked())
            self.assertEqual(panel._min_pid_lines_spin.value(), 2)
            panel._min_pid_lines_spin.setValue(2)
            self.assertEqual(self.db.get_config('pid_min_line_width'), '2')
            panel._min_pid_lines_chk.setChecked(False)
            self.assertEqual(self.db.get_config('pid_min_line_width_enabled'), '0')
            panel._min_pid_lines_chk.setChecked(True)
            self.assertEqual(self.db.get_config('pid_min_line_width_enabled'), '1')
        finally:
            panel.deleteLater()

    def test_tag_strip_spaces_checkbox_still_works(self):
        from hazop import SettingsPanel
        panel = SettingsPanel(self.db)
        try:
            panel._strip_spaces_chk.setChecked(False)
            self.assertEqual(self.db.get_config('tag_strip_spaces'), '0')
            panel._strip_spaces_chk.setChecked(True)
            self.assertEqual(self.db.get_config('tag_strip_spaces'), '1')
        finally:
            panel.deleteLater()

    def test_ocr_default_engine_setting_persists(self):
        from hazop import SettingsPanel
        panel = SettingsPanel(self.db)
        try:
            idx = panel._ocr_default_combo.findData('auto')
            self.assertGreaterEqual(idx, 0, "'Automatiskt' option should always be present")
            panel._ocr_default_combo.setCurrentIndex(idx)
            self.assertEqual(self.db.get_config('ocr_default_engine'), 'auto')
        finally:
            panel.deleteLater()

    def test_ocr_default_engine_reloads_from_config(self):
        from hazop import SettingsPanel
        self.db.set_config('ocr_default_engine', 'auto')
        panel = SettingsPanel(self.db)
        try:
            self.assertEqual(panel._ocr_default_combo.currentData(), 'auto')
        finally:
            panel.deleteLater()

    def test_page_orientation_setting_persists(self):
        from hazop import SettingsPanel
        panel = SettingsPanel(self.db)
        try:
            idx = panel._page_orientation_combo.findData('landscape')
            self.assertGreaterEqual(idx, 0)
            panel._page_orientation_combo.setCurrentIndex(idx)
            self.assertEqual(self.db.get_config('pid_page_orientation_hint'), 'landscape')
        finally:
            panel.deleteLater()

    def test_tag_search_timeout_default_is_two_seconds(self):
        """"Standard ska vara 2 sekunder" (2026-08-18, see NOTES.md
        "kombinerad placeringsmeny")."""
        from hazop import SettingsPanel
        panel = SettingsPanel(self.db)
        try:
            self.assertAlmostEqual(panel._tag_search_timeout_spin.value(), 2.0)
            self.assertEqual(self.db.get_config('equipment_tag_search_timeout_ms', '2000'), '2000')
        finally:
            panel.deleteLater()

    def test_tag_search_timeout_setting_persists_and_reloads(self):
        from hazop import SettingsPanel
        panel = SettingsPanel(self.db)
        try:
            panel._tag_search_timeout_spin.setValue(3.5)
            self.assertEqual(self.db.get_config('equipment_tag_search_timeout_ms'), '3500')
        finally:
            panel.deleteLater()

        panel2 = SettingsPanel(self.db)
        try:
            self.assertAlmostEqual(panel2._tag_search_timeout_spin.value(), 3.5)
        finally:
            panel2.deleteLater()

    def test_ocr_default_engine_skips_prompt_when_configured(self):
        """resolve_ocr_scan_choice() (pid_viewer.py) is the actual wiring
        behind the OCR-standardval setting: with a specific, available
        engine configured, it must return that engine directly without
        showing the Yes/No prompt (QMessageBox.question must not be called)."""
        from pid_viewer import resolve_ocr_scan_choice
        with unittest.mock.patch('equipment_detection.ocr_status',
                                  return_value={'tesseract': True, 'easyocr': False,
                                                'rapidocr': False, 'pil': True}), \
             unittest.mock.patch('pid_viewer.ocr_status',
                                  return_value={'tesseract': True, 'easyocr': False,
                                                'rapidocr': False, 'pil': True}), \
             unittest.mock.patch('pid_viewer.QMessageBox.question') as mock_question:
            self.db.set_config('ocr_default_engine', 'tesseract')
            use_ocr, engine = resolve_ocr_scan_choice(self.db, None)
            self.assertTrue(use_ocr)
            self.assertEqual(engine, 'tesseract')
            mock_question.assert_not_called()

    def test_ask_default_falls_back_to_prompt(self):
        """The default 'ask' setting must preserve the original behaviour
        exactly: still show the Yes/No prompt."""
        from pid_viewer import resolve_ocr_scan_choice
        with unittest.mock.patch('pid_viewer.ocr_status',
                                  return_value={'tesseract': True, 'easyocr': False,
                                                'rapidocr': False, 'pil': True}), \
             unittest.mock.patch('pid_viewer.QMessageBox.question',
                                  return_value=hazop.QMessageBox.StandardButton.Yes) as mock_question:
            self.db.set_config('ocr_default_engine', 'ask')
            use_ocr, engine = resolve_ocr_scan_choice(self.db, None)
            mock_question.assert_called_once()
            self.assertTrue(use_ocr)




if __name__ == "__main__":
    unittest.main()
