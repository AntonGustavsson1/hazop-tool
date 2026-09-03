"""Workbook verification for the revision-safe LOPA Excel export."""

import os
import shutil
import sys
import tempfile
import unittest
from pathlib import Path

_TEST_DIR = Path(__file__).resolve().parent
_HAZOP_DIR = _TEST_DIR.parent
for _path in (_HAZOP_DIR, _TEST_DIR):
    if str(_path) not in sys.path:
        sys.path.insert(0, str(_path))

from database import Database
from lopa_export import export_lopa_excel


class LopaExportTests(unittest.TestCase):
    def setUp(self):
        self._tmpdir = tempfile.mkdtemp(prefix='hazop_lopa_export_')
        self.db = Database(path=os.path.join(self._tmpdir, 'project.db'))

    def tearDown(self):
        self.db.conn.close()
        shutil.rmtree(self._tmpdir, ignore_errors=True)

    def test_export_contains_revision_snapshot_and_lopa_sections(self):
        node_id = self.db.add_node()
        deviation_id = self.db.deviations(node_id)[0]['id']
        cause_id = self.db.add_cause(deviation_id)
        self.db.update_cause(cause_id, description='LT-101 HH', base_frequency=0.1)
        consequence_id = self.db.add_consequence(cause_id)
        self.db.update_consequence(consequence_id, 'Överfyllnad', 3, '')
        category = self.db.consequence_categories()[0]
        self.db.set_consequence_severity(consequence_id, category['id'], 3)
        sensor = self.db.add_safeguard(consequence_id)
        self.db.update_safeguard(sensor, description='LSHH', rrf=100, sg_type='SIS')
        barrier = self.db.add_safeguard(consequence_id)
        self.db.update_safeguard(barrier, description='PSV', rrf=10, sg_type='Mekanisk')
        equipment_id = self.db.add_equipment_item('LT-101', 'LT-101', 'LT', 0, 'Nivågivare', '', 0)
        self.db.add_safeguard_equipment_link(sensor, equipment_id, 'HH')
        created = self.db.create_lopa(display_number='017', sif_number='SIF-017',
                                      sif_name='Överfyllnadsskydd', sis_name='SIS-A')
        imported = self.db.add_lopa_source_from_safeguard(created['lopa_id'], sensor)
        self.db.update_lopa_source_analysis_details(
            imported['source_id'], control_frequency='Årligen', assumption_percent=10)
        final_group = self.db.add_lopa_final_group(created['revision_id'])
        self.db.add_lopa_final_member(
            created['revision_id'], equipment_id=equipment_id, action_text='Stäng ventil',
            group_id=final_group)
        self.db.update_lopa_revision_details(
            created['revision_id'], document_date='2026-09-02',
            additional_actions='Verifiera ändlägesbrytare', process_safety_time='8')
        self.db.add_lopa_comment(created['revision_id'], 'Exportkommentar', 'Anton')

        path = os.path.join(self._tmpdir, 'lopa.xlsx')
        ok, error = export_lopa_excel(self.db, path, [(created['lopa_id'], created['revision_id'])])
        self.assertTrue(ok, error)
        import openpyxl
        workbook = openpyxl.load_workbook(path, data_only=True)
        self.assertEqual(['LOPA sammanfattning', 'LOPA 017 R00'], workbook.sheetnames)
        summary = workbook['LOPA sammanfattning']
        self.assertEqual('LOPA-nr', summary['A3'].value)
        self.assertEqual('017', summary['A4'].value)
        sheet = workbook['LOPA 017 R00']
        values = [cell.value for row in sheet.iter_rows() for cell in row if cell.value]
        for required in ('Skyddsbarriäranalys (LOPA)', 'Källscenarier från HAZOP',
                         'Givardel', 'Manöverdel', 'Oberoende barriärer',
                         'Eskalering och beräkningsspår', 'Kommentarer',
                         'Exportkommentar', 'SIF-017', 'Stäng ventil',
                         'Konsekvens', 'Överfyllnad'):
            self.assertIn(required, values)

    def test_export_accepts_multiple_explicit_revisions(self):
        first = self.db.create_lopa(display_number='001', sif_name='Första SIF')
        second = self.db.create_lopa(display_number='002', sif_name='Andra SIF')
        path = os.path.join(self._tmpdir, 'flera_lopa.xlsx')
        ok, error = export_lopa_excel(
            self.db, path,
            [(first['lopa_id'], first['revision_id']),
             (second['lopa_id'], second['revision_id'])])
        self.assertTrue(ok, error)
        import openpyxl
        workbook = openpyxl.load_workbook(path, data_only=True)
        self.assertEqual(['LOPA sammanfattning', 'LOPA 001 R00', 'LOPA 002 R00'], workbook.sheetnames)
        self.assertEqual('001', workbook['LOPA sammanfattning']['A4'].value)
        self.assertEqual('002', workbook['LOPA sammanfattning']['A5'].value)


if __name__ == '__main__':
    unittest.main()
